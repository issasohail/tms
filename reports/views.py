from invoices.models import InvoiceItem, Invoice
from .utils import date_preset_range
from .forms import ExpenseFilterForm
from invoices.models import ItemCategory
from django.shortcuts import render
from django.db.models.functions import Coalesce
from django.db.models import Sum, F, Value, CharField
from django.shortcuts import render, get_object_or_404
from django.views.generic import ListView, DetailView, CreateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django_tables2 import SingleTableView
from django.utils import timezone
from datetime import timedelta
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.http import HttpResponse
from tenants.models import Tenant
from properties.models import Property
from payments.models import Payment
from invoices.models import Invoice
from expenses.models import Expense
from .models import FinancialReport
from .models import Report
from .tables import FinancialReportTable
from .forms import FinancialReportForm
from properties.models import Property
import csv
from io import StringIO
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum, F, Value, CharField, DecimalField, Q
from django.db.models.functions import Coalesce


class ReportListView(ListView):
    model = Report
    template_name = 'reports/report_list.html'
    context_object_name = 'reports'


class FinancialReportListView(LoginRequiredMixin, SingleTableView):
    model = FinancialReport
    table_class = FinancialReportTable
    template_name = 'reports/financial_report_list.html'
    context_object_name = 'reports'


class FinancialReportDetailView(LoginRequiredMixin, DetailView):
    model = FinancialReport
    template_name = 'reports/financial_report_detail.html'
    context_object_name = 'report'


class FinancialReportCreateView(LoginRequiredMixin, CreateView):
    model = FinancialReport
    form_class = FinancialReportForm
    template_name = 'reports/financial_report_form.html'

    def get_success_url(self):
        return reverse_lazy('financial_report_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, 'Financial report created successfully.')
        return response


def generate_property_report(request, property_id):
    property = get_object_or_404(Property, pk=property_id)

    # Create a file-like buffer to receive PDF data.
    buffer = BytesIO()

    # Create the PDF object, using the buffer as its "file."
    doc = SimpleDocTemplate(buffer, pagesize=letter)

    # Prepare data for the report
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(
        Paragraph(f"Property Report: {property.property_name}", styles['Title']))
    elements.append(Paragraph(
        f"Generated on: {timezone.now().strftime('%Y-%m-%d')}", styles['Normal']))
    elements.append(Paragraph(" ", styles['Normal']))  # Spacer

    # Property Details
    elements.append(Paragraph("Property Details", styles['Heading2']))
    property_data = [
        ["Name:", property.property_name],
        ["Address:", property.address],
        ["Total Units:", property.total_units],
    ]
    property_table = Table(property_data, colWidths=[100, 300])
    property_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(property_table)
    elements.append(Paragraph(" ", styles['Normal']))  # Spacer

    # Financial Summary
    elements.append(Paragraph("Financial Summary", styles['Heading2']))

    # Get date range (last 30 days)
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)

    # Get financial data
    total_rent = Invoice.objects.filter(
        tenant__unit__property=property,
        issue_date__range=[start_date, end_date],
        description__contains='Monthly Rent'
    ).aggregate(total=models.Sum('amount'))['total'] or 0

    total_expenses = Expense.objects.filter(
        property=property,
        date__range=[start_date, end_date]
    ).aggregate(total=models.Sum('amount'))['total'] or 0

    total_payments = Payment.objects.filter(
        tenant__unit__property=property,
        payment_date__range=[start_date, end_date]
    ).aggregate(total=models.Sum('amount'))['total'] or 0

    financial_data = [
        ["Metric", "Amount"],
        ["Total Rent Invoiced", f"${total_rent:,.2f}"],
        ["Total Expenses", f"${total_expenses:,.2f}"],
        ["Total Payments Received", f"${total_payments:,.2f}"],
        ["Net Income", f"${(total_payments - total_expenses):,.2f}"],
    ]

    financial_table = Table(financial_data, colWidths=[200, 200])
    financial_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
    ]))
    elements.append(financial_table)

    # Build the PDF
    doc.build(elements)

    # File response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="property_report_{property.property_name}_{end_date}.pdf"'
    return response


def generate_tenant_statement(request, tenant_id):
    tenant = get_object_or_404(Tenant, pk=tenant_id)

    # Create a file-like buffer to receive PDF data.
    buffer = BytesIO()

    # Create the PDF object, using the buffer as its "file."
    doc = SimpleDocTemplate(buffer, pagesize=letter)

    # Prepare data for the report
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(
        f"Tenant Statement: {tenant.first_name} {tenant.last_name}", styles['Title']))
    elements.append(Paragraph(
        f"Generated on: {timezone.now().strftime('%Y-%m-%d')}", styles['Normal']))
    elements.append(Paragraph(" ", styles['Normal']))  # Spacer

    # Tenant Details
    elements.append(Paragraph("Tenant Details", styles['Heading2']))
    tenant_data = [
        ["Name:", f"{tenant.first_name} {tenant.last_name}"],
        ["Unit:", f"{tenant.unit.property.propertry_name} - Unit {tenant.unit.unit_number}"],
        ["Email:", tenant.email],
        ["Phone:", tenant.phone],
        ["Move-in Date:", tenant.move_in_date.strftime('%Y-%m-%d')],
        ["Lease End Date:", tenant.lease_end_date.strftime('%Y-%m-%d')],
        ["Current Balance:", f"${tenant.current_balance:,.2f}"],
    ]
    tenant_table = Table(tenant_data, colWidths=[100, 300])
    tenant_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(tenant_table)
    elements.append(Paragraph(" ", styles['Normal']))  # Spacer

    # Transaction History
    elements.append(Paragraph("Transaction History", styles['Heading2']))

    # Get invoices and payments
    invoices = Invoice.objects.filter(tenant=tenant).order_by('-issue_date')
    payments = Payment.objects.filter(tenant=tenant).order_by('-payment_date')

    transaction_data = [["Date", "Type", "Description", "Amount", "Balance"]]

    # Combine and sort all transactions by date
    all_transactions = []
    for invoice in invoices:
        all_transactions.append({
            'date': invoice.issue_date,
            'type': 'Invoice',
            'description': invoice.description,
            'amount': -invoice.amount,
            'balance': None  # Will calculate later
        })

    for payment in payments:
        all_transactions.append({
            'date': payment.payment_date,
            'type': 'Payment',
            'description': f"Payment ({payment.get_payment_method_display()})",
            'amount': payment.amount,
            'balance': None  # Will calculate later
        })

    # Sort by date descending
    all_transactions.sort(key=lambda x: x['date'], reverse=True)

    # Calculate running balance
    balance = tenant.current_balance
    for transaction in all_transactions:
        transaction['balance'] = balance
        # For invoices (negative), this adds to balance
        balance -= transaction['amount']

    # Add to table data
    for transaction in all_transactions:
        transaction_data.append([
            transaction['date'].strftime('%Y-%m-%d'),
            transaction['type'],
            transaction['description'],
            f"${transaction['amount']:,.2f}" if transaction[
                'amount'] >= 0 else f"(${-transaction['amount']:,.2f})",
            f"${transaction['balance']:,.2f}",
        ])

    transaction_table = Table(
        transaction_data, colWidths=[80, 60, 180, 80, 80])
    transaction_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ]))
    elements.append(transaction_table)

    # Build the PDF
    doc.build(elements)

    # File response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="tenant_statement_{tenant.last_name}_{tenant.first_name}_{timezone.now().date()}.pdf"'
    return response


class ReportCreateView(LoginRequiredMixin, CreateView):
    model = Report
    fields = ['title', 'content']  # Exclude created_by from form

    def form_valid(self, form):
        form.instance.created_by = self.request.user  # Set the user automatically

        def form_valid(self, form):
            form.instance.created_by = self.request.user  # Set current user
        return super().form_valid(form)


# reports/views.py


def _date_range_from_request(request):
    preset = request.GET.get('preset') or 'this_month'
    rng = date_preset_range(preset)
    if rng:
        return rng
    # custom
    start = request.GET.get('start')
    end = request.GET.get('end')
    return (start or None, end or None)


def expense_summary_report(request):
    form = ExpenseFilterForm(request.GET or None)
    start, end = _date_range_from_request(request)
    qs = Expense.objects.select_related('property', 'category').all()
    if start:
        qs = qs.filter(date__gte=start)
    if end:
        qs = qs.filter(date__lte=end)
    if form.is_valid():
        if form.cleaned_data.get('property'):
            qs = qs.filter(property=form.cleaned_data['property'])
        if form.cleaned_data.get('unit'):
            # match either direct unit or via distributions
            qs = qs.filter(distributions__unit=form.cleaned_data['unit'])
        if form.cleaned_data.get('category'):
            qs = qs.filter(category=form.cleaned_data['category'])

    # group by Property, Category (and optionally Unit label)
    annotated = (
        qs
        .annotate(unit_lbl=Value('', output_field=CharField()))
        .values('property__property_name', 'category__name')
        .annotate(
            total=Coalesce(
                Sum('amount'),
                Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        .order_by('property__property_name', 'category__name')
    )


    return render(request, 'reports/expense_summary.html', {
        'form': form, 'rows': annotated, 'start': start, 'end': end,
    })


def expense_detail_report(request):
    form = ExpenseFilterForm(request.GET or None)
    start, end = _date_range_from_request(request)
    qs = Expense.objects.select_related(
        'property', 'category').prefetch_related('receipts')
    if start:
        qs = qs.filter(date__gte=start)
    if end:
        qs = qs.filter(date__lte=end)
    if form.is_valid():
        if form.cleaned_data.get('property'):
            qs = qs.filter(property=form.cleaned_data['property'])
        if form.cleaned_data.get('unit'):
            qs = qs.filter(distributions__unit=form.cleaned_data['unit'])
        if form.cleaned_data.get('category'):
            qs = qs.filter(category=form.cleaned_data['category'])
    qs = qs.distinct().order_by('property__property_name', 'date', 'pk')
    return render(request, 'reports/expense_detail.html', {'form': form, 'rows': qs, 'start': start, 'end': end})


# reports/views.py (continued)


def profit_and_loss_report(request):
    form = ExpenseFilterForm(request.GET or None)
    start, end = _date_range_from_request(request)

    # Revenue: join InvoiceItem -> Invoice -> Lease -> Unit -> Property
    rev = (InvoiceItem.objects
           .select_related('invoice', 'category', 'invoice__lease')
           .filter(invoice__issue_date__gte=start if start else None)
           .filter(invoice__issue_date__lte=end if end else None))

    # If filters:
    if form.is_valid():
        if form.cleaned_data.get('property'):
            rev = rev.filter(
                invoice__lease__unit__property=form.cleaned_data['property'])
        if form.cleaned_data.get('unit'):
            rev = rev.filter(invoice__lease__unit=form.cleaned_data['unit'])
        if form.cleaned_data.get('category'):
            rev = rev.filter(category=form.cleaned_data['category'])

    rev_g = (
        rev.values('invoice__lease__unit__property__property_name',
                'invoice__lease__unit__unit_number', 'category__name')
        .annotate(
            total_rev=Coalesce(
                Sum('amount'),
                Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
    )


    # Expenses grouped the same way
    exp = Expense.objects.select_related('property', 'category')
    if start:
        exp = exp.filter(date__gte=start)
    if end:
        exp = exp.filter(date__lte=end)
    if form.is_valid():
        if form.cleaned_data.get('property'):
            exp = exp.filter(property=form.cleaned_data['property'])
        if form.cleaned_data.get('unit'):
            exp = exp.filter(distributions__unit=form.cleaned_data['unit'])
        if form.cleaned_data.get('category'):
            exp = exp.filter(category=form.cleaned_data['category'])

    exp_g = (
        exp.values('property__property_name', 'category__name')
        .annotate(
            total_exp=Coalesce(
                Sum('amount'),
                Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
    )
    # Merge to a dict keyed by (property, unit, category)
    from collections import defaultdict
    rows = defaultdict(lambda: {'rev': 0, 'exp': 0})
    for r in rev_g:
        key = (r['invoice__lease__unit__property__property_name'],
               r['invoice__lease__unit__unit_number'] or 'All Unit',
               r['category__name'])
        rows[key]['rev'] = r['total_rev'] or 0
    for e in exp_g:
        # for expenses, unit may be “Multiple” or “All Unit”; summarize at property/category
        key = (e['property__property_name'], 'All Unit', e['category__name'])
        rows[key]['exp'] = e['total_exp'] or 0

    final = []
    for (prop, unit, cat), agg in sorted(rows.items()):
        pl = (agg['rev'] or 0) - (agg['exp'] or 0)
        final.append({'property': prop, 'unit': unit, 'category': cat,
                      'revenue': agg['rev'], 'expense': agg['exp'], 'profit_loss': pl})

    return render(request, 'reports/profit_loss.html', {
        'form': form, 'rows': final, 'start': start, 'end': end
    })
# reports/views.py
from datetime import timedelta
from decimal import Decimal

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.utils import timezone
from django.views.generic import TemplateView

from leases.models import Lease
from properties.models import Property
from invoices.services import security_deposit_totals  # where you put it
from collections import OrderedDict
from django.urls import reverse

import csv
import io

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
except ImportError:
    canvas = None

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    Image = None


class SecurityDepositReportView(LoginRequiredMixin, TemplateView):
    template_name = "reports/security_deposit_report.html"

    # --- helpers ---------------------------------------------------------

    def get_base_queryset(self):
        """
        Base leases to include in the report:
        - have a unit & property
        - have a tenant
        - security_deposit > 0 (you can relax if you want)
        """
        return (
            Lease.objects
            .select_related("tenant", "unit", "unit__property")
            .filter(security_deposit__gt=Decimal("0.00"))
        )

    def filter_queryset(self, qs):
        req = self.request
        prop_id = req.GET.get("property")
        active_only = req.GET.get("active_only") == "on"

        if prop_id:
            qs = qs.filter(unit__property_id=prop_id)

        if active_only:
            # adjust if your Lease.status choices are different
            qs = qs.filter(status="ACTIVE")

        return qs

    def build_rows(self, qs):
        """
        Turn leases into row dicts:
        property_name, unit, tenant, phone, end_date,
        required, paid, refunded, balance, flags, etc.
        Also attach a continuous serial number (sn) and lease_id.
        """
        rows = []
        today = timezone.localdate()
        soon = today + timedelta(days=30)

        for idx, lease in enumerate(
            qs.order_by("unit__property__property_name", "unit__unit_number")
        ):
            totals = security_deposit_totals(lease)

            end_date = lease.end_date
            is_expiring = bool(end_date and end_date <= soon)
            is_expired = bool(end_date and end_date < today)

            rows.append({
                "sn": idx + 1,
                "lease_id": lease.pk,  # 👈 for links
                "property_name": lease.unit.property.property_name,
                "unit_number": lease.unit.unit_number,
                "tenant_name": lease.tenant.get_full_name(),
                "phone": lease.tenant.phone or "",
                "end_date": end_date,
                "is_expiring": is_expiring,
                "is_expired": is_expired,
                "required": totals["required"],
                "paid": totals["paid_in"],
                "refunded": totals["refunded"],
                "balance": totals["balance_to_collect"],
            })

        return rows

    # --- export helpers --------------------------------------------------

    def export_csv(self, rows, property_totals, grand_totals):
        from collections import OrderedDict
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        header = [
            "SN", "Property", "Unit", "Tenant", "Phone",
            "Lease End", "Sec. Required", "Paid", "Refunded", "Held (Balance)",
        ]
        writer.writerow(header)

        # regroup rows by property (same as template)
        grouped = OrderedDict()
        for r in rows:
            grouped.setdefault(r["property_name"], []).append(r)

        for pname, plist in grouped.items():
            # Property header line (just label)
            writer.writerow([f"Property: {pname}"] + [""] * (len(header) - 1))

            for r in plist:
                writer.writerow([
                    r["sn"],
                    r["property_name"],
                    r["unit_number"],
                    r["tenant_name"],
                    r["phone"],
                    r["end_date"].strftime("%Y-%m-%d") if r["end_date"] else "",
                    f"{r['required']:.2f}",
                    f"{r['paid']:.2f}",
                    f"{r['refunded']:.2f}",
                    f"{r['balance']:.2f}",
                ])

            # Property subtotal row
            pt = property_totals[pname]
            writer.writerow([
                "", f"Subtotal {pname}", "", "", "",
                "",
                f"{pt['required']:.2f}",
                f"{pt['paid']:.2f}",
                f"{pt['refunded']:.2f}",
                f"{pt['balance']:.2f}",
            ])
            writer.writerow([])  # blank line between groups

        # Grand total row
        gt = grand_totals
        writer.writerow([
            "", "GRAND TOTAL", "", "", "",
            "",
            f"{gt['required']:.2f}",
            f"{gt['paid']:.2f}",
            f"{gt['refunded']:.2f}",
            f"{gt['balance']:.2f}",
        ])

        resp = HttpResponse(buffer.getvalue(), content_type="text/csv")
        resp["Content-Disposition"] = 'attachment; filename="security_deposits.csv"'
        return resp

    def export_xlsx(self, rows, property_totals, grand_totals):
        if not openpyxl:
            return HttpResponse("openpyxl not installed", status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Security Deposits"

        header = [
            "SN", "Property", "Unit", "Tenant", "Phone",
            "Lease End", "Sec. Required", "Paid", "Refunded", "Held (Balance)",
        ]
        ws.append(header)

        from collections import OrderedDict
        grouped = OrderedDict()
        for r in rows:
            grouped.setdefault(r["property_name"], []).append(r)

        for pname, plist in grouped.items():
            ws.append([f"Property: {pname}"] + [""] * (len(header) - 1))

            for r in plist:
                ws.append([
                    r["sn"],
                    r["property_name"],
                    r["unit_number"],
                    r["tenant_name"],
                    r["phone"],
                    r["end_date"].strftime("%Y-%m-%d") if r["end_date"] else "",
                    float(r["required"]),
                    float(r["paid"]),
                    float(r["refunded"]),
                    float(r["balance"]),
                ])

            pt = property_totals[pname]
            ws.append([
                "", f"Subtotal {pname}", "", "", "",
                "",
                float(pt["required"]),
                float(pt["paid"]),
                float(pt["refunded"]),
                float(pt["balance"]),
            ])
            ws.append([])

        gt = grand_totals
        ws.append([
            "", "GRAND TOTAL", "", "", "",
            "",
            float(gt["required"]),
            float(gt["paid"]),
            float(gt["refunded"]),
            float(gt["balance"]),
        ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        resp = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="security_deposits.xlsx"'
        return resp

    def export_pdf(self, rows, property_totals, grand_totals):
        if not canvas:
            return HttpResponse("reportlab not installed", status=500)

        from collections import OrderedDict

        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=landscape(A4))
        width, height = landscape(A4)

        def fmt_money(x):
            return f"{x:,.0f}"

        def new_page():
            c.showPage()

        y = height - 40
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, y, "Security Deposit Report")
        y -= 20

        headers = [
            "SN", "Property", "Unit", "Tenant", "Phone",
            "Lease End", "Req", "Paid", "Refunded", "Held",
        ]
        col_x = [40, 80, 230, 330, 520, 630, 710, 760, 810, 870]

        c.setFont("Helvetica-Bold", 8)
        for x, h in zip(col_x, headers):
            c.drawString(x, y, h)
        y -= 14
        c.line(40, y, width - 40, y)
        y -= 10

        c.setFont("Helvetica", 7)

        grouped = OrderedDict()
        for r in rows:
            grouped.setdefault(r["property_name"], []).append(r)

        def ensure_space(lines_needed=3):
            nonlocal y
            if y < 40 + lines_needed * 12:
                new_page()
                y_reset_header()

        def y_reset_header():
            nonlocal y
            y = height - 40
            c.setFont("Helvetica-Bold", 14)
            c.drawString(40, y, "Security Deposit Report (cont.)")
            y -= 20
            c.setFont("Helvetica-Bold", 8)
            for x, h in zip(col_x, headers):
                c.drawString(x, y, h)
            y -= 14
            c.line(40, y, width - 40, y)
            y -= 10
            c.setFont("Helvetica", 7)

        # regroup & draw
        for pname, plist in grouped.items():
            ensure_space()
            c.setFont("Helvetica-Bold", 8)
            c.drawString(40, y, f"Property: {pname}")
            y -= 12
            c.setFont("Helvetica", 7)

            for r in plist:
                ensure_space()
                data = [
                    r["sn"],
                    r["property_name"],
                    r["unit_number"],
                    r["tenant_name"],
                    r["phone"],
                    r["end_date"].strftime("%Y-%m-%d") if r["end_date"] else "",
                    fmt_money(r["required"]),
                    fmt_money(r["paid"]),
                    fmt_money(r["refunded"]),
                    fmt_money(r["balance"]),
                ]
                for x, val in zip(col_x, data):
                    c.drawString(x, y, str(val))
                y -= 12

            # subtotal
            pt = property_totals[pname]
            ensure_space()
            c.setFont("Helvetica-Bold", 8)
            c.drawString(40, y, f"Subtotal {pname}")
            c.drawRightString(col_x[6] + 40, y, fmt_money(pt["required"]))
            c.drawRightString(col_x[7] + 40, y, fmt_money(pt["paid"]))
            c.drawRightString(col_x[8] + 40, y, fmt_money(pt["refunded"]))
            c.drawRightString(col_x[9] + 40, y, fmt_money(pt["balance"]))
            y -= 16
            c.setFont("Helvetica", 7)

        # grand total
        gt = grand_totals
        ensure_space()
        c.setFont("Helvetica-Bold", 9)
        c.drawString(40, y, "GRAND TOTAL")
        c.drawRightString(col_x[6] + 40, y, fmt_money(gt["required"]))
        c.drawRightString(col_x[7] + 40, y, fmt_money(gt["paid"]))
        c.drawRightString(col_x[8] + 40, y, fmt_money(gt["refunded"]))
        c.drawRightString(col_x[9] + 40, y, fmt_money(gt["balance"]))

        c.showPage()
        c.save()
        pdf = buf.getvalue()
        buf.close()

        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = 'attachment; filename="security_deposits.pdf"'
        return resp

    def compute_totals(self, rows):
        """
        Returns (property_totals, grand_totals)

        property_totals: OrderedDict keyed by property_name:
          {
            "My Property": {
               "required": Decimal,
               "paid": Decimal,
               "refunded": Decimal,
               "balance": Decimal,
            },
            ...
          }

        grand_totals: same keys but for all rows.
        """
        property_totals = OrderedDict()
        grand = {
            "required": Decimal("0.00"),
            "paid": Decimal("0.00"),
            "refunded": Decimal("0.00"),
            "balance": Decimal("0.00"),
        }

        for r in rows:
            pname = r["property_name"]
            if pname not in property_totals:
                property_totals[pname] = {
                    "required": Decimal("0.00"),
                    "paid": Decimal("0.00"),
                    "refunded": Decimal("0.00"),
                    "balance": Decimal("0.00"),
                }

            pt = property_totals[pname]

            for key in ("required", "paid", "refunded", "balance"):
                val = r[key] or Decimal("0.00")
                pt[key] += val
                grand[key] += val

        return property_totals, grand

    def export_jpg(self, rows, property_totals, grand_totals):
        if not Image:
            return HttpResponse("Pillow not installed", status=500)

        from collections import OrderedDict

        def fmt_money(x):
            return f"{x:,.0f}"

        row_h = 22
        header_h = 30
        margin = 10

        grouped = OrderedDict()
        for r in rows:
            grouped.setdefault(r["property_name"], []).append(r)

        # estimate height: header + for each row + subtotal + grand total
        total_lines = 2  # title + header row
        for pname, plist in grouped.items():
            total_lines += 1  # property title
            total_lines += len(plist)
            total_lines += 1  # subtotal
        total_lines += 1  # grand total

        height = margin * 2 + header_h + row_h * total_lines
        width = 1400

        img = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(img)

        try:
            font_title = ImageFont.truetype("arial.ttf", 18)
            font_hdr = ImageFont.truetype("arial.ttf", 12)
            font_row = ImageFont.truetype("arial.ttf", 11)
        except Exception:
            font_title = font_hdr = font_row = ImageFont.load_default()

        x_positions = [10, 60, 260, 360, 620, 760, 860, 960, 1060, 1160]

        y = margin
        draw.text((x_positions[0], y), "Security Deposit Report", font=font_title, fill="black")
        y += header_h

        headers = ["SN", "Property", "Unit", "Tenant", "Phone",
                   "Lease End", "Req", "Paid", "Refunded", "Held"]
        for x, h in zip(x_positions, headers):
            draw.text((x, y), h, font=font_hdr, fill="black")
        y += row_h

        # rows
        for pname, plist in grouped.items():
            draw.text((x_positions[0], y), f"Property: {pname}", font=font_hdr, fill="black")
            y += row_h

            for r in plist:
                data = [
                    str(r["sn"]),
                    r["property_name"],
                    r["unit_number"],
                    r["tenant_name"],
                    r["phone"],
                    r["end_date"].strftime("%Y-%m-%d") if r["end_date"] else "",
                    fmt_money(r["required"]),
                    fmt_money(r["paid"]),
                    fmt_money(r["refunded"]),
                    fmt_money(r["balance"]),
                ]
                for x, val in zip(x_positions, data):
                    draw.text((x, y), val, font=font_row, fill="black")
                y += row_h

            pt = property_totals[pname]
            line = f"Subtotal {pname}"
            draw.text((x_positions[0], y), line, font=font_hdr, fill="black")
            draw.text((x_positions[6], y), fmt_money(pt["required"]), font=font_hdr, fill="black")
            draw.text((x_positions[7], y), fmt_money(pt["paid"]), font=font_hdr, fill="black")
            draw.text((x_positions[8], y), fmt_money(pt["refunded"]), font=font_hdr, fill="black")
            draw.text((x_positions[9], y), fmt_money(pt["balance"]), font=font_hdr, fill="black")
            y += row_h

        gt = grand_totals
        draw.text((x_positions[0], y), "GRAND TOTAL", font=font_hdr, fill="black")
        draw.text((x_positions[6], y), fmt_money(gt["required"]), font=font_hdr, fill="black")
        draw.text((x_positions[7], y), fmt_money(gt["paid"]), font=font_hdr, fill="black")
        draw.text((x_positions[8], y), fmt_money(gt["refunded"]), font=font_hdr, fill="black")
        draw.text((x_positions[9], y), fmt_money(gt["balance"]), font=font_hdr, fill="black")

        output = io.BytesIO()
        img.save(output, format="JPEG", quality=90)
        output.seek(0)

        resp = HttpResponse(output.read(), content_type="image/jpeg")
        resp["Content-Disposition"] = 'attachment; filename="security_deposits.jpg"'
        return resp

    # --- main GET --------------------------------------------------------

    def get(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_base_queryset())
        rows = self.build_rows(qs)
        property_totals, grand_totals = self.compute_totals(rows)

        export = request.GET.get("export")
        if export == "csv":
            return self.export_csv(rows, property_totals, grand_totals)
        elif export == "xlsx":
            return self.export_xlsx(rows, property_totals, grand_totals)
        elif export == "pdf":
            return self.export_pdf(rows, property_totals, grand_totals)
        elif export == "jpg":
            return self.export_jpg(rows, property_totals, grand_totals)

        context = self.get_context_data(
            rows=rows,
            properties=Property.objects.all().order_by("property_name"),
            current_property=request.GET.get("property", ""),
            active_only=request.GET.get("active_only") == "on",
            property_totals=property_totals,
            grand_totals=grand_totals,
        )
        return self.render_to_response(context)


class MaintenanceCollectionReportView(LoginRequiredMixin, TemplateView):
    template_name = "reports/maintenance_collection_report.html"

    def _filters(self):
        today = timezone.localdate()
        start = self.request.GET.get("start") or today.replace(day=1).isoformat()
        end = self.request.GET.get("end") or today.isoformat()
        return {
            "start": start,
            "end": end,
            "building": self.request.GET.get("building") or "",
            "unit": self.request.GET.get("unit") or "",
            "tenant_status": self.request.GET.get("tenant_status") or "all",
        }

    def build_rows(self):
        filters = self._filters()
        qs = (
            InvoiceItem.objects
            .select_related(
                "invoice", "invoice__lease", "invoice__lease__tenant",
                "invoice__lease__unit", "invoice__lease__unit__property", "category",
            )
            .filter(invoice__issue_date__gte=filters["start"])
            .filter(invoice__issue_date__lte=filters["end"])
            .exclude(invoice__status="cancelled")
            .filter(Q(category__name__icontains="maintenance") | Q(description__icontains="maintenance"))
            .order_by("invoice__lease__unit__property__property_name", "invoice__lease__unit__unit_number", "invoice__issue_date")
        )
        if filters["building"]:
            qs = qs.filter(invoice__lease__unit__property_id=filters["building"])
        if filters["unit"]:
            qs = qs.filter(invoice__lease__unit_id=filters["unit"])
        if filters["tenant_status"] == "active":
            qs = qs.filter(invoice__lease__tenant__is_active=True)
        elif filters["tenant_status"] == "inactive":
            qs = qs.filter(invoice__lease__tenant__is_active=False)

        rows = []
        totals = {
            "charged": Decimal("0.00"),
            "paid": Decimal("0.00"),
            "balance": Decimal("0.00"),
        }
        for item in qs:
            invoice = item.invoice
            lease = invoice.lease
            charged = item.amount or Decimal("0.00")
            paid = charged if invoice.status == "paid" else Decimal("0.00")
            balance = charged - paid
            rows.append({
                "date": invoice.issue_date,
                "invoice": invoice,
                "building": lease.unit.property if lease and lease.unit_id else None,
                "unit": lease.unit if lease else None,
                "tenant": lease.tenant if lease else None,
                "tenant_active": lease.tenant.is_active if lease and lease.tenant_id else False,
                "description": item.description or item.category.name,
                "charged": charged,
                "paid": paid,
                "balance": balance,
                "status": invoice.get_status_display(),
            })
            totals["charged"] += charged
            totals["paid"] += paid
            totals["balance"] += balance
        return rows, totals, filters

    def export_csv(self, rows, totals):
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["Date", "Building", "Unit", "Tenant", "Tenant Active", "Invoice", "Description", "Charged", "Paid", "Balance", "Status"])
        for row in rows:
            writer.writerow([
                row["date"].isoformat(),
                row["building"] or "",
                row["unit"] or "",
                row["tenant"] or "",
                "Yes" if row["tenant_active"] else "No",
                row["invoice"].invoice_number,
                row["description"],
                f"{row['charged']:.2f}",
                f"{row['paid']:.2f}",
                f"{row['balance']:.2f}",
                row["status"],
            ])
        writer.writerow(["", "", "", "", "", "", "TOTAL", f"{totals['charged']:.2f}", f"{totals['paid']:.2f}", f"{totals['balance']:.2f}", ""])
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="maintenance_collection.csv"'
        return response

    def export_xlsx(self, rows, totals):
        if not openpyxl:
            return HttpResponse("openpyxl not installed", status=500)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Maintenance Collection"
        ws.append(["Date", "Building", "Unit", "Tenant", "Tenant Active", "Invoice", "Description", "Charged", "Paid", "Balance", "Status"])
        for row in rows:
            ws.append([
                row["date"].isoformat(),
                str(row["building"] or ""),
                str(row["unit"] or ""),
                str(row["tenant"] or ""),
                "Yes" if row["tenant_active"] else "No",
                row["invoice"].invoice_number,
                row["description"],
                float(row["charged"]),
                float(row["paid"]),
                float(row["balance"]),
                row["status"],
            ])
        ws.append(["", "", "", "", "", "", "TOTAL", float(totals["charged"]), float(totals["paid"]), float(totals["balance"]), ""])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="maintenance_collection.xlsx"'
        return response

    def get(self, request, *args, **kwargs):
        from properties.models import Unit

        rows, totals, filters = self.build_rows()
        if request.GET.get("export") == "csv":
            return self.export_csv(rows, totals)
        if request.GET.get("export") == "xlsx":
            return self.export_xlsx(rows, totals)
        return self.render_to_response(self.get_context_data(
            rows=rows,
            totals=totals,
            filters=filters,
            properties=Property.objects.order_by("property_name"),
            units=Unit.objects.select_related("property").order_by("property__property_name", "unit_number"),
        ))


class SimpleTmsReportView(LoginRequiredMixin, TemplateView):
    template_name = "reports/simple_report.html"
    report_type = ""
    title = ""

    def money(self, value):
        return Decimal(value or 0)

    def collection_rows(self, keyword):
        today = timezone.localdate()
        start = self.request.GET.get("start") or today.replace(day=1).isoformat()
        end = self.request.GET.get("end") or today.isoformat()
        qs = (
            InvoiceItem.objects
            .select_related("invoice", "invoice__lease", "invoice__lease__tenant", "invoice__lease__unit", "invoice__lease__unit__property", "category")
            .filter(invoice__issue_date__gte=start, invoice__issue_date__lte=end)
            .exclude(invoice__status="cancelled")
            .filter(Q(category__name__icontains=keyword) | Q(description__icontains=keyword))
            .order_by("invoice__issue_date")
        )
        rows = []
        totals = {"charged": Decimal("0.00"), "paid": Decimal("0.00"), "balance": Decimal("0.00")}
        for item in qs:
            lease = item.invoice.lease
            charged = self.money(item.amount)
            paid = charged if item.invoice.status == "paid" else Decimal("0.00")
            balance = charged - paid
            rows.append([
                item.invoice.issue_date,
                lease.unit.property.property_name,
                lease.unit.unit_number,
                lease.tenant.get_full_name(),
                item.invoice.invoice_number,
                item.description or item.category.name,
                charged,
                paid,
                balance,
                item.invoice.get_status_display(),
            ])
            totals["charged"] += charged
            totals["paid"] += paid
            totals["balance"] += balance
        return ["Date", "Building", "Unit", "Tenant", "Invoice", "Description", "Charged", "Paid", "Balance", "Status"], rows, totals, {"start": start, "end": end}

    def tenant_balance_rows(self):
        qs = Lease.objects.select_related("tenant", "unit", "unit__property").order_by("unit__property__property_name", "unit__unit_number")
        status = self.request.GET.get("status") or "active"
        if status != "all":
            qs = qs.filter(status=status)
        rows = []
        total = Decimal("0.00")
        for lease in qs:
            balance = self.money(lease.get_balance)
            total += balance
            rows.append([
                lease.unit.property.property_name,
                lease.unit.unit_number,
                lease.tenant.get_full_name(),
                lease.get_status_display(),
                lease.start_date,
                lease.end_date,
                balance,
            ])
        return ["Building", "Unit", "Tenant", "Lease Status", "Start", "End", "Balance"], rows, {"balance": total}, {"status": status}

    def renewal_due_rows(self):
        days = int(self.request.GET.get("days") or 60)
        today = timezone.localdate()
        due = today + timedelta(days=days)
        qs = (
            Lease.objects.select_related("tenant", "unit", "unit__property")
            .filter(status="active", end_date__gte=today, end_date__lte=due)
            .order_by("end_date")
        )
        rows = [[
            lease.end_date,
            lease.unit.property.property_name,
            lease.unit.unit_number,
            lease.tenant.get_full_name(),
            lease.monthly_rent,
            lease.security_deposit,
        ] for lease in qs]
        return ["End Date", "Building", "Unit", "Tenant", "Rent", "Security Deposit"], rows, {}, {"days": days}

    def vacant_unit_rows(self):
        from properties.models import Unit
        qs = Unit.objects.select_related("property").filter(status="vacant").order_by("property__property_name", "unit_number")
        rows = [[
            unit.property.property_name,
            unit.unit_number,
            unit.bedrooms,
            unit.bathrooms,
            unit.monthly_rent,
            unit.society_maintenance,
            unit.water_charges,
        ] for unit in qs]
        return ["Building", "Unit", "Bedrooms", "Bathrooms", "Rent", "Maintenance", "Water"], rows, {}, {}

    def meter_rows(self):
        from smart_meter.models import Meter
        qs = Meter.objects.select_related("unit", "unit__property", "live").order_by("unit__property__property_name", "unit__unit_number")
        rows = []
        for meter in qs:
            live = getattr(meter, "live", None)
            lease = meter.current_lease
            rows.append([
                meter.unit.property.property_name,
                meter.unit.unit_number,
                lease.tenant.get_full_name() if lease else "",
                f"Lease #{lease.id}" if lease else "",
                meter.meter_number,
                meter.get_billing_mode_display(),
                live.total_energy if live else "",
                live.ts if live else "",
                live.balance if live else "",
                meter.get_power_status_display(),
            ])
        return ["Building", "Unit", "Tenant", "Active Lease", "Meter", "Mode", "Latest Reading", "Reading Date", "Balance", "Power"], rows, {}, {}

    def build_report(self):
        if self.report_type == "rent_collection":
            return self.collection_rows("rent")
        if self.report_type == "tenant_balance":
            return self.tenant_balance_rows()
        if self.report_type == "renewal_due":
            return self.renewal_due_rows()
        if self.report_type == "vacant_units":
            return self.vacant_unit_rows()
        if self.report_type == "late_fees":
            return self.collection_rows("late fee")
        if self.report_type == "meter":
            return self.meter_rows()
        return [], [], {}, {}

    def export_csv(self, headers, rows):
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        writer.writerows(rows)
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{self.report_type}.csv"'
        return response

    def export_xlsx(self, headers, rows):
        if not openpyxl:
            return HttpResponse("openpyxl not installed", status=500)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.title[:31]
        ws.append(headers)
        for row in rows:
            ws.append(list(row))
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{self.report_type}.xlsx"'
        return response

    def get(self, request, *args, **kwargs):
        headers, rows, totals, filters = self.build_report()
        if request.GET.get("export") == "csv":
            return self.export_csv(headers, rows)
        if request.GET.get("export") == "xlsx":
            return self.export_xlsx(headers, rows)
        return self.render_to_response(self.get_context_data(
            title=self.title,
            report_type=self.report_type,
            headers=headers,
            rows=rows,
            totals=totals,
            filters=filters,
        ))


class RentCollectionReportView(SimpleTmsReportView):
    report_type = "rent_collection"
    title = "Rent Collection Report"


class TenantBalanceReportView(SimpleTmsReportView):
    report_type = "tenant_balance"
    title = "Tenant Balance Report"


class LeaseRenewalDueReportView(SimpleTmsReportView):
    report_type = "renewal_due"
    title = "Lease Renewal Due Report"


class VacantUnitReportView(SimpleTmsReportView):
    report_type = "vacant_units"
    title = "Vacant Unit Report"


class LateFeeReportView(SimpleTmsReportView):
    report_type = "late_fees"
    title = "Late Fee Report"


class MeterReportView(SimpleTmsReportView):
    report_type = "meter"
    title = "Meter Report"
