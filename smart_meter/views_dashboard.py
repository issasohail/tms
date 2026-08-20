from __future__ import annotations
from openpyxl.styles import Alignment, Font
from .models import Lease
from .models import Lease
from django.conf import settings
from urllib.parse import urlencode
import traceback
from invoices.models import Invoice, InvoiceItem
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import permission_required, login_required
import calendar
from .models import Lease
from django.http import JsonResponse, HttpResponse
from properties.models import Property
from invoices.models import Invoice, InvoiceItem, ItemCategory  # invoices app
from .models import Lease  # leases.Lease
from django.urls import reverse, NoReverseMatch
from django.http import JsonResponse
from django.http import HttpResponse, JsonResponse, Http404
from io import BytesIO
from django.views.decorators.http import require_GET
from django.http import JsonResponse, Http404
import re
from collections import defaultdict
# adjust app label if different
from invoices.models import Invoice, InvoiceItem, ItemCategory
from properties.models import Unit
from django.shortcuts import render
from django.db.models import Sum, Q, OuterRef, Subquery, F, Window
from datetime import date
from leases.models import Lease
# adjust app label if needed
from invoices.models import Invoice, InvoiceItem, ItemCategory
from smart_meter.models import Meter, MeterReading
from django.db.models import Sum
from calendar import monthrange
from datetime import date, datetime
from django.urls import reverse
from django.utils.http import urlencode
from decimal import Decimal
from weasyprint import HTML
from collections import OrderedDict, defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal, ROUND_HALF_UP
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.decorators import permission_required

from django.db.models import Min, Max
from django.db.models.functions import RowNumber, TruncDay, TruncHour, TruncMonth
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from properties.models import Property, Unit
from smart_meter.models import Meter, MeterReading, LiveReading, MeterBalance, MeterInstallation
from smart_meter.status import online_threshold_minutes
from smart_meter.utils.display import attach_active_meter_counts, display_labels_for_units
import json
from django.utils.safestring import mark_safe

KW_RE = re.compile(r"total\s*usage\s*[:=]\s*([0-9]+(?:\.[0-9]+)?)", re.I)
try:
    from leases.models import Lease, LeaseUnitOccupancy
except Exception:
    Lease = None
    LeaseUnitOccupancy = None
try:
    from .models import Category  # adjust if Category lives elsewhere
except Exception:
    Category = None

def _billing_currency():
    from core.currency import currency_symbol
    from core.models import GlobalSettings

    return currency_symbol(GlobalSettings.get_solo())
def _aware_range(start_d: date, end_d: date):
    tz = timezone.get_current_timezone()
    start_naive = datetime.combine(start_d, time.min)
    end_naive = datetime.combine(end_d, time.max)
    start_dt = timezone.make_aware(start_naive, tz)
    end_dt = timezone.make_aware(end_naive, tz) + timedelta(microseconds=1)
    return start_dt, end_dt, tz


def _period_key_and_label(ts_local: datetime, granularity: str):
    if granularity == "hourly":
        key = ts_local.replace(minute=0, second=0, microsecond=0)
        label = key.strftime("%b %d %H:00")
    elif granularity == "monthly":
        key = date(ts_local.year, ts_local.month, 1)
        label = key.strftime("%b %Y")
    else:  # daily
        key = ts_local.date()
        label = key.strftime("%b %d, %Y")
    return key, label


def _reading_period_expression(granularity: str, tz):
    """
    Return the DB-side period bucket matching ``_period_key_and_label``.

    The report range still filters directly on ``ts`` so MySQL can use the
    existing (meter_id, ts) index. Only the much smaller grouped result is
    returned to Python.
    """
    if granularity == "hourly":
        return TruncHour("ts", tzinfo=tz)
    if granularity == "monthly":
        return TruncMonth("ts", tzinfo=tz)
    return TruncDay("ts", tzinfo=tz)


def _boundary_readings_by_meter(
    meter_ids,
    start_dt: datetime,
    end_dt: datetime,
    tz,
    granularity: str,
):
    """
    Return only the first and last MeterReading needed for each meter/period.

    Historically the dashboard fetched *every* reading in the selected window
    and then kept only the first/last value for each period in Python. On a
    15-minute history table that can transfer tens of thousands of rows for a
    monthly report.

    This implementation preserves the same first/last semantics while using:
      1. one GROUP BY query to identify first_ts/last_ts per meter + period;
      2. one indexed lookup for the corresponding reading values.

    Duplicate timestamps are resolved deterministically: the lowest id is used
    for a period's first reading and the highest id for its last reading.
    """
    readings_by_meter = defaultdict(list)
    if not meter_ids:
        return readings_by_meter

    period_expr = _reading_period_expression(granularity, tz)
    boundaries = list(
        MeterReading.objects
        .filter(meter_id__in=meter_ids, ts__gte=start_dt, ts__lt=end_dt)
        .annotate(_report_period=period_expr)
        .values("meter_id", "_report_period")
        .annotate(first_ts=Min("ts"), last_ts=Max("ts"))
        .order_by("meter_id", "_report_period")
    )
    if not boundaries:
        return readings_by_meter

    candidate_timestamps = {
        ts
        for boundary in boundaries
        for ts in (boundary["first_ts"], boundary["last_ts"])
        if ts is not None
    }
    if not candidate_timestamps:
        return readings_by_meter

    candidates_by_key = defaultdict(list)
    candidate_rows = (
        MeterReading.objects
        .filter(
            meter_id__in=meter_ids,
            ts__in=candidate_timestamps,
        )
        .order_by("meter_id", "ts", "id")
        .values("id", "meter_id", "ts", "total_energy")
    )
    for row in candidate_rows:
        candidates_by_key[(row["meter_id"], row["ts"])].append(row)

    for boundary in boundaries:
        meter_id = boundary["meter_id"]
        first_ts = boundary["first_ts"]
        last_ts = boundary["last_ts"]

        first_candidates = candidates_by_key.get((meter_id, first_ts), ())
        last_candidates = candidates_by_key.get((meter_id, last_ts), ())
        if not first_candidates or not last_candidates:
            continue

        first_row = first_candidates[0]
        last_row = last_candidates[-1]

        readings_by_meter[meter_id].append(first_row)
        if last_row["id"] != first_row["id"]:
            readings_by_meter[meter_id].append(last_row)

    return readings_by_meter


# smart_meter/views_dashboard.py


def _chain_rows_from_groups(groups, prev_end: Decimal | None = None):
    """
    Build rows so that each row's start_kwh equals the previous row's end_kwh.
    If prev_end is None (no earlier reading), we fall back to the group's min.
    """
    rows = []
    last_end = prev_end  # carry-forward baseline

    for _, g in groups.items():
        gmin = Decimal(g["min"])
        gmax = Decimal(g["max"])

        start_kwh = gmin if last_end is None else last_end
        end_kwh = gmax

        usage = end_kwh - start_kwh
        if usage < 0:
            # handle meter reset/replacement gracefully
            usage = Decimal("0")

        rows.append({
            "period_label": g["label"],
            "start_kwh": start_kwh,
            "end_kwh": end_kwh,
            "usage": usage,
        })
        last_end = end_kwh

    return rows


def _fmt0(n: Decimal | float | int) -> str:
    """Round half up to 0 decimals and format with comma."""
    d = Decimal(str(n)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    return f"{int(d):,}"


def _tenant_display_name(tenant):
    if not tenant:
        return ""
    return tenant.get_full_name() if hasattr(tenant, "get_full_name") else str(tenant)


def _tenant_names_for_units(unit_ids, start_d: date, end_d: date):
    tenant_names = {}
    if not Lease or not unit_ids:
        return tenant_names

    try:
        overlap_qs = (
            Lease.objects
            .filter(unit_id__in=unit_ids, start_date__lte=end_d, end_date__gte=start_d)
            .select_related("tenant")
            .order_by("unit_id", "-start_date")
        )
        for lease in overlap_qs:
            if lease.unit_id in tenant_names:
                continue
            tenant_name = _tenant_display_name(getattr(lease, "tenant", None))
            if tenant_name:
                tenant_names[lease.unit_id] = tenant_name

        missing_unit_ids = [unit_id for unit_id in unit_ids if unit_id not in tenant_names]
        if missing_unit_ids:
            today = timezone.localdate()
            active_qs = (
                Lease.objects
                .filter(unit_id__in=missing_unit_ids, start_date__lte=today, end_date__gte=today)
                .select_related("tenant")
                .order_by("unit_id", "-start_date")
            )
            for lease in active_qs:
                if lease.unit_id in tenant_names:
                    continue
                tenant_name = _tenant_display_name(getattr(lease, "tenant", None))
                if tenant_name:
                    tenant_names[lease.unit_id] = tenant_name
    except Exception:
        return {}

    return tenant_names


def _tenant_names_for_unit_dates(unit_dates):
    tenant_names = {}
    if not Lease or not unit_dates:
        return tenant_names

    all_dates = [
        used_date
        for dates in unit_dates.values()
        for used_date in dates
        if used_date
    ]
    if not all_dates:
        return tenant_names

    unit_ids = [unit_id for unit_id in unit_dates.keys() if unit_id]
    min_date = min(all_dates)
    max_date = max(all_dates)

    occupancies = []
    if LeaseUnitOccupancy:
        occupancies = list(
            LeaseUnitOccupancy.objects.filter(
                unit_id__in=unit_ids,
                move_in_date__lte=max_date,
            )
            .filter(Q(move_out_date__isnull=True) | Q(move_out_date__gte=min_date))
            .select_related("lease", "lease__tenant")
            .order_by("unit_id", "-move_in_date", "-id")
        )

    leases = list(
        Lease.objects.filter(
            unit_id__in=unit_ids,
            start_date__lte=max_date,
        )
        .filter(Q(end_date__isnull=True) | Q(end_date__gte=min_date))
        .select_related("tenant")
        .order_by("unit_id", "-start_date", "-id")
    )

    for unit_id, dates in unit_dates.items():
        for used_date in dates:
            name = "Vacant"
            for occupancy in occupancies:
                if occupancy.unit_id != unit_id:
                    continue
                if occupancy.move_in_date <= used_date and (
                    occupancy.move_out_date is None or occupancy.move_out_date >= used_date
                ):
                    name = _tenant_display_name(occupancy.lease.tenant) or "Vacant"
                    break
            if name == "Vacant":
                for lease in leases:
                    if lease.unit_id != unit_id:
                        continue
                    if lease.start_date <= used_date and (
                        lease.end_date is None or lease.end_date >= used_date
                    ):
                        name = _tenant_display_name(lease.tenant) or "Vacant"
                        break
            tenant_names[(unit_id, used_date)] = name

    return tenant_names


def _per_meter_series(meters_qs, start_d: date, end_d: date, granularity: str):
    """
    Returns:
      labels_asc, datasets, combined_rows_asc, totals_dict
    """
    start_dt, end_dt, tz = _aware_range(start_d, end_d)

    meters = attach_active_meter_counts(meters_qs)
    meter_ids = [m.id for m in meters]
    meter_to_rows = {}
    key_union = set()

    readings_by_meter = _boundary_readings_by_meter(
        meter_ids,
        start_dt,
        end_dt,
        tz,
        granularity,
    )

    prev_end_by_meter = {}
    if meter_ids:
        # Fetch the latest timestamp before the report window per meter using
        # GROUP BY + MAX(ts). This avoids ranking/sorting the entire historical
        # MeterReading set, which is expensive on MySQL for long-running meters.
        previous_points = list(
            MeterReading.objects
            .filter(meter_id__in=meter_ids, ts__lt=start_dt)
            .order_by()
            .values("meter_id")
            .annotate(last_ts=Max("ts"))
        )
        latest_ts_by_meter = {
            row["meter_id"]: row["last_ts"]
            for row in previous_points
            if row.get("last_ts") is not None
        }
        if latest_ts_by_meter:
            # Restrict the second query to the small set of candidate timestamps.
            # Duplicate timestamps are resolved deterministically by highest id.
            candidate_rows = (
                MeterReading.objects
                .filter(
                    meter_id__in=latest_ts_by_meter.keys(),
                    ts__in=set(latest_ts_by_meter.values()),
                )
                .order_by("meter_id", "ts", "id")
                .values("id", "meter_id", "ts", "total_energy")
            )
            latest_row_by_meter = {}
            for row in candidate_rows:
                meter_id = row["meter_id"]
                if row["ts"] != latest_ts_by_meter.get(meter_id):
                    continue
                current = latest_row_by_meter.get(meter_id)
                if current is None or row["id"] > current["id"]:
                    latest_row_by_meter[meter_id] = row
            for meter_id, row in latest_row_by_meter.items():
                if row["total_energy"] is not None:
                    prev_end_by_meter[meter_id] = Decimal(str(row["total_energy"]))

    for m in meters:
        prev_end = prev_end_by_meter.get(m.id)

        groups = OrderedDict()
        for r in readings_by_meter.get(m.id, []):
            ts_local = timezone.localtime(r["ts"], tz)
            key, label = _period_key_and_label(ts_local, granularity)
            val = Decimal(str(r["total_energy"] or "0"))
            if key not in groups:
                groups[key] = {"min": val, "max": val, "label": label}
            else:
                groups[key]["max"] = val

        base_rows = _chain_rows_from_groups(groups, prev_end=prev_end)

        rows_with_key = []
        for k, g in groups.items():
            row = next(
                (x for x in base_rows if x["period_label"] == g["label"]), None)
            if row:
                r2 = dict(row)
                r2["period_key"] = k
                rows_with_key.append(r2)

        meter_to_rows[m] = rows_with_key
        for r in rows_with_key:
            key_union.add(r["period_key"])

    if not key_union:
        return [], [], [], {
            "total_kwh": Decimal("0"),
            "usage_charges": Decimal("0"),
            "service_charges": Decimal("0"),
            "grand_total": Decimal("0"),
        }

    # Oldest → Latest
    keys_sorted = sorted(key_union)

    def _date_from_key(k):
        if isinstance(k, datetime):
            return k.date()
        if isinstance(k, date):
            return k
        return start_d

    unit_dates = defaultdict(set)
    for m in meters:
        if not m.unit_id:
            continue
        for k in keys_sorted:
            unit_dates[m.unit_id].add(_date_from_key(k))
    tenant_names = _tenant_names_for_unit_dates(unit_dates)

    def _label_from_key(k):
        if isinstance(k, datetime):
            return k.strftime("%b %d %H:00")
        elif isinstance(k, date):
            if k.day == 1:
                return k.strftime("%b %Y")
            return k.strftime("%b %d, %Y")
        return str(k)

    labels_sorted = [_label_from_key(k) for k in keys_sorted]

    datasets = []
    combined_rows = []
    total_kwh = Decimal("0")
    usage_charges = Decimal("0")
    service_total = Decimal("0")

    for m in meters:
        unit_number = m.display_location_name
        legend_label = f"{unit_number} / {m.meter_number}"
        key_to_row = {r["period_key"]: r for r in meter_to_rows.get(m, [])}
        series = []
        for k in keys_sorted:
            row = key_to_row.get(k)
            series.append(float(row["usage"]) if row else None)

        datasets.append({
            "label": legend_label,
            "data": series,
            "meterNumber": m.meter_number,
            "unitNumber": unit_number,
            "propertyName": getattr(getattr(m.unit, "property", None), "property_name", ""),
            "meterRole": m.meter_role,
            "meterRoleDisplay": m.get_meter_role_display(),
        })

        rate = Decimal(str(m.effective_unit_rate or "0"))
        svc = Decimal(str(m.service_charges or "0"))

        for k in keys_sorted:
            row = key_to_row.get(k)
            if not row:
                continue
            tenant_name = tenant_names.get((m.unit_id, _date_from_key(k)), "Vacant")
            usage_amt = (row["usage"] * rate).quantize(Decimal("0.01"))
            service_amt = svc if granularity == "monthly" else Decimal("0.00")
            total_amt = (usage_amt + service_amt).quantize(Decimal("0.01"))

            combined_rows.append({
                "meter_id": m.id,
                "meter_number": m.meter_number,
                "meter_role": m.meter_role,
                "meter_role_display": m.get_meter_role_display(),
                "unit_number": unit_number,
                "unit_id": m.unit_id,
                "property_name": getattr(getattr(m.unit, "property", None), "property_name", ""),
                "tenant_name": tenant_name,
                "period_label": _label_from_key(k),
                "period_key": k,
                "start_kwh": row["start_kwh"],
                "end_kwh": row["end_kwh"],
                "usage": row["usage"],
                "unit_rate": rate,
                "usage_amount": usage_amt,
                "service_charges": service_amt,
                "total_amount": total_amt,
            })
            total_kwh += row["usage"]
            usage_charges += usage_amt
            service_total += service_amt

    # Keep Billing and Audit meters grouped consistently.
    combined_rows.sort(key=lambda x: (
        0 if x["meter_role"] == Meter.METER_ROLE_BILLING else 1,
        x["period_key"], x["property_name"], x["unit_number"], x["meter_number"]))

    totals = {
        "total_kwh": total_kwh,
        "usage_charges": usage_charges,
        "service_charges": service_total,
        "grand_total": (usage_charges + service_total).quantize(Decimal("0.01")),
    }
    return labels_sorted, datasets, combined_rows, totals


def _export_filename(request, base: str, ext: str) -> str:
    meter_id = (request.GET.get("meter") or "").strip()
    start = (request.GET.get("start") or "").strip()
    end = (request.GET.get("end") or "").strip()

    if meter_id:
        try:
            mn = Meter.objects.values_list(
                "meter_number", flat=True).get(id=meter_id)
        except Meter.DoesNotExist:
            mn = "unknown"
        return f"meter-data-{mn}-{start or 'start'}_to_{end or 'end'}.{ext}"
    return f"All-meters-{start or 'start'}_to_{end or 'end'}.{ext}"


def _export_rows(request):
    today = timezone.localdate()
    start_date = date(today.year, today.month, 1)
    end_date = today
    if request.GET.get("start") and request.GET.get("end"):
        try:
            start_date = date.fromisoformat(request.GET["start"])
            end_date = date.fromisoformat(request.GET["end"])
        except Exception:
            pass

    report_type = (request.GET.get("report_type") or "daily").strip().lower()
    if report_type not in ("hourly", "daily", "monthly"):
        report_type = "daily"

    prop_id = (request.GET.get("property") or "").strip()
    unit_id = (request.GET.get("unit") or "").strip()
    meter_id = (request.GET.get("meter") or "").strip()
    meter_role = (request.GET.get("role") or "").strip().lower()

    units_qs = Unit.objects.all()
    if prop_id:
        units_qs = units_qs.filter(property_id=prop_id)

    meters_qs = Meter.objects.select_related("unit", "unit__property")
    if unit_id:
        meters_qs = meters_qs.filter(unit_id=unit_id)
    elif prop_id:
        meters_qs = meters_qs.filter(unit__property_id=prop_id)

    selected_meters = meters_qs.order_by(
        "meter_role", "unit__property__property_name", "unit__unit_number", "meter_number"
    )
    if meter_id:
        selected_meters = selected_meters.filter(id=meter_id)

    valid_roles = {Meter.METER_ROLE_BILLING, Meter.METER_ROLE_CHECK, "all"}
    if meter_role not in valid_roles:
        include_check = (request.GET.get("include_check") or "").lower() in ("1", "true", "yes", "on")
        selected_role = selected_meters.values_list("meter_role", flat=True).first() if meter_id else None
        meter_role = "all" if include_check else selected_role or Meter.METER_ROLE_BILLING
    if meter_role != "all":
        selected_meters = selected_meters.filter(meter_role=meter_role)

    billing_meters = selected_meters.filter(meter_role=Meter.METER_ROLE_BILLING)
    if meter_role == "all":
        labels, datasets, rows, _display_totals = _per_meter_series(
            selected_meters, start_date, end_date, report_type)
        _billing_labels, _billing_datasets, _billing_rows, totals = _per_meter_series(
            billing_meters, start_date, end_date, report_type)
    else:
        labels, datasets, rows, totals = _per_meter_series(
            selected_meters, start_date, end_date, report_type)
    return report_type, rows, totals, start_date, end_date, prop_id, meter_id


def _user_whatsapp(request):
    # Adjust these attribute names to your auth/profile model
    u = getattr(request, 'user', None)
    if not u:
        return ''
    for attr in ("whatsapp", "whatsapp_number", "phone_number", "mobile", "profile.whatsapp", "profile.phone_number"):
        try:
            parts = attr.split('.')
            obj = u
            for p in parts:
                obj = getattr(obj, p)
            if obj:
                return str(obj)
        except Exception:
            continue
    return ''


@login_required
def energy_dashboard(request):
    wa = ""
    if request.user.is_authenticated:
        wa = getattr(request.user, "whatsapp_number", "") or ""

    today = timezone.localdate()
    start_date = today
    end_date = today

    if request.GET.get("start") and request.GET.get("end"):
        try:
            start_date = date.fromisoformat(request.GET["start"])
            end_date = date.fromisoformat(request.GET["end"])
        except Exception:
            pass

    report_type = (request.GET.get("report_type") or "daily").strip().lower()
    if report_type not in ("hourly", "daily", "monthly"):
        report_type = "daily"

    prop_id = (request.GET.get("property") or "").strip()
    unit_id = (request.GET.get("unit") or "").strip()
    meter_id = (request.GET.get("meter") or "").strip()
    meter_role = (request.GET.get("role") or "").strip().lower()

    all_properties = Property.objects.only("id", "property_name").order_by("property_name")

    units_qs = Unit.objects.only("id", "property_id", "unit_number")
    if prop_id:
        units_qs = units_qs.filter(property_id=prop_id)
    filtered_units = units_qs.order_by("unit_number")

    meters_qs = Meter.objects.select_related("unit", "unit__property")
    if unit_id:
        meters_qs = meters_qs.filter(unit_id=unit_id)
    elif prop_id:
        meters_qs = meters_qs.filter(unit__property_id=prop_id)
    filtered_meters = meters_qs.order_by(
        "meter_role", "unit__property__property_name", "unit__unit_number", "meter_number"
    )

    valid_roles = {Meter.METER_ROLE_BILLING, Meter.METER_ROLE_CHECK, "all"}
    if meter_role not in valid_roles:
        include_check = (request.GET.get("include_check") or "").lower() in ("1", "true", "yes", "on")
        selected_role = filtered_meters.filter(id=meter_id).values_list("meter_role", flat=True).first() if meter_id else None
        meter_role = "all" if include_check else selected_role or Meter.METER_ROLE_BILLING
    if meter_role != "all":
        filtered_meters = filtered_meters.filter(meter_role=meter_role)

    selected_meters = filtered_meters
    if meter_id:
        selected_meters = selected_meters.filter(id=meter_id)

    per_meter_mode = bool(meter_id)

    selected_meter = None
    live_panel = {}
    if per_meter_mode:
        selected_meter = get_object_or_404(Meter, id=meter_id)
        attach_active_meter_counts([selected_meter])
        lr = LiveReading.objects.filter(meter=selected_meter).first()
        if lr:
            last_ts = lr.ts
            voltage_a = lr.voltage_a
            current_a = lr.current_a
            total_energy = lr.total_energy
        else:
            snap = MeterReading.objects.filter(
                meter=selected_meter).order_by("-ts").first()
            last_ts = snap.ts if snap else None
            voltage_a = getattr(snap, "voltage_a", None) if snap else None
            current_a = getattr(snap, "current_a", None) if snap else None
            total_energy = getattr(snap, "total_energy",
                                   None) if snap else None

        cutoff_dt = timezone.now() - timedelta(minutes=online_threshold_minutes())
        is_online = bool(lr and lr.ts >= cutoff_dt)
        bal_obj = MeterBalance.objects.filter(unit=selected_meter.unit).first()

        live_panel = {
            "unit_name": selected_meter.display_location_name,
            "meter_number": selected_meter.meter_number,
            "is_online": is_online,
            "last_ts": last_ts,
            "voltage_a": voltage_a,
            "current_a": current_a,
            "total_energy": total_energy,
            "unit_rate": selected_meter.effective_unit_rate,
            "balance": getattr(bal_obj, "balance", None),
        }

    if meter_role == "all":
        # Build the selected-meter series once. Billing-only totals are derived
        # from those already-built rows instead of repeating all reading queries.
        labels, datasets, table_rows, _display_totals = _per_meter_series(
            selected_meters, start_date, end_date, report_type
        )
        billing_rows = [
            row for row in table_rows
            if row["meter_role"] == Meter.METER_ROLE_BILLING
        ]
        total_kwh = sum((row["usage"] for row in billing_rows), Decimal("0"))
        usage_charges = sum(
            (row["usage_amount"] for row in billing_rows), Decimal("0")
        )
        service_charges = sum(
            (row["service_charges"] for row in billing_rows), Decimal("0")
        )
        totals = {
            "total_kwh": total_kwh,
            "usage_charges": usage_charges,
            "service_charges": service_charges,
            "grand_total": (usage_charges + service_charges).quantize(
                Decimal("0.01")
            ),
        }
        display_meters = selected_meters
    else:
        labels, datasets, table_rows, totals = _per_meter_series(
            selected_meters, start_date, end_date, report_type
        )
        display_meters = selected_meters

    online_count = offline_count = 0
    online_set = set()
    if not per_meter_mode:
        cutoff = timezone.now() - timedelta(minutes=online_threshold_minutes())
        # get latest live-reading timestamps for all selected meters
        live_map = {lr.meter_id: lr.ts for lr in LiveReading.objects.filter(
            meter__in=display_meters)}
        for mid in display_meters.values_list("id", flat=True):
            ts = live_map.get(mid)
            if ts and ts >= cutoff:
                online_count += 1
                online_set.add(mid)
            else:
                offline_count += 1

    # Monthly heading for all meters
    month_heading = ""
    if report_type == "monthly":
        if start_date.year == end_date.year and start_date.month == end_date.month:
            month_heading = start_date.strftime("%B %Y")
        else:
            month_heading = f"{start_date.strftime('%b %Y')} – {end_date.strftime('%b %Y')}"

    # Property name for heading when filtering by property
    property_name = None
    if prop_id:
        try:
            property_name = Property.objects.values_list(
                "property_name", flat=True).get(id=prop_id)
        except Property.DoesNotExist:
            property_name = None

    # Build display-safe rows with rounding (0 decimals) and comma formatting
    wa = _user_whatsapp(request)
    rows_disp = []
    for r in table_rows:
        rows_disp.append({
            **r,
            "whatsapp": wa,
            "is_online": (r["meter_id"] in online_set) if not per_meter_mode else None,
            "start_kwh": (r["start_kwh"]),
            "end_kwh": (r["end_kwh"]),
            "usage": (r["usage"]),
            "unit_rate": _fmt0(r["unit_rate"]),
            "usage_amount": _fmt0(r["usage_amount"]),
            "service_charges": _fmt0(r["service_charges"]),
            "total_amount": _fmt0(r["total_amount"]),
        })

    totals_disp = {
        "total_kwh": (totals["total_kwh"]),
        "usage_charges": _fmt0(totals["usage_charges"]),
        "service_charges": _fmt0(totals["service_charges"]),
        "grand_total": _fmt0(totals["grand_total"]),
    }

    context = {
        "all_properties": all_properties,
        "filtered_units": filtered_units,
        "filtered_meters": filtered_meters,
        "meters": filtered_meters,
        "current_property": prop_id,
        "current_unit": unit_id,
        "current_meter": meter_id,
        "current_role": meter_role,
        "meter_role_choices": Meter.METER_ROLE_CHOICES,
        "report_type": report_type,
        "start_date": start_date,
        "end_date": end_date,
        "online_minutes": online_threshold_minutes(),

        "labels": labels,
        "datasets": datasets,
        "rows": table_rows,
        "rows_disp": rows_disp,
        "totals": totals,
        "totals_disp": totals_disp,

        "selected_meter": selected_meter,
        "live_panel": live_panel,
        "online_count": online_count,
        "offline_count": offline_count,
        "currency": _billing_currency(),
        "month_heading": month_heading,
        "property_name": property_name,
    }
    return render(request, "smart_meter/dashboard.html", context)


# ---------- Exports ----------
@permission_required("accounts.can_export_energy", raise_exception=True)
def energy_export_csv(request):
    report_type, rows, totals, start_date, end_date, prop_id, meter_id = _export_rows(
        request)
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{_export_filename(request, "energy", "csv")}"'

    import csv
    w = csv.writer(resp)
    header = [
        "S/N", "Meter #", "Unit", "Property", "Tenant", "WhatsApp", "Period",
        "Begin (kWh)", "End (kWh)", "Usage (kWh)", "Rate (Rs/kWh)", "Usage Charges (Rs)"
    ]
    if report_type == "monthly":
        header += ["Service Charges (Rs)", "Total (Rs)"]
    w.writerow(header)

    wa = _user_whatsapp(request)
    for i, r in enumerate(rows, start=1):
        base = [
            i,
            r["meter_number"],
            r["unit_number"],
            r["property_name"],
            r.get("tenant_name", ""),
            wa,
            r["period_label"],
            _fmt0(r["start_kwh"]),
            _fmt0(r["end_kwh"]),
            _fmt0(r["usage"]),
            _fmt0(r["unit_rate"]),
            _fmt0(r["usage_amount"]),
        ]
        if report_type == "monthly":
            base += [_fmt0(r["service_charges"]), _fmt0(r["total_amount"])]
        w.writerow(base)
    return resp


@permission_required("accounts.can_export_energy", raise_exception=True)
def energy_export_xlsx(request):
    report_type, rows, totals, start_date, end_date, prop_id, meter_id = _export_rows(
        request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Energy"

    header = [
        "S/N", "Meter #", "Unit", "Property", "Tenant", "WhatsApp", "Period",
        "Begin (kWh)", "End (kWh)", "Usage (kWh)", "Rate (Rs/kWh)", "Usage Charges (Rs)"
    ]
    if report_type == "monthly":
        header += ["Service Charges (Rs)", "Total (Rs)"]
    ws.append(header)

    wa = _user_whatsapp(request)
    for i, r in enumerate(rows, start=1):
        base = [
            i,
            r["meter_number"],
            r["unit_number"],
            r["property_name"],
            r.get("tenant_name", ""),
            wa,
            r["period_label"],
            _fmt0(r["start_kwh"]),
            _fmt0(r["end_kwh"]),
            _fmt0(r["usage"]),
            _fmt0(r["unit_rate"]),
            _fmt0(r["usage_amount"]),
        ]
        if report_type == "monthly":
            base += [_fmt0(r["service_charges"]), _fmt0(r["total_amount"])]
        ws.append(base)

    for col in ws.columns:
        try:
            max_len = max(len(str(c.value))
                          if c.value is not None else 0 for c in col)
        except ValueError:
            max_len = 10
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 32)

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{_export_filename(request, "energy", "xlsx")}"'
    return resp


@permission_required("accounts.can_export_energy", raise_exception=True)
def energy_export_pdf(request):
    # Build the same dataset used by the dashboard, but locally for PDF
    report_type, rows, totals, start_date, end_date, prop_id, meter_id = _export_rows(
        request)

    # Heading bits
    month_heading = ""
    if report_type == "monthly":
        if start_date.year == end_date.year and start_date.month == end_date.month:
            month_heading = start_date.strftime("%B %Y").upper()
        else:
            month_heading = f"{start_date.strftime('%b %Y').upper()} – {end_date.strftime('%b %Y').upper()}"

    property_name = None
    if prop_id:
        property_name = (
            Property.objects.filter(id=prop_id)
            .values_list("property_name", flat=True)
            .first()
        )

    # Display-ready rows (0 decimals + comma formatting) + WhatsApp
    wa = _user_whatsapp(request)
    rows_disp = [
        {
            **r,
            "whatsapp": wa,
            "start_kwh": _fmt0(r["start_kwh"]),
            "end_kwh": _fmt0(r["end_kwh"]),
            "usage": _fmt0(r["usage"]),
            "unit_rate": _fmt0(r["unit_rate"]),
            "usage_amount": _fmt0(r["usage_amount"]),
            "service_charges": _fmt0(r["service_charges"]),
            "total_amount": _fmt0(r["total_amount"]),
        }
        for r in rows
    ]

    totals_disp = {
        "total_kwh": _fmt0(totals["total_kwh"]),
        "usage_charges": _fmt0(totals["usage_charges"]),
        "service_charges": _fmt0(totals["service_charges"]),
        "grand_total": _fmt0(totals["grand_total"]),
    }

    context = {
        "rows_disp": rows_disp,
        "totals_disp": totals_disp,
        "currency": _billing_currency(),
        "report_type": report_type,
        "start_date": start_date,
        "end_date": end_date,
        "property_name": property_name,
        "month_heading": month_heading,
        "current_meter": meter_id,  # template checks this in a heading condition
    }

    html = render_to_string(
        "smart_meter/dashboard_pdf.html", context, request=request)
    pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf()

    # Reuse the standard filename helper
    fname = _export_filename(request, "energy", "pdf")
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


def energy_chart_page(request):
    # Reuse the same data as dashboard
    # You already compute everything in energy_dashboard; replicate the core by calling helpers
    today = timezone.localdate()
    start_date = date(today.year, today.month, 1)
    end_date = today
    try:
        if request.GET.get("start"):
            start_date = date.fromisoformat(request.GET["start"])
        if request.GET.get("end"):
            end_date = date.fromisoformat(request.GET["end"])
    except Exception:
        pass

    report_type = (request.GET.get("report_type") or "daily").strip().lower()
    if report_type not in ("hourly", "daily", "monthly"):
        report_type = "daily"

    prop_id = (request.GET.get("property") or "").strip()
    unit_id = (request.GET.get("unit") or "").strip()
    meter_id = (request.GET.get("meter") or "").strip()

    all_properties = Property.objects.all().order_by("property_name")
    units_qs = Unit.objects.all()
    if prop_id:
        units_qs = units_qs.filter(property_id=prop_id)
    filtered_units = units_qs.order_by("unit_number")

    meters_qs = Meter.objects.select_related("unit", "unit__property")
    if unit_id:
        meters_qs = meters_qs.filter(unit_id=unit_id)
    elif prop_id:
        meters_qs = meters_qs.filter(unit__property_id=prop_id)
    filtered_meters = meters_qs.order_by("meter_number")

    selected_meters = filtered_meters
    if meter_id:
        selected_meters = selected_meters.filter(id=meter_id)

    labels, datasets, table_rows, totals = _per_meter_series(
        selected_meters, start_date, end_date, report_type
    )

    # Optional monthly heading + property name (same as dashboard)
    month_heading = ""
    if report_type == "monthly":
        if start_date.year == end_date.year and start_date.month == end_date.month:
            month_heading = start_date.strftime("%B %Y")
        else:
            month_heading = f"{start_date.strftime('%b %Y')} – {end_date.strftime('%b %Y')}"

    property_name = None
    if prop_id:
        try:
            property_name = Property.objects.values_list(
                "property_name", flat=True).get(id=prop_id)
        except Property.DoesNotExist:
            property_name = None

    # Display-friendly rows/totals
    wa = _user_whatsapp(request)
    rows_disp = [{
        **r,
        "whatsapp": wa,
        "start_kwh": _fmt0(r["start_kwh"]),
        "end_kwh": _fmt0(r["end_kwh"]),
        "usage": _fmt0(r["usage"]),
        "unit_rate": _fmt0(r["unit_rate"]),
        "usage_amount": _fmt0(r["usage_amount"]),
        "service_charges": _fmt0(r["service_charges"]),
        "total_amount": _fmt0(r["total_amount"]),
    } for r in table_rows]

    totals_disp = {
        "total_kwh": _fmt0(totals["total_kwh"]),
        "usage_charges": _fmt0(totals["usage_charges"]),
        "service_charges": _fmt0(totals["service_charges"]),
        "grand_total": _fmt0(totals["grand_total"]),
    }

    # Back link with same filters
    back_qs = request.GET.urlencode()
    back_url = f"{reverse('smart_meter:energy_dashboard')}?{back_qs}" if back_qs else reverse(
        'smart_meter:energy_dashboard')

    ctx = {
        "all_properties": all_properties,
        "filtered_units": filtered_units,
        "filtered_meters": filtered_meters,
        "meters": filtered_meters,
        "current_property": prop_id,
        "current_unit": unit_id,
        "current_meter": meter_id,
        "report_type": report_type,
        "start_date": start_date,
        "end_date": end_date,

        "labels": labels,
        "datasets": datasets,
        "rows": table_rows,
        "rows_disp": rows_disp,
        "totals": totals,
        "totals_disp": totals_disp,
        "currency": _billing_currency(),
        "month_heading": month_heading,
        "property_name": property_name,

        "back_url": back_url,
    }
    return render(request, "smart_meter/chart_only.html", ctx)

# views_dashboard.py




def _month_bounds(yyyy_mm: str | None):
    """Return (start_date, end_date) for selected month; defaults to current month."""
    today = timezone.localdate()
    if not yyyy_mm:
        y, m = today.year, today.month
    else:
        y, m = map(int, yyyy_mm.split("-"))
    start = date(y, m, 1)
    end = date(y, m, monthrange(y, m)[1])
    return start, end


def _prev_month_bounds(any_day: date):
    m = any_day.month - 1 or 12
    y = any_day.year - 1 if any_day.month == 1 else any_day.year
    start = date(y, m, 1)
    end = date(y, m, monthrange(y, m)[1])
    return start, end


# smart_meter/views_dashboard.py


BILLING_SUMMARY_DEFAULT_COLUMNS = [
    "Unit",
    "Tenant",
    "Month",
    "Rent",
    "Society Maintenance",
    "Water",
    "Internet",
    "Electric (prev)",
    "Other",
    "Total",
]


def _first_day(d: date) -> date:
    return date(d.year, d.month, 1)


def _last_day(d: date) -> date:
    return date(d.year, d.month, monthrange(d.year, d.month)[1])


def _prev_month(d: date) -> date:
    y = d.year
    m = d.month - 1
    if m == 0:
        m = 12
        y -= 1
    return date(y, m, 1)


def _cat_lookup():
    """
    Build a case-insensitive lookup of ItemCategory names -> id.
    """
    by_name = {}
    for c in ItemCategory.objects.filter(is_active=True):
        by_name[c.name.strip().lower()] = c.id
    return by_name


def _money(x):
    return Decimal(x or 0).quantize(Decimal("0.01"))

# views_dashboard.py


# PDF (WeasyPrint already used elsewhere in this file)
try:
    from weasyprint import HTML
except Exception:
    HTML = None  # graceful fallback

from smart_meter.models import Bill  # electric bills (prev month)  # noqa



def _ym_from_qs(request):
    """Read ?month=YYYY-MM (defaults to current month)."""
    today = timezone.localdate()
    y = today.year
    m = today.month
    mstr = (request.GET.get("month") or "").strip()
    if mstr:
        try:
            y, m = [int(x) for x in mstr.split("-")]
        except Exception:
            pass
    return y, m


def _bounds_for_month(y, m):
    start = timezone.datetime(y, m, 1).date()
    end = timezone.datetime(y, m, monthrange(y, m)[1]).date()
    return start, end


def _prev_year_month(y, m):
    return (y - 1, 12) if m == 1 else (y, m - 1)


def _zero():
    return Decimal("0.00")


def _fmt0(n: Decimal | float | int) -> str:
    """Round half-up to 0 decimals and format with comma (for HTML/PDF only)."""
    d = Decimal(str(n or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return f"{int(d):,}"


def _sum_items(invoice_qs, categories: list[str]) -> Decimal:
    """Sum invoice items in given categories for invoices queryset."""
    if not categories:
        return _zero()
    cats = [c.strip().lower() for c in categories]
    qs = (InvoiceItem.objects
          .filter(invoice__in=invoice_qs, category__name__in=categories)
          .aggregate(t=Sum("amount")))
    return qs["t"] or _zero()


def _sum_items_excluding(invoice_qs, excluded: list[str]) -> Decimal:
    exc = excluded or []
    qs = (InvoiceItem.objects
          .filter(invoice__in=invoice_qs)
          .exclude(category__name__in=exc)
          .aggregate(t=Sum("amount")))
    return qs["t"] or _zero()


def _lease_balance_upto(lease: Lease, upto_date):
    """Invoices - Payments up to a date."""
    inv = (lease.invoices
           .filter(issue_date__lte=upto_date)
           .aggregate(t=Sum("amount"))["t"] or _zero())

    # Try a general payments relation first; if a Payments app exists, we’ll use it
    pay = _zero()
    try:
        pay = (lease.payments.filter(date__lte=upto_date)
               .aggregate(t=Sum("amount"))["t"] or _zero())
    except Exception:
        try:
            from payments.models import Payment as LeasePayment
            pay = (LeasePayment.objects.filter(lease=lease, date__lte=upto_date)
                   .aggregate(t=Sum("amount"))["t"] or _zero())
        except Exception:
            pass
    return inv - pay


# smart_meter/views_dashboard.py


try:
    from weasyprint import HTML
except Exception:
    HTML = None

# adjust imports to your app structure

# views_dashboard.py


ZERO = Decimal("0.00")
DEFAULT_CAT_IDS = [1, 2, 12, 4, 7]  # Rent, Society, Water, Internet, Electric


def _first_last_of_month(y: int, m: int):
    first = date(y, m, 1)
    last = date(y, m, calendar.monthrange(y, m)[1])
    return first, last


def _ym_from_qs(request):
    mstr = (request.GET.get("month") or "").strip()
    if mstr and re.match(r"^\d{4}-\d{2}$", mstr):
        y, m = map(int, mstr.split("-"))
    else:
        today = timezone.localdate()
        y, m = today.year, today.month
    return y, m


def _parse_cols_param(request):
    extras = []
    tokens = (request.GET.get("cols") or "").strip()
    if tokens:
        for t in tokens.split(","):
            t = t.strip()
            if not t:
                continue
            try:
                v = int(t)
                if v not in extras and v not in DEFAULT_CAT_IDS:
                    extras.append(v)
            except Exception:
                pass
    return extras


def _tenant_display(tenant):
    if not tenant:
        return "Vacant", ""
    if hasattr(tenant, "get_full_name"):
        full_name = (tenant.get_full_name() or "").strip()
    else:
        first_name = (getattr(tenant, "first_name", "") or "").strip()
        last_name = (getattr(tenant, "last_name", "") or "").strip()
        full_name = (first_name + " " + last_name).strip()
    full_name = full_name or getattr(tenant, "name", "") or str(tenant)
    phone = (
        getattr(tenant, "phone", None)
        or getattr(tenant, "phone2", None)
        or getattr(tenant, "phone3", None)
        or ""
    )
    return full_name, phone


def _billing_month_units_and_leases(start, end):
    """
    Return one row source per unit/tenant record that overlaps the selected month.
    A unit with two leases in the month gets two records; a unit with no lease
    gets one Vacant record.
    """
    month_overlap = Q(start_date__lte=end) & (Q(end_date__isnull=True) | Q(end_date__gte=start))
    occupancy_overlap = Q(move_in_date__lte=end) & (Q(move_out_date__isnull=True) | Q(move_out_date__gte=start))

    unit_ids = set(
        MeterInstallation.objects.filter(
            start_date__lte=end
        ).filter(
            Q(end_date__isnull=True) | Q(end_date__gte=start)
        ).values_list("unit_id", flat=True)
    )

    unit_ids.update(
        Lease.objects.filter(month_overlap).values_list("unit_id", flat=True)
    )

    occupancies = list(
        LeaseUnitOccupancy.objects.filter(occupancy_overlap)
        .select_related("lease", "lease__tenant", "unit", "unit__property")
        .order_by("unit__property__property_name", "unit__unit_number", "move_in_date", "id")
    )
    unit_ids.update(o.unit_id for o in occupancies)

    units = list(
        Unit.objects.filter(id__in=unit_ids)
        .select_related("property")
        .order_by("property__property_name", "unit_number", "id")
    )

    rows_by_unit = defaultdict(list)
    occupancy_lease_ids = set()
    for occ in occupancies:
        rows_by_unit[occ.unit_id].append({"unit": occ.unit, "lease": occ.lease})
        occupancy_lease_ids.add(occ.lease_id)

    leases = (
        Lease.objects.filter(month_overlap)
        .exclude(id__in=occupancy_lease_ids)
        .select_related("tenant", "unit", "unit__property")
        .order_by("unit__property__property_name", "unit__unit_number", "start_date", "id")
    )
    for lease in leases:
        rows_by_unit[lease.unit_id].append({"unit": lease.unit, "lease": lease})

    row_sources = []
    for unit in units:
        unit_rows = rows_by_unit.get(unit.id)
        if unit_rows:
            row_sources.extend(unit_rows)
        else:
            row_sources.append({"unit": unit, "lease": None})

    return row_sources


def build_billing_summary_context(request):
    y, m = _ym_from_qs(request)
    start, end = _first_last_of_month(y, m)

    extra_ids = _parse_cols_param(request)
    active_only = request.GET.get("active_only") == "1"
    active_only_param = "&active_only=1" if active_only else ""

    month_full = date(y, m, 1).strftime("%B %Y")
    generated_at = timezone.localtime(
        timezone.now()).strftime("%d %b %Y, %I:%M %p")

    # Load categories once; reused for header labels and the picker modal.
    all_categories = list(ItemCategory.objects.order_by("name").values("id", "name"))
    cat_map = {c["id"]: c["name"] for c in all_categories}

    # Invoices in month
    inv_qs = (
        Invoice.objects
        .filter(issue_date__gte=start, issue_date__lte=end)
        .select_related("lease", "lease__tenant", "lease__unit", "lease__unit__property")
        .order_by("lease__unit__property__property_name", "lease__unit__unit_number", "id")
    )
    invoices = list(
        inv_qs.only(
            "id", "lease_id", "issue_date",
            "lease__id", "lease__tenant__id",
            "lease__unit__id", "lease__unit__unit_number",
            "lease__unit__property__id", "lease__unit__property__property_name",
        )
    )
    inv_ids = [invoice.id for invoice in invoices]

    items_by_inv = defaultdict(list)
    if inv_ids:
        for it in (
            InvoiceItem.objects
            .filter(invoice_id__in=inv_ids)
            .only("invoice_id", "category_id", "amount")
        ):
            items_by_inv[it.invoice_id].append(it)

    inv_by_lease = defaultdict(list)
    for inv in invoices:
        inv_by_lease[inv.lease_id].append(inv)

    # Rows are based on the selected billing month, not the current lease status.
    # This keeps historical tenant names correct and shows Vacant when no lease
    # overlaps the month.
    row_sources = _billing_month_units_and_leases(start, end)
    unit_display_labels = display_labels_for_units(
        source["unit"] for source in row_sources
    )
    row_lease_ids = sorted({
        source["lease"].id
        for source in row_sources
        if source.get("lease")
    })

    balance_by_lease = defaultdict(Decimal)
    if row_lease_ids:
        from django.db.models import Case, DecimalField, F, When
        from django.db.models.functions import Coalesce
        from payments.models import Payment

        invoice_totals = {
            row["lease_id"]: Decimal(row["total"] or 0)
            for row in (
                Invoice.objects
                .filter(lease_id__in=row_lease_ids)
                .values("lease_id")
                .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
            )
        }

        payment_totals = {
            row["lease_id"]: Decimal(row["total"] or 0)
            for row in (
                Payment.objects
                .filter(lease_id__in=row_lease_ids)
                .values("lease_id")
                .annotate(
                    total=Coalesce(
                        Sum(
                            Case(
                                When(
                                    detail__isnull=False,
                                    then=F("detail__lease_amount"),
                                ),
                                default=F("amount"),
                                output_field=DecimalField(max_digits=12, decimal_places=2),
                            )
                        ),
                        Decimal("0.00"),
                    )
                )
            )
        }

        for lease_id in row_lease_ids:
            balance_by_lease[lease_id] = invoice_totals.get(lease_id, ZERO) - payment_totals.get(lease_id, ZERO)

    # visible cats = defaults + extras (no dups)
    visible_cat_ids = []
    for cid in DEFAULT_CAT_IDS + extra_ids:
        if cid not in visible_cat_ids:
            visible_cat_ids.append(cid)

    # Header columns (order matters):
    #   [cats (defaults first, then extras)], Other, Total Invoice, Total Balance, WhatsApp
    header_cols = []
    for cid in visible_cat_ids:
        header_cols.append({
            "kind": "cat",
            "key": f"cat:{cid}",
            "cat_id": cid,
            "label": cat_map.get(cid, f"Cat {cid}"),
        })
    header_cols.append({"kind": "other", "key": "other", "label": "Other"})
    header_cols.append(
        {"kind": "total", "key": "total", "label": "Total\nInvoice"})
    header_cols.append(
        {"kind": "balance", "key": "total_balance", "label": "Total\nBalance"})
    header_cols.append(
        {"kind": "whatsapp", "key": "whatsapp", "label": "WhatsApp\nReminder"})

    groups = defaultdict(list)
    grand = defaultdict(Decimal)

    for source in row_sources:
        lease = source["lease"]
        unit = source["unit"]
        prop = unit.property

        # Bucket amounts by category ID
        bucket = defaultdict(Decimal)
        row_total = ZERO
        if lease:
            for inv in inv_by_lease.get(lease.id, []):
                for it in items_by_inv.get(inv.id, []):
                    amt = Decimal(it.amount or 0)
                    bucket[it.category_id] += amt
                    row_total += amt

        vis_values = {cid: bucket.get(cid, ZERO) for cid in visible_cat_ids}

        other_sum = ZERO
        for cid, amt in bucket.items():
            if cid not in vis_values:
                other_sum += amt

        total_balance = balance_by_lease.get(lease.id, ZERO) if lease else ZERO
        total_balance = Decimal(total_balance or 0)

        t = lease.tenant if lease else None
        full_name, phone = _tenant_display(t)

        cells = []
        for col in header_cols:
            k = col["kind"]
            if k == "cat":
                v = vis_values.get(col["cat_id"], ZERO)
                cells.append(
                    {"key": col["key"], "kind": "cat", "cat_id": col["cat_id"], "value": v})
            elif k == "other":
                cells.append(
                    {"key": col["key"], "kind": "other", "value": other_sum})
            elif k == "total":
                cells.append(
                    {"key": col["key"], "kind": "total", "value": row_total})
            elif k == "balance":
                cells.append(
                    {"key": col["key"], "kind": "balance", "value": total_balance})
            else:  # whatsapp
                cells.append({
                    "key": col["key"], "kind": "whatsapp", "value": "",
                    "phone": phone, "tenant_name": full_name,
                    "property_name": prop.property_name, "unit_number": unit.unit_number,
                    "balance": total_balance,
                })

        # accumulate grand numeric
        for cell in cells:
            if cell["kind"] in ("cat", "other", "total", "balance"):
                grand[cell["key"]] = grand.get(
                    cell["key"], ZERO) + (cell["value"] or ZERO)

        groups[prop.id].append({
            "sn": None,
            "lease": lease,
            "property": prop,
            "unit": unit,
            "unit_display_label": unit_display_labels.get(unit.id, unit.unit_number),
            "tenant_name": full_name,
            "tenant_phone": phone,
            "cells": cells,
        })

    # Order properties; add subtotals aligned to header_cols
    props = Property.objects.filter(
        id__in=groups.keys()).order_by("property_name")

    grouped = []
    sn = 1
    for prop in props:
        rows = groups[prop.id]
        subtotal_map = defaultdict(Decimal)
        for r in rows:
            for cell in r["cells"]:
                if cell["kind"] in ("cat", "other", "total", "balance"):
                    subtotal_map[cell["key"]] += (cell["value"] or ZERO)
        for r in rows:
            r["sn"] = sn
            sn += 1
        subtotal_row = []
        for col in header_cols:
            if col["kind"] == "whatsapp":
                subtotal_row.append("")
            else:
                subtotal_row.append(subtotal_map.get(col["key"], ZERO))
        # ### NEW: build clickable subtotal_cells matching header_cols
        subtotal_cells = []
        for col in header_cols:
            if col["kind"] == "whatsapp":
                subtotal_cells.append({"kind": "whatsapp", "value": ""})
            elif col["kind"] == "cat":
                subtotal_cells.append(
                    {"kind": "cat", "cat_id": col["cat_id"], "value": subtotal_map.get(col["key"], ZERO)})
            else:
                # other / total / balance
                subtotal_cells.append(
                    {"kind": col["kind"], "value": subtotal_map.get(col["key"], ZERO)})

        grouped.append({
            "property": prop,
            "rows": rows,
            "subtotal_row": subtotal_row,
            "subtotal_cells": subtotal_cells,   # ### NEW
        })

    # Grand aligned row
    grand_row = []
    for col in header_cols:
        if col["kind"] == "whatsapp":
            grand_row.append("")
        else:
            grand_row.append(grand.get(col["key"], ZERO))

     # ### NEW: clickable grand_cells
    grand_cells = []
    for col in header_cols:
        if col["kind"] == "whatsapp":
            grand_cells.append({"kind": "whatsapp", "value": ""})
        elif col["kind"] == "cat":
            grand_cells.append(
                {"kind": "cat", "cat_id": col["cat_id"], "value": grand.get(col["key"], ZERO)})
        else:
            grand_cells.append(
                {"kind": col["kind"], "value": grand.get(col["key"], ZERO)})

    ctx = {
        "grand_row": grand_row,
        "grand_cells": grand_cells,
        "year": y, "month": m,
        "start": start, "end": end,
        "header_cols": header_cols,
        "groups": grouped,
        "grand_row": grand_row,
        "all_categories": all_categories,
        "visible_cat_ids": visible_cat_ids,
        "has_extras": bool(extra_ids),  # for PDF orientation
        "active_only": active_only,
        "active_only_param": active_only_param,
    }
    return ctx


@login_required
def billing_summary(request):
    ctx = build_billing_summary_context(request)
    return render(request, "smart_meter/billing_summary.html", ctx)


def _two_line_cat(name: str, max_len: int = 13):
    s = (name or "")[:max_len]
    if len(s) > 6:
        a, b = s[:6], s[6:]
    else:
        sp = s.find(" ")
        if sp != -1:
            a, b = s[:sp], s[sp+1:]
        else:
            a, b = s, ""
    short_html = f"{a}<br>{b}" if b else a      # for HTML/PDF
    short_text = f"{a}\n{b}" if b else a       # for Excel wrap
    return s, short_html, short_text


def _unit_key(u):
    s = str(u or "")
    m = re.search(r"\d+", s)
    return (int(m.group()) if m else 10**9, s)


@login_required
def billing_summary_items(request):
    """
    AJAX (JSON) + optional export (?export=excel|pdf)
    Params: month, scope(unit|property|grand), unit_id/property_id, category(*all*|other|<cat_id>), cols
    """
    status_on_error = 200 if request.GET.get("debug") == "1" else 500

    try:
        y, m = _ym_from_qs(request)
        start, end = _first_last_of_month(y, m)
        month_label = date(y, m, 1).strftime("%b %Y")
        full_month_label = date(y, m, 1).strftime("%B %Y")

        scope = (request.GET.get("scope") or "").strip().lower()
        if scope not in ("unit", "property", "grand"):
            return JsonResponse({"error": f"Invalid scope '{scope}'."}, status=400)

        # visible cat ids (for "other")
        visible_ids = set(DEFAULT_CAT_IDS)
        cols = (request.GET.get("cols") or "").strip()
        if cols:
            for t in cols.split(","):
                t = t.strip()
                if not t:
                    continue
                try:
                    visible_ids.add(int(t))
                except Exception:
                    pass

        unit_id = request.GET.get("unit_id")
        prop_id = request.GET.get("property_id")
        active_only = request.GET.get("active_only") == "1"

        # derive heading property label
        scope_prop_name = ""
        if scope == "unit":
            if not unit_id:
                return JsonResponse({"error": "Missing unit_id for scope=unit."}, status=400)
            lease_qs = Lease.objects.filter(
                unit_id=unit_id).select_related("unit__property")
            if active_only:
                lease_qs = lease_qs.filter(status="active")
            lease_ids = list(lease_qs.values_list("id", flat=True))
            u0 = lease_qs.first()
            if u0 and getattr(u0.unit, "property", None):
                scope_prop_name = u0.unit.property.property_name
        elif scope == "property":
            if not prop_id:
                return JsonResponse({"error": "Missing property_id for scope=property."}, status=400)
            lease_qs = Lease.objects.filter(
                unit__property_id=prop_id).select_related("unit__property")
            if active_only:
                lease_qs = lease_qs.filter(status="active")
            lease_ids = list(lease_qs.values_list("id", flat=True))
            u0 = lease_qs.first()
            if u0 and getattr(u0.unit, "property", None):
                scope_prop_name = u0.unit.property.property_name
        else:
            lease_qs = Lease.objects.all()
            if active_only:
                lease_qs = lease_qs.filter(status="active")
            lease_ids = list(lease_qs.values_list("id", flat=True))

        # list invoices
        invs = (
            Invoice.objects.filter(
                lease_id__in=lease_ids, issue_date__gte=start, issue_date__lte=end)
            .select_related("lease", "lease__tenant", "lease__unit", "lease__unit__property")
            .order_by("lease__unit")
        )

        # category filter + heading label
        # --- category filter + heading label (robust) ---
        category_param = (request.GET.get("category") or "").strip()

        # default
        category_label = "all category"  # used when "*all*" or empty

        if category_param not in ("", "*all*"):
            if category_param == "other":
                category_label = "Other"
            else:
                # Try to resolve numeric id to the real category name
                cid = None
                try:
                    cid = int(category_param)
                except Exception:
                    cid = None

                resolved = None
                if cid is not None:
                    # 1) Prefer Category table (works even if there are no items this month)
                    if Category:
                        resolved = Category.objects.filter(
                            id=cid).values_list("name", flat=True).first()

                    # 2) Fallback: resolve via any invoice item (if Category model import not available)
                    if not resolved:
                        resolved = (InvoiceItem.objects
                                    .filter(category_id=cid)
                                    .values_list("category__name", flat=True)
                                    .first())

                category_label = resolved or "Category"  # final fallback

        lines = []
        for inv in invs:
            tenant_obj = getattr(inv.lease, "tenant", None)
            tenant_name = getattr(tenant_obj, "name", "") or (
                str(tenant_obj) if tenant_obj else "")
            unit_number = getattr(inv.lease.unit, "unit_number", "")
            prop_obj = getattr(inv.lease.unit, "property", None)
            property_name = getattr(
                prop_obj, "property_name", "") if prop_obj else ""

            for it in InvoiceItem.objects.filter(invoice_id=inv.id).select_related("category"):
                # include filter
                if category_param == "*all*" or category_param == "":
                    include = True
                elif category_param == "other":
                    include = (it.category_id not in visible_ids)
                else:
                    try:
                        include = (it.category_id == int(category_param))
                    except Exception:
                        include = False
                if not include:
                    continue

                try:
                    inv_url = reverse(
                        "invoices:invoice_detail", kwargs={"pk": inv.id})
                    if it.category_id:
                        inv_url += f"?category_id={it.category_id}"
                except NoReverseMatch:
                    inv_url = "#"

                full_cat_name = (getattr(it.category, "name", "")
                                 or "") if it.category_id else ""
                cat_name = full_cat_name[:13]  # <= single line, max 13 chars

                lines.append({
                    "sn": 0,  # fill later
                    "invoice_id": inv.id,
                    "issue_date": inv.issue_date,
                    "property": property_name,
                    "tenant": tenant_name,
                    "unit": unit_number,
                    # use in UI/Excel (one line ≤ 13 chars)
                    "category": cat_name,
                    "category_short": cat_name,      # alias if you reference this name elsewhere
                    "category_full": full_cat_name,  # keep full for headings/tooltips
                    "amount": Decimal(it.amount or 0),
                    "invoice_url": inv_url,
                })

        # sort and serial
        lines.sort(key=lambda it: (
            _unit_key(it.get("unit")),
            (it.get("tenant") or "").lower(),
            (it.get("category") or "").lower(),
            it.get("issue_date") or date(1900, 1, 1),
        ))
        for i, l in enumerate(lines, start=1):
            l["sn"] = i

        # group by property for UI subtotals
        from collections import defaultdict
        by_prop = defaultdict(list)
        for l in lines:
            by_prop[l["property"]].append(l)

        groups_items = []
        for prop, rows in by_prop.items():
            subtotal = sum((x["amount"] for x in rows), ZERO)
            groups_items.append(
                {"property": prop, "rows": rows, "subtotal": subtotal})

        total_amount = sum((x["amount"] for x in lines), ZERO)

        # unified heading text
        prop_label = scope_prop_name if scope in (
            "property", "unit") and scope_prop_name else "ALL Property"
        heading_text = f"Detail for {prop_label}, for {category_label} for the month of {full_month_label}"

        # build qs for export links
        base_params = {}
        for k in ("month", "scope", "category", "unit_id", "property_id", "cols", "active_only"):
            v = request.GET.get(k)
            if v:
                base_params[k] = v
        qs = urlencode(base_params)

        # ===== exports
        export = (request.GET.get("export") or "").lower().strip()
        if export in ("excel", "pdf"):
            filename_base = "Detail-All" if scope == "grand" else (
                f"Detail-{scope_prop_name}" if scope in (
                    "property", "unit") and scope_prop_name else "Detail"
            )

            if export == "excel":
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Alignment, Font, numbers

                wb = Workbook()
                ws = wb.active
                ws.title = f"Details {month_label}"

                # merged heading (A..G)
                ws.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=7)
                hcell = ws.cell(row=1, column=1, value=heading_text)
                hcell.font = Font(size=13, bold=True)
                hcell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions[1].height = 24

                # headers row (row 2) — Inv#
                ws.append(["S.N.", "Date", "Inv#", "Tenant",
                          "Unit", "Category", "Amount"])

                # data rows start row 3
                for it in lines:
                    ws.append([
                        it["sn"],
                        it["issue_date"],         # date
                        it["invoice_id"],         # Inv#
                        it["tenant"],
                        it["unit"],
                        it["category_short"],   # 2-line category
                        float(it["amount"]),
                    ])
                DATE_FMT = "DD-MMM-YYYY"  # e.g. 07-Oct-2025

                # format Date col (B)
                for r in range(3, ws.max_row + 1):
                    ws.cell(row=r, column=2).number_format = DATE_FMT

                # total
                ws.append(["", "", "", "", "", "Total", float(total_amount)])

                # widths
                widths = [6, 12, 12, 24, 15, 16, 12]
                for idx, w in enumerate(widths, start=1):
                    ws.column_dimensions[get_column_letter(idx)].width = w

                resp = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                resp["Content-Disposition"] = f'attachment; filename="{filename_base}-{month_label}.xlsx"'
                wb.save(resp)
                return resp

            else:  # PDF
                if HTML is None:
                    return HttpResponse("WeasyPrint not installed on server.", status=500)

                generated_at = timezone.localtime(
                    timezone.now()).strftime("%d %b %Y, %I:%M %p")
                html = render_to_string(
                    "smart_meter/partials/billing_summary_items_export.html",
                    {
                        "heading_text": heading_text,
                        "lines": lines,
                        "total_amount": total_amount,
                        "generated_at": generated_at,
                    },
                    request=request,
                )
                pdf = HTML(
                    string=html, base_url=request.build_absolute_uri()).write_pdf()
                resp = HttpResponse(pdf, content_type="application/pdf")
                resp["Content-Disposition"] = f'attachment; filename="{filename_base}-{month_label}.pdf"'
                return resp

        # ===== AJAX fragment
        html = render_to_string(
            "smart_meter/partials/billing_summary_items.html",
            {
                "heading_text": heading_text,
                "qs": qs,
                "groups": groups_items,
                "total_amount": total_amount,
            },
            request=request,
        )
        return JsonResponse({"html": html})

    except Exception as e:
        return JsonResponse(
            {"error": f"{type(e).__name__}: {e}",
             "trace": traceback.format_exc()[:4000]},
            status=status_on_error,
        )


@permission_required("accounts.can_export_energy", raise_exception=True)
def billing_summary_export_excel(request):
    ctx = build_billing_summary_context(request)
    y, m = ctx["year"], ctx["month"]
    month_full = date(y, m, 1).strftime("%B %Y")

    # helper: 13-char, break after first word (TEXT version for Excel headers)
    def _short_13_text(label: str):
        s = (label or "").strip()
        if len(s) > 13:
            s = s[:13]
        i = s.find(" ")
        if i > 0:
            return s[:i] + "\n" + s[i+1:]
        cut = min(6, len(s))
        return s[:cut] + ("\n" + s[cut:] if len(s) > cut else "")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Billing {y}-{m:02d}"

    # Build Excel headers (skip WhatsApp)
    headers = ["S.N.", "Unit", "Tenant", "Phone"]
    excel_kinds = []  # keep kinds to set widths/wrap later for each added column
    for col in ctx["header_cols"]:
        if col["kind"] == "whatsapp":
            continue
        if col["kind"] in ("cat", "other"):
            headers.append(_short_13_text(col["label"]))
        else:
            headers.append(col["label"].replace("\n", " "))
        excel_kinds.append(col["kind"])

    # Row 1: merged title
    title = f"Billing Summary for the Month of {month_full}"
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(headers))
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(size=13, bold=True)
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2: column headers
    ws.append(headers)
    # Bold + wrap only cat/other header cells
    # first 4 fixed cols already exist
    for c_idx, kind in enumerate(excel_kinds, start=5):
        hcell = ws.cell(row=2, column=c_idx)
        hcell.font = Font(bold=True)
        if kind in ("cat", "other"):
            hcell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center")
        else:
            hcell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for g in ctx["groups"]:
        # Property header row (merged)
        ws.append([f"Property: {g['property'].property_name}"])
        ws.merge_cells(start_row=ws.max_row, start_column=1,
                       end_row=ws.max_row, end_column=len(headers))

        for r in g["rows"]:
            row = [r["sn"], r["unit_display_label"],
                   r["tenant_name"], r["tenant_phone"]]
            for cell in r["cells"]:
                if cell["kind"] == "whatsapp":
                    continue
                if cell["kind"] in ("cat", "other", "total", "balance"):
                    row.append(float(cell["value"]))
                else:
                    row.append("")
            ws.append(row)

        # Subtotal row
        st = ["", f"Subtotal for {g['property'].property_name}", "", ""]
        for v in g["subtotal_row"]:
            if isinstance(v, (int, float, Decimal)):
                st.append(float(v))
            else:
                st.append("")
        ws.append(st)

    # Grand total
    gt = ["", "GRAND TOTAL", "", ""]
    for v in ctx["grand_row"]:
        if isinstance(v, (int, float, Decimal)):
            gt.append(float(v))
        else:
            gt.append("")
    ws.append(gt)

    # Tighter, readable widths (no WhatsApp col)
    # first 4 fixed:
    widths = [6, 16, 18, 15]
    # then dynamic for each added header
    for kind in excel_kinds:
        if kind in ("cat", "other"):
            widths.append(10)  # narrow category cols
        elif kind in ("total", "balance"):
            widths.append(12)
        else:
            widths.append(10)

    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Output
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fn = f"billing_summary_{y}-{m:02d}.xlsx"
    resp = HttpResponse(out.read(
    ), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fn}"'
    return resp


@permission_required("accounts.can_export_energy", raise_exception=True)
def billing_summary_export_pdf(request):
    if HTML is None:
        return HttpResponse("WeasyPrint not installed.", status=500)

    ctx = build_billing_summary_context(request)

    # --- month + timestamp for footer ---
    y, m = ctx["year"], ctx["month"]
    ctx["month_full"] = date(y, m, 1).strftime("%B %Y")
    ctx["generated_at"] = timezone.localtime(
        timezone.now()).strftime("%d %b %Y, %I:%M %p")

    # --- orientation rule you already had ---
    ctx["orientation"] = "landscape" if ctx.get("has_extras") else "portrait"

    # --- helper: 13-char, break after first word (HTML version for PDF headers) ---
    def _short_13_html(label: str):
        s = (label or "").strip()
        if len(s) > 13:
            s = s[:13]
        i = s.find(" ")
        if i > 0:
            return f"{s[:i]}<br>{s[i+1:]}"
        cut = min(6, len(s))
        return s[:cut] + ("<br>" + s[cut:] if len(s) > cut else "")

    # Build a PDF-specific header list (skip WhatsApp; shorten category-like headers)
    header_cols_pdf = []
    for col in ctx["header_cols"]:
        if col["kind"] == "whatsapp":
            continue
        newcol = dict(col)
        if col["kind"] in ("cat", "other"):
            newcol["label_html"] = _short_13_html(col["label"])
        else:
            newcol["label_html"] = col["label"]
        header_cols_pdf.append(newcol)
    ctx["header_cols_pdf"] = header_cols_pdf

    # Render
    pdf_html = render_to_string(
        "smart_meter/billing_summary_pdf.html", ctx, request=request)
    pdf = HTML(string=pdf_html,
               base_url=request.build_absolute_uri("/")).write_pdf()
    fn = f"billing_summary_{y}-{m:02d}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{fn}"'
    return resp
