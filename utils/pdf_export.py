from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import letter, portrait
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import io
from html import unescape
from datetime import datetime
from decimal import Decimal
from xml.sax.saxutils import escape as xml_escape
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django_tables2.export.export import TableExport
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, portrait, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.platypus import BaseDocTemplate, PageTemplate, Frame
from django.contrib.auth import get_user_model
import logging
from reportlab.platypus import Spacer  # Add this import
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.DEBUG)  # Add this line

User = get_user_model()

# Standard PDF formatting for the entire app
STANDARD_PDF_FORMAT = {
    'page_size': 'A4',
    'orientation': 'portrait',  # Default orientation
    'header': {
        'font': 'Helvetica-Bold',
        'size': 14,
        'color': colors.black,
        'spacing': 20
    },
    'table': {
        'header_background': colors.HexColor(0xf2f2f2),
        'header_text_color': colors.black,
        'text_color': colors.black,
        'font': 'Helvetica',
        'header_size': 10,
        'body_size': 8,
        'grid_color': colors.black,
        'padding': 6
    },
    'footer': {
        'font': 'Helvetica',
        'size': 8,
        'color': colors.black
    },
    'margins': {
        'left': 36,
        'right': 36,
        'top': 36,
        'bottom': 36
    }
}


def get_document_title(request, default_title):
    """Get title from table or use default"""
    try:
        # Try to get from resolved view name
        app_name = request.resolver_match.app_name
        return f"{app_name.replace('_', ' ').title()} - {default_title}"
    except Exception as e:
        logger.warning(f"Could not extract title: {str(e)}")
        return f"TMS - {default_title}"


def _table_total_balance(table):
    total = Decimal("0.00")
    found = False
    for item in table.data:
        value = getattr(item, "list_balance", None)
        if value is None:
            value = getattr(item, "get_balance", None)
        if callable(value):
            value = value()
        if value is None:
            continue
        total += Decimal(str(value or 0))
        found = True
    return total if found else None


def _table_property_label(table):
    property_names = set()
    for item in table.data:
        unit = getattr(item, "unit", None)
        property_obj = getattr(unit, "property", None)
        property_name = getattr(property_obj, "property_name", "")
        if property_name:
            property_names.add(str(property_name).strip())

    if len(property_names) == 1:
        return f"Property: {next(iter(property_names))}"
    return "Property: All Properties"


def _plain_export_value(value):
    if value is None:
        return ""
    return " ".join(unescape(strip_tags(str(value))).split())


def _apply_export_verbose_names(table):
    for column_name, verbose_name in getattr(
        table, "export_verbose_names", {}
    ).items():
        if column_name in table.columns:
            table.columns[column_name].column.verbose_name = verbose_name


def _renumber_exporter_dataset(exporter):
    from tablib import Dataset

    dataset = exporter.dataset
    headers = list(dataset.headers or [])
    serial_index = next(
        (
            index
            for index, header in enumerate(headers)
            if str(header).strip().lower().replace("#", "")
            in {"sn", "s.n.", "s.n"}
        ),
        None,
    )
    if serial_index is None:
        return exporter

    renumbered = Dataset(headers=headers)
    for serial_number, row in enumerate(dataset, start=1):
        values = list(row)
        values[serial_index] = serial_number
        renumbered.append(values)
    exporter.dataset = renumbered
    return exporter


def add_header_footer(canvas, doc, title, request, compact_layout=False):
    """Add header and footer with page numbers, timestamp, and user info"""
    canvas.saveState()

    styles = getSampleStyleSheet()
    footer_style = ParagraphStyle(
        'Footer',
        fontName=STANDARD_PDF_FORMAT['footer']['font'],
        fontSize=STANDARD_PDF_FORMAT['footer']['size'],
        textColor=STANDARD_PDF_FORMAT['footer']['color'],
        alignment=TA_CENTER
    )

    # Get current user info
    user_info = ""
    if request and request.user.is_authenticated:
        full_name_getter = getattr(request.user, "get_full_name", None)
        full_name = full_name_getter() if callable(full_name_getter) else ""
        username_getter = getattr(request.user, "get_username", None)
        username = (
            username_getter()
            if callable(username_getter)
            else getattr(request.user, "username", "")
        )
        user_info = f" | Generated by: {full_name or username}"

    page_num = canvas.getPageNumber()
    if compact_layout:
        footer_y = doc.bottomMargin / 2
        canvas.setFillColor(STANDARD_PDF_FORMAT['footer']['color'])
        canvas.setFont(
            STANDARD_PDF_FORMAT['footer']['font'],
            STANDARD_PDF_FORMAT['footer']['size'],
        )
        canvas.drawString(
            doc.leftMargin,
            footer_y,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}{user_info}",
        )
        canvas.drawCentredString(doc.pagesize[0] / 2, footer_y, title)
        canvas.drawRightString(
            doc.pagesize[0] - doc.rightMargin,
            footer_y,
            f"Page {page_num}",
        )
        canvas.restoreState()
        return

    # Footer text with page number, timestamp, and user info
    footer_text = (f"Page {page_num} of {doc.page} | "
                   f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                   f"{user_info}")

    p = Paragraph(footer_text, footer_style)
    w, h = p.wrap(doc.width, doc.bottomMargin/2)
    p.drawOn(canvas, doc.leftMargin, h)

    canvas.restoreState()


class PDFTableExport:
    def __init__(self, table, export_name='export', request=None):
        self.table = table
        self.export_name = export_name
        self.title = getattr(table, 'pdf_export_title',
                             get_document_title(request, export_name.replace('_', ' ').title()))
        self.request = request

       # Set default PDF options
        self.pdf_options = {
            'orientation': 'portrait',  # Default value
            'column_widths': None,
            'format': 'A4'
        }

        # Check for attributes at both class and Meta levels
        table_attrs = {}

        # Check class-level attributes first
        if hasattr(table, 'pdf_export_attrs'):
            table_attrs.update(getattr(table, 'pdf_export_attrs', {}))

        # Then check Meta attributes
        if hasattr(table, '_meta') and hasattr(table._meta, 'pdf_export_attrs'):
            table_attrs.update(table._meta.pdf_export_attrs)

        # Merge with defaults
        self.pdf_options.update(table_attrs)

        logger.debug(f"Merged PDF options: {self.pdf_options}")

        # Set title from table or default
        self.title = self.title or getattr(table, 'pdf_export_title',
                                           get_document_title(request, export_name.replace('_', ' ').title()))

        # Always exclude actions column
        export_exclude = tuple(getattr(table, 'export_exclude', ()))
        if 'actions' in table.base_columns:
            export_exclude += ('actions',)
        table.exclude = tuple(
            dict.fromkeys((*tuple(table.exclude or ()), *export_exclude))
        )
        _apply_export_verbose_names(table)

    def export_pdf(self):
        try:
            # Setup PDF document
            buffer = io.BytesIO()
            orientation = self.pdf_options.get(
                'orientation', 'portrait').lower()
            page_size = landscape(
                letter) if orientation == 'landscape' else portrait(letter)

            doc = BaseDocTemplate(
                buffer,
                pagesize=page_size,
                leftMargin=STANDARD_PDF_FORMAT['margins']['left'],
                rightMargin=STANDARD_PDF_FORMAT['margins']['right'],
                topMargin=STANDARD_PDF_FORMAT['margins']['top'],
                bottomMargin=STANDARD_PDF_FORMAT['margins']['bottom'],
                title=self.title,
                author="Tenant Management System",
                creator="TMS App"
            )

            # Create frame for content
            frame = Frame(
                doc.leftMargin,
                doc.bottomMargin,
                doc.width,
                doc.height,
                id='normal'
            )

            # Build story elements
            elements = []

            compact_heading = bool(
                getattr(self.table, 'compact_export_heading', False)
            )

            # Add document title
            title_style = ParagraphStyle(
                'Title',
                fontName=STANDARD_PDF_FORMAT['header']['font'],
                fontSize=STANDARD_PDF_FORMAT['header']['size'],
                textColor=STANDARD_PDF_FORMAT['header']['color'],
                alignment=TA_CENTER,
                spaceAfter=STANDARD_PDF_FORMAT['header']['spacing']
            )
            total_balance = _table_total_balance(self.table)
            if compact_heading:
                heading_table = Table(
                    [[
                        Paragraph(f"<b>{xml_escape(_table_property_label(self.table))}</b>", ParagraphStyle(
                            'HeadingLeft', parent=title_style, fontSize=10,
                            alignment=TA_LEFT, spaceAfter=0,
                        )),
                        Paragraph(self.title, ParagraphStyle(
                            'HeadingCenter', parent=title_style, fontSize=12,
                            alignment=TA_CENTER, spaceAfter=0,
                        )),
                        Paragraph(
                            f"<b>Total Balance: {total_balance:,.2f}</b>" if total_balance is not None else "",
                            ParagraphStyle(
                                'HeadingRight', parent=title_style, fontSize=10,
                                alignment=TA_RIGHT, spaceAfter=0,
                            ),
                        ),
                    ]],
                    colWidths=[doc.width / 3] * 3,
                )
                heading_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ]))
                elements.append(heading_table)
            else:
                elements.append(Paragraph(self.title, title_style))

            # Get table data using proper CSV parsing
            exporter = TableExport('csv', self.table)
            _renumber_exporter_dataset(exporter)
            data = exporter.export()

            # Use csv module to properly parse rows with commas in data
            import csv
            from io import StringIO
            csv_data = StringIO(data)
            reader = csv.reader(csv_data)
            rows = list(reader)

            # Get column names excluding actions
            col_names = [
                col.name for col in self.table.columns if col.name != 'actions']

            # Create styles for wrapping text
            styles = getSampleStyleSheet()
            table_header_size = (
                9 if compact_heading else STANDARD_PDF_FORMAT['table']['header_size']
            )
            table_padding = (
                4 if compact_heading else STANDARD_PDF_FORMAT['table']['padding']
            )
            header_style = ParagraphStyle(
                'Header',
                parent=styles['Normal'],
                fontName=STANDARD_PDF_FORMAT['table']['font'],
                fontSize=table_header_size,
                textColor=STANDARD_PDF_FORMAT['table']['header_text_color'],
                leading=table_header_size + 2,
                spaceBefore=0,
                spaceAfter=0,
                alignment=TA_CENTER
            )

            cell_style = ParagraphStyle(
                'Cell',
                parent=styles['Normal'],
                fontName=STANDARD_PDF_FORMAT['table']['font'],
                fontSize=STANDARD_PDF_FORMAT['table']['body_size'],
                textColor=STANDARD_PDF_FORMAT['table']['text_color'],
                leading=STANDARD_PDF_FORMAT['table']['body_size'] + 2,
                spaceBefore=0,
                spaceAfter=0,
                alignment=TA_LEFT
            )

            # Convert all cells to Paragraphs for text wrapping
            wrapped_rows = []
            for i, row in enumerate(rows):
                # Skip empty rows
                if not any(row):
                    continue

                wrapped_row = []
                for j, cell in enumerate(row):
                    # Use header style for header row, cell style for others
                    style = header_style if i == 0 else cell_style
                    # Remove any extra quotes that might have been added during CSV export
                    cell_value = cell.strip('"\'') if cell else ''
                    cell_value = xml_escape(unescape(strip_tags(cell_value)))
                    wrapped_row.append(Paragraph(cell_value, style))
                wrapped_rows.append(wrapped_row)

            # Determine column widths
            column_widths = self.pdf_options.get('column_widths', {})

            if column_widths:
                # Calculate total requested width
                total_requested = sum(column_widths.values())
                available_width = doc.width * 0.95  # Leave 5% margin

                if total_requested > available_width:
                    # Scale down proportionally
                    scale_factor = available_width / total_requested
                    col_widths = [column_widths.get(
                        name, 40) * scale_factor for name in col_names]
                else:
                    # Use exact widths if they fit
                    col_widths = [column_widths.get(name, 40)
                                  for name in col_names]
            else:
                # Distribute evenly if no custom widths
                col_widths = None

            # Create table with calculated widths
            pdf_table = Table(wrapped_rows, colWidths=col_widths, repeatRows=1)
            pdf_table.hAlign = 'LEFT'

            # Apply table styling
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0),
                 STANDARD_PDF_FORMAT['table']['header_background']),
                ('TEXTCOLOR', (0, 0), (-1, 0),
                 STANDARD_PDF_FORMAT['table']['header_text_color']),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0),
                 STANDARD_PDF_FORMAT['table']['font']),
                ('FONTSIZE', (0, 0), (-1, 0),
                 table_header_size),
                ('FONTSIZE', (0, 1), (-1, -1),
                 STANDARD_PDF_FORMAT['table']['body_size']),
                ('BOTTOMPADDING', (0, 0), (-1, 0),
                 table_padding),
                ('TOPPADDING', (0, 1), (-1, -1),
                 table_padding),
                ('BOTTOMPADDING', (0, 1), (-1, -1),
                 table_padding),
                ('GRID', (0, 0), (-1, -1), 1,
                 STANDARD_PDF_FORMAT['table']['grid_color']),
                ('WORDWRAP', (0, 0), (-1, -1)),
                ('LEADING', (0, 0), (-1, -1),
                 STANDARD_PDF_FORMAT['table']['body_size'] + 2),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ])

            pdf_table.setStyle(table_style)

            # Calculate and show total on top for the standard stacked heading.
            if not compact_heading:
                try:
                    if total_balance is None:
                        raise ValueError("No balance field is available for this export.")
                    total_para = Paragraph(
                        f"<b>Total Balance: {total_balance:,.2f}</b>", title_style)
                    elements.append(total_para)
                    elements.append(Spacer(1, 12))
                except Exception as e:
                    logger.warning(f"Failed to calculate total at top: {e}")

            elements.append(pdf_table)

            # Add page template with header/footer
            doc.addPageTemplates([
                PageTemplate(
                    id='AllPages',
                    frames=frame,
                    onPage=lambda canvas, doc: add_header_footer(
                        canvas,
                        doc,
                        self.title,
                        self.request,
                        compact_layout=compact_heading,
                    )
                )
            ])

            doc.build(elements)
            pdf = buffer.getvalue()
            buffer.close()

            # Create response with proper filename
            filename = f"{self.title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            logger.error(f"PDF export failed: {str(e)}", exc_info=True)
            return HttpResponse(f"PDF generation failed: {str(e)}", status=500)


def _pdf_response_to_jpg(pdf_response, title):
    from io import BytesIO

    import fitz
    from PIL import Image

    document = fitz.open(stream=pdf_response.content, filetype="pdf")
    pages = []
    try:
        for page in document:
            pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            pages.append(
                Image.frombytes("RGB", (pixmap.width, pixmap.height), pixmap.samples)
            )

        if not pages:
            raise ValueError("The PDF export did not contain any pages.")

        gap = 20
        canvas_width = max(page.width for page in pages)
        canvas_height = sum(page.height for page in pages) + gap * (len(pages) - 1)
        combined = Image.new("RGB", (canvas_width, canvas_height), "white")

        y_position = 0
        for page in pages:
            x_position = (canvas_width - page.width) // 2
            combined.paste(page, (x_position, y_position))
            y_position += page.height + gap

        output = BytesIO()
        combined.save(output, format="JPEG", quality=92, optimize=True)
        filename = f"{title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.jpg"
        response = HttpResponse(output.getvalue(), content_type="image/jpeg")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
    finally:
        document.close()
        for page in pages:
            page.close()


def handle_export(request, table, export_name, title=None, allowed_formats=('csv', 'xlsx', 'pdf', 'jpg')):
    try:
        export_format = request.GET.get('_export')
        if not export_format or export_format not in allowed_formats:
            return None

        # Exclude actions and table-specific export-only columns.
        export_exclude = tuple(getattr(table, 'export_exclude', ()))
        if 'actions' in table.base_columns:
            export_exclude += ('actions',)
        table.exclude = tuple(
            dict.fromkeys((*tuple(table.exclude or ()), *export_exclude))
        )
        _apply_export_verbose_names(table)

        # Get title from table or default
        title = getattr(table, 'pdf_export_title',
                        get_document_title(request, export_name.replace('_', ' ').title()))



        if export_format in ('pdf', 'jpg'):
            exporter = PDFTableExport(
                table=table, export_name=export_name, request=request)
            pdf_response = exporter.export_pdf()
            if export_format == 'pdf' or pdf_response.status_code != 200:
                return pdf_response
            return _pdf_response_to_jpg(pdf_response, title)

        # Handle CSV/Excel exports
        exporter = TableExport(export_format, table)
        _renumber_exporter_dataset(exporter)

        if export_format == 'xlsx':
            import openpyxl
            from openpyxl.utils import get_column_letter
            from io import BytesIO

            # Create in-memory workbook
            output = BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Export"
            ws.sheet_view.showGridLines = False
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            compact_heading = bool(getattr(table, 'compact_export_heading', False))
            include_property = bool(getattr(table, 'excel_include_property', False))
            column_count = len(table.columns) + (1 if include_property else 0)
            total_balance = _table_total_balance(table)
            if compact_heading:
                left_end = max(1, column_count // 3)
                right_start = column_count - left_end + 1
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=left_end)
                ws.merge_cells(
                    start_row=1,
                    start_column=left_end + 1,
                    end_row=1,
                    end_column=right_start - 1,
                )
                ws.merge_cells(
                    start_row=1,
                    start_column=right_start,
                    end_row=1,
                    end_column=column_count,
                )
                heading_values = (
                    (1, _table_property_label(table), "left", 10),
                    (left_end + 1, title, "center", 12),
                    (
                        right_start,
                        f"Total Balance: {total_balance:,.2f}" if total_balance is not None else "",
                        "right",
                        10,
                    ),
                )
                for column, value, alignment, size in heading_values:
                    cell = ws.cell(row=1, column=column, value=value)
                    cell.font = openpyxl.styles.Font(bold=True, size=size)
                    cell.alignment = openpyxl.styles.Alignment(horizontal=alignment)
                header_row = 2
                data_start_row = 3
                ws.print_title_rows = "1:2"
                ws.oddFooter.left.text = "Generated: &D &T"
                ws.oddFooter.center.text = title
                ws.oddFooter.right.text = "Page &P of &N"
                for section in (ws.oddFooter.left, ws.oddFooter.center, ws.oddFooter.right):
                    section.font = "Arial"
                    section.size = 8
                    section.color = "000000"
            else:
                ws.merge_cells(
                    start_row=1,
                    start_column=1,
                    end_row=1,
                    end_column=column_count,
                )
                title_cell = ws.cell(row=1, column=1, value=title)
                title_cell.font = openpyxl.styles.Font(bold=True, size=14)
                title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                if total_balance is not None:
                    ws.merge_cells(
                        start_row=2,
                        start_column=1,
                        end_row=2,
                        end_column=column_count,
                    )
                    total_cell = ws.cell(
                        row=2,
                        column=1,
                        value=f"Total Balance: {total_balance:,.2f}",
                    )
                    total_cell.font = openpyxl.styles.Font(bold=True, size=12)
                    total_cell.alignment = openpyxl.styles.Alignment(horizontal="center")
                header_row = 3
                data_start_row = 4
                ws.print_title_rows = "1:3"
                ws.oddFooter.center.text = "Page &P of &N"
                ws.oddFooter.center.font = "Arial"
                ws.oddFooter.center.size = 8
                ws.oddFooter.center.color = "000000"

            # Add column headers in row 2
            headers = []
            for column in table.columns:
                # Get verbose_name if available, otherwise use the field name
                header = getattr(column, 'verbose_name', column.header)
                headers.append(str(header))

            column_names = [column.name for column in table.columns]
            property_index = None
            if include_property:
                property_index = (
                    column_names.index("unit") + 1
                    if "unit" in column_names
                    else len(column_names)
                )
                column_names.insert(property_index, "property")
                headers.insert(property_index, "Property")

            for col_idx, header in enumerate(headers, start=1):
                header_cell = ws.cell(row=header_row, column=col_idx, value=header)
                header_cell.font = openpyxl.styles.Font(bold=True)
                header_cell.fill = openpyxl.styles.PatternFill(
                    start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                header_cell.alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            thin_side = openpyxl.styles.Side(style="thin", color="000000")
            table_border = openpyxl.styles.Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side,
            )
            for header_cell in ws[header_row]:
                header_cell.border = table_border

            # Add plain-text table data without HTML links, forms, or buttons.
            nowrap_columns = {
                "tenant", "property", "unit", "vehicle_info", "monthly_payments",
                "balance", "security_due",
            }
            records = list(table.data) if include_property else []
            for dataset_index, row in enumerate(exporter.dataset):
                row_idx = data_start_row + dataset_index
                values = list(row)
                if property_index is not None:
                    record = records[dataset_index]
                    property_obj = getattr(getattr(record, "unit", None), "property", None)
                    property_name = getattr(property_obj, "property_name", "")
                    values.insert(property_index, property_name)
                for col_idx, value in enumerate(values, start=1):
                    column_name = column_names[col_idx - 1]
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=(
                            dataset_index + 1
                            if column_name == "sn"
                            else _plain_export_value(value)
                        ),
                    )
                    cell.border = table_border
                    cell.alignment = openpyxl.styles.Alignment(
                        vertical="top",
                        wrap_text=column_name not in nowrap_columns,
                    )

            last_data_row = data_start_row + len(exporter.dataset) - 1
            filter_end_row = max(header_row, last_data_row)
            ws.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(len(headers))}{filter_end_row}"
            )
            ws.freeze_panes = f"A{data_start_row}"

            # Add total row at the end
            try:
                amount_values = [
                    item.amount
                    for item in table.data
                    if getattr(item, "amount", None) is not None
                ]
                if amount_values:
                    total = sum(amount_values)
                    ws.append([''] * (len(headers) - 2) + ['Total', float(total)])
            except Exception as e:
                logger.warning(f"Excel total row failed: {e}")

            # Auto-size columns (excluding merged title cell)
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)

                # Check header width first
                if len(headers[col_idx - 1]) > max_length:
                    max_length = len(headers[col_idx - 1])

                # Check data rows width
                for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            cell_length = len(
                                str(cell.value)) if cell.value else 0
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass

                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = min(
                    adjusted_width, 50)  # Cap at 50

            if compact_heading:
                compact_widths = {
                    "sn": 5,
                    "tenant": 24,
                    "property": 18,
                    "unit": 16,
                    "family_members": 8,
                    "bill_water_charges": 9,
                    "vehicle_info": 10,
                    "monthly_payments": 14,
                    "due_date": 9,
                    "status": 10,
                    "start_date": 12,
                    "end_date": 12,
                    "balance": 14,
                    "security_due": 14,
                }
                for col_idx, column_name in enumerate(column_names, start=1):
                    width = compact_widths.get(column_name)
                    if width:
                        ws.column_dimensions[get_column_letter(col_idx)].width = width

            wb.save(output)
            data = output.getvalue()
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            # For CSV, use the original exporter
            data = exporter.export()
            content_type_value = exporter.content_type
            content_type = (
                content_type_value()
                if callable(content_type_value)
                else content_type_value
            )

        filename = f"{title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.{export_format}"
        response = HttpResponse(data, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        return HttpResponse(f"Export failed: {str(e)}", status=500)

# Add this to your existing pdf_export.py


# Add this to your pdf_export.py

logger = logging.getLogger(__name__)


class PaymentPDFGenerator:
    @staticmethod
    def generate_payment_pdf(payment, request=None):
        """Generate a PDF receipt for a single payment"""
        from io import BytesIO
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, portrait

        buffer = BytesIO()

        # Create document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=portrait(letter),
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=12
        )

        # Build elements
        elements = []

        # Add title
        elements.append(Paragraph("PAYMENT RECEIPT", title_style))

        # Add payment details
        elements.append(
            Paragraph(f"Reference: {payment.reference_number}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Add more payment details as needed...

        # Build PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        filename = f"payment_receipt_{payment.id}.pdf"
        return pdf, filename


class PaymentReceiptPDF1:
    @staticmethod
    def generate(payment, request=None):
        """Generate a PDF receipt for a single payment"""
        from io import BytesIO
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, portrait

        buffer = BytesIO()

        # Create document with smaller width (half page)
        doc = SimpleDocTemplate(
            buffer,
            pagesize=(4.25*72, 11*72),  # Half letter width
            leftMargin=18,
            rightMargin=18,
            topMargin=18,
            bottomMargin=18
        )

        styles = getSampleStyleSheet()

        # Custom styles to match your receipt design
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=6,
            fontName='Helvetica-Bold'
        )

        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Helvetica-Bold'
        )

        normal_style = styles['Normal']
        normal_style.fontSize = 9
        normal_style.leading = 11

        # Build elements
        elements = []

        # Title Section
        elements.append(Paragraph("PAYMENT RECEIPT", title_style))
        elements.append(
            Paragraph(f"#{payment.reference_number or payment.id}", normal_style))
        elements.append(Spacer(1, 12))

        # Payment Info
        payment_data = [
            ["Date:", payment.payment_date.strftime("%b %d, %Y")],
            ["Method:", payment.get_payment_method_display()],
            ["Amount:", f"${payment.amount:.2f}"],
        ]

        if payment.notes:
            payment_data.append(["Notes:", payment.notes])

        payment_table = Table(payment_data, colWidths=[100, 200])
        payment_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(payment_table)
        elements.append(Spacer(1, 12))

        # Tenant and Property Info
        tenant_info = [
            ["Tenant:",
                f"{payment.lease.tenant.first_name} {payment.lease.tenant.last_name}"],
            ["Phone:", payment.lease.tenant.phone],
            ["Property:", payment.lease.unit.property.property_name],
            ["Unit:", payment.lease.unit.unit_number],
            ["Balance:", f"${payment.lease.get_balance:.2f}"]
        ]

        info_table = Table(tenant_info, colWidths=[100, 200])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('PADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(info_table)

        # Footer
        elements.append(Spacer(1, 24))
        elements.append(Paragraph(
            "Thank you for your payment-this file is from utils forlder!", normal_style))
        elements.append(Paragraph(f"Generated on {datetime.now().strftime('%b %d, %Y')}",
                                  ParagraphStyle(
            'Footer',
            parent=normal_style,
            fontSize=8,
            textColor=colors.grey
        )))

        # Build PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        filename = f"payment_receipt_{payment.id}_{payment.payment_date.strftime('%Y%m%d')}.pdf"
        return pdf, filename


class PaymentReceiptPDF:
    @staticmethod
    def generate(payment, request=None):
        """Generate a PDF receipt matching payment_detail.html exactly"""
        from io import BytesIO
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.platypus.flowables import Image

        buffer = BytesIO()

        # Create document with similar dimensions to your HTML template
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )

        styles = getSampleStyleSheet()

        # Custom styles to match your HTML template
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            fontName='Helvetica-Bold',
            alignment=TA_LEFT
        )

        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Helvetica-Bold',
            textColor=colors.black
        )

        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            fontName='Helvetica'
        )

        small_style = ParagraphStyle(
            'Small',
            parent=styles['Normal'],
            fontSize=8,
            leading=9,
            fontName='Helvetica',
            textColor=colors.grey
        )

        bold_style = ParagraphStyle(
            'Bold',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            fontName='Helvetica-Bold'
        )

        # Build elements
        elements = []

        # Title Section (matches your HTML)
        elements.append(Paragraph("PAYMENT RECEIPT", title_style))
        elements.append(
            Paragraph(f"#{payment.reference_number or payment.id}", normal_style))
        elements.append(Spacer(1, 12))

        # Payment Info Section
        payment_info = [
            ["<b>Date:</b>", payment.payment_date.strftime("%b %d, %Y")],
            ["<b>Method:</b>", payment.get_payment_method_display()],
            ["<b>Amount:</b>", f"${payment.amount:.2f}"],
        ]

        if payment.notes:
            payment_info.append(["<b>Notes:</b>", payment.notes])

        payment_table = Table(payment_info, colWidths=[100, 200])
        payment_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(payment_table)
        elements.append(Spacer(1, 12))

        # Tenant and Property Info (matches your HTML)
        tenant_info = [
            ["<b>Tenant:</b>",
                f"{payment.lease.tenant.first_name} {payment.lease.tenant.last_name}"],
            ["<b>Phone:</b>", payment.lease.tenant.phone],
            ["<b>Property:</b>", payment.lease.unit.property.property_name],
            ["<b>Unit:</b>", payment.lease.unit.unit_number],
            ["<b>Balance:</b>", f"${payment.lease.get_balance:.2f}"]
        ]

        info_table = Table(tenant_info, colWidths=[100, 200])
        info_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 12))

        # Payment Details Table (matches your HTML template)
        payment_details = [
            ["<b>Description</b>", "<b>Amount</b>"],
            [f"Payment for {payment.get_payment_method_display()}",
             f"${payment.amount:.2f}"]
        ]

        details_table = Table(payment_details, colWidths=[300, 100])
        details_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(0xf8f9fa)),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))
        elements.append(details_table)
        elements.append(Spacer(1, 12))

        # Lease Information Section
        lease_info = [
            ["<b>Lease ID:</b>", payment.lease.id],
            ["<b>Period:</b>",
                f"{payment.lease.start_date.strftime('%b %d, %Y')} to {payment.lease.end_date.strftime('%b %d, %Y')}"],
            ["<b>Security Deposit:</b>",
                f"${payment.lease.security_deposit:.2f} ({'Paid' if payment.lease.security_deposit_paid else 'Pending'})"]
        ]

        lease_table = Table(lease_info, colWidths=[100, 200])
        lease_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(lease_table)
        elements.append(Spacer(1, 12))

        # Footer (matches your HTML template)
        elements.append(Paragraph("Thank you for your payment!", normal_style))
        elements.append(
            Paragraph(f"Generated on {datetime.now().strftime('%b %d, %Y')}", small_style))

        # Build PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        filename = f"payment_receipt_{payment.id}_{payment.payment_date.strftime('%Y%m%d')}.pdf"
        return pdf, filename

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO
from datetime import datetime


class PaymentDetailReceiptPDF:
    @staticmethod
    def generate(payment_detail, request=None):
        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=letter,
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )
        styles = getSampleStyleSheet()

        title = ParagraphStyle(
            "t", parent=styles["Heading2"],
            fontName="Helvetica-Bold", fontSize=13, spaceAfter=10
        )
        normal = ParagraphStyle("n", parent=styles["Normal"], fontSize=9, leading=11)

        p = payment_detail.payment
        lease = p.lease

        is_security_refund = (payment_detail.security_type or "").upper() == "REFUND"
        is_lease_refund = (payment_detail.lease_amount or 0) < 0
        is_refund = is_security_refund or is_lease_refund
        receipt_title = (
            "LEASE REFUND RECEIPT"
            if is_lease_refund
            else "SECURITY REFUND RECEIPT"
            if is_security_refund
            else "PAYMENT DETAIL RECEIPT"
        )
        total_label = "TOTAL REFUNDED" if is_refund else "TOTAL RECEIVED"
        security_label = "Security Refund" if is_security_refund else "Security Amount"
        security_value = f"Rs. {payment_detail.security_amount:,.2f}"
        total_value = f"Rs. {payment_detail.total_received():,.2f}"
        lease_value = f"Rs. {payment_detail.lease_amount:,.2f}"
        if is_lease_refund:
            lease_value = f"-Rs. {abs(payment_detail.lease_amount):,.2f}"
            total_value = f"-Rs. {abs(payment_detail.lease_amount):,.2f}"
        elif is_security_refund:
            security_value = f"-Rs. {payment_detail.security_amount:,.2f}"
            total_value = f"-Rs. {payment_detail.security_amount:,.2f}"

        elems = [
            Paragraph(receipt_title, title),
            Paragraph(f"Payment Detail #: {payment_detail.pk}", normal),
            Paragraph(f"Payment Ref: {p.reference_number or p.pk}", normal),
            Paragraph(f"Date: {p.payment_date.strftime('%b %d, %Y')}", normal),
            Spacer(1, 10),
        ]

        meta = [
            ["Tenant", f"{lease.tenant.first_name} {lease.tenant.last_name}".strip()],
            ["Property", lease.unit.property.property_name],
            ["Unit", lease.unit.unit_number],
            ["Payment Method", str(p.payment_method) if p.payment_method else "—"],
        ]
        t = Table(meta, colWidths=[120, 360])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(0xF8F9FA)),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elems += [t, Spacer(1, 12)]

        details = [
            ["Description", "Amount"],
            ["Lease Refund" if is_lease_refund else "Lease Amount", lease_value],
            [security_label, security_value],
            ["Security Type", payment_detail.security_type],
            [total_label, total_value],
        ]
        d = Table(details, colWidths=[300, 180])
        d.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(0xF2F2F2)),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ]))
        elems.append(d)

        elems.append(Spacer(1, 10))
        elems.append(Paragraph(
            f"Generated on {datetime.now().strftime('%b %d, %Y')}",
            ParagraphStyle("f", parent=normal, fontSize=8, textColor=colors.grey)
        ))

        doc.build(elems)
        pdf = buf.getvalue()
        buf.close()

        filename = f"payment_detail_{payment_detail.pk}_{p.payment_date.strftime('%Y%m%d')}.pdf"
        return pdf, filename
