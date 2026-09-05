from accounts.access import restrict_queryset_to_properties
# same helper you use for Cut/Restore

import calendar
import csv
import inspect
import logging
from collections import OrderedDict, defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO

from django import forms
from django.conf import settings
from django.conf import settings as dj_settings  # at top of file

# You will write these
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator

# smart_meter/views.py
# smart_meter/views.py
from django.db import transaction
from django.db.models import (
    BooleanField,
    Case,
    F,
    Max,
    Min,
    OuterRef,
    Q,  # keep if you still use search elsewhere
    Subquery,
    Value,
    When,
)
from django.db.models.functions import Lower, TruncMonth
from django.http import (
    Http404,
    HttpResponse,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.timezone import now
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.currency import format_money
from core.models import GlobalSettings
from leases.models import (
    Lease,  # adjust if lease is in another app
    LeaseUnitOccupancy,
)
from properties.models import Property, Unit  # adjust if different
from smart_meter.forms import MeterPrepaidSettingsForm
from smart_meter.meter_client import send_restore_command  # ✅ we'll add this below

# make sure this is imported
from smart_meter.models import (
    Bill,
    LiveReading,
    Meter,
    MeterBalance,
    MeterEvent,
    MeterCommand,
    MeterPrepaidSettings,
    MeterReading,
    MeterSettings,
)
from smart_meter.services.billing import generate_bill_for_unit
from smart_meter.services.relay_status import (
    parse_authoritative_relay_state,
    reconcile_live_relay_command_state,
)
from smart_meter.status import (
    online_threshold_minutes,
    resolve_meter_online_status,
    resolve_meter_online_statuses,
)

# You will write these
from smart_meter.utils import send_cutoff_command, send_restore_command
from smart_meter.utils.commands import (
    refresh_live,
    request_instant_live_reading,
    send_cutoff_command,
    send_restore_command,
)

# vendor helper
# vendor frame builder
from smart_meter.utils.db_send import send_via_db as _db_send  # fallback sender
from smart_meter.utils.display import attach_active_meter_counts
from smart_meter.utils.messaging import build_whatsapp_url
from smart_meter.utils.tenants import (
    active_tenant_info_for_units,
    attach_active_tenant_names,
    attach_tenant_names_for_dates,
)
from smart_meter.utils.vpn import public_ip, vpn_connected
from smart_meter.services.prepaid_parameters import build_parameter_frame
from smart_meter.vendor.switch_OnOff import (
    RELAY_CLOSE_COMMAND,
    RELAY_OPEN_COMMAND,
    frame_command as build_switch_frame,
)

def _local_date_bounds(day):
    """Return timezone-aware [local midnight, next local midnight) bounds."""
    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime.combine(day, time.min), tz)
    end = timezone.make_aware(datetime.combine(day + timedelta(days=1), time.min), tz)
    return start, end


from .forms import (
    AssignMeterForm,
    CloseMeterInstallationForm,
    InstallMeterToUnitForm,
    MeterCheckGroupForm,
    MeterCheckGroupMembershipForm,
    MeterForm,
    MeterReadingProfileForm,
    MeterReadingForm,
    MeterSettingsForm,
    MoveLeaseUnitForm,
    ReadingManualForm,
    SwitchLabForm,
    SwitchMeterForm,
    UnknownToMeterForm,
)

# smart_meter/views.py
from .models import (
    Bill,
    Lease,
    LiveReading,
    Meter,
    MeterBalance,
    MeterCheckGroup,
    MeterCheckGroupMembership,
    MeterEvent,
    MeterInstallation,
    MeterReading,
    MeterRawFrame,
    MeterRoleHistory,
    Unit,
    UnknownMeter,
)
from .services.billing import generate_bill_for_unit
from .tasks import poll_all_meters

DISABLE_CUTOFFS = getattr(settings, "DISABLE_CUTOFFS", False)

# optional sender (if you have the listener); we fall back gracefully if missing
try:
    # Prefer the control client if it exists and imports cleanly
    from smart_meter.utils.control_client import send_via_db as _control_send
except Exception:
    _control_send = None


send_via_db = _control_send or _db_send
logger = logging.getLogger("meter_control")

if not callable(send_via_db):
    # You can log and set a no-op that returns a clear error for callers
    import logging

    logger = logging.getLogger(__name__)
    logger.error("send_via_db is not callable; control sender unavailable")

    def send_via_db(*args, **kwargs):
        return {"ok": False, "error": "control sender unavailable", "payload": None}

    # --- END robust sender import ---

try:
    logger.info(
        "Active send_via_db line 201: %s.%s",
        getattr(send_via_db, "__module__", "?"),
        getattr(send_via_db, "__name__", "?"),
    )
except Exception:
    pass


# Detect which parameter the active sender supports
try:
    _SEND_SIG = inspect.signature(send_via_db)
    _SUPPORTS_FRAME = "frame" in _SEND_SIG.parameters
    _SUPPORTS_FRAME_HEX = "frame_hex" in _SEND_SIG.parameters
except Exception:
    _SUPPORTS_FRAME = False
    _SUPPORTS_FRAME_HEX = True  # fall back to hex path


def _as_hex(frame):
    try:
        return frame.hex().upper()
    except AttributeError:
        return str(frame).upper()


# universal caller that adapts to the available parameters
_SIG = None
try:
    _SIG = inspect.signature(send_via_db)
except Exception:
    _SIG = None

# Inspect active sender once
try:
    _SEND_SIG = inspect.signature(send_via_db)
    _PARAMS = set(_SEND_SIG.parameters.keys())
except Exception:
    _SEND_SIG = None
    _PARAMS = set()


def _call_send(*, meter_number, frame=None, frame_hex=None, **kwargs):
    """
    Universal sender:
    - Accepts either `frame` (bytes) or `frame_hex` (str).
    - Maps to the active send_via_db signature.
    - Drops unknown kwargs (like allow_switch) safely.
    """
    # Decide which payload param to use
    if "frame" in _PARAMS and frame is not None:
        kwargs["frame"] = frame
    else:
        # We must use hex
        if frame_hex is None:
            if frame is None:
                raise ValueError("Provide either frame or frame_hex")
            frame_hex = _as_hex(frame)
        kwargs["frame_hex"] = frame_hex

    # Keep only kwargs the sender supports
    if _PARAMS:
        kwargs = {
            k: v for k, v in kwargs.items() if k in _PARAMS or k == "meter_number"
        }

    return send_via_db(meter_number=meter_number, **kwargs)


def _send_switch(meter_number, frame, **kwargs):
    """Call send_via_db regardless of whether it wants bytes or hex."""
    if _SUPPORTS_FRAME:
        return send_via_db(meter_number=meter_number, frame=frame, **kwargs)
    # default to hex
    hex_str = frame.hex().upper() if hasattr(frame, "hex") else str(frame)
    # some backends may not accept extra flags; drop ones they likely don't know about
    kwargs = {k: v for k, v in kwargs.items() if k in _SEND_SIG.parameters}
    return send_via_db(meter_number=meter_number, frame_hex=hex_str, **kwargs)


# If you have these helpers; otherwise we’ll just log the event
try:
    from smart_meter.meter_client import send_cutoff_command, send_restore_command
except Exception:
    send_cutoff_command = None
    send_restore_command = None


def assign_meter(request):
    if request.method == "POST":
        form = AssignMeterForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("assign_meter")
    else:
        form = AssignMeterForm()

    return render(request, "smart_meter/assign_meter.html", {"form": form})


def _first_active_meter_for_unit(unit):
    installation = (
        unit.meter_installations.filter(is_active=True, end_date__isnull=True)
        .select_related("meter")
        .order_by("start_date", "id")
        .first()
    )
    return installation.meter if installation else None


# views.py (replace daily_report)


def daily_report(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    meter = _first_active_meter_for_unit(unit)
    if not meter:
        return render(
            request,
            "smart_meter/daily.html",
            {
                "unit": unit,
                "rows": [],
                "chart_labels": [],
                "chart_data": [],
                "start": None,
                "end": None,
            },
        )

    # date range (default last 7 days)
    try:
        start_str = request.GET.get("start")
        end_str = request.GET.get("end")
        if start_str and end_str:
            start = date.fromisoformat(start_str)
            end = date.fromisoformat(end_str)
        else:
            end = timezone.localdate()
            start = end - timedelta(days=6)
    except Exception:
        end = timezone.localdate()
        start = end - timedelta(days=6)

    # Build aware datetimes for the range [start, end+1day)
    tz = timezone.get_current_timezone()
    sdt = timezone.make_aware(datetime.combine(start, time.min), tz)
    edt = timezone.make_aware(datetime.combine(end + timedelta(days=1), time.min), tz)

    # Pull readings sorted by ts
    qs = (
        MeterReading.objects.filter(meter=meter, ts__gte=sdt, ts__lt=edt)
        .order_by("ts")
        .values("ts", "total_energy")
    )

    # Group by LOCAL date and keep min/max per day
    by_day = OrderedDict()  # {date: {"min": Decimal, "max": Decimal}}
    for r in qs:
        ts_local = timezone.localtime(r["ts"], tz)
        d = ts_local.date()
        val = Decimal(str(r["total_energy"] or "0"))
        if d not in by_day:
            by_day[d] = {"min": val, "max": val}
        else:
            by_day[d]["max"] = val

    rows = []
    chart_labels = []
    chart_data = []

    if by_day:
        # Stitch continuity: start of first day = that day's min; subsequent starts = previous day's end
        days = list(by_day.keys())
        first_day = days[0]
        prev_end = by_day[first_day]["min"] or Decimal(0)

        for d in days:
            end_kwh = by_day[d]["max"] if by_day[d]["max"] is not None else prev_end
            start_kwh = prev_end
            usage = end_kwh - start_kwh
            if usage < 0:
                usage = Decimal(0)

            rows.append(
                {
                    "date": d,
                    "start_kwh": start_kwh,
                    "end_kwh": end_kwh,
                    "units": usage,
                }
            )

            chart_labels.append(d.strftime("%b %d, %Y"))
            chart_data.append(float(usage))
            prev_end = end_kwh

    return render(
        request,
        "smart_meter/daily.html",
        {
            "unit": unit,
            "rows": rows,
            "chart_labels": chart_labels,
            "chart_data": chart_data,
            "start": start,
            "end": end,
        },
    )


def monthly_report(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    meter = _first_active_meter_for_unit(unit)
    if not meter:
        return render(request, "smart_meter/monthly.html", {"unit": unit, "rows": []})

    qs = (
        MeterReading.objects.filter(meter=meter)
        .annotate(month=TruncMonth("ts"))
        .values("month")
        .annotate(start_kwh=Min("total_energy"), end_kwh=Max("total_energy"))
        .order_by("month")
    )

    rows = []
    for r in qs:
        if r["start_kwh"] is None or r["end_kwh"] is None:
            continue
        used = Decimal(r["end_kwh"]) - Decimal(r["start_kwh"])
        rows.append(
            {
                "month": r["month"],
                "start_kwh": r["start_kwh"],
                "end_kwh": r["end_kwh"],
                "units": max(used, Decimal("0.000")),
            }
        )
    return render(request, "smart_meter/monthly.html", {"unit": unit, "rows": rows})


def generate_bill_view(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    if request.method == "POST":
        from datetime import date

        month = request.POST.get("month")  # YYYY-MM
        y, m = map(int, month.split("-"))
        # simple month end
        from calendar import monthrange

        period_start = date(y, m, 1)
        period_end = date(y, m, monthrange(y, m)[1])
        bill = generate_bill_for_unit(unit, period_start, period_end)
        return redirect("admin:smart_meter_bill_change", bill.id)
    return render(request, "smart_meter/generate_bill.html", {"unit": unit})


def view_bills(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    bills = Bill.objects.filter(unit=unit).order_by("-period_start")
    return render(request, "smart_meter/bills.html", {"unit": unit, "bills": bills})


def meter_dashboard(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id, is_smart_meter=True)

    # Get latest reading
    latest = MeterReading.objects.filter(unit=unit).order_by("-timestamp").first()

    # Get balance
    balance, _ = MeterBalance.objects.get_or_create(unit=unit)

    # Get tenant phone number (via lease)
    try:
        lease = Lease.objects.filter(unit=unit).latest("start_date")
        tenant = lease.tenant
        phone = tenant.phone  # assumes your tenant model has this
    except:
        tenant = None
        phone = None

    # If balance is low, build WhatsApp alert URL
    wa_url = None
    if balance.balance <= 100 and phone:
        message = f"⚠️ Dear {tenant.name}, your electricity meter balance is ₹{balance.balance}. Please recharge soon to avoid disconnection."
        wa_url = build_whatsapp_url(phone, message)

    # Last 7 days usage for chart
    start_date = now() - timedelta(days=7)
    readings = (
        MeterReading.objects.filter(unit=unit, timestamp__gte=start_date)
        .order_by("timestamp")
        .values("timestamp", "total_energy")
    )

    labels = [r["timestamp"].strftime("%d %b %H:%M") for r in readings]
    values = [float(r["total_energy"] or 0) for r in readings]

    # Monthly total for billing
    current_month = now().replace(day=1)
    month_readings = MeterReading.objects.filter(
        unit=unit, timestamp__gte=current_month
    ).order_by("timestamp")

    start_kwh = month_readings.first().total_energy if month_readings.exists() else 0
    end_kwh = month_readings.last().total_energy if month_readings.exists() else 0
    total_kwh = round((end_kwh or 0) - (start_kwh or 0), 2)

    context = {
        "unit": unit,
        "latest": latest,
        "labels": labels,
        "values": values,
        "total_kwh": total_kwh,
        "peak": latest.peak_hour if latest else False,
        "wa_url": wa_url,  # ✅ pass WhatsApp link to template
        "balance": balance,
    }

    return render(request, "smart_meter/dashboard.html", context)


BILLING_RATE = Decimal("7.50")  # ₹7.50 per kWh


def view_bills(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    bills = Bill.objects.filter(unit=unit).order_by("-month")
    return render(request, "smart_meter/bill_list.html", {"unit": unit, "bills": bills})


# smart_meter/views.py


def meter_status(request, meter_id: int):
    """
    Return transport reachability and measurement freshness independently.
    """
    try:
        meter = Meter.objects.get(pk=meter_id)
    except Meter.DoesNotExist:
        raise Http404("Meter not found")

    minutes = online_threshold_minutes()

    # prefer OneToOne 'live' row if present
    lr = getattr(meter, "live", None)
    ts = lr.ts if isinstance(lr, LiveReading) else None

    status = resolve_meter_online_status(meter, lr)

    return JsonResponse(
        {
            "online": status["is_online"],
            "is_online": status["is_online"],
            "connection_state": status["connection_state"],
            "is_connected": status["is_connected"],
            "measurement_is_fresh": status["measurement_is_fresh"],
            "last_contact_at": _ts_iso(status["last_contact_at"]),
            "last_measurement_at": _ts_iso(status["last_measurement_at"]),
            "last_reading_ts": ts.isoformat() if ts else None,
            "minutes_window": minutes,
        }
    )


def meter_settings(request):
    settings, _ = MeterSettings.objects.get_or_create(id=1)

    if request.method == "POST":
        form = MeterSettingsForm(request.POST, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, "Settings updated.")
    else:
        form = MeterSettingsForm(instance=settings)

    return render(request, "smart_meter/settings.html", {"form": form})


# views.py

# views.py


# smart_meter/views.py (relevant parts)


# smart_meter/views.py


def _meters_annotated_qs(request, online_minutes: int = 10):
    prop_id = (request.GET.get("property") or "").strip()
    unit_id = (request.GET.get("unit") or "").strip()
    meter_id = (request.GET.get("meter") or "").strip()
    q = (request.GET.get("q") or "").strip()
    role = (request.GET.get("role") or "").strip().lower()
    active_filter = (request.GET.get("active") or "").strip().lower()

    qs = Meter.objects.select_related("unit", "unit__property")

    if prop_id:
        qs = qs.filter(unit__property_id=prop_id)
    if unit_id:
        qs = qs.filter(unit_id=unit_id)
    if meter_id:
        qs = qs.filter(id=meter_id)
    if role in (Meter.METER_ROLE_BILLING, Meter.METER_ROLE_CHECK):
        qs = qs.filter(meter_role=role)
    if active_filter == "active":
        qs = qs.filter(is_active=True)
    elif active_filter == "inactive":
        qs = qs.filter(is_active=False)
    if q:
        qs = qs.filter(
            Q(meter_number__icontains=q)
            | Q(name__icontains=q)
            | Q(unit__unit_number__icontains=q)
            | Q(unit__property__property_name__icontains=q)
        )

    # Live row per meter (if duplicates exist, use the newest)
    live_qs = LiveReading.objects.filter(meter=OuterRef("pk")).order_by(
        "-ts"
    )  # if you use ts, change to "-ts"

    # Current balance by unit
    bal_qs = MeterBalance.objects.filter(unit=OuterRef("unit_id")).values("balance")[:1]

    # 👇 annotate from **LiveReading** instead of MeterReading
    qs = qs.annotate(
        balance=Subquery(bal_qs),
        last_ts=Subquery(live_qs.values("ts")[:1]),  # or "ts"
        last_voltage_a=Subquery(live_qs.values("voltage_a")[:1]),
        last_current_a=Subquery(live_qs.values("current_a")[:1]),
        last_total_energy=Subquery(live_qs.values("total_energy")[:1]),
        last_forward_energy=Subquery(live_qs.values("forward_active_energy_kwh")[:1]),
        last_reverse_energy=Subquery(live_qs.values("reverse_active_energy_kwh")[:1]),
        last_status_word=Subquery(live_qs.values("status_word")[:1]),
    )

    # Online/Offline from **live** timestamp
    cutoff_dt = timezone.now() - timedelta(minutes=online_minutes)
    qs = qs.annotate(
        is_online=Case(
            When(last_ts__gte=cutoff_dt, then=Value(True)),
            default=Value(False),
            output_field=BooleanField(),
        ),
        is_cutoff_flag=Case(
            When(
                Q(balance__isnull=False) & Q(balance__lte=F("min_balance_cutoff")),
                then=Value(True),
            ),
            default=Value(False),
            output_field=BooleanField(),
        ),
        is_low=Case(
            When(
                Q(balance__isnull=False) & Q(balance__lt=F("min_balance_alert")),
                then=Value(True),
            ),
            default=Value(False),
            output_field=BooleanField(),
        ),
    )

    return qs.order_by("-last_ts", "meter_number")


SMART_METER_CHIPS = OrderedDict(
    [
        (
            "total",
            {
                "label": "Total",
                "class": "primary",
                "help": "All meters in the current property, unit, meter and role filters.",
            },
        ),
        (
            "active",
            {"label": "Active", "class": "success", "help": "Meters marked active."},
        ),
        (
            "offline",
            {
                "label": "Offline",
                "class": "secondary",
                "help": "No live reading inside the online time window.",
            },
        ),
        (
            "vacant",
            {
                "label": "Vacant",
                "class": "dark",
                "help": "Meters without a currently active lease on their unit.",
            },
        ),
        (
            "negative",
            {
                "label": "Negative",
                "class": "danger",
                "help": "Meters whose unit balance is below zero.",
            },
        ),
        (
            "low",
            {
                "label": "Low Balance",
                "class": "warning",
                "help": "Meters below their configured minimum balance alert.",
            },
        ),
        (
            "cutoff",
            {
                "label": "Cut Off",
                "class": "danger",
                "help": "Meters whose saved power status is off.",
            },
        ),
        (
            "needs_attention",
            {
                "label": "Needs Attention",
                "class": "warning",
                "help": "Inactive, offline, low/negative balance, or cut-off meters.",
            },
        ),
        (
            "billing_issues",
            {
                "label": "Billing Issues",
                "class": "info",
                "help": "Billing-role meters missing a unit/rate, or occupied meters with no reading.",
            },
        ),
        (
            "check",
            {
                "label": "Audit Meters",
                "class": "info",
                "help": "Meters assigned the Audit role.",
            },
        ),
    ]
)


def _normalized_meter_chip(request):
    chip = (request.GET.get("chip") or "active").strip().lower()
    return chip if chip in SMART_METER_CHIPS else "total"


def _with_meter_operational_flags(qs):
    """
    Add reusable flags used by the clickable Smart Meter chips.
    Keep this backend-only so counts, filters, exports and refreshes use one rule.
    """
    from django.db.models import Exists

    today = timezone.localdate()
    active_lease_qs = Lease.objects.filter(
        unit_id=OuterRef("unit_id"),
        status="active",
        start_date__lte=today,
        end_date__gte=today,
    )
    return qs.annotate(has_active_lease=Exists(active_lease_qs))


def _meter_chip_q(chip):
    vacant_q = Q(unit__isnull=True) | Q(has_active_lease=False)
    offline_q = Q(is_online=False)
    negative_q = Q(balance__isnull=False, balance__lt=0)
    low_q = Q(is_low=True)
    cutoff_q = Q(power_status__iexact="off")
    inactive_q = Q(is_active=False)
    billing_issue_q = Q(meter_role=Meter.METER_ROLE_BILLING) & (
        Q(unit__isnull=True) | Q(has_active_lease=True, last_ts__isnull=True)
    )

    if chip == "active":
        return Q(is_active=True)
    if chip == "offline":
        return offline_q
    if chip == "vacant":
        return vacant_q
    if chip == "low":
        return low_q
    if chip == "negative":
        return negative_q
    if chip == "cutoff":
        return cutoff_q
    if chip == "needs_attention":
        return inactive_q | offline_q | low_q | negative_q | cutoff_q
    if chip == "billing_issues":
        return billing_issue_q
    if chip == "check":
        return Q(meter_role=Meter.METER_ROLE_CHECK)
    return Q()


def _apply_meter_chip_filter(qs, chip):
    chip = chip if chip in SMART_METER_CHIPS else "total"
    if chip == "total":
        return qs
    return qs.filter(_meter_chip_q(chip))


def _meter_chip_counts(base_qs):
    from django.db.models import Count

    aggregates = {"total": Count("pk")}
    aggregates.update(
        {
            key: Count("pk", filter=_meter_chip_q(key))
            for key in SMART_METER_CHIPS.keys()
            if key != "total"
        }
    )
    return base_qs.aggregate(**aggregates)


def _meter_chip_cards(request, base_qs, url_name):
    current_chip = _normalized_meter_chip(request)
    counts = _meter_chip_counts(base_qs)
    base_url = reverse(url_name)
    cards = []

    for key, meta in SMART_METER_CHIPS.items():
        qd = request.GET.copy()
        for remove_key in ("chip", "page", "offline"):
            qd.pop(remove_key, None)
        qd["chip"] = key
        query = qd.urlencode()
        cards.append(
            {
                "key": key,
                "label": meta["label"],
                "count": counts.get(key, 0),
                "url": f"{base_url}?{query}" if query else base_url,
                "class": meta["class"],
                "help": meta.get("help", ""),
                "active": current_chip == key,
            }
        )
    return cards


# smart_meter/views.py


def meter_list(request):
    online_minutes = online_threshold_minutes()
    active_filter = (request.GET.get("active") or "").strip().lower()

    # Your existing helper builds the base queryset with flags/filters.
    # Add chip flags/counts before applying the selected chip filter so the chips
    # always show the full count for the current property/unit/meter/search filter.
    meters_base_qs = _with_meter_operational_flags(
        _meters_annotated_qs(request, online_minutes=online_minutes)
    )
    meters_base_qs = restrict_queryset_to_properties(meters_base_qs, request.user, "unit__property")
    current_chip = _normalized_meter_chip(request)
    if active_filter in {"active", "inactive"} and "chip" not in request.GET:
        current_chip = "total"
    chip_cards = _meter_chip_cards(request, meters_base_qs, "smart_meter:meter_list")
    meters_qs = _apply_meter_chip_filter(meters_base_qs, current_chip)

    # ✅ Make it efficient for template access (no N+1)
    meters_qs = meters_qs.select_related("unit", "unit__property").only(
        "id",
        "meter_number",
        "name",
        "billing_mode",
        "meter_role",
        "unit_rate",
        "min_balance_alert",
        "min_balance_cutoff",
        "is_active",
        "installed_at",
        "power_status",
        "unit_id",
        "unit__id",
        "unit__property_id",
        "unit__unit_number",
        "unit__electricity_unit_rate",
        "unit__property__id",
        "unit__property__property_name",
        "unit__property__electricity_unit_rate",
    )

    # (Optional) ensure each row has a balance value for scripts like "auto-select negative"
    latest_balance = LiveReading.objects.filter(meter=OuterRef("pk")).values("balance")[
        :1
    ]
    meters_qs = meters_qs.annotate(balance=Subquery(latest_balance))

    # ---- last LiveReading per meter (no fragile reverse join) ----
    latest_lr = LiveReading.objects.filter(meter_id=OuterRef("pk")).order_by("-id")

    meters_qs = meters_qs.annotate(
        # fields for default & tiebreak ordering
        prop_name=Lower("unit__property__property_name"),
        unit_num=F("unit__unit_number"),
        meter_num=F("meter_number"),
        power_val=Subquery(latest_lr.values("total_power")[:1]),
        last_read_at=Subquery(latest_lr.values("ts")[:1]),
    )

    # ---- determine sort target from query string; default: property → unit ----
    sort = (request.GET.get("sort") or "property").lower()
    dir_ = (request.GET.get("dir") or "asc").lower()
    sort_map = {
        "property": "prop_name",
        "unit": "unit_num",
        "meter": "meter_num",
        "power": "power_status",
        "last": "last_ts",
    }
    order_field = sort_map.get(sort, "prop_name")
    if dir_ == "desc":
        order_field = "-" + order_field

    # main order + stable tie-breakers
    meters_qs = meters_qs.order_by(order_field, "prop_name", "unit_num", "meter_num")

    # ---- build header links here (so template doesn’t need parentheses/logic) ----
    qd = request.GET.copy()
    for k in ["sort", "dir", "page"]:
        qd.pop(k, None)
    base_qs = qd.urlencode()

    def link_for(col):
        next_dir = "desc" if (sort == col and dir_ == "asc") else "asc"
        return (
            f"?{base_qs}&sort={col}&dir={next_dir}"
            if base_qs
            else f"?sort={col}&dir={next_dir}"
        )

    header_links = {
        "meter": link_for("meter"),
        "property": link_for("property"),
        "unit": link_for("unit"),
        "power": link_for("power"),
        "last": link_for("last"),
    }
    paginator = Paginator(meters_qs, 50)
    page_obj = paginator.get_page(request.GET.get("page"))
    page_obj.object_list = attach_active_meter_counts(page_obj.object_list)
    from smart_meter.rates import resolve_electricity_rate

    page_statuses = resolve_meter_online_statuses(
        (meter, meter.last_ts) for meter in page_obj.object_list
    )
    for meter in page_obj.object_list:
        meter_status = page_statuses[meter.pk]
        meter.is_online = meter_status["is_online"]
        meter.is_connected = meter_status["is_connected"]
        meter.measurement_is_fresh = meter_status["measurement_is_fresh"]
        meter.connection_state = meter_status["connection_state"]
        meter.last_contact_at = meter_status["last_contact_at"]
        meter.display_electricity_rate = resolve_electricity_rate(meter=meter)
        meter.confirmed_relay_state = parse_authoritative_relay_state(
            meter.last_status_word
        )
        meter.relay_confirmed_at = meter.last_ts if meter.confirmed_relay_state else None
    attach_active_tenant_names(
        page_obj.object_list,
        lambda meter: meter.unit_id,
    )
    total_count = paginator.count

    ctx = {
        "meters": page_obj,
        "page_obj": page_obj,
        "paginator": paginator,
        "total_count": total_count,
        # ...your existing filters...
        "online_minutes": online_minutes,
        "chip_cards": chip_cards,
        "current_chip": current_chip,
    }

    # add to context (before render)
    ctx.update(
        {
            "current_sort": sort,
            "current_dir": dir_,
            "header_links": header_links,
            "base_qs": base_qs,
        }
    )

    try:
        (
            selected_meters,
            all_properties,
            filtered_units,
            filtered_meters,
            prop_id,
            unit_id,
            meter_param,
        ) = _filtered_meter_sets(request)
    except NameError:
        # Fallback (if helper not present in your repo)
        prop_id = (request.GET.get("property") or "").strip()
        unit_id = (request.GET.get("unit") or "").strip()
        meter_param = (request.GET.get("meter") or "").strip()

        from properties.models import Property, Unit

        all_properties = Property.objects.all().order_by("property_name")

        units_qs = Unit.objects.all()
        if prop_id:
            units_qs = units_qs.filter(property_id=prop_id)
        filtered_units = units_qs.order_by("unit_number")

        meters_qs = Meter.objects.select_related("unit", "unit__property")
        if unit_id:
            meters_qs = meters_qs.filter(unit_id=unit_id)
        elif prop_id:
            meters_qs = meters_qs.filter(unit__property_id=prop_id)
        filtered_meters = meters_qs.order_by("meter_number")

    # Keep the search box value in the UI
    q = (request.GET.get("q") or "").strip()

    # Expose everything the filter partial needs
    ctx.update(
        {
            "all_properties": all_properties,
            "filtered_units": filtered_units,
            "filtered_meters": filtered_meters,
            "current_property": prop_id,
            "current_unit": unit_id,
            "current_meter": meter_param,
            "current_role": (request.GET.get("role") or "").strip().lower(),
            "current_active": active_filter,
            "q": q,
            # Backward-compat alias if the partial uses a different key
            "properties": all_properties,
        }
    )

    return render(request, "smart_meter/meter_list.html", ctx)


def add_meter(request):
    if request.method == "POST":
        form = MeterForm(request.POST)
        if form.is_valid():
            meter = form.save()
            from smart_meter.models import MeterAssignmentHistory

            MeterRoleHistory.objects.create(
                meter=meter,
                role=meter.meter_role,
                start_date=timezone.localdate(),
                is_active=True,
                changed_by=request.user if request.user.is_authenticated else None,
                reason="Meter created.",
            )
            MeterAssignmentHistory.objects.create(
                meter=meter,
                unit=meter.unit,
                lease=meter.current_lease,
                old_meter=None,
                new_meter=meter,
                old_unit=None,
                new_unit=meter.unit,
                old_lease=None,
                new_lease=meter.current_lease,
                changed_by=request.user if request.user.is_authenticated else None,
                notes="Meter created and assigned.",
            )
            return redirect("smart_meter:meter_list")
    else:
        form = MeterForm()

    return render(request, "smart_meter/meter_form.html", {"form": form})


@transaction.atomic
def meter_edit(request, pk):
    meter = get_object_or_404(Meter, pk=pk)
    old_unit = meter.unit
    old_lease = meter.current_lease
    old_role = meter.meter_role
    if request.method == "POST":
        form = MeterForm(request.POST, instance=meter)
        if form.is_valid():
            new_role = form.cleaned_data["meter_role"]
            form.instance.meter_role = old_role
            if new_role != old_role:
                try:
                    with transaction.atomic():
                        effective_date = timezone.localdate()
                        if (
                            old_role == Meter.METER_ROLE_CHECK
                            and new_role == Meter.METER_ROLE_BILLING
                        ):
                            group = (
                                MeterCheckGroup.objects.select_for_update()
                                .filter(check_meter=meter)
                                .first()
                            )
                            if group:
                                has_active_memberships = group.memberships.filter(
                                    is_active=True,
                                    end_date__isnull=True,
                                ).exists()
                                if has_active_memberships:
                                    group.check_meter = form.cleaned_data["replacement_check_meter"]
                                    group.save(update_fields=["check_meter"])
                                elif group.is_active:
                                    group.is_active = False
                                    group.save(update_fields=["is_active"])
                        elif (
                            old_role == Meter.METER_ROLE_BILLING
                            and new_role == Meter.METER_ROLE_CHECK
                        ):
                            memberships = list(
                                meter.check_group_memberships.select_for_update().filter(
                                    is_active=True,
                                    end_date__isnull=True,
                                )
                            )
                            for membership in memberships:
                                membership.close(
                                    end_date=max(effective_date, membership.start_date),
                                    notes="Ended automatically because the meter role changed to Audit.",
                                )
                        meter.change_role(
                            new_role,
                            effective_date=effective_date,
                            user=request.user if request.user.is_authenticated else None,
                            reason=request.POST.get("notes", ""),
                        )
                except ValidationError as exc:
                    if hasattr(exc, "message_dict"):
                        error_messages = [
                            message
                            for messages_list in exc.message_dict.values()
                            for message in messages_list
                        ]
                    else:
                        error_messages = exc.messages
                    for error_message in error_messages:
                        form.add_error("meter_role", error_message)
                    return render(
                        request,
                        "smart_meter/meter_form.html",
                        {"form": form, "edit": True, "original_meter_role": old_role},
                    )
                messages.success(
                    request,
                    f"Meter role changed to {meter.get_meter_role_display()} and history recorded.",
                )
            meter = form.save()
            if old_unit_id := getattr(old_unit, "id", None):
                unit_changed = old_unit_id != meter.unit_id
            else:
                unit_changed = True
            if unit_changed:
                from smart_meter.models import MeterAssignmentHistory

                effective_date = timezone.localdate()
                active_installation = (
                    MeterInstallation.objects.select_for_update()
                    .filter(meter=meter, is_active=True, end_date__isnull=True)
                    .first()
                )
                if active_installation:
                    effective_date = max(effective_date, active_installation.start_date)

                _effective_start, effective_end = _local_date_bounds(effective_date)
                latest_reading = (
                    MeterReading.objects.filter(meter=meter, ts__lt=effective_end)
                    .order_by("-ts", "-id")
                    .first()
                )
                transfer_reading = (
                    latest_reading.total_energy
                    if latest_reading and latest_reading.total_energy is not None
                    else Decimal(0)
                )
                assignment_note = (
                    request.POST.get("notes", "") or ""
                ).strip() or "Meter reassigned through meter edit."

                if active_installation:
                    active_installation.close(
                        end_date=effective_date,
                        end_reading=transfer_reading,
                        notes=assignment_note,
                    )

                new_lease = None
                if meter.unit_id:
                    new_lease = (
                        Lease.objects.filter(
                            unit_id=meter.unit_id,
                            status="active",
                            start_date__lte=effective_date,
                            end_date__gte=effective_date,
                        )
                        .order_by("-start_date", "-id")
                        .first()
                        or Lease.objects.filter(unit_id=meter.unit_id, status="active")
                        .order_by("-start_date", "-id")
                        .first()
                    )
                    MeterInstallation.objects.create(
                        meter=meter,
                        unit=meter.unit,
                        lease=new_lease,
                        start_date=effective_date,
                        start_reading=transfer_reading,
                        installed_by=(
                            request.user if request.user.is_authenticated else None
                        ),
                        reason="Meter reassigned through meter edit",
                        notes=assignment_note,
                    )

                MeterAssignmentHistory.objects.create(
                    meter=meter,
                    unit=meter.unit,
                    lease=new_lease,
                    old_meter=meter,
                    new_meter=meter,
                    old_unit=old_unit,
                    new_unit=meter.unit,
                    old_lease=old_lease,
                    new_lease=new_lease,
                    changed_by=request.user if request.user.is_authenticated else None,
                    notes=assignment_note,
                )
                messages.success(
                    request,
                    "Meter assignment and installation history updated.",
                )
            return redirect("smart_meter:meter_detail", pk=meter.pk)
    else:
        form = MeterForm(instance=meter)
    return render(
        request,
        "smart_meter/meter_form.html",
        {"form": form, "edit": True, "original_meter_role": old_role},
    )


def _raw_frame_history_queryset(meter, search_term=""):
    """Return the safe, searchable audit-frame queryset for one meter."""
    frames = MeterRawFrame.objects.filter(meter=meter).only(
        "id", "received_at", "source_ip", "source_port", "control_code",
        "data_identifier", "data_length", "raw_frame_hex", "checksum_style",
        "decoded_data", "trust_classification", "parser_version",
    )
    if search_term:
        frames = frames.filter(
            Q(data_identifier__icontains=search_term)
            | Q(source_ip__icontains=search_term)
            | Q(raw_frame_hex__icontains=search_term)
        )
    return frames.order_by("-received_at", "-id")


def _raw_frame_decoded_value(decoded_data, *names):
    """Accept parser aliases without inventing a reading when none was decoded."""
    decoded_data = decoded_data or {}
    for name in names:
        value = decoded_data.get(name)
        if value not in (None, ""):
            return value
    return None


def _prepare_raw_frame_for_display(frame):
    decoded_data = frame.decoded_data or {}
    frame.reading_values = {
        "balance": _raw_frame_decoded_value(decoded_data, "balance"),
        "forward_energy": _raw_frame_decoded_value(
            decoded_data,
            "forward_active_energy_kwh",
            "forward_active_energy",
            "total_energy",
        ),
        "reverse_energy": _raw_frame_decoded_value(
            decoded_data,
            "reverse_active_energy_kwh",
            "reverse_active_energy",
            "reverse_energy",
        ),
        "total_power": _raw_frame_decoded_value(decoded_data, "total_power"),
        "pf_total": _raw_frame_decoded_value(decoded_data, "pf_total"),
        "voltage": _raw_frame_decoded_value(decoded_data, "voltage_a"),
        "current": _raw_frame_decoded_value(decoded_data, "current_a"),
        "status_word": _raw_frame_decoded_value(decoded_data, "status_word"),
        "voltage_a": _raw_frame_decoded_value(decoded_data, "voltage_a"),
        "voltage_b": _raw_frame_decoded_value(decoded_data, "voltage_b"),
        "voltage_c": _raw_frame_decoded_value(decoded_data, "voltage_c"),
        "current_a": _raw_frame_decoded_value(decoded_data, "current_a"),
        "current_b": _raw_frame_decoded_value(decoded_data, "current_b"),
        "current_c": _raw_frame_decoded_value(decoded_data, "current_c"),
        "power_a": _raw_frame_decoded_value(decoded_data, "power_a"),
        "power_b": _raw_frame_decoded_value(decoded_data, "power_b"),
        "power_c": _raw_frame_decoded_value(decoded_data, "power_c"),
    }
    return frame


@login_required
@permission_required("smart_meter.view_raw_dlt645_frames", raise_exception=True)
def meter_raw_frame_history(request, pk):
    """Read-only, searchable audit history for the meter's persisted frames."""
    meter = get_object_or_404(
        Meter.objects.select_related("unit", "unit__property"),
        pk=pk,
    )
    search_term = (request.GET.get("q") or "").strip()
    frames = _raw_frame_history_queryset(meter, search_term)
    page_obj = Paginator(frames, 50).get_page(request.GET.get("page"))
    for frame in page_obj:
        _prepare_raw_frame_for_display(frame)
    return render(
        request,
        "smart_meter/meter_raw_frame_history.html",
        {
            "meter": meter,
            "page_obj": page_obj,
            "search_term": search_term,
            "is_three_phase": (
                meter.reading_profile == Meter.READING_PROFILE_TOTAL_AND_PER_PHASE
            ),
        },
    )


@login_required
@permission_required("smart_meter.view_raw_dlt645_frames", raise_exception=True)
def meter_raw_frame_history_xlsx(request, pk):
    """Export exactly the currently filtered raw-reading audit history."""
    meter = get_object_or_404(
        Meter.objects.select_related("unit", "unit__property"), pk=pk
    )
    search_term = (request.GET.get("q") or "").strip()
    is_three_phase = meter.reading_profile == Meter.READING_PROFILE_TOTAL_AND_PER_PHASE

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Raw Readings"
    headers = [
        "Received",
        "Meter #",
        "Current Unit",
        "Property",
        "Balance",
        "Forward Energy (kWh)",
        "Reverse Energy (kWh)",
        "Voltage (V)",
        "Current (A)",
        "Total Power (kW)",
        "PF Total",
        "Relay Status",
        "Data Identifier",
        "Control Code",
        "Source IP",
        "Source Port",
        "Trust",
        "Parser Version",
        "Raw Frame Hex",
    ]
    if is_three_phase:
        headers[8:8] = [
            "Voltage B (V)", "Voltage C (V)",
            "Current B (A)", "Current C (A)",
            "Power A (kW)", "Power B (kW)", "Power C (kW)",
        ]
    worksheet.append(headers)

    unit_name = meter.display_location_name or ""
    property_name = getattr(getattr(meter.unit, "property", None), "property_name", "")
    for frame in _raw_frame_history_queryset(meter, search_term).iterator(chunk_size=500):
        values = _prepare_raw_frame_for_display(frame).reading_values
        row = [
            timezone.localtime(frame.received_at).strftime("%Y-%m-%d %H:%M:%S"),
            meter.meter_number,
            unit_name,
            property_name,
            values["balance"],
            values["forward_energy"],
            values["reverse_energy"],
            values["voltage"],
            values["current"],
            values["total_power"],
            values["pf_total"],
            values["status_word"],
            frame.data_identifier,
            f"0x{frame.control_code:02X}",
            frame.source_ip,
            frame.source_port,
            frame.get_trust_classification_display(),
            frame.parser_version,
            frame.raw_frame_hex,
        ]
        if is_three_phase:
            row[8:8] = [
                values["voltage_b"], values["voltage_c"],
                values["current_b"], values["current_c"],
                values["power_a"], values["power_b"], values["power_c"],
            ]
        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = cell.font.copy(bold=True)
    for column in worksheet.columns:
        letter = get_column_letter(column[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[letter].width = min(max(12, max_length + 2), 42)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="raw_readings_{meter.meter_number}_{timezone.now():%Y%m%d_%H%M%S}.xlsx"'
    )
    return response


@require_POST
@login_required
def meter_role_update(request, pk):
    meter = get_object_or_404(Meter, pk=pk)
    new_role = (request.POST.get("meter_role") or "").strip().lower()
    if new_role not in (Meter.METER_ROLE_BILLING, Meter.METER_ROLE_CHECK):
        return JsonResponse(
            {"success": False, "error": "Select a valid meter role."}, status=400
        )

    try:
        with transaction.atomic():
            if (
                meter.meter_role == Meter.METER_ROLE_CHECK
                and new_role == Meter.METER_ROLE_BILLING
            ):
                group = (
                    MeterCheckGroup.objects.select_for_update()
                    .filter(check_meter=meter, is_active=True)
                    .first()
                )
                if group and not group.memberships.filter(
                    is_active=True,
                    end_date__isnull=True,
                ).exists():
                    group.is_active = False
                    group.save(update_fields=["is_active"])
            if (
                meter.meter_role == Meter.METER_ROLE_BILLING
                and new_role == Meter.METER_ROLE_CHECK
            ):
                memberships = list(
                    meter.check_group_memberships.select_for_update().filter(
                        is_active=True,
                        end_date__isnull=True,
                    )
                )
                for membership in memberships:
                    membership.close(
                        end_date=max(timezone.localdate(), membership.start_date),
                        notes="Ended automatically by inline role change to Audit.",
                    )
            meter.change_role(
                new_role,
                effective_date=timezone.localdate(),
                user=request.user,
                reason="Inline role update.",
            )
    except ValidationError as exc:
        if hasattr(exc, "message_dict"):
            error = " ".join(
                message
                for messages_list in exc.message_dict.values()
                for message in messages_list
            )
        else:
            error = " ".join(exc.messages)
        return JsonResponse({"success": False, "error": error}, status=400)
    except Exception:
        logger.exception(
            "Meter role update failed for meter_id=%s role=%s user=%s",
            meter.pk,
            new_role,
            getattr(request.user, "username", ""),
        )
        return JsonResponse(
            {
                "success": False,
                "error": "The role could not be saved. Please refresh and try again.",
            },
            status=500,
        )

    meter.refresh_from_db(fields=["meter_role"])

    return JsonResponse(
        {
            "success": True,
            "meter_id": meter.pk,
            "role": meter.meter_role,
            "label": meter.get_meter_role_display(),
        }
    )


@login_required
def meter_check_group_list(request):
    from django.db.models import Count

    today = timezone.localdate()
    show_linked = request.GET.get("show_linked") == "1"
    group_queryset = MeterCheckGroup.objects.all()
    if not show_linked:
        group_queryset = group_queryset.filter(superseded_by_energy_system__isnull=True)
    groups = list(
        group_queryset.select_related(
            "property",
            "check_meter",
            "check_meter__unit",
            "check_meter__unit__property",
            "superseded_by_energy_system",
            "energy_system",
        )
        .annotate(
            membership_record_count=Count("memberships", distinct=True),
            active_billing_meter_count=Count(
                "memberships",
                filter=(
                    Q(memberships__is_active=True)
                    & Q(memberships__start_date__lte=today)
                    & (
                        Q(memberships__end_date__isnull=True)
                        | Q(memberships__end_date__gte=today)
                    )
                    & Q(memberships__billing_meter__meter_role=Meter.METER_ROLE_BILLING)
                    & Q(memberships__billing_meter__is_active=True)
                ),
                distinct=True,
            )
        )
        .order_by("property__property_name", "name")
    )
    coverage_by_group = defaultdict(set)
    coverage_rows = MeterCheckGroupMembership.objects.filter(
        group_id__in=[group.pk for group in groups],
        is_active=True,
        start_date__lte=today,
        billing_meter__is_active=True,
    ).filter(
        Q(end_date__isnull=True) | Q(end_date__gte=today)
    ).values_list("group_id", "billing_meter__unit__property__property_name")
    for group_id, property_name in coverage_rows:
        if property_name:
            coverage_by_group[group_id].add(property_name)
    for group in groups:
        group.covered_property_names = ", ".join(sorted(coverage_by_group[group.pk]))
    return render(
        request,
        "smart_meter/check_group_list.html",
        {"groups": groups, "show_linked": show_linked},
    )


@require_POST
@login_required
def meter_check_group_name_update(request, pk):
    name = (request.POST.get("name") or "").strip()
    if not name:
        return JsonResponse(
            {"success": False, "error": "Check Group name is required."}, status=400
        )
    if len(name) > MeterCheckGroup._meta.get_field("name").max_length:
        return JsonResponse(
            {"success": False, "error": "Check Group name cannot exceed 100 characters."},
            status=400,
        )

    with transaction.atomic():
        group = get_object_or_404(
            MeterCheckGroup.objects.select_for_update(), pk=pk
        )
        group.name = name
        group.save(update_fields=["name"])
    return JsonResponse({"success": True, "name": group.name})


def _check_group_active_membership_count(group_id, as_of=None):
    as_of = as_of or timezone.localdate()
    return (
        MeterCheckGroupMembership.objects.filter(
            group_id=group_id,
            is_active=True,
            start_date__lte=as_of,
            billing_meter__is_active=True,
            billing_meter__meter_role=Meter.METER_ROLE_BILLING,
        )
        .filter(Q(end_date__isnull=True) | Q(end_date__gte=as_of))
        .count()
    )


def _check_group_delete_info(group):
    memberships = list(
        group.memberships.select_related(
            "billing_meter",
            "billing_meter__unit",
            "billing_meter__unit__property",
        ).order_by(
            "billing_meter__unit__property__property_name",
            "billing_meter__unit__unit_number",
            "billing_meter__meter_number",
            "start_date",
        )
    )
    attach_active_meter_counts(memberships, lambda membership: membership.billing_meter)
    targets = (
        MeterCheckGroup.objects.filter(
            is_active=True,
            check_meter__is_active=True,
            check_meter__meter_role=Meter.METER_ROLE_CHECK,
        )
        .exclude(pk=group.pk)
        .select_related("check_meter")
        .order_by("check_meter__meter_number", "name")
    )
    return {
        "group": {"id": group.pk, "name": group.name},
        "memberships": [
            {
                "id": membership.pk,
                "meter_number": membership.billing_meter.meter_number,
                "location": (
                    f"{membership.billing_meter.unit.property.property_name} / "
                    f"{membership.billing_meter.display_location_name}"
                    if membership.billing_meter.unit_id
                    else "Unassigned"
                ),
                "start_date": membership.start_date.isoformat(),
                "end_date": membership.end_date.isoformat() if membership.end_date else None,
                "status": "Active" if membership.is_active else "Ended",
            }
            for membership in memberships
        ],
        "targets": [
            {
                "id": target.pk,
                "label": f"{target.check_meter.meter_number} — {target.name}",
            }
            for target in targets
        ],
        "membership_count": len(memberships),
        "active_membership_count": _check_group_active_membership_count(group.pk),
    }


@login_required
def meter_check_group_delete_manage(request, pk):
    group = get_object_or_404(MeterCheckGroup, pk=pk)
    if request.method == "GET":
        return JsonResponse({"success": True, **_check_group_delete_info(group)})
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Unsupported request."}, status=405)

    action = (request.POST.get("action") or "").strip().lower()
    if action == "delete":
        with transaction.atomic():
            locked_group = MeterCheckGroup.objects.select_for_update().get(pk=group.pk)
            membership_ids = list(
                locked_group.memberships.select_for_update().values_list("pk", flat=True)
            )
            if membership_ids:
                return JsonResponse(
                    {
                        "success": False,
                        "error": (
                            f"Reassign all {len(membership_ids)} membership record(s) "
                            "before deleting this Check Group."
                        ),
                    },
                    status=409,
                )
            group_name = locked_group.name
            locked_group.delete()
        return JsonResponse(
            {"success": True, "deleted": True, "message": f'Check Group "{group_name}" was permanently deleted.'}
        )

    if action not in {"reassign", "reassign_all"}:
        return JsonResponse({"success": False, "error": "Select a valid action."}, status=400)

    try:
        target_group_id = int(request.POST.get("target_group_id") or 0)
    except (TypeError, ValueError):
        target_group_id = 0
    if not target_group_id or target_group_id == group.pk:
        return JsonResponse(
            {"success": False, "error": "Select another Audit meter."}, status=400
        )

    try:
        with transaction.atomic():
            source_group = MeterCheckGroup.objects.select_for_update().get(pk=group.pk)
            target_group = MeterCheckGroup.objects.select_for_update().get(
                pk=target_group_id,
                is_active=True,
                check_meter__is_active=True,
                check_meter__meter_role=Meter.METER_ROLE_CHECK,
            )
            memberships = source_group.memberships.select_for_update()
            if action == "reassign":
                try:
                    membership_id = int(request.POST.get("membership_id") or 0)
                except (TypeError, ValueError):
                    membership_id = 0
                memberships = memberships.filter(pk=membership_id)
            memberships = list(memberships.order_by("start_date", "pk"))
            if not memberships:
                return JsonResponse(
                    {"success": False, "error": "The selected membership no longer exists."},
                    status=404,
                )
            for membership in memberships:
                membership.group = target_group
                audit_note = (
                    f"Reassigned from Check Group {source_group.name} to "
                    f"{target_group.name} by {request.user} on {timezone.localdate().isoformat()}."
                )
                membership.notes = (membership.notes + "\n" + audit_note).strip()
                membership.save(update_fields=["group", "notes"])
    except MeterCheckGroup.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": "The selected target Audit meter is unavailable."},
            status=400,
        )
    except ValidationError as exc:
        if hasattr(exc, "message_dict"):
            error = " ".join(
                message
                for field_messages in exc.message_dict.values()
                for message in field_messages
            )
        else:
            error = " ".join(exc.messages)
        return JsonResponse({"success": False, "error": error}, status=400)

    return JsonResponse(
        {
            "success": True,
            "message": (
                f"{len(memberships)} membership record(s) reassigned to "
                f"Audit meter {target_group.check_meter.meter_number}."
            ),
            "counts": {
                str(source_group.pk): _check_group_active_membership_count(source_group.pk),
                str(target_group.pk): _check_group_active_membership_count(target_group.pk),
            },
            **_check_group_delete_info(source_group),
        }
    )


@login_required
def meter_check_group_form(request, pk=None):
    group = get_object_or_404(MeterCheckGroup, pk=pk) if pk else None
    if request.method == "POST":
        form = MeterCheckGroupForm(request.POST, instance=group)
        if form.is_valid():
            group = form.save()
            messages.success(request, "Check group saved.")
            return redirect("smart_meter:meter_check_group_detail", pk=group.pk)
    else:
        form = MeterCheckGroupForm(instance=group)
    return render(
        request,
        "smart_meter/check_group_form.html",
        {
            "form": form,
            "group": group,
        },
    )


@login_required
def meter_check_group_detail(request, pk):
    group = get_object_or_404(
        MeterCheckGroup.objects.select_related(
            "property",
            "check_meter",
            "check_meter__unit",
            "check_meter__unit__property",
        ),
        pk=pk,
    )
    today = timezone.localdate()
    start_date = today.replace(day=1)
    end_date = today
    quick_range = (request.GET.get("range") or "").strip().lower()
    if quick_range == "this_month":
        start_date = today.replace(day=1)
    elif quick_range == "last_month":
        end_date = today.replace(day=1) - timedelta(days=1)
        start_date = end_date.replace(day=1)
    elif quick_range == "this_week":
        start_date = today - timedelta(days=today.weekday())
    elif quick_range == "last_week":
        end_date = today - timedelta(days=today.weekday() + 1)
        start_date = end_date - timedelta(days=6)
    else:
        quick_range = ""
        try:
            if request.GET.get("start"):
                start_date = date.fromisoformat(request.GET["start"])
            if request.GET.get("end"):
                end_date = date.fromisoformat(request.GET["end"])
        except ValueError:
            messages.warning(request, "Invalid date range; the current month is shown.")
    if end_date < start_date:
        start_date, end_date = end_date, start_date

    if request.method == "POST":
        membership_form = MeterCheckGroupMembershipForm(request.POST, group=group)
        if membership_form.is_valid():
            membership_form.save()
            messages.success(request, "Billing meter added to the check group.")
            return redirect("smart_meter:meter_check_group_detail", pk=group.pk)
    else:
        membership_form = MeterCheckGroupMembershipForm(
            group=group,
            initial={"start_date": today},
        )

    from smart_meter.services.reconciliation import calculate_check_group_period

    reconciliation = calculate_check_group_period(group, start_date, end_date)
    check_rows = reconciliation["check_rows"]
    billing_rows = reconciliation["billing_rows"]
    effective_memberships = reconciliation["effective_memberships"]
    audit_summary_start_kwh = reconciliation["audit_summary_start_kwh"]
    audit_summary_end_kwh = reconciliation["audit_summary_end_kwh"]
    check_kwh = reconciliation["check_kwh"]
    billing_kwh = reconciliation["billing_kwh"]
    variance_kwh = reconciliation["variance_kwh"]
    variance_rs = reconciliation["variance_rs"]
    leakage_percent = reconciliation["leakage_percent"]

    memberships = list(group.memberships.select_related(
        "billing_meter", "billing_meter__unit", "billing_meter__unit__property"
    ).order_by(
        "billing_meter__unit__property__property_name",
        "billing_meter__unit__unit_number",
        "billing_meter__meter_number",
        "-is_active",
    ))
    attach_active_meter_counts(memberships, lambda membership: membership.billing_meter)
    _report_start, report_end_exclusive = _local_date_bounds(end_date)
    report_target_at = report_end_exclusive - timedelta(microseconds=1)

    def closest_reading_at(target_at, previous_at, next_at):
        candidates = [value for value in (previous_at, next_at) if value is not None]
        if not candidates:
            return None
        return min(candidates, key=lambda value: abs(value - target_at))

    audit_previous_reading_at = (
        MeterReading.objects.filter(
            meter_id=group.check_meter_id,
            ts__lte=report_target_at,
        )
        .order_by("-ts", "-id")
        .values_list("ts", flat=True)
        .first()
    )
    audit_next_reading_at = (
        MeterReading.objects.filter(
            meter_id=group.check_meter_id,
            ts__gt=report_target_at,
        )
        .order_by("ts", "id")
        .values_list("ts", flat=True)
        .first()
    )
    audit_last_reading_at = closest_reading_at(
        report_target_at, audit_previous_reading_at, audit_next_reading_at
    )
    audit_last_total_energy = None
    if audit_last_reading_at:
        audit_last_total_energy = (
            MeterReading.objects.filter(
                meter_id=group.check_meter_id,
                ts=audit_last_reading_at,
            )
            .order_by("-id")
            .values_list("total_energy", flat=True)
            .first()
        )
    reconciliation_target_at = audit_last_reading_at or report_target_at
    previous_ts_subquery = (
        MeterReading.objects.filter(
            meter_id=OuterRef("pk"),
            ts__lte=reconciliation_target_at,
        )
        .order_by("-ts", "-id")
        .values("ts")[:1]
    )
    next_ts_subquery = (
        MeterReading.objects.filter(
            meter_id=OuterRef("pk"),
            ts__gt=reconciliation_target_at,
        )
        .order_by("ts", "id")
        .values("ts")[:1]
    )
    membership_meter_ids = {
        membership.billing_meter_id for membership in memberships
    }
    reading_candidate_rows = (
        Meter.objects.filter(pk__in=membership_meter_ids)
        .annotate(
            report_previous_reading_at=Subquery(previous_ts_subquery),
            report_next_reading_at=Subquery(next_ts_subquery),
        )
        .values_list(
            "pk", "report_previous_reading_at", "report_next_reading_at"
        )
    )
    reading_candidates_by_meter = {
        meter_id: (previous_at, next_at)
        for meter_id, previous_at, next_at in reading_candidate_rows
    }

    def reading_freshness(reading_at, target_at):
        if reading_at is None:
            return "No reading", "bg-danger"
        delta_seconds = (reading_at - target_at).total_seconds()
        absolute_minutes = max(0, round(abs(delta_seconds) / 60))
        if absolute_minutes == 0:
            label = "On target"
        elif absolute_minutes < 60:
            label = f"{absolute_minutes}m"
        elif absolute_minutes < 1440:
            hours, minutes = divmod(absolute_minutes, 60)
            label = f"{hours}h" + (f" {minutes}m" if minutes else "")
        else:
            days, remaining_minutes = divmod(absolute_minutes, 1440)
            hours = remaining_minutes // 60
            label = f"{days}d" + (f" {hours}h" if hours else "")
        if absolute_minutes:
            label = f"{label} {'after' if delta_seconds > 0 else 'before'}"
        if absolute_minutes <= 60:
            css_class = "bg-success"
        elif absolute_minutes <= 1440:
            css_class = "bg-warning text-dark"
        else:
            css_class = "bg-danger"
        return label, css_class

    for membership in memberships:
        reading_at = closest_reading_at(
            reconciliation_target_at,
            *reading_candidates_by_meter.get(membership.billing_meter_id, (None, None)),
        )
        membership.billing_meter.report_last_reading_at = reading_at
        membership.billing_meter.report_last_reading_date = (
            timezone.localtime(reading_at).date() if reading_at else None
        )
        (
            membership.billing_meter.report_reading_status,
            membership.billing_meter.report_reading_status_class,
        ) = reading_freshness(reading_at, reconciliation_target_at)

    audit_reading_status, audit_reading_status_class = reading_freshness(
        audit_last_reading_at, report_target_at
    )

    chart_periods = sorted({row["period_key"] for row in billing_rows})
    chart_meter_rows = {}
    sorted_chart_memberships = sorted(
        effective_memberships,
        key=lambda membership: (
            getattr(
                getattr(membership.billing_meter.unit, "property", None),
                "property_name",
                "",
            ),
            membership.billing_meter.display_location_name,
            membership.billing_meter.meter_number,
        ),
    )
    for membership in sorted_chart_memberships:
        meter = membership.billing_meter
        property_name = getattr(
            getattr(meter.unit, "property", None), "property_name", ""
        )
        chart_meter_rows.setdefault(
            meter.pk,
            {
                "unit_label": f"{property_name[:8]} / {meter.display_location_name}",
                "unit_number": meter.display_location_name,
                "meter_number": meter.meter_number,
                "usage_by_period": defaultdict(lambda: Decimal("0")),
            },
        )
    for row in billing_rows:
        meter_row = chart_meter_rows.setdefault(
            row["meter_id"],
            {
                "unit_label": f'{row["property_name"][:8]} / {row["unit_number"]}',
                "unit_number": row["unit_number"],
                "meter_number": row["meter_number"],
                "usage_by_period": defaultdict(lambda: Decimal("0")),
            },
        )
        meter_row["usage_by_period"][row["period_key"]] += Decimal(str(row["usage"]))
    billing_chart_data = {
        "labels": [period.isoformat() for period in chart_periods],
        "datasets": [
            {
                "serial": serial,
                "meterId": meter_id,
                "unitLabel": meter_row["unit_label"],
                "unitNumber": meter_row["unit_number"],
                "meterNumber": meter_row["meter_number"],
                "label": meter_row["unit_label"],
                "data": [
                    (
                        float(meter_row["usage_by_period"][period])
                        if period in meter_row["usage_by_period"]
                        else None
                    )
                    for period in chart_periods
                ],
            }
            for serial, (meter_id, meter_row) in enumerate(
                chart_meter_rows.items(), start=1
            )
        ],
        "auditLastReading": {
            "totalEnergy": (
                float(audit_last_total_energy)
                if audit_last_total_energy is not None
                else None
            ),
            "at": (
                timezone.localtime(audit_last_reading_at).strftime("%Y-%m-%d %H:%M")
                if audit_last_reading_at
                else ""
            ),
        },
    }
    billing_meter_count = len(chart_meter_rows)
    average_billing_kwh = (
        billing_kwh / billing_meter_count if billing_meter_count else Decimal("0")
    )
    variance_type = "Loss" if variance_kwh >= 0 else "Gain"
    audit_group_options = (
        MeterCheckGroup.objects.filter(
            is_active=True,
            check_meter__is_active=True,
            check_meter__meter_role=Meter.METER_ROLE_CHECK,
        )
        .select_related("check_meter", "check_meter__unit", "check_meter__unit__property")
        .order_by("check_meter__meter_number", "name")
    )
    covered_property_names = sorted({
        membership.billing_meter.unit.property.property_name
        for membership in memberships
        if membership.is_active
        and membership.end_date is None
        and membership.start_date <= today
        and membership.billing_meter.is_active
        and membership.billing_meter.unit_id
        and membership.billing_meter.unit.property_id
    })
    return render(
        request,
        "smart_meter/check_group_detail.html",
        {
            "group": group,
            "start_date": start_date,
            "end_date": end_date,
            "quick_range": quick_range,
            "check_kwh": check_kwh,
            "billing_kwh": billing_kwh,
            "variance_kwh": variance_kwh,
            "variance_rs": variance_rs,
            "leakage_percent": leakage_percent,
            "check_rows": check_rows,
            "audit_summary_start_kwh": audit_summary_start_kwh,
            "audit_summary_end_kwh": audit_summary_end_kwh,
            "billing_rows": billing_rows,
            "memberships": memberships,
            "membership_form": membership_form,
            "covered_property_names": covered_property_names,
            "audit_last_reading_at": audit_last_reading_at,
            "audit_last_total_energy": audit_last_total_energy,
            "audit_reading_status": audit_reading_status,
            "audit_reading_status_class": audit_reading_status_class,
            "report_target_at": report_target_at,
            "reconciliation_target_at": reconciliation_target_at,
            "billing_chart_data": billing_chart_data,
            "billing_meter_count": billing_meter_count,
            "average_billing_kwh": average_billing_kwh,
            "variance_type": variance_type,
            "variance_absolute_kwh": abs(variance_kwh),
            "audit_group_options": audit_group_options,
        },
    )


@require_POST
@login_required
def meter_check_group_membership_end(request, pk, membership_id):
    group = get_object_or_404(MeterCheckGroup, pk=pk)
    membership = get_object_or_404(
        MeterCheckGroupMembership,
        pk=membership_id,
        group=group,
        is_active=True,
    )
    membership.close(
        end_date=timezone.localdate(),
        notes=request.POST.get("notes", "Membership ended."),
    )
    messages.success(request, "Billing-meter membership ended.")
    return redirect("smart_meter:meter_check_group_detail", pk=group.pk)


@require_POST
@login_required
def meter_check_group_membership_manage(request, pk, membership_id):
    group = get_object_or_404(MeterCheckGroup, pk=pk)
    action = (request.POST.get("action") or "").strip().lower()
    if action not in {"update", "delete"}:
        return JsonResponse(
            {"success": False, "error": "Select a valid action."}, status=400
        )

    try:
        with transaction.atomic():
            membership = get_object_or_404(
                MeterCheckGroupMembership.objects.select_for_update(),
                pk=membership_id,
                group=group,
            )
            meter_number = membership.billing_meter.meter_number
            if action == "delete":
                membership.delete()
                return JsonResponse(
                    {
                        "success": True,
                        "deleted": True,
                        "message": f"Membership for meter {meter_number} was deleted.",
                    }
                )

            try:
                start_date = date.fromisoformat((request.POST.get("start_date") or "").strip())
            except ValueError:
                return JsonResponse(
                    {"success": False, "error": "Enter a valid start date."}, status=400
                )
            end_date_value = (request.POST.get("end_date") or "").strip()
            try:
                end_date = date.fromisoformat(end_date_value) if end_date_value else None
            except ValueError:
                return JsonResponse(
                    {"success": False, "error": "Enter a valid end date."}, status=400
                )

            membership.start_date = start_date
            membership.end_date = end_date
            membership.is_active = end_date is None
            membership.notes = (request.POST.get("notes") or "").strip()
            membership.save()
    except ValidationError as exc:
        if hasattr(exc, "message_dict"):
            error = " ".join(
                str(message)
                for field_messages in exc.message_dict.values()
                for message in field_messages
            )
        else:
            error = " ".join(str(message) for message in exc.messages)
        return JsonResponse({"success": False, "error": error}, status=400)

    return JsonResponse(
        {
            "success": True,
            "message": f"Membership for meter {meter_number} was updated.",
        }
    )


def meter_delete(request, pk):
    meter = get_object_or_404(Meter, pk=pk)
    if request.method == "POST":
        meter.delete()
        return redirect("smart_meter:meter_list")
    return render(request, "smart_meter/meter_confirm_delete.html", {"meter": meter})


def meter_detail(request, pk):
    meter = get_object_or_404(
        Meter.objects.select_related("unit", "unit__property"),
        pk=pk,
    )
    installation_history = list(
        MeterInstallation.objects.filter(meter=meter)
        .select_related("unit", "unit__property", "lease", "lease__tenant")
        .order_by("-is_active", "-start_date", "-id")
    )
    active_installations = [
        installation
        for installation in installation_history
        if installation.is_active and installation.end_date is None
    ]
    if meter.unit_id:
        current_installation = next(
            (
                installation
                for installation in active_installations
                if installation.unit_id == meter.unit_id
            ),
            None,
        )
    else:
        current_installation = active_installations[0] if active_installations else None

    current_lease = None
    if meter.unit_id:
        today = date.today()
        current_lease = (
            Lease.objects.filter(
                unit_id=meter.unit_id,
                status="active",
                start_date__lte=today,
                end_date__gte=today,
            )
            .select_related("tenant")
            .order_by("-start_date", "-id")
            .first()
        )
        if current_lease is None:
            current_lease = (
                Lease.objects.filter(unit_id=meter.unit_id, status="active")
                .select_related("tenant")
                .order_by("-start_date", "-id")
                .first()
            )

    if current_lease is None and current_installation and current_installation.lease_id:
        current_lease = current_installation.lease

    latest_reading = MeterReading.objects.filter(meter=meter).order_by("-ts").first()
    latest_live = meter.latest_live
    display_reading = latest_live or latest_reading
    display_forward_energy = (
        getattr(display_reading, "forward_active_energy_kwh", None)
        if display_reading
        else None
    )
    if display_forward_energy is None and display_reading:
        display_forward_energy = display_reading.total_energy
    display_reverse_energy = (
        getattr(display_reading, "reverse_active_energy_kwh", None)
        if display_reading
        else None
    )
    display_net_grid_energy = (
        display_forward_energy - display_reverse_energy
        if display_forward_energy is not None and display_reverse_energy is not None
        else None
    )
    display_reading_is_fresh = bool(
        display_reading
        and display_reading.ts
        and display_reading.ts
        >= timezone.now() - timedelta(minutes=online_threshold_minutes())
    )
    current_energy = (
        getattr(latest_live, "total_energy", None)
        if latest_live and getattr(latest_live, "total_energy", None) is not None
        else (latest_reading.total_energy if latest_reading else None)
    )

    display_installation_history = []

    # Resolve each assignment-change boundary reading in the assignment-history
    # query itself. With the existing (meter, ts) reading index this avoids one
    # latest-reading query for every history row.
    change_energy_sq = (
        MeterReading.objects.filter(meter_id=OuterRef("meter_id"), ts__lte=OuterRef("change_date"))
        .order_by("-ts", "-id")
        .values("total_energy")[:1]
    )
    assignment_changes = list(
        meter.assignment_history.select_related(
            "old_unit",
            "old_unit__property",
            "new_unit",
            "new_unit__property",
            "old_lease",
            "old_lease__tenant",
            "new_lease",
            "new_lease__tenant",
        )
        .annotate(_change_energy=Subquery(change_energy_sq))
        .order_by("change_date", "id")
    )

    # History rendering repeatedly asks which lease occupied a unit on a given
    # date. Load the small relevant lease set once rather than issuing a query
    # from inside each installation/change loop.
    history_unit_ids = {
        unit_id
        for unit_id in (
            [installation.unit_id for installation in installation_history]
            + [change.old_unit_id for change in assignment_changes]
            + [change.new_unit_id for change in assignment_changes]
            + [meter.unit_id]
        )
        if unit_id
    }
    history_dates = [installation.start_date for installation in installation_history]
    history_dates += [
        timezone.localtime(change.change_date).date() for change in assignment_changes
    ]
    history_min_date = min(history_dates) if history_dates else date.today()
    history_max_date = max(
        [installation.end_date or date.today() for installation in installation_history]
        + [date.today()]
    )
    leases_by_unit = defaultdict(list)
    if history_unit_ids:
        history_leases = (
            Lease.objects.filter(
                unit_id__in=history_unit_ids,
                start_date__lte=history_max_date,
                end_date__gte=history_min_date,
            )
            .select_related("tenant", "unit", "unit__property")
            .order_by("unit_id", "-start_date", "-id")
        )
        for history_lease in history_leases:
            leases_by_unit[history_lease.unit_id].append(history_lease)

    def active_lease_for_unit(unit, on_date):
        if not unit:
            return None
        for lease in leases_by_unit.get(unit.id, ()):
            if lease.start_date <= on_date <= lease.end_date:
                return lease
        return None

    def lease_for_history_row(unit, on_date, fallback=None):
        lease = active_lease_for_unit(unit, on_date)
        if lease:
            return lease
        if fallback and getattr(fallback, "unit_id", None) == getattr(unit, "id", None):
            return fallback
        return None

    for installation in installation_history:
        segment_unit = installation.unit
        segment_lease = installation.lease
        segment_from = installation.start_date
        segment_start_reading = installation.start_reading
        relevant_changes = [
            change
            for change in assignment_changes
            if timezone.localtime(change.change_date).date() >= installation.start_date
            and (
                installation.end_date is None
                or timezone.localtime(change.change_date).date()
                <= installation.end_date
            )
            and change.new_unit_id
            and (
                change.old_unit_id == getattr(segment_unit, "id", None)
                or change.new_unit_id == meter.unit_id
                or change.meter_id == meter.id
            )
        ]

        for change in relevant_changes:
            change_day = timezone.localtime(change.change_date).date()
            change_energy = change._change_energy
            display_installation_history.append(
                {
                    "unit": segment_unit,
                    "lease": lease_for_history_row(
                        segment_unit,
                        segment_from,
                        segment_lease or change.old_lease,
                    ),
                    "start_date": segment_from,
                    "display_to_date": change_day,
                    "start_reading": segment_start_reading,
                    "display_end_reading": change_energy,
                    "is_active": False,
                }
            )
            segment_unit = change.new_unit
            segment_lease = lease_for_history_row(
                change.new_unit, change_day, change.new_lease
            )
            segment_from = change_day
            segment_start_reading = change_energy

        segment_start_dt, _segment_start_end = _local_date_bounds(segment_from)
        _segment_end_start, segment_end_dt = _local_date_bounds(
            installation.end_date or date.today()
        )
        last_installation_reading = (
            MeterReading.objects.filter(
                meter=meter,
                ts__gte=segment_start_dt,
                ts__lt=segment_end_dt,
            )
            .order_by("-ts", "-id")
            .first()
        )
        display_to_date = installation.end_date
        display_end_reading = installation.end_reading
        segment_is_active = bool(
            installation.is_active
            and installation.end_date is None
            and meter.is_active
            and getattr(segment_unit, "id", None) == meter.unit_id
        )
        if display_end_reading is None:
            if segment_is_active:
                display_end_reading = current_energy
            elif last_installation_reading:
                display_end_reading = last_installation_reading.total_energy
        if (
            not segment_is_active
            and display_to_date is None
            and last_installation_reading
        ):
            display_to_date = timezone.localtime(last_installation_reading.ts).date()
        display_installation_history.append(
            {
                "unit": segment_unit,
                "lease": lease_for_history_row(
                    segment_unit, segment_from, segment_lease
                ),
                "start_date": segment_from,
                "display_to_date": display_to_date,
                "start_reading": segment_start_reading,
                "display_end_reading": display_end_reading,
                "is_active": segment_is_active,
            }
        )

    display_installation_history.sort(
        key=lambda row: (row["is_active"], row["start_date"]),
        reverse=True,
    )

    readings = list(MeterReading.objects.filter(meter=meter).order_by("-ts")[:500])
    recent_daily_readings = []
    latest_by_day = OrderedDict()
    max_current_by_day = {}
    max_voltage_by_day = {}

    def phase_max(reading, names):
        values = [getattr(reading, name, None) for name in names]
        values = [value for value in values if value is not None]
        return max(values) if values else None

    for reading in readings:
        local_ts = timezone.localtime(reading.ts)
        reading.local_day = local_ts.date()
        reading.max_current_value = phase_max(
            reading, ("current_a", "current_b", "current_c")
        )
        reading.max_voltage_value = phase_max(
            reading, ("voltage_a", "voltage_b", "voltage_c")
        )
        latest_by_day.setdefault(reading.local_day, reading)
        current_best = max_current_by_day.get(reading.local_day)
        if reading.max_current_value is not None and (
            current_best is None
            or reading.max_current_value > current_best.max_current_value
        ):
            max_current_by_day[reading.local_day] = reading
        voltage_best = max_voltage_by_day.get(reading.local_day)
        if reading.max_voltage_value is not None and (
            voltage_best is None
            or reading.max_voltage_value > voltage_best.max_voltage_value
        ):
            max_voltage_by_day[reading.local_day] = reading

    daily_latest = list(latest_by_day.values())[:7]
    for index, reading in enumerate(daily_latest):
        previous = daily_latest[index + 1] if index + 1 < len(daily_latest) else None
        reading.day_difference = None
        if (
            previous
            and reading.total_energy is not None
            and previous.total_energy is not None
        ):
            reading.day_difference = reading.total_energy - previous.total_energy
        reading.day_max_current = max_current_by_day.get(reading.local_day)
        reading.day_max_voltage = max_voltage_by_day.get(reading.local_day)
        recent_daily_readings.append(reading)

    from smart_meter.models import MeterCommand, MeterCreditAccount, MeterPrepaidPilot

    credit_account = (
        MeterCreditAccount.objects.filter(meter=meter)
        .select_related("installation", "lease")
        .order_by("-is_enabled", "-created_at")
        .first()
    )
    recent_commands = MeterCommand.objects.filter(meter=meter).order_by("-created_at")[
        :20
    ]
    prepaid_pilot = MeterPrepaidPilot.objects.filter(meter=meter).first()
    meter_feature_flags = {
        "credit_eval": bool(
            getattr(settings, "METER_ENABLE_AUTOMATIC_CREDIT_EVALUATION", False)
        ),
        "notifications": bool(
            getattr(settings, "METER_ENABLE_AUTOMATIC_NOTIFICATIONS", False)
        ),
        "auto_cutoff": bool(getattr(settings, "METER_ENABLE_AUTOMATIC_CUTOFF", False)),
        "auto_restore": bool(
            getattr(settings, "METER_ENABLE_AUTOMATIC_RESTORE", False)
        ),
        "prepaid_reads": bool(getattr(settings, "METER_ENABLE_PREPAID_READS", False)),
        "prepaid_writes": bool(getattr(settings, "METER_ENABLE_PREPAID_WRITES", False)),
        "prepaid_allowlisted": meter.pk
        in set(getattr(settings, "METER_PREPAID_ALLOWED_METER_IDS", ()) or ()),
        "credit_allowlisted": meter.pk
        in set(getattr(settings, "METER_CREDIT_ALLOWED_METER_IDS", ()) or ()),
        "emergency_stop": bool(getattr(settings, "METER_EMERGENCY_STOP", False)),
    }
    from smart_meter.rates import resolve_electricity_rate

    electricity_rate = resolve_electricity_rate(
        meter=meter,
        lease=current_lease,
    )

    return render(
        request,
        "smart_meter/meter_detail.html",
        {
            "meter": meter,
            "reading_profile_form": MeterReadingProfileForm(instance=meter),
            "electricity_rate": electricity_rate,
            "current_installation": current_installation,
            "current_lease": current_lease,
            "current_tenant": current_lease.tenant if current_lease else None,
            "installation_history": display_installation_history,
            "latest_reading": latest_reading,
            "latest_live": latest_live,
            "display_reading": display_reading,
            "display_forward_energy": display_forward_energy,
            "display_reverse_energy": display_reverse_energy,
            "display_net_grid_energy": display_net_grid_energy,
            "display_reading_is_fresh": display_reading_is_fresh,
            "recent_daily_readings": recent_daily_readings,
            "credit_account": credit_account,
            "recent_commands": recent_commands,
            "prepaid_pilot": prepaid_pilot,
            "meter_feature_flags": meter_feature_flags,
        },
    )


@require_POST
@login_required
@permission_required("smart_meter.change_meter", raise_exception=True)
def meter_reading_profile_update(request, pk):
    meter = get_object_or_404(Meter, pk=pk)
    form = MeterReadingProfileForm(request.POST, instance=meter)
    if form.is_valid():
        form.save()
        messages.success(request, "Meter reading profile updated.")
    else:
        messages.error(request, "; ".join(
            message for errors in form.errors.values() for message in errors
        ))

    system_id = request.POST.get("energy_system_id")
    if system_id:
        from smart_meter.models import EnergySystem

        if EnergySystem.objects.filter(pk=system_id, grid_interface_meter=meter).exists():
            return redirect("smart_meter:energy_system_detail", pk=system_id)
    return redirect("smart_meter:meter_detail", pk=meter.pk)


def install_meter_to_unit(request, unit_id):
    unit = get_object_or_404(Unit, pk=unit_id)
    if request.method == "POST":
        form = InstallMeterToUnitForm(request.POST, unit=unit, user=request.user)
        if form.is_valid():
            installation = form.save()
            messages.success(
                request,
                f"Meter {installation.meter.meter_number} installed on {unit}.",
            )
            return redirect("properties:unit_detail", pk=unit.pk)
    else:
        form = InstallMeterToUnitForm(unit=unit, user=request.user)
    return render(
        request,
        "smart_meter/meter_installation_form.html",
        {
            "form": form,
            "unit": unit,
            "title": "Install Meter",
            "submit_label": "Install Meter",
            "cancel_url": reverse("properties:unit_detail", args=[unit.pk]),
        },
    )


def switch_meter(request, unit_id):
    unit = get_object_or_404(Unit, pk=unit_id)
    if request.method == "POST":
        form = SwitchMeterForm(request.POST, unit=unit)
        if form.is_valid():
            with transaction.atomic():
                old_installation = (
                    MeterInstallation.objects.select_for_update()
                    .select_related("meter")
                    .get(pk=form.cleaned_data["old_installation"].pk)
                )
                old_installation.close(
                    end_date=form.cleaned_data["switch_date"],
                    end_reading=form.cleaned_data["old_end_reading"],
                    notes=form.cleaned_data.get("notes", ""),
                )
                if old_installation.meter.unit_id == unit.pk:
                    Meter.objects.filter(pk=old_installation.meter_id).update(unit=None)

                new_installation = MeterInstallation.objects.create(
                    meter=form.cleaned_data["new_meter"],
                    unit=unit,
                    lease=form.cleaned_data.get("lease"),
                    start_date=form.cleaned_data["switch_date"],
                    start_reading=form.cleaned_data["new_start_reading"],
                    installed_by=request.user
                    if request.user.is_authenticated
                    else None,
                    reason=form.cleaned_data.get("reason") or "Meter switched",
                    notes=form.cleaned_data.get("notes", ""),
                )

            messages.success(
                request,
                f"Switched {old_installation.meter.meter_number} to {new_installation.meter.meter_number}.",
            )
            return redirect("properties:unit_detail", pk=unit.pk)
    else:
        form = SwitchMeterForm(unit=unit)
    return render(
        request,
        "smart_meter/meter_switch_form.html",
        {
            "form": form,
            "unit": unit,
            "cancel_url": reverse("properties:unit_detail", args=[unit.pk]),
        },
    )


def close_meter_installation(request, installation_id):
    installation = get_object_or_404(
        MeterInstallation.objects.select_related("meter", "unit"),
        pk=installation_id,
        is_active=True,
    )
    if request.method == "POST":
        form = CloseMeterInstallationForm(request.POST, installation=installation)
        if form.is_valid():
            installation.close(
                end_date=form.cleaned_data["end_date"],
                end_reading=form.cleaned_data.get("end_reading"),
                notes=form.cleaned_data.get("notes", ""),
            )
            if installation.meter.unit_id == installation.unit_id:
                Meter.objects.filter(pk=installation.meter_id).update(unit=None)
            messages.success(
                request, f"Closed installation for {installation.meter.meter_number}."
            )
            return redirect("smart_meter:meter_detail", pk=installation.meter.pk)
    else:
        form = CloseMeterInstallationForm(installation=installation)
    return render(
        request,
        "smart_meter/meter_installation_close_form.html",
        {
            "form": form,
            "installation": installation,
            "cancel_url": reverse(
                "smart_meter:meter_detail", args=[installation.meter.pk]
            ),
        },
    )


def move_lease_unit(request, lease_id):
    lease = get_object_or_404(Lease, pk=lease_id)
    if request.method == "POST":
        form = MoveLeaseUnitForm(request.POST, lease=lease)
        if form.is_valid():
            move_date = form.cleaned_data["move_date"]
            new_unit = form.cleaned_data["new_unit"]
            notes = form.cleaned_data.get("notes", "")
            with transaction.atomic():
                active = (
                    LeaseUnitOccupancy.objects.select_for_update()
                    .filter(lease=lease, move_out_date__isnull=True)
                    .order_by("-move_in_date", "-id")
                    .first()
                )
                if active:
                    active.move_out_date = (
                        move_date - timedelta(days=1)
                        if move_date > active.move_in_date
                        else move_date
                    )
                    if notes:
                        active.notes = (active.notes + "\n" + notes).strip()
                    active.save()
                elif lease.unit_id:
                    move_out_date = (
                        move_date - timedelta(days=1)
                        if move_date > lease.start_date
                        else move_date
                    )
                    LeaseUnitOccupancy.objects.create(
                        lease=lease,
                        unit=lease.unit,
                        move_in_date=lease.start_date,
                        move_out_date=move_out_date,
                        notes="Created from existing lease unit during tenant move.",
                    )

                LeaseUnitOccupancy.objects.create(
                    lease=lease,
                    unit=new_unit,
                    move_in_date=move_date,
                    notes=notes,
                )
                lease.unit = new_unit
                lease.save(update_fields=["unit"])

            messages.success(request, f"Moved lease #{lease.pk} to {new_unit}.")
            return redirect("leases:lease_detail", pk=lease.pk)
    else:
        form = MoveLeaseUnitForm(lease=lease)
    return render(
        request,
        "smart_meter/lease_unit_move_form.html",
        {
            "form": form,
            "lease": lease,
            "cancel_url": reverse("leases:lease_detail", args=[lease.pk]),
        },
    )


def edit_reading(request, pk):
    reading = get_object_or_404(MeterReading, pk=pk)
    if request.method == "POST":
        form = MeterReadingForm(request.POST, instance=reading)
        if form.is_valid():
            form.save()
            return redirect("reading_list")
    else:
        form = MeterReadingForm(instance=reading)
    return render(
        request, "smart_meter/reading_form.html", {"form": form, "edit": True}
    )


def delete_reading(request, pk):
    reading = get_object_or_404(MeterReading, pk=pk)
    if request.method == "POST":
        reading.delete()
        return redirect("reading_list")
    return render(
        request, "smart_meter/reading_confirm_delete.html", {"reading": reading}
    )


def meter_readings(request, meter_id):
    meter = get_object_or_404(Meter, id=meter_id)
    readings = MeterReading.objects.filter(meter=meter).order_by("-timestamp")
    return render(
        request,
        "smart_meter/meter_readings.html",
        {"meter": meter, "readings": readings},
    )


def fetch_meter_data(request):
    # Fetch data for all meters
    poll_all_meters()  # Call your polling function here to fetch the data
    return JsonResponse({"status": "success"})


def toggle_power(request, meter_id):
    meter = get_object_or_404(Meter, pk=meter_id)
    # Logic to toggle power, e.g., sending a TCP command to turn off the meter
    # send_tcp_command_to_toggle_power(meter)
    meter.is_active = not meter.is_active  # Simulate power toggling
    meter.save()
    return redirect("smart_meter:meter_list")


def recharge_balance(request, meter_id):
    meter = get_object_or_404(Meter, pk=meter_id)
    if request.method == "POST":
        amount = float(request.POST.get("amount", 0))
        meter.balance += amount
        meter.save()
    return redirect("smart_meter:meter_list")


def refund_balance(request, meter_id):
    meter = get_object_or_404(Meter, pk=meter_id)
    if request.method == "POST":
        amount = float(request.POST.get("amount", 0))
        meter.balance -= amount
        meter.save()
    return redirect("smart_meter:meter_list")


# views.py

# ... keep your other imports


def live_custom(request):
    online_minutes = online_threshold_minutes()
    # 1) Pull the filter values from query string
    q = (request.GET.get("q") or "").strip()
    offline_only = request.GET.get("offline") == "1"
    active_filter = (request.GET.get("active") or "active").strip()

    # 2) Reuse the same cascading filter sets as meter list
    #    selected_meters = the final meter set based on property/unit/meter GET params
    (
        selected_meters,
        all_properties,
        filtered_units,
        filtered_meters,
        prop_id,
        unit_id,
        meter_id,
    ) = _filtered_meter_sets(request, include_meter_property=False)
    if active_filter == "active":
        filtered_meters = filtered_meters.filter(is_active=True)
    elif active_filter == "inactive":
        filtered_meters = filtered_meters.filter(is_active=False)
    # The dropdown renders ``display_location_name`` too. Attach the same
    # bulk unit counts used by the live rows so that property never falls back
    # to two per-unit database lookups.
    filtered_meters = attach_active_meter_counts(filtered_meters)
    active_meter_count_by_id = {
        meter.id: meter._active_unit_meter_count for meter in filtered_meters
    }

    meter_scope_qs = _with_meter_operational_flags(
        _meters_annotated_qs(request, online_minutes=online_minutes)
    )
    if active_filter == "active":
        meter_scope_qs = meter_scope_qs.filter(is_active=True)
    elif active_filter == "inactive":
        meter_scope_qs = meter_scope_qs.filter(is_active=False)
    current_chip = _normalized_meter_chip(request)
    chip_cards = _meter_chip_cards(
        request, meter_scope_qs, "smart_meter:smart_meter_live_custom"
    )
    meter_scope_qs = _apply_meter_chip_filter(meter_scope_qs, current_chip)

    # 3) Base queryset: only readings for the selected meters
    qs = (
        LiveReading.objects.select_related(
            "meter", "meter__unit", "meter__unit__property"
        )
        .only(
            "id",
            "meter",
            "ts",
            "source_ip",
            "source_port",
            "balance",
            "total_energy",
            "forward_active_energy_kwh",
            "reverse_active_energy_kwh",
            "voltage_a",
            "current_a",
            "total_power",
            "pf_total",
            "status_word",
            "meter__id",
            "meter__unit",
            "meter__meter_number",
            "meter__power_status",
            "meter__name",
            "meter__is_active",
            "meter__meter_role",
            "meter__unit__id",
            "meter__unit__property",
            "meter__unit__unit_number",
            "meter__unit__property__id",
            "meter__unit__property__property_name",
        )
        .order_by(
            "meter__unit__property__property_name",
            "meter__unit__unit_number",
            "meter__meter_number",
        )
    )
    if meter_id:
        qs = qs.filter(meter_id=meter_id)
    elif unit_id:
        qs = qs.filter(meter__unit_id=unit_id)
    elif prop_id:
        qs = qs.filter(meter__unit__property_id=prop_id)
    if active_filter == "active":
        qs = qs.filter(meter__is_active=True)
    elif active_filter == "inactive":
        qs = qs.filter(meter__is_active=False)

    # 4) Optional free-text search across property / unit / meter
    if q:
        qs = qs.filter(
            Q(meter__unit__unit_number__icontains=q)
            | Q(meter__meter_number__icontains=q)
            | Q(meter__unit__property__property_name__icontains=q)
        )

    qs = qs.filter(meter_id__in=meter_scope_qs.values("id"))

    # 5) Compute 'is_online' and apply offline-only filter if requested
    rows = []
    status_by_meter_id = resolve_meter_online_statuses(
        (reading.meter, reading) for reading in qs
    )
    qs = list(qs)
    for r in qs:
        status = status_by_meter_id[r.meter_id]
        for field, value in status.items():
            setattr(r, field, value)
        if offline_only and r.is_online:
            continue
        rows.append(r)
    tenant_info = active_tenant_info_for_units(
        reading.meter.unit_id
        for reading in rows
        if reading.meter and reading.meter.unit_id
    )
    for reading in rows:
        unit_key = reading.meter.unit_id if reading.meter else None
        info = tenant_info.get(unit_key)
        reading.tenant_name = info["name"] if info else "Vacant"
        reading.tenant_id = info["tenant_id"] if info else None
    latest_relay_commands = {}
    for command in MeterCommand.objects.filter(
        meter_id__in=[reading.meter_id for reading in rows],
        command_type="relay",
    ).order_by("meter_id", "-created_at"):
        latest_relay_commands.setdefault(command.meter_id, command)
    for reading in rows:
        reading.latest_relay_command = latest_relay_commands.get(reading.meter_id)
        relay_state = reconcile_live_relay_command_state(
            reading.meter,
            reading.latest_relay_command,
            reading.status_word,
            reading.ts,
            is_fresh=reading.measurement_is_fresh,
        )
        reading.confirmed_relay_state = relay_state["confirmed_state"]
        reading.relay_confirmed_at = (
            reading.ts if reading.confirmed_relay_state else None
        )
        reading.relay_command_status = relay_state["status"]
        reading.relay_command_desired_state = relay_state["desired_state"]
        reading.relay_command_error = relay_state["error"]
        reading.relay_operation_label = relay_state["operation_label"]
        reading.relay_indicator_label = relay_state["indicator_label"]
        reading.relay_indicator_class = relay_state["indicator_class"]
    readings_missing_counts = []
    for reading in rows:
        if reading.meter_id in active_meter_count_by_id:
            reading.meter._active_unit_meter_count = active_meter_count_by_id[
                reading.meter_id
            ]
        else:
            readings_missing_counts.append(reading)
    if readings_missing_counts:
        attach_active_meter_counts(
            readings_missing_counts,
            lambda reading: reading.meter,
        )

    # Mark selected flags (NO template comparison needed)
    # formatter-proof flags (avoid template comparisons)
    cp = str(prop_id or "")
    cu = str(unit_id or "")
    cm = str(meter_id or "")

    for p in all_properties:
        p.is_selected = str(p.id) == cp

    for u in filtered_units:
        u.is_selected = str(u.id) == cu

    for m in filtered_meters:
        m.is_selected = str(m.id) == cm

    offline_checked = request.GET.get("offline") == "1"
    energy_end_date = timezone.localdate()
    energy_start_date = energy_end_date.replace(day=1)

    # 6) Render with everything the filter bar needs
    return render(
        request,
        "smart_meter/live_custom.html",
        {
            "rows": rows,
            "online_minutes": online_minutes,
            "q": q,
            "offline_only": offline_only,
            "active_filter": active_filter,
            "chip_cards": chip_cards,
            "current_chip": current_chip,
            "current_active": active_filter,
            # dropdown data (same as meter list)
            "all_properties": all_properties,
            "filtered_units": filtered_units,
            "filtered_meters": filtered_meters,
            "current_property": prop_id,
            "current_unit": unit_id,
            "current_meter": meter_id,
            "current_role": (request.GET.get("role") or "").strip().lower(),
            "energy_start_date": energy_start_date,
            "energy_end_date": energy_end_date,
            "vpn_connected": vpn_connected(),
            "public_ip": public_ip(),
        },
    )


@require_POST
def recharge_meter(request, meter_id):
    meter = get_object_or_404(Meter, id=meter_id)
    try:
        amt = Decimal(request.POST.get("amount", "0") or "0")
    except Exception:
        messages.error(request, "Invalid amount.")
        return redirect("smart_meter_live_custom")

    if amt <= 0:
        messages.error(request, "Amount must be greater than 0.")
        return redirect("smart_meter_live_custom")

    bal, _ = MeterBalance.objects.get_or_create(unit=meter.unit)
    # Add credit to balance
    bal.balance = (bal.balance or Decimal("0.00")) + amt
    bal.save()

    MeterEvent.objects.create(
        unit=meter.unit, event_type="recharge", note=f"Recharge via live page: +₹{amt}"
    )
    messages.success(request, f"Recharged meter {meter.meter_number} by ₹{amt}.")
    return redirect("smart_meter_live_custom")


# smart_meter/views.py


def _redirect_back(request, fallback_name="smart_meter:meter_list"):
    return redirect(request.META.get("HTTP_REFERER") or reverse(fallback_name))


@require_POST
def cutoff_meter(request, meter_id):
    """Cut OFF (open relay) for a single meter — meter-number based, no IP needed."""
    meter = get_object_or_404(Meter, pk=meter_id)

    byCmd = RELAY_OPEN_COMMAND  # OFF
    frame = build_switch_frame(meter.meter_number, byCmd)
    frame_hex = _as_hex(frame)
    cmd_name = "OFF"

    # Audit: exactly what the user requested

    try:
        frame_hex = frame.hex().upper()
    except AttributeError:
        frame_hex = str(frame)

    # Entry audit (this is specifically what you asked for when pressing the ON/OFF button)
    logger.info(
        "REQUEST TX CUT-OFF from cutoff_meter user=%s path=%s meter=%s "
        "cmd=%s(0x%02X) frame=%s disable_cutoffs=%s sender=%s.%s",
        getattr(request.user, "username", "anonymous"),  # user=%s
        getattr(request, "path", ""),  # path=%s
        meter.meter_number,  # meter=%s
        cmd_name,  # cmd=%s
        byCmd,  # 0x%02X  (int!)
        frame_hex,  # frame=%s
        DISABLE_CUTOFFS,  # disable_cutoffs=%s
        getattr(send_via_db, "__module__", "?"),  # sender=%s
        getattr(send_via_db, "__name__", "?"),  # .%s
    )

    from smart_meter.models import MeterEvent

    MeterEvent.objects.create(
        unit=meter.unit,
        event_type="cutoff_tx",
        note=f"frame={frame_hex} by={getattr(request.user, 'username', 'anonymous')}",
    )
    # blank line separator
    logger.info("-------------------------------------")

    # Send (honor feature flag if you kept it)
    if DISABLE_CUTOFFS:
        res = {
            "ok": False,
            "error": "switching disabled by DISABLE_CUTOFFS",
            "payload": "skipped:DISABLE_CUTOFFS",
        }
        logger.info(
            "RESPONSE meter=%s cmd=%s ok=%s error=%s payload=%s",
            meter.meter_number,
            cmd_name,
            res.get("ok"),
            res.get("error"),
            res.get("payload"),
        )
    else:
        try:
            secret = getattr(settings, "METER_CTRL_SECRET", None)  # optional

            res = _call_send(
                meter_number=meter.meter_number,
                frame=frame,
                timeout=32.0,
                expect_di=None,
                allow_switch=True,  # <-- explicit
                initiated_by=request.user.get_username(),  # <-- who clicked
                reason="manual switch from UI",  # <-- audit
                command_type="relay",
                desired_state="off",
                source="manual",
                auth=secret,  # <-- optional shared secret
            )
            logger.info(
                "RESPONSE RX CUT-OFF meter=%s cmd=%s ok=%s error=%s payload=%s",
                meter.meter_number,
                cmd_name,
                res.get("ok"),
                res.get("error"),
                res.get("payload"),
            )

            MeterEvent.objects.create(
                unit=meter.unit,
                event_type="cutoff_rx",
                note=f"ok={res.get('ok')} error={res.get('error')} payload={res.get('payload')}",
            )

            import time

            if res.get("ok"):
                time.sleep(5)  # brief settle
                try:
                    refresh_live(meter.meter_number)
                    lr = LiveReading.objects.filter(meter=meter).order_by("-ts").first()
                    amps = float(getattr(lr, "current_a", 0) or 0)
                    watts = float(getattr(lr, "total_power", 0) or 0)
                    logger.info(
                        "POST-CUTOFF VERIFY meter=%s I=%.3fA P=%.3fW ts=%s",
                        meter.meter_number,
                        amps,
                        watts,
                        getattr(lr, "ts", None),
                    )
                    if amps > 0.02 or watts > 5:
                        messages.warning(
                            request,
                            f"Cutoff sent, but current is {amps:.3f}A / {watts:.0f}W — relay may be closed, bypassed, or stuck.",
                        )
                except Exception:
                    pass

        except Exception as e:
            logger.exception(
                "SEND_FAILED meter=%s cmd=%s error=%s", meter.meter_number, cmd_name, e
            )
            # For AJAX callers, return JSON error
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                # blank line separator

                return JsonResponse({"success": False, "error": str(e)}, status=500)
            messages.error(request, f"Cut off failed for {meter.meter_number}: {e}")
            # blank line separator

            return _redirect_back(request)

    # Sending/acknowledging a command is not authoritative relay state.
    # ``Meter.power_status`` is updated only by a valid 0x028011FF readback.
    success = bool(res.get("ok"))

    # Respond depending on caller
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        # blank line separator

        return JsonResponse(
            {
                "success": success,
                "error": res.get("error"),
                "command_id": res.get("command_id"),
                "command_status": res.get("status"),
                "desired_state": "off",
            }
        )
    else:
        if success:
            messages.success(request, f"Cut off sent to {meter.meter_number}.")
        else:
            messages.error(
                request,
                f"Cut off failed for {meter.meter_number}: {res.get('error', 'no reply')}",
            )

        return _redirect_back(request)


@require_POST
def restore_meter(request, meter_id):
    """Restore (close relay) for a single meter — meter-number based, no IP needed."""
    meter = get_object_or_404(Meter, pk=meter_id)

    byCmd = RELAY_CLOSE_COMMAND  # ON (DL/T645 close/permit-close command)
    frame = build_switch_frame(meter.meter_number, byCmd)
    frame_hex = _as_hex(frame)
    cmd_name = "ON"

    try:
        frame_hex = frame.hex().upper()
    except AttributeError:
        frame_hex = str(frame)

    logger.info(
        "REQUEST from RESTORE_METER user=%s path=%s meter=%s cmd=%s(0x%02X) frame=%s",
        getattr(request.user, "username", "anonymous"),
        getattr(request, "path", ""),
        meter.meter_number,
        cmd_name,
        byCmd,
        frame_hex,
    )

    logger.info("-------------------------------------")

    if DISABLE_CUTOFFS:
        res = {
            "ok": False,
            "error": "switching disabled by DISABLE_CUTOFFS",
            "payload": "skipped:DISABLE_CUTOFFS",
        }
        logger.info(
            "RESPONSE meter=%s cmd=%s ok=%s error=%s payload=%s",
            meter.meter_number,
            cmd_name,
            res.get("ok"),
            res.get("error"),
            res.get("payload"),
        )
    else:
        try:
            secret = getattr(settings, "METER_CTRL_SECRET", None)  # optional
            res = _call_send(
                meter_number=meter.meter_number,
                frame=frame,
                timeout=32.0,
                expect_di=None,
                allow_switch=True,  # <-- explicit
                initiated_by=request.user.get_username(),  # <-- who clicked
                reason="manual switch from UI",  # <-- audit
                command_type="relay",
                desired_state="on",
                source="manual",
                auth=secret,  # <-- optional shared secret
            )
            logger.info(
                "RESPONSE meter=%s cmd=%s ok=%s error=%s payload=%s",
                meter.meter_number,
                cmd_name,
                res.get("ok"),
                res.get("error"),
                res.get("payload"),
            )
        except Exception as e:
            logger.exception(
                "SEND_FAILED meter=%s cmd=%s error=%s", meter.meter_number, cmd_name, e
            )
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                # blank line separator

                return JsonResponse({"success": False, "error": str(e)}, status=500)
            messages.error(request, f"Restore failed for {meter.meter_number}: {e}")
            # blank line separator

            return _redirect_back(request)

    success = bool(res.get("ok"))

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse(
            {
                "success": success,
                "error": res.get("error"),
                "command_id": res.get("command_id"),
                "command_status": res.get("status"),
                "desired_state": "on",
            }
        )
    else:
        if success:
            messages.success(request, f"Restore sent to {meter.meter_number}.")
        else:
            messages.error(
                request,
                f"Restore failed for {meter.meter_number}: {res.get('error', 'no reply')}",
            )

        return _redirect_back(request)


def unknown_meter_list(request):
    q = request.GET.get("q", "").strip()
    qs = UnknownMeter.objects.filter(status="new").order_by("-last_seen")
    if q:
        qs = qs.filter(meter_number__icontains=q)
    return render(
        request, "smart_meter/unknown_meter_list.html", {"unknown_meters": qs, "q": q}
    )


@transaction.atomic
def unknown_meter_convert(request, pk):
    um = get_object_or_404(UnknownMeter, pk=pk)
    if request.method == "POST":
        form = UnknownToMeterForm(
            request.POST, initial={"meter_number": um.meter_number}
        )
        if form.is_valid():
            meter = form.save(commit=False)
            meter.meter_number = um.meter_number  # enforce
            meter.save()
            # unit is not a field on Meter in some schemas; if your Meter has FK unit, then save above already covered it
            um.status = "added"
            um.save(update_fields=["status"])
            messages.success(request, f"Meter {meter.meter_number} created.")
            return redirect("smart_meter:unknown_meter_list")
    else:
        form = UnknownToMeterForm(initial={"meter_number": um.meter_number})
    return render(
        request, "smart_meter/unknown_meter_convert.html", {"um": um, "form": form}
    )


def unknown_meter_ignore(request, pk):
    um = get_object_or_404(UnknownMeter, pk=pk)
    um.status = "ignored"
    um.save(update_fields=["status"])
    messages.info(request, f"Ignored {um.meter_number}.")
    return redirect("smart_meter:unknown_meter_list")


# smart_meter/views.py


@transaction.atomic
def unknown_meter_quick_add(request, pk):
    um = get_object_or_404(UnknownMeter, pk=pk)
    # Create Meter with just the number if it doesn't exist
    meter, created = Meter.objects.get_or_create(
        meter_number=um.meter_number,
        defaults={
            # optional defaults—adjust to your Meter fields
            "power_status": "on",  # or your model’s default/choice
        },
    )
    # mark unknown as added
    um.status = "added"
    um.save(update_fields=["status"])
    if created:
        messages.success(request, f"✅ Meter {meter.meter_number} created.")
    else:
        messages.info(
            request, f"ℹ️ Meter {meter.meter_number} already existed; marked as added."
        )
    return redirect("smart_meter:unknown_meter_list")


# smart_meter/views.py

try:
    from leases.models import Lease
except Exception:
    Lease = None

try:
    from smart_meter.utils import build_whatsapp_url
except Exception:
    build_whatsapp_url = None


# smart_meter/views.py


try:
    from leases.models import Lease
except Exception:
    Lease = None

try:
    from smart_meter.utils import build_whatsapp_url
except Exception:
    build_whatsapp_url = None


BILLING_RATE = Decimal("7.50")


def _filtered_meter_sets(request, include_meter_property=True):
    """Return (meters_qs, all_properties, filtered_units, filtered_meters, current ids) for meter_filters.html."""
    prop_id = (request.GET.get("property") or "").strip()
    unit_id = (request.GET.get("unit") or "").strip()
    meter_id = (request.GET.get("meter") or "").strip()
    role = (request.GET.get("role") or "").strip().lower()

    all_properties = Property.objects.only("id", "property_name").order_by(
        "property_name"
    )

    units_qs = Unit.objects.only("id", "property_id", "unit_number")
    if prop_id:
        units_qs = units_qs.filter(property_id=prop_id)
    filtered_units = units_qs.order_by("unit_number")

    meter_related = ("unit", "unit__property") if include_meter_property else ("unit",)
    meters_qs = Meter.objects.select_related(*meter_related)
    if include_meter_property:
        meters_qs = meters_qs.only(
            "id",
            "meter_number",
            "name",
            "unit",
            "unit__id",
            "unit__unit_number",
            "unit__property",
            "unit__property__id",
            "unit__property__property_name",
        )
    else:
        meters_qs = meters_qs.only(
            "id",
            "meter_number",
            "name",
            "unit",
            "unit__id",
            "unit__unit_number",
        )
    if unit_id:
        meters_qs = meters_qs.filter(unit_id=unit_id)
    elif prop_id:
        meters_qs = meters_qs.filter(unit__property_id=prop_id)
    if role in (Meter.METER_ROLE_BILLING, Meter.METER_ROLE_CHECK):
        meters_qs = meters_qs.filter(meter_role=role)
    filtered_meters = meters_qs.order_by("meter_number")

    # Final meter set (what charts/tables use)
    selected_meters = filtered_meters
    if meter_id:
        selected_meters = selected_meters.filter(id=meter_id)

    return (
        selected_meters,
        all_properties,
        filtered_units,
        filtered_meters,
        prop_id,
        unit_id,
        meter_id,
    )


# smart_meter/views.py — replace energy_dashboard with this version


# smart_meter/views.py


try:
    from leases.models import Lease
except Exception:
    Lease = None

try:
    from smart_meter.utils import build_whatsapp_url
except Exception:
    build_whatsapp_url = None


def _aware_midnight(d: date):
    tz = timezone.get_current_timezone()
    return timezone.make_aware(datetime.combine(d, time.min), tz)


def _meter_q(param: str) -> Q:
    """Allow meter dropdown to use either ID or meter_number."""
    if not param:
        return Q()
    if param.isdigit():
        return Q(id=int(param)) | Q(meter_number=param)
    return Q(meter_number=param)


def energy_dashboard(request):
    """
    - Default: ALL meters, current month, daily usage lines (one line per meter).
    - If monthly + ALL meters: grouped bars (one bar per meter per month).
    - If ONE meter + hourly: hourly usage line for the chosen range.
    - Data labels are shown on the chart (values on points/bars).
    """
    # ---------- date window (default = current month) ----------
    today = date.today()
    start_date = date(today.year, today.month, 1)
    end_date = today
    if request.GET.get("start") and request.GET.get("end"):
        try:
            start_date = date.fromisoformat(request.GET["start"])
            end_date = date.fromisoformat(request.GET["end"])
        except Exception:
            pass

    dt_start = _aware_midnight(start_date)  # inclusive
    dt_end_excl = _aware_midnight(end_date + timedelta(days=1))  # exclusive

    report_type = request.GET.get("report_type", "daily")  # daily | monthly | hourly

    # ---------- filters (same keys as meter_filters.html) ----------
    prop_id = (request.GET.get("property") or "").strip()
    unit_id = (request.GET.get("unit") or "").strip()
    meter_param = (request.GET.get("meter") or "").strip()  # id or meter_number

    all_properties = Property.objects.all().order_by("property_name")

    units_qs = Unit.objects.all()
    if prop_id:
        units_qs = units_qs.filter(property_id=prop_id)
    filtered_units = units_qs.order_by("unit_number")

    meters_qs = Meter.objects.select_related("unit", "unit__property")
    if unit_id:
        meters_qs = meters_qs.filter(unit_id=unit_id)
    elif prop_id:
        meters_qs = meters_qs.filter(unit__property_id=prop_id)
    filtered_meters = meters_qs.order_by("meter_number")

    if meter_param:
        filtered_meters = filtered_meters.filter(_meter_q(meter_param))

    selected_meters = filtered_meters
    per_meter_mode = bool(meter_param)

    # ---------- live/per-meter status card ----------
    selected_meter = None
    unit = None
    latest_ts = latest_voltage_a = latest_current_a = latest_total_energy = None
    balance_obj = None
    wa_url = None

    if per_meter_mode:
        selected_meter = (
            Meter.objects.select_related("unit", "unit__property")
            .filter(_meter_q(meter_param))
            .first()
        )
        if selected_meter:
            unit = selected_meter.unit
            # Prefer live
            live = (
                LiveReading.objects.filter(meter=selected_meter).order_by("-ts").first()
            )
            if live:
                latest_ts = live.ts
                latest_voltage_a = live.voltage_a
                latest_current_a = live.current_a
                latest_total_energy = live.total_energy
            else:
                snap = (
                    MeterReading.objects.filter(meter=selected_meter)
                    .order_by("-ts")
                    .first()
                )
                if snap:
                    latest_ts = snap.ts
                    latest_voltage_a = snap.voltage_a
                    latest_current_a = snap.current_a
                    latest_total_energy = snap.total_energy

            # Balance + optional WA alert link
            balance_obj, _ = MeterBalance.objects.get_or_create(unit=unit)
            if (
                balance_obj
                and balance_obj.balance is not None
                and Lease
                and build_whatsapp_url
                and balance_obj.balance <= selected_meter.min_balance_alert
            ):
                try:
                    lease = Lease.objects.filter(unit=unit).latest("start_date")
                    if getattr(lease, "tenant", None) and lease.tenant.phone:
                        msg = (
                            f"⚠️ Dear {lease.tenant.get_full_name()}, your meter balance is "
                            f"{format_money(balance_obj.balance, GlobalSettings.get_solo())}. Please recharge soon to avoid disconnection."
                        )
                        wa_url = build_whatsapp_url(lease.tenant.phone, msg)
                except Exception:
                    pass

    # ---------- pull snapshots once. bucket in Python (no DB tz deps) ----------
    tz = timezone.get_current_timezone()

    # Special case: hourly buckets only make sense for ONE meter;
    # if hourly with ALL meters, we’ll return empty series and the template will nudge to pick a meter.
    hourly_mode = report_type == "hourly"

    base_qs = (
        MeterReading.objects.filter(
            meter__in=selected_meters, ts__gte=dt_start, ts__lt=dt_end_excl
        )
        .values("meter_id", "ts", "total_energy")
        .order_by("meter_id", "ts")
    )

    debug_snapshot_count_in_window = base_qs.count()

    # Fallback: if window empty, try the last 7 days with data
    if debug_snapshot_count_in_window == 0:
        latest = (
            MeterReading.objects.filter(meter__in=selected_meters)
            .order_by("-ts")
            .values_list("ts", flat=True)
            .first()
        )
        if latest:
            last_day = latest.astimezone(tz).date()
            fb_start = _aware_midnight(last_day - timedelta(days=6))
            fb_end = _aware_midnight(last_day + timedelta(days=1))
            base_qs = (
                MeterReading.objects.filter(
                    meter__in=selected_meters, ts__gte=fb_start, ts__lt=fb_end
                )
                .values("meter_id", "ts", "total_energy")
                .order_by("meter_id", "ts")
            )
            debug_snapshot_count_in_window = base_qs.count()

    # ---------- Python bucketing ----------
    # Per meter per period min/max => usage
    per_meter_period_minmax = defaultdict(lambda: {"min": None, "max": None})
    # Per meter whole window min/max for totals
    per_meter_window_minmax = defaultdict(lambda: {"min": None, "max": None})
    # For chart datasets: usage per meter per period (preserve insertion order of periods)
    per_meter_usage = defaultdict(lambda: OrderedDict())

    for row in base_qs:
        mid = row["meter_id"]
        ts = row["ts"].astimezone(tz)
        val = row["total_energy"]
        if val is None:
            continue

        if hourly_mode and per_meter_mode:
            # bucket to exact hour
            period_key = ts.replace(minute=0, second=0, microsecond=0)
        elif report_type == "monthly":
            period_key = date(ts.year, ts.month, 1)
        else:  # daily (default)
            period_key = ts.date()

        # update min/max for that (meter, period)
        key = (mid, period_key)
        mm = per_meter_period_minmax[key]
        mm["min"] = val if mm["min"] is None or val < mm["min"] else mm["min"]
        mm["max"] = val if mm["max"] is None or val > mm["max"] else mm["max"]

        # whole-window min/max
        mw = per_meter_window_minmax[mid]
        mw["min"] = val if mw["min"] is None or val < mw["min"] else mw["min"]
        mw["max"] = val if mw["max"] is None or val > mw["max"] else mw["max"]

    # Build per meter usage map & a sorted list of periods
    # (we want consistent x-axis across datasets; fill missing with 0)
    period_set = set()
    for (mid, p), mm in per_meter_period_minmax.items():
        if mm["min"] is None or mm["max"] is None:
            continue
        use = Decimal(mm["max"]) - Decimal(mm["min"])
        if use < 0:
            use = Decimal(0)
        per_meter_usage[mid][p] = use
        period_set.add(p)

    periods_sorted = sorted(period_set)
    if hourly_mode and per_meter_mode:
        x_labels = [p.strftime("%d %b %H:00") for p in periods_sorted]
    elif report_type == "monthly":
        x_labels = [p.strftime("%b %Y") for p in periods_sorted]
    else:
        x_labels = [p.strftime("%b %d") for p in periods_sorted]

    # Chart datasets:
    id_to_number = dict(selected_meters.values_list("id", "meter_number"))
    series_datasets = []
    for mid in selected_meters.values_list("id", flat=True):
        # keep series order by current queryset order
        data = [float(per_meter_usage[mid].get(p, Decimal(0))) for p in periods_sorted]
        series_datasets.append(
            {
                "label": id_to_number.get(mid, f"Meter {mid}"),
                "data": data,
                # No explicit colors; Chart.js picks defaults. (User asked for values displayed; we do via datalabels plugin.)
            }
        )

    # Totals & cost use the resolved Global -> Property -> Unit -> Meter rate.
    monthly_total = Decimal(0)
    monthly_cost = Decimal(0)
    from smart_meter.rates import resolve_electricity_rate

    rate_map = {
        meter.id: resolve_electricity_rate(meter=meter).rate
        for meter in selected_meters.select_related("unit", "unit__property")
    }
    for mid, mm in per_meter_window_minmax.items():
        if mm["min"] is None or mm["max"] is None:
            continue
        use = Decimal(mm["max"]) - Decimal(mm["min"])
        if use < 0:
            use = Decimal(0)
        monthly_total += use
        monthly_cost += use * Decimal(rate_map.get(mid) or 0)

    # “billing_rate” display: single meter => that meter’s rate; all meters => only show if all rates are same
    if per_meter_mode and selected_meter:
        billing_rate = resolve_electricity_rate(meter=selected_meter).rate
    else:
        distinct_rates = set(rate_map.values())
        billing_rate = next(iter(distinct_rates)) if len(distinct_rates) == 1 else None

    # Fleet online/offline counts
    online_count = offline_count = 0
    if not per_meter_mode:
        cutoff = timezone.now() - timedelta(minutes=online_threshold_minutes())
        latest_live = {
            lr.meter_id: lr.ts
            for lr in LiveReading.objects.filter(meter__in=selected_meters)
        }
        for mid in selected_meters.values_list("id", flat=True):
            ts_live = latest_live.get(mid)
            if ts_live and ts_live >= cutoff:
                online_count += 1
            else:
                offline_count += 1

    context = {
        # filters & choices
        "all_properties": all_properties,
        "filtered_units": filtered_units,
        "filtered_meters": filtered_meters,  # used by template dropdown
        "meters": filtered_meters,  # alias for backward‐compat
        "current_property": prop_id,
        "current_unit": unit_id,
        "current_meter": meter_param,
        # live/per-meter card
        "selected_meter": selected_meter,
        "unit": unit,
        "latest_ts": latest_ts,
        "latest_voltage_a": latest_voltage_a,
        "latest_current_a": latest_current_a,
        "latest_total_energy": latest_total_energy,
        "balance": balance_obj,
        "wa_url": wa_url,
        # chart series
        "report_type": report_type,
        "start_date": start_date,
        "end_date": end_date,
        "series_labels": x_labels,
        "series_datasets": series_datasets,
        # summaries
        "monthly_total": monthly_total,
        "monthly_cost": monthly_cost,
        "billing_rate": billing_rate,  # configured currency/kWh (None if mixed)
        "online_count": online_count,
        "offline_count": offline_count,
        "online_minutes": online_threshold_minutes(),
        "selected_property_id": selected_property_id,
        "selected_unit_id": selected_unit_id,
        "selected_meter_id": selected_meter_id,
        # debug (optional)
        "debug_selected_meters_count": selected_meters.count(),
        "debug_snapshot_count_in_window": debug_snapshot_count_in_window,
        "debug_meter_param": meter_param or "(all)",
    }
    return render(request, "smart_meter/dashboard.html", context)


def _filtered_meter_sets(request, include_meter_property=True):
    """
    Reuse the exact same GET keys your meter filters use: ?property=, ?unit=, ?meter=
    Returns (selected_meters, all_properties, filtered_units,
             filtered_meters, prop_id, unit_id, meter_id)
    """
    prop_id = (request.GET.get("property") or "").strip()
    unit_id = (request.GET.get("unit") or "").strip()
    meter_id = (request.GET.get("meter") or "").strip()
    role = (request.GET.get("role") or "").strip().lower()

    all_properties = Property.objects.only("id", "property_name").order_by(
        "property_name"
    )

    units_qs = Unit.objects.only("id", "property_id", "unit_number")
    if prop_id:
        units_qs = units_qs.filter(property_id=prop_id)
    filtered_units = units_qs.order_by("unit_number")

    meter_related = ("unit", "unit__property") if include_meter_property else ("unit",)
    meters_qs = Meter.objects.select_related(*meter_related)
    if include_meter_property:
        meters_qs = meters_qs.only(
            "id",
            "meter_number",
            "name",
            "unit",
            "unit__id",
            "unit__unit_number",
            "unit__property",
            "unit__property__id",
            "unit__property__property_name",
        )
    else:
        meters_qs = meters_qs.only(
            "id",
            "meter_number",
            "name",
            "unit",
            "unit__id",
            "unit__unit_number",
        )
    if unit_id:
        meters_qs = meters_qs.filter(unit_id=unit_id)
    elif prop_id:
        meters_qs = meters_qs.filter(unit__property_id=prop_id)
    if role in (Meter.METER_ROLE_BILLING, Meter.METER_ROLE_CHECK):
        meters_qs = meters_qs.filter(meter_role=role)
    filtered_meters = meters_qs.order_by("meter_number")

    # Final meter set for charts/tables:
    selected_meters = filtered_meters
    if meter_id:
        selected_meters = selected_meters.filter(id=meter_id)

    return (
        selected_meters,
        all_properties,
        filtered_units,
        filtered_meters,
        prop_id,
        unit_id,
        meter_id,
    )


def _parse_meter_param(meter_param: str):
    """
    Accept both a numeric meter PK (id) or a meter_number string.
    Return a Q object to filter meters accordingly.
    """
    if not meter_param:
        return Q()
    # If it looks like an integer PK
    if meter_param.isdigit():
        return Q(id=int(meter_param)) | Q(meter_number=meter_param)
    # Otherwise treat it as meter_number
    return Q(meter_number=meter_param)


@login_required
@require_POST
def fetch_meter_data(request):
    try:
        # Simulate data fetching - in real app, this would call your API
        # Update last_updated timestamp for all meters
        from .models import Meter

        Meter.objects.update(last_updated=timezone.now())

        return JsonResponse(
            {"status": "success", "message": "Meter data refreshed successfully"}
        )
    except Exception as e:
        return JsonResponse(
            {"status": "error", "message": f"Failed to fetch data: {e!s}"}, status=500
        )


# views.py

QUICK_RANGES = {
    "today",
    "yesterday",
    "this_week",
    "last_week",
    "this_month",
    "last_month",
    "custom",
    "",
}


def _range_to_dates(range_key: str):
    t = timezone.localdate()
    if range_key == "today":
        return t, t
    if range_key == "yesterday":
        y = t - timedelta(days=1)
        return y, y
    if range_key == "this_week":
        s = t - timedelta(days=t.weekday())
        e = s + timedelta(days=6)
        return s, e
    if range_key == "last_week":
        s = t - timedelta(days=t.weekday() + 7)
        e = s + timedelta(days=6)
        return s, e
    if range_key == "this_month":
        s = t.replace(day=1)
        e = t.replace(day=calendar.monthrange(t.year, t.month)[1])
        return s, e
    if range_key == "last_month":
        y, m = (t.year - 1, 12) if t.month == 1 else (t.year, t.month - 1)
        s = date(y, m, 1)
        e = date(y, m, calendar.monthrange(y, m)[1])
        return s, e
    return None, None


def _reading_date_window_from_request(request):
    range_key = (request.GET.get("range") or "").strip()
    if range_key not in QUICK_RANGES:
        range_key = ""

    start_str = (request.GET.get("start") or "").strip()
    end_str = (request.GET.get("end") or "").strip()

    start_date = end_date = None
    if range_key and range_key != "custom":
        start_date, end_date = _range_to_dates(range_key)
    else:
        try:
            if start_str:
                start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
            if end_str:
                end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
        except ValueError:
            start_date = end_date = None

    tz = timezone.get_current_timezone()

    def aware_start(d: date | None):
        if not d:
            return None
        return timezone.make_aware(datetime.combine(d, time.min), tz)

    def aware_end_exclusive(d: date | None):
        if not d:
            return None
        next_day = d + timedelta(days=1)
        return timezone.make_aware(datetime.combine(next_day, time.min), tz)

    return (
        range_key,
        start_date,
        end_date,
        aware_start(start_date),
        aware_end_exclusive(end_date),
    )


def _reading_local_date(reading):
    ts = getattr(reading, "ts", None)
    if not ts:
        return None
    if timezone.is_aware(ts):
        return timezone.localtime(ts).date()
    return ts.date()


def reading_list(request):
    page_size = 100
    prop_id = request.GET.get("property") or ""
    unit_id = request.GET.get("unit") or ""
    meter_id = request.GET.get("meter") or ""
    role = (request.GET.get("role") or "").strip().lower()
    active_filter = (request.GET.get("active") or "").strip().lower()

    range_key = (request.GET.get("range") or "").strip()
    if range_key not in QUICK_RANGES:
        range_key = ""  # treat unknown as "All time"

    start_str = (request.GET.get("start") or "").strip()
    end_str = (request.GET.get("end") or "").strip()

    # ---------- Build date window as DATES first ----------
    # Priority:
    #   1) If range_key is a known preset (and not "custom"): use it
    #   2) else parse start/end strings (custom/manual)
    start_date = end_date = None
    if range_key and range_key != "custom":
        # must return date objects (not datetimes)
        start_date, end_date = _range_to_dates(range_key)
    else:
        try:
            if start_str:
                start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
            if end_str:
                end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
        except ValueError:
            start_date = end_date = None

    # Normalize: if only one side provided, allow open-ended
    #   (leave the other as None and we’ll only apply the side we have)
    # ---------- Convert to TZ-AWARE DATETIMES ----------
    tz = timezone.get_current_timezone()

    def aware_start(d: date | None):
        if not d:
            return None
        return timezone.make_aware(datetime.combine(d, time.min), tz)

    def aware_end_exclusive(d: date | None):
        if not d:
            return None
        # end is inclusive by date -> exclusive at next day's midnight
        next_day = d + timedelta(days=1)
        return timezone.make_aware(datetime.combine(next_day, time.min), tz)

    start_dt = aware_start(start_date)
    end_dt_excl = aware_end_exclusive(end_date)

    # ---------- Base queryset ----------
    readings = MeterReading.objects.select_related(
        "meter", "meter__unit", "meter__unit__property"
    ).only(
        "id",
        "meter_id",
        "ts",
        "source_ip",
        "source_port",
        "total_energy",
        "forward_active_energy_kwh",
        "reverse_active_energy_kwh",
        "total_power",
        "pf_total",
        "voltage_a",
        "voltage_b",
        "voltage_c",
        "current_a",
        "current_b",
        "current_c",
        "power_a",
        "power_b",
        "power_c",
        "meter__id",
        "meter__meter_number",
        "meter__name",
        "meter__meter_role",
        "meter__reading_profile",
        "meter__is_active",
        "meter__unit_id",
        "meter__unit__id",
        "meter__unit__property_id",
        "meter__unit__unit_number",
        "meter__unit__property__id",
        "meter__unit__property__property_name",
    )

    if meter_id:
        readings = readings.filter(meter_id=meter_id)
    elif unit_id:
        readings = readings.filter(meter__unit_id=unit_id)
    elif prop_id:
        readings = readings.filter(meter__unit__property_id=prop_id)
    if role in (Meter.METER_ROLE_BILLING, Meter.METER_ROLE_CHECK):
        readings = readings.filter(meter__meter_role=role)
    if active_filter == "active":
        readings = readings.filter(meter__is_active=True)
    elif active_filter == "inactive":
        readings = readings.filter(meter__is_active=False)

    # ---------- Date filters (use datetime bounds; robust across time zones) ----------
    if start_dt:
        readings = readings.filter(ts__gte=start_dt)
    if end_dt_excl:
        readings = readings.filter(ts__lt=end_dt_excl)

    readings = readings.order_by("-ts")

    try:
        page_number = max(1, int(request.GET.get("page") or 1))
    except (TypeError, ValueError):
        page_number = 1

    offset = (page_number - 1) * page_size
    page_items = list(readings[offset : offset + page_size + 1])
    has_next = len(page_items) > page_size
    rows = page_items[:page_size]
    total_results = readings.count()
    total_pages = max(1, (total_results + page_size - 1) // page_size)
    page_numbers = range(
        max(1, page_number - 2),
        min(total_pages, page_number + 2) + 1,
    )
    for reading in rows:
        reading.display_forward_energy = (
            reading.forward_active_energy_kwh
            if reading.forward_active_energy_kwh is not None
            else reading.total_energy
        )
        reading.display_net_energy = (
            reading.display_forward_energy - reading.reverse_active_energy_kwh
            if reading.display_forward_energy is not None
            and reading.reverse_active_energy_kwh is not None
            else None
        )
    attach_active_meter_counts(rows, lambda reading: reading.meter)
    attach_tenant_names_for_dates(
        rows,
        lambda reading: reading.meter.unit_id if reading.meter else None,
        _reading_local_date,
    )

    class ReadingPage:
        def __init__(self, object_list, number, per_page, has_next_page):
            self.object_list = object_list
            self.number = number
            self.per_page = per_page
            self._has_next = has_next_page

        def __len__(self):
            return len(self.object_list)

        @property
        def has_previous(self):
            return self.number > 1

        @property
        def has_next(self):
            return self._has_next

        def previous_page_number(self):
            return max(1, self.number - 1)

        def next_page_number(self):
            return self.number + 1

        def start_index(self):
            if not self.object_list:
                return 0
            return ((self.number - 1) * self.per_page) + 1

    qs = request.GET.copy()
    qs.pop("page", None)

    meter_options = Meter.objects.only("id", "meter_number", "meter_role", "unit_id")
    filtered_meters_ctx = (
        meter_options.filter(unit_id=unit_id)
        if unit_id
        else meter_options.filter(unit__property_id=prop_id)
        if prop_id
        else meter_options
    )
    if role in (Meter.METER_ROLE_BILLING, Meter.METER_ROLE_CHECK):
        filtered_meters_ctx = filtered_meters_ctx.filter(meter_role=role)
    if active_filter == "active":
        filtered_meters_ctx = filtered_meters_ctx.filter(is_active=True)
    elif active_filter == "inactive":
        filtered_meters_ctx = filtered_meters_ctx.filter(is_active=False)

    ctx = dict(
        all_properties=Property.objects.only("id", "property_name").order_by(
            "property_name"
        ),
        filtered_units=(
            Unit.objects.only("id", "property_id", "unit_number").filter(
                property_id=prop_id
            )
            if prop_id
            else Unit.objects.only("id", "property_id", "unit_number")
        ).order_by("unit_number"),
        filtered_meters=filtered_meters_ctx.order_by("meter_number"),
        current_property=prop_id,
        current_unit=unit_id,
        current_meter=meter_id,
        current_role=role,
        current_active=active_filter,
        rows=rows,
        page_obj=ReadingPage(rows, page_number, page_size, has_next),
        page_numbers=page_numbers,
        total_pages=total_pages,
        range=range_key,  # keeps the dropdown state
        # still dates for the template's value="{{ start|date:'Y-m-d' }}"
        start=start_date,
        end=end_date,
        qs=qs.urlencode(),
    )
    return render(request, "smart_meter/reading_list.html", ctx)


# --- Reuse the exact same filtering logic for both list & exports ---


def _filtered_readings_qs(request):
    prop_id = request.GET.get("property") or ""
    unit_id = request.GET.get("unit") or ""
    meter_id = request.GET.get("meter") or ""
    q = request.GET.get("q") or ""
    role = (request.GET.get("role") or "").strip().lower()
    active_filter = (request.GET.get("active") or "").strip().lower()

    qs = MeterReading.objects.select_related(
        "meter", "meter__unit", "meter__unit__property"
    )

    if meter_id:
        qs = qs.filter(meter_id=meter_id)
    elif unit_id:
        qs = qs.filter(meter__unit_id=unit_id)
    elif prop_id:
        qs = qs.filter(meter__unit__property_id=prop_id)
    if role in (Meter.METER_ROLE_BILLING, Meter.METER_ROLE_CHECK):
        qs = qs.filter(meter__meter_role=role)
    if active_filter == "active":
        qs = qs.filter(meter__is_active=True)
    elif active_filter == "inactive":
        qs = qs.filter(meter__is_active=False)

    _, _, _, start_dt, end_dt_excl = _reading_date_window_from_request(request)
    if start_dt:
        qs = qs.filter(ts__gte=start_dt)
    if end_dt_excl:
        qs = qs.filter(ts__lt=end_dt_excl)

    if q:
        qs = qs.filter(
            Q(meter__meter_number__icontains=q)
            | Q(meter__unit__unit_number__icontains=q)
            | Q(meter__unit__property__property_name__icontains=q)
        )
    return qs


# smart_meter/views.py (replace the headers/rows in both exporters)


def _reading_export_chunks(qs, chunk_size=2000):
    offset = 0
    while True:
        rows = list(qs[offset : offset + chunk_size])
        if not rows:
            break
        attach_tenant_names_for_dates(
            rows,
            lambda reading: reading.meter.unit_id if reading.meter else None,
            _reading_local_date,
        )
        attach_active_meter_counts(rows, lambda reading: reading.meter)
        yield rows
        offset += chunk_size


def export_meter_readings_csv(request):
    qs = _filtered_readings_qs(request).order_by("-ts")
    now_str = timezone.now().strftime("%Y%m%d_%H%M%S")
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="meter_readings_{now_str}.csv"'
    writer = csv.writer(resp)

    headers = [
        "Timestamp",
        "Property",
        "Unit / Meter Name",
        "Tenant",
        "Meter",
        "Voltage_A(V)",
        "Current_A(A)",
        "Total_Power(W)",
        "Total_Energy(kWh)",
        "PF_Total",
    ]
    writer.writerow(headers)

    for rows in _reading_export_chunks(qs):
        for r in rows:
            ts = getattr(r, "ts", None)
            row = [
                ts.strftime("%Y-%m-%d %H:%M:%S") if ts else "",
                getattr(r.meter.unit.property, "property_name", ""),
                r.meter.display_location_name,
                getattr(r, "tenant_name", "Vacant"),
                r.meter.meter_number,
                r.voltage_a if r.voltage_a is not None else "",
                r.current_a if r.current_a is not None else "",
                r.total_power if r.total_power is not None else "",
                r.total_energy if r.total_energy is not None else "",
                r.pf_total if r.pf_total is not None else "",
            ]
            writer.writerow(row)
    return resp


def export_meter_readings_xlsx(request):
    qs = _filtered_readings_qs(request).order_by("-ts")
    now_str = timezone.now().strftime("%Y%m%d_%H%M%S")

    wb = Workbook()
    ws = wb.active
    ws.title = "Readings"

    headers = [
        "Timestamp",
        "Property",
        "Unit / Meter Name",
        "Tenant",
        "Meter",
        "Voltage_A(V)",
        "Current_A(A)",
        "Total_Power(W)",
        "Total_Energy(kWh)",
        "PF_Total",
    ]
    ws.append(headers)

    for rows in _reading_export_chunks(qs):
        for r in rows:
            ts = getattr(r, "ts", None)
            row = [
                ts.strftime("%Y-%m-%d %H:%M:%S") if ts else "",
                getattr(r.meter.unit.property, "property_name", ""),
                r.meter.display_location_name,
                getattr(r, "tenant_name", "Vacant"),
                r.meter.meter_number,
                r.voltage_a if r.voltage_a is not None else None,
                r.current_a if r.current_a is not None else None,
                r.total_power if r.total_power is not None else None,
                r.total_energy if r.total_energy is not None else None,
                r.pf_total if r.pf_total is not None else None,
            ]
            ws.append(row)

    # Auto width
    for col in ws.columns:
        try:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        except ValueError:
            max_len = 10
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 32)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = (
        f'attachment; filename="meter_readings_{now_str}.xlsx"'
    )
    return resp


# smart_meter/views.py (exporters)


def meters_export_csv(request):
    qs = _meters_annotated_qs(request, online_minutes=10)
    now_str = timezone.now().strftime("%Y%m%d_%H%M%S")

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="meters_{now_str}.csv"'
    w = csv.writer(resp)

    headers = [
        "Status",
        "Meter #",
        "Name",
        "Property",
        "Unit / Meter Name",
        "Power",
        "Unit Rate",
        "Min Alert",
        "Min Cutoff",
        "Active",
        "Installed",
        "Balance",
        "Last Reading",
        "Voltage_A(V)",
        "Current_A(A)",
        "Total_kWh",
    ]
    w.writerow(headers)

    for offset in range(0, qs.count(), 1000):
        meters = attach_active_meter_counts(qs[offset : offset + 1000])
        for m in meters:
            w.writerow(
                [
                    "Online" if m.is_online else "Offline",
                    m.meter_number,
                    m.name,
                    getattr(m.unit.property, "property_name", ""),
                    m.display_location_name,
                    m.power_status,
                    float(m.unit_rate) if m.unit_rate is not None else "",
                    float(m.min_balance_alert)
                    if m.min_balance_alert is not None
                    else "",
                    float(m.min_balance_cutoff)
                    if m.min_balance_cutoff is not None
                    else "",
                    "Yes" if m.is_active else "No",
                    m.installed_at.strftime("%Y-%m-%d") if m.installed_at else "",
                    float(m.balance) if m.balance is not None else "",
                    m.last_ts.strftime("%Y-%m-%d %H:%M:%S") if m.last_ts else "",
                    float(m.last_voltage_a) if m.last_voltage_a is not None else "",
                    float(m.last_current_a) if m.last_current_a is not None else "",
                    float(m.last_total_energy)
                    if m.last_total_energy is not None
                    else "",
                ]
            )
    return resp


def meters_export_xlsx(request):
    qs = _meters_annotated_qs(request, online_minutes=10)
    now_str = timezone.now().strftime("%Y%m%d_%H%M%S")

    wb = Workbook()
    ws = wb.active
    ws.title = "Meters"

    headers = [
        "Status",
        "Meter #",
        "Name",
        "Property",
        "Unit / Meter Name",
        "Power",
        "Unit Rate",
        "Min Alert",
        "Min Cutoff",
        "Active",
        "Installed",
        "Balance",
        "Last Reading",
        "Voltage_A(V)",
        "Current_A(A)",
        "Total_kWh",
    ]
    ws.append(headers)

    for offset in range(0, qs.count(), 1000):
        meters = attach_active_meter_counts(qs[offset : offset + 1000])
        for m in meters:
            ws.append(
                [
                    "Online" if m.is_online else "Offline",
                    m.meter_number,
                    m.name,
                    getattr(m.unit.property, "property_name", ""),
                    m.display_location_name,
                    m.power_status,
                    float(m.unit_rate) if m.unit_rate is not None else None,
                    float(m.min_balance_alert)
                    if m.min_balance_alert is not None
                    else None,
                    float(m.min_balance_cutoff)
                    if m.min_balance_cutoff is not None
                    else None,
                    "Yes" if m.is_active else "No",
                    m.installed_at.strftime("%Y-%m-%d") if m.installed_at else None,
                    float(m.balance) if m.balance is not None else None,
                    m.last_ts.strftime("%Y-%m-%d %H:%M:%S") if m.last_ts else None,
                    float(m.last_voltage_a) if m.last_voltage_a is not None else None,
                    float(m.last_current_a) if m.last_current_a is not None else None,
                    float(m.last_total_energy)
                    if m.last_total_energy is not None
                    else None,
                ]
            )

    # auto widths
    for col in ws.columns:
        try:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        except ValueError:
            max_len = 10
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 32)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="meters_{now_str}.xlsx"'
    return resp


# --- Hourly Report ---


def hourly_report(request):
    """
    Shows per-hour usage (ΔkWh within each hour) & max current for the selected day.
    If a single meter is chosen -> one dataset; otherwise one line per meter.
    Avoids DB time zone funcs by grouping in Python.
    """
    # ----- filters -----
    prop_id = (request.GET.get("property") or "").strip()
    unit_id = (request.GET.get("unit") or "").strip()
    meter_id = (request.GET.get("meter") or "").strip()
    day_str = (request.GET.get("day") or "").strip()

    # default = today
    tz = timezone.get_current_timezone()
    today = timezone.localtime(timezone.now(), tz).date()
    target_day = today
    if day_str:
        try:
            y, m, d = map(int, day_str.split("-"))
            target_day = datetime(y, m, d).date()
        except Exception:
            target_day = today

    # dropdown datasets (same pattern as elsewhere)
    all_properties = Property.objects.all().order_by("property_name")

    units_qs = Unit.objects.all()
    if prop_id:
        units_qs = units_qs.filter(property_id=prop_id)
    filtered_units = units_qs.order_by("unit_number")

    meters_qs = Meter.objects.select_related("unit", "unit__property")
    if unit_id:
        meters_qs = meters_qs.filter(unit_id=unit_id)
    elif prop_id:
        meters_qs = meters_qs.filter(unit__property_id=prop_id)
    filtered_meters = meters_qs.order_by("meter_number")

    # which meters to plot
    selected_meters = filtered_meters
    if meter_id:
        selected_meters = selected_meters.filter(id=meter_id)

    selected_meters = list(selected_meters)
    if not selected_meters:
        return render(
            request,
            "smart_meter/hourly_report.html",
            {
                "all_properties": all_properties,
                "filtered_units": filtered_units,
                "filtered_meters": filtered_meters,
                "current_property": prop_id,
                "current_unit": unit_id,
                "current_meter": meter_id,
                "target_day": target_day,
                "labels": [f"{h:02d}:00" for h in range(24)],
                "usage_series": [],
                "current_series": [],
                "table_rows": [],
            },
        )

    # ----- pull snapshots for that day -----
    day_start = timezone.make_aware(
        datetime(target_day.year, target_day.month, target_day.day, 0, 0, 0), tz
    )
    day_end = day_start + timedelta(days=1)

    qs = (
        MeterReading.objects.filter(
            meter__in=selected_meters, ts__gte=day_start, ts__lt=day_end
        )
        .values("meter_id", "ts", "total_energy", "current_a")
        .order_by("meter_id", "ts")
    )

    # group per meter → per hour
    # usage_per[meter_id][hour_index] = ΔkWh in that hour
    usage_per = {m.id: [0.0] * 24 for m in selected_meters}
    current_per = {m.id: [0.0] * 24 for m in selected_meters}

    # accumulate min/max kWh & max current per hour
    by_meter_hour = defaultdict(
        lambda: defaultdict(lambda: {"min": None, "max": None, "max_i": None})
    )

    for row in qs:
        mid = row["meter_id"]
        ts = row["ts"]
        # normalize to local tz and hour bucket
        ts_local = timezone.localtime(ts, tz)
        h = ts_local.hour

        kwh = Decimal(row["total_energy"] or 0)
        amp = Decimal(row["current_a"] or 0)

        bucket = by_meter_hour[mid][h]
        if bucket["min"] is None or kwh < bucket["min"]:
            bucket["min"] = kwh
        if bucket["max"] is None or kwh > bucket["max"]:
            bucket["max"] = kwh
        if bucket["max_i"] is None or amp > bucket["max_i"]:
            bucket["max_i"] = amp

    for m in selected_meters:
        for h in range(24):
            b = by_meter_hour[m.id].get(h)
            if not b:
                continue
            start = b["min"] or Decimal(0)
            end = b["max"] or Decimal(0)
            delta = end - start
            if delta < 0:
                delta = Decimal(0)
            usage_per[m.id][h] = float(delta)
            current_per[m.id][h] = float(b["max_i"] or 0)

    labels = [f"{h:02d}:00" for h in range(24)]

    # build chart series: one dataset per meter
    usage_series = []
    current_series = []
    for m in selected_meters:
        label = f"{m.meter_number} — {getattr(m.unit, 'unit_number', '')}"
        usage_series.append({"label": label, "data": usage_per[m.id]})
        current_series.append({"label": label, "data": current_per[m.id]})

    # simple table: hour + each meter’s usage (kWh)
    table_rows = []
    for idx, lab in enumerate(labels):
        row = {"hour": lab, "vals": []}
        for m in selected_meters:
            row["vals"].append(
                {
                    "label": m.meter_number,
                    "usage": usage_per[m.id][idx],
                    "current": current_per[m.id][idx],
                }
            )
        table_rows.append(row)

    ctx = {
        "all_properties": all_properties,
        "filtered_units": filtered_units,
        "filtered_meters": filtered_meters,
        "current_property": prop_id,
        "current_unit": unit_id,
        "current_meter": meter_id,
        "target_day": target_day,
        "labels": labels,
        "usage_series": usage_series,
        "current_series": current_series,
        "table_rows": table_rows,
    }
    return render(request, "smart_meter/hourly_report.html", ctx)


# --- add imports at the top of views.py ---

# vendor helpers


class SwitchPowerForm(forms.Form):
    meter = forms.ModelChoiceField(
        queryset=Meter.objects.order_by("meter_number"),
        label="Meter",
        widget=forms.Select(attrs={"class": "form-select"}),
    )
    action = forms.ChoiceField(
        choices=[("on", "Turn ON"), ("off", "Turn OFF")], widget=forms.RadioSelect
    )


def meter_switch(request):
    """Send ON/OFF to a meter using the vendor frame helper."""
    VIEW_NAME = "meter_switch"
    TEMPLATE_NAME = "smart_meter/control_switch.html"

    form = SwitchPowerForm(request.POST or None)
    context = {"form": form, "result": None}

    if request.method == "POST" and form.is_valid():
        meter = form.cleaned_data["meter"]
        on = form.cleaned_data["action"] == "on"
        byCmd = RELAY_CLOSE_COMMAND if on else RELAY_OPEN_COMMAND
        cmd_name = "ON" if on else "OFF"

        # Build frame
        frame = build_switch_frame(meter.meter_number, byCmd)
        try:
            frame_hex = frame.hex().upper()
        except AttributeError:
            frame_hex = str(frame)

        # ---- AUDIT: request
        logger.info(
            "REQUEST FROM METER_SWITCH view=%s template=%s method=%s user=%s path=%s meter=%s cmd=%s(0x%02X) frame=%s",
            VIEW_NAME,
            TEMPLATE_NAME,
            request.method,
            getattr(request.user, "username", "anonymous"),
            getattr(request, "path", ""),
            meter.meter_number,
            cmd_name,
            byCmd,
            frame_hex,
        )
        # blank line separator

        # Optional feature flag (set DISABLE_CUTOFFS=False in settings for real switching)
        if DISABLE_CUTOFFS:
            res = {
                "ok": False,
                "error": "switching disabled by DISABLE_CUTOFFS",
                "payload": "skipped:DISABLE_CUTOFFS",
            }
            ok = True
            logger.info(
                "RESPONSE view=%s meter=%s cmd=%s ok=%s error=%s payload=%s",
                VIEW_NAME,
                meter.meter_number,
                cmd_name,
                res.get("ok"),
                res.get("error"),
                res.get("payload"),
            )
        else:
            try:
                secret = getattr(settings, "METER_CTRL_SECRET", None)  # optional
                res = _send_switch(
                    meter_number=meter.meter_number,
                    frame=frame,
                    timeout=32.0,
                    expect_di=None,
                    allow_switch=True,  # <-- explicit
                    initiated_by=request.user.get_username(),  # <-- who clicked
                    reason="manual switch from UI",  # <-- audit
                    command_type="relay",
                    desired_state="on" if on else "off",
                    source="manual",
                    auth=secret,  # <-- optional shared secret
                )
                ok = bool(res.get("ok"))
                logger.info(
                    "RESPONSE view=%s meter=%s cmd=%s ok=%s error=%s payload=%s",
                    VIEW_NAME,
                    meter.meter_number,
                    cmd_name,
                    res.get("ok"),
                    res.get("error"),
                    res.get("payload"),
                )
            except Exception as e:
                logger.exception(
                    "SEND_FAILED view=%s meter=%s cmd=%s error=%s",
                    VIEW_NAME,
                    meter.meter_number,
                    cmd_name,
                    e,
                )
                messages.error(request, f"Failed: {e}")

                return render(
                    request,
                    TEMPLATE_NAME,
                    {**context, "result": {"ok": False, "error": str(e)}},
                )

        # Update state and feedback
        if ok:
            try:
                refresh_live(meter.meter_number)  # best-effort
            except Exception:
                pass
            messages.success(request, f"Command sent. Reply: {res.get('reply', '')}")
        else:
            messages.error(
                request, f"Failed: {res.get('error', 'no reply')} (meter may be busy)"
            )

        context["result"] = res
        logger.info("-------------------------------------")

    return render(request, TEMPLATE_NAME, context)


class PrepaidParamsForm(forms.Form):
    meter = forms.ModelChoiceField(
        queryset=Meter.objects.order_by("meter_number"),
        label="Meter",
        widget=forms.Select(attrs={"class": "form-select"}),
    )
    # keep it practical: two prices + two alarm levels + overdraft
    rate1_price = forms.DecimalField(
        label="Rate 1 price (Rs/kWh)", decimal_places=4, max_digits=10, initial=0
    )
    rate2_price = forms.DecimalField(
        label="Rate 2 price (Rs/kWh)", decimal_places=4, max_digits=10, initial=0
    )
    alarm1 = forms.DecimalField(
        label="Alarm amount 1 (Rs)", decimal_places=2, max_digits=10, initial=0
    )
    alarm2 = forms.DecimalField(
        label="Alarm amount 2 (Rs)", decimal_places=2, max_digits=10, initial=0
    )
    overdraft = forms.DecimalField(
        label="Overdraft limit (Rs)", decimal_places=2, max_digits=10, initial=0
    )


# smart_meter/views.py

# vendor
# your control client


def prepaid_params(request):
    if request.method == "POST":
        post_meter_id = request.POST.get("meter")
        instance = None
        if post_meter_id:
            try:
                instance = MeterPrepaidSettings.objects.select_related("meter").get(
                    meter_id=post_meter_id
                )
            except MeterPrepaidSettings.DoesNotExist:
                instance = None

        form = MeterPrepaidSettingsForm(request.POST, instance=instance)
        if form.is_valid():
            pps = form.save()  # don't shadow django settings
            frame = build_parameter_frame(
                pps.meter.meter_number, pps.to_vendor_parameters()
            )["frame"]
            frame_hex = _as_hex(frame)

            secret = getattr(
                dj_settings, "METER_CTRL_SECRET", None
            )  # read from Django settings

            # ---- DEFENSIVE GUARD: ensure we have a callable
            if send_via_db is None:
                messages.error(request, "Control sender unavailable")
                return redirect(
                    request.META.get("HTTP_REFERER")
                    or reverse("smart_meter:prepaid_params")
                )

            res = _call_send(
                meter_number=pps.meter.meter_number,  # <-- FIX
                frame=frame,
                timeout=32.0,
                expect_di=None,
                allow_switch=True,
                initiated_by=getattr(
                    request.user, "get_username", lambda: "anonymous"
                )(),
                reason="manual prepaid Parameter 1 update",
                auth=secret,
                command_type="prepaid_write",
                max_attempts=1,
            )
            if res.get("ok"):
                messages.success(
                    request, f"Prepaid parameters sent to {pps.meter.meter_number}."
                )
            else:
                messages.error(
                    request, f"Failed to send: {res.get('error', 'no reply')}"
                )
            return redirect("smart_meter:prepaid_params")
    else:
        form = MeterPrepaidSettingsForm()

    return render(request, "smart_meter/prepaid_params.html", {"form": form})


@require_POST
def bulk_power_action(request):
    """
    Bulk ON/OFF using the same logic as `meter_switch`.
    POST:
      - action: 'cutoff' | 'restore'
      - scope : 'selected' | 'negative'
      - meters: repeated meter IDs (when scope='selected')
    """
    action = (request.POST.get("action") or "").lower()
    scope = (request.POST.get("scope") or "selected").lower()
    ids = request.POST.getlist("meters")  # <input name="meters" ...>

    if action not in ("cutoff", "restore"):
        messages.error(request, "Invalid bulk action.")
        return redirect(
            request.META.get("HTTP_REFERER") or reverse("smart_meter:meter_list")
        )

    # Determine target meters strictly by meter_number (no IP dependency)
    if scope == "negative":
        meters_qs = Meter.objects.filter(unit__meterbalance__balance__lt=0)
    elif scope == "low":
        meters_qs = Meter.objects.filter(
            unit__meterbalance__balance__lt=F("min_balance_alert")
        )
    else:
        meters_qs = Meter.objects.filter(id__in=ids)

    total = meters_qs.count()
    if total == 0:
        messages.warning(request, "No meters to process.")
        return redirect(
            request.META.get("HTTP_REFERER") or reverse("smart_meter:meter_list")
        )

    byCmd = RELAY_CLOSE_COMMAND if action == "restore" else RELAY_OPEN_COMMAND
    cmd_name = "ON" if byCmd == RELAY_CLOSE_COMMAND else "OFF"
    ok_count = 0
    failures = []

    # ---- AUDIT: bulk header
    logger.info(
        "BULK_REQUEST FROM BULK_POWER_ACTION  user=%s path=%s action=%s scope=%s count=%s ids=%s",
        getattr(request.user, "username", "anonymous"),
        getattr(request, "path", ""),
        cmd_name,
        scope,
        total,
        ",".join(map(str, ids)) if ids else "",
    )
    # blank line separator

    for m in meters_qs.iterator():
        try:
            frame = build_switch_frame(m.meter_number, byCmd)
            try:
                frame_hex = frame.hex().upper()
            except AttributeError:
                frame_hex = str(frame)

            # per-meter REQUEST
            logger.info(
                "REQUEST user=%s meter=%s cmd=%s(0x%02X) frame=%s",
                getattr(request.user, "username", "anonymous"),
                m.meter_number,
                cmd_name,
                byCmd,
                frame_hex,
            )
            # blank line separator

            # optional skip in dev/safety
            if DISABLE_CUTOFFS:
                res = {
                    "ok": False,
                    "error": "switching disabled by DISABLE_CUTOFFS",
                    "payload": "skipped:DISABLE_CUTOFFS",
                }
                logger.info(
                    "RESPONSE meter=%s cmd=%s ok=%s error=%s payload=%s",
                    m.meter_number,
                    cmd_name,
                    res.get("ok"),
                    res.get("error"),
                    res.get("payload"),
                )
                # blank line separator

            else:
                secret = getattr(settings, "METER_CTRL_SECRET", None)  # optional

                # ---- DEFENSIVE GUARD: ensure we have a callable
                if send_via_db is None:
                    failures.append(f"{m.meter_number}: control sender unavailable")
                    logger.error(
                        "RESPONSE meter=%s cmd=%s ok=%s error=%s payload=%s",
                        m.meter_number,
                        cmd_name,
                        False,
                        "control sender unavailable",
                        None,
                    )
                    continue  # move on to the next meter

                res = _call_send(
                    meter_number=m.meter_number,
                    frame=frame,
                    timeout=32.0,
                    expect_di=None,
                    allow_switch=True,  # <-- explicit
                    initiated_by=request.user.get_username(),  # <-- who clicked
                    reason="manual switch from UI",  # <-- audit
                    command_type="relay",
                    desired_state="on" if action == "restore" else "off",
                    source="manual",
                    auth=secret,  # <-- optional shared secret
                )
                logger.info(
                    "RESPONSE meter=%s cmd=%s ok=%s error=%s payload=%s",
                    m.meter_number,
                    cmd_name,
                    res.get("ok"),
                    res.get("error"),
                    res.get("payload"),
                )
                # blank line separator

            if res.get("ok"):
                ok_count += 1
                # best-effort; don't count as failure
                try:
                    refresh_live(m.meter_number)
                except Exception:
                    pass
            else:
                failures.append(f"{m.meter_number}: {res.get('error', 'no reply')}")

        except Exception as e:
            logger.exception(
                "SEND_FAILED meter=%s cmd=%s error=%s", m.meter_number, cmd_name, e
            )
            failures.append(f"{m.meter_number}: {e}")

    # ---- AUDIT: bulk summary
    logger.info(
        "BULK_RESPONSE action=%s total=%s ok=%s failed=%s",
        cmd_name,
        total,
        ok_count,
        len(failures),
    )
    # blank line separator
    logger.info("-------------------------------------")

    if ok_count:
        messages.success(
            request, f"{action.title()} sent to {ok_count}/{total} meter(s)."
        )
    if failures:
        preview = "; ".join(failures[:5])
        more = f" (+{len(failures) - 5} more)" if len(failures) > 5 else ""
        messages.error(request, f"Failed for {len(failures)} meter(s): {preview}{more}")

    return redirect(
        request.META.get("HTTP_REFERER") or reverse("smart_meter:meter_list")
    )


# Import the vendor frame builder (support both layouts)
try:
    from smart_meter.vendor.switch_OnOff import frame_command as build_switch_frame
except Exception:
    # fallback if no vendor/ folder
    from smart_meter.switch_OnOff import frame_command as build_switch_frame


def switch_lab(request):
    """
    Build (and optionally send) the vendor ON/OFF frame purely from meter_number.
    """
    VIEW_NAME = "switch_lab"
    TEMPLATE_NAME = "smart_meter/switch_lab.html"

    form = SwitchLabForm(request.POST or None)
    result = None
    send_result = None

    if request.method == "POST" and form.is_valid():
        meter_hex = form.cleaned_data["meter_number"]
        byCmd = (
            RELAY_CLOSE_COMMAND
            if form.cleaned_data["action"] == "on"
            else RELAY_OPEN_COMMAND
        )
        cmd_name = "ON" if byCmd == RELAY_CLOSE_COMMAND else "OFF"

        # Build frame via vendor function
        frame = build_switch_frame(meter_hex, byCmd)
        try:
            frame_hex = frame.hex().upper()
        except AttributeError:
            frame_hex = str(frame)

        # ---- AUDIT: request
        logger.info(
            "REQUEST FROM SWITCH_LAB view=%s template=%s method=%s user=%s path=%s meter=%s cmd=%s(0x%02X) frame=%s preview_only=%s",
            VIEW_NAME,
            TEMPLATE_NAME,
            request.method,
            getattr(request.user, "username", "anonymous"),
            getattr(request, "path", ""),
            meter_hex,
            cmd_name,
            byCmd,
            frame_hex,
            form.cleaned_data.get("preview_only"),
        )
        # blank line separator

        # Prepare preview info for the page
        result = {
            "cmd": cmd_name,
            "meter": meter_hex,
            "length": len(frame),
            "hex": frame_hex,
        }

        # Optionally send via listener (unless preview_only)
        if not form.cleaned_data.get("preview_only"):
            if DISABLE_CUTOFFS:
                send_result = {
                    "ok": False,
                    "error": "switching disabled by DISABLE_CUTOFFS",
                    "payload": "skipped:DISABLE_CUTOFFS",
                }
                logger.info(
                    "RESPONSE view=%s meter=%s cmd=%s ok=%s error=%s payload=%s",
                    VIEW_NAME,
                    meter_hex,
                    cmd_name,
                    send_result.get("ok"),
                    send_result.get("error"),
                    send_result.get("payload"),
                )
                # blank line separator

            else:
                try:
                    send_result = send_via_db(
                        meter_number=meter_hex,
                        frame_hex=frame_hex,
                        timeout=32.0,
                        initiated_by=request.user.get_username(),
                        reason="manual switch from switch lab",
                        command_type="relay",
                        desired_state="on" if cmd_name == "ON" else "off",
                        source="manual",
                    )
                    logger.info(
                        "RESPONSE view=%s meter=%s cmd=%s ok=%s error=%s payload=%s",
                        VIEW_NAME,
                        meter_hex,
                        cmd_name,
                        send_result.get("ok"),
                        send_result.get("error"),
                        send_result.get("payload"),
                    )
                    # blank line separator

                except Exception as e:
                    send_result = {"ok": False, "error": str(e)}
                    logger.exception(
                        "SEND_FAILED view=%s meter=%s cmd=%s error=%s",
                        VIEW_NAME,
                        meter_hex,
                        cmd_name,
                        e,
                    )

            # Flash UI messages (optional)
            if send_result.get("ok"):
                messages.success(request, "Command sent successfully.")
            else:
                messages.error(
                    request, f"Send failed: {send_result.get('error', 'no reply')}"
                )

        # blank line separator
        logger.info("-------------------------------------")

    return render(
        request,
        TEMPLATE_NAME,
        {
            "form": form,
            "result": result,
            "send_result": send_result,
        },
    )


# smart_meter/views.py


@login_required
# Optional: restrict who can add readings
@permission_required("smart_meter.add_MeterReading", raise_exception=True)
def meter_reading_create(request):
    """
    Create a manual reading. Redirect back to the listing, preserving filters.
    """
    # Preserve filters / return path
    next_qs = request.GET.urlencode() or request.META.get("QUERY_STRING", "")
    # explicit ?next=/smart_meter/readings...
    return_to = request.GET.get("next") or reverse("smart_meter:reading_list")
    if request.method == "POST":
        form = ReadingManualForm(request.POST, request=request)
        if form.is_valid():
            obj = form.save(commit=False)
            # obj.created_by = request.user  # if you have this field
            obj.save()
            messages.success(request, "Manual reading added.")
            return redirect(return_to)
    else:
        form = ReadingManualForm(request=request)

    return render(
        request, "smart_meter/reading_form.html", {"form": form, "return_to": return_to}
    )


# smart_meter/views.py


@login_required
@permission_required("smart_meter.change_meterreading", raise_exception=True)
def meter_reading_row_edit(request, pk):
    r = get_object_or_404(MeterReading, pk=pk)

    # Cancel returns display row
    if request.GET.get("cancel"):
        return render(
            request, "smart_meter/partials/reading_row_display.html", {"r": r}
        )

    if request.method == "POST":
        form = ReadingManualForm(
            request.POST, instance=r, request=request, prefix=f"r{r.pk}"
        )
        if form.is_valid():
            form.save()
            # return the DISPLAY <tr> so the row snaps back
            return render(
                request, "smart_meter/partials/reading_row_display.html", {"r": r}
            )
        # invalid -> return EDIT <tr> with errors
        return render(
            request,
            "smart_meter/partials/reading_row_edit.html",
            {"r": r, "form": form},
            status=400,
        )

    # GET -> return EDIT <tr>
    form = ReadingManualForm(instance=r, request=request, prefix=f"r{r.pk}")
    return render(
        request, "smart_meter/partials/reading_row_edit.html", {"r": r, "form": form}
    )


@login_required
@permission_required("smart_meter.delete_meterreading", raise_exception=True)
@require_POST
def meter_reading_delete(request, pk):
    r = get_object_or_404(MeterReading, pk=pk)
    r.delete()
    # Redirect back to the same page (with the same filters)
    current_url = (
        request.headers.get("HX-Current-URL")
        or request.META.get("HTTP_REFERER")
        or reverse("smart_meter:reading_list")
    )
    resp = HttpResponse(status=204)  # no content needed
    resp["HX-Redirect"] = current_url  # HTMX does a client-side redirect
    return resp


@login_required
@permission_required("smart_meter.change_meterreading", raise_exception=True)
def meter_reading_row(request, pk):
    r = get_object_or_404(MeterReading, pk=pk)
    return render(request, "smart_meter/partials/reading_row_display.html", {"r": r})


# views.py


@require_POST
def reset_meter_display_balance(request, meter_id):
    meter = get_object_or_404(Meter, pk=meter_id)
    messages.error(
        request,
        "Display-balance reset is disabled until the meter's ESAM/MAC and "
        f"transaction requirements are verified ({meter.meter_number}).",
    )
    return redirect(request.META.get("HTTP_REFERER", "/"))


@require_POST
def set_meter_display_balance(request, meter_id):
    meter = get_object_or_404(Meter, pk=meter_id)

    try:
        amt = Decimal(request.POST.get("amount", "0"))
    except Exception:
        messages.error(request, "Invalid amount.")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    if amt < 0:
        messages.error(request, "Amount must be ≥ 0.")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    messages.error(
        request,
        "Display-balance writes are disabled until the meter's ESAM/MAC and "
        f"transaction requirements are verified ({meter.meter_number}, requested {amt:.2f}).",
    )
    return redirect(request.META.get("HTTP_REFERER", "/"))


# smart_meter/views.py
from django.db.models import Q


def _fmt(v, decimals=None):
    if v is None:
        return ""
    try:
        if decimals is None:
            return str(v)
        return f"{float(v):.{decimals}f}"
    except Exception:
        return str(v)


def _ts_iso(dt):
    return dt.isoformat() if dt else ""


@login_required
@require_POST
def instant_live_reading(request, meter_id):
    """Request one fresh DL/T645 0x028011FF reading without changing the normal reader."""
    import time as time_module

    meter = get_object_or_404(Meter, pk=meter_id)
    before = (
        LiveReading.objects.filter(meter=meter).values_list("ts", flat=True).first()
    )
    result = request_instant_live_reading(meter.meter_number, timeout=8.0)
    if not result.get("ok"):
        return JsonResponse(
            {"ok": False, "error": result.get("error") or "Instant reading failed."},
            status=409,
        )

    # The listener remains responsible for parsing and persisting the reading.
    deadline = time_module.time() + 2.0
    reading = None
    while time_module.time() < deadline:
        reading = LiveReading.objects.filter(meter=meter).first()
        if reading and (before is None or reading.ts > before):
            break
        time_module.sleep(0.1)
    if not reading or (before is not None and reading.ts <= before):
        return JsonResponse(
            {
                "ok": False,
                "error": "The meter replied, but a fresh live reading was not persisted yet. Refresh and try again.",
            },
            status=409,
        )

    relay_state = getattr(meter, "relay_state", None)
    return JsonResponse(
        {
            "ok": True,
            "meter_id": meter.pk,
            "meter_number": meter.meter_number,
            "updated_ts": _ts_iso(reading.ts),
            "total_energy": _fmt(reading.total_energy, 3),
            "forward_energy": _fmt(
                reading.forward_active_energy_kwh
                if reading.forward_active_energy_kwh is not None else reading.total_energy, 3
            ),
            "reverse_energy": _fmt(reading.reverse_active_energy_kwh, 3),
            "balance": _fmt(reading.balance, 2),
            "voltage_a": _fmt(reading.voltage_a, 1),
            "current_a": _fmt(reading.current_a, 3),
            "total_power": _fmt(reading.total_power, 3),
            "relay_state": relay_state or "unknown",
        }
    )


def live_custom_data(request):
    # keep filters identical to live_custom
    q = (request.GET.get("q") or "").strip()
    offline_only = request.GET.get("offline") == "1"
    active_filter = (request.GET.get("active") or "active").strip()

    (
        selected_meters,
        all_properties,
        filtered_units,
        filtered_meters,
        prop_id,
        unit_id,
        meter_id,
    ) = _filtered_meter_sets(request, include_meter_property=False)

    meter_scope_qs = _with_meter_operational_flags(
        _meters_annotated_qs(request, online_minutes=online_threshold_minutes())
    )
    if active_filter == "active":
        meter_scope_qs = meter_scope_qs.filter(is_active=True)
    elif active_filter == "inactive":
        meter_scope_qs = meter_scope_qs.filter(is_active=False)
    meter_scope_qs = _apply_meter_chip_filter(
        meter_scope_qs, _normalized_meter_chip(request)
    )

    qs = (
        LiveReading.objects.select_related(
            "meter", "meter__unit", "meter__unit__property"
        )
        .only(
            "id",
            "meter",
            "ts",
            "source_ip",
            "source_port",
            "balance",
            "total_energy",
            "forward_active_energy_kwh",
            "reverse_active_energy_kwh",
            "voltage_a",
            "current_a",
            "total_power",
            "pf_total",
            "status_word",
            "meter__id",
            "meter__unit",
            "meter__meter_number",
            "meter__power_status",
            "meter__name",
            "meter__is_active",
            "meter__meter_role",
            "meter__unit__id",
            "meter__unit__property",
            "meter__unit__unit_number",
            "meter__unit__property__id",
            "meter__unit__property__property_name",
        )
        .order_by(
            "meter__unit__property__property_name",
            "meter__unit__unit_number",
            "meter__meter_number",
        )
    )
    if meter_id:
        qs = qs.filter(meter_id=meter_id)
    elif unit_id:
        qs = qs.filter(meter__unit_id=unit_id)
    elif prop_id:
        qs = qs.filter(meter__unit__property_id=prop_id)
    if active_filter == "active":
        qs = qs.filter(meter__is_active=True)
    elif active_filter == "inactive":
        qs = qs.filter(meter__is_active=False)

    if q:
        qs = qs.filter(
            Q(meter__unit__unit_number__icontains=q)
            | Q(meter__meter_number__icontains=q)
            | Q(meter__unit__property__property_name__icontains=q)
        )

    qs = qs.filter(meter_id__in=meter_scope_qs.values("id"))

    tenant_info = active_tenant_info_for_units(
        qs.values_list("meter__unit_id", flat=True)
    )

    payload = []
    rows = attach_active_meter_counts(qs, lambda reading: reading.meter)
    latest_relay_commands = {}
    for command in MeterCommand.objects.filter(
        meter_id__in=[reading.meter_id for reading in rows],
        command_type="relay",
    ).order_by("meter_id", "-created_at"):
        latest_relay_commands.setdefault(command.meter_id, command)
    status_by_meter_id = resolve_meter_online_statuses(
        (reading.meter, reading) for reading in rows
    )
    for r in rows:
        status = status_by_meter_id[r.meter_id]
        if offline_only and status["is_online"]:
            continue

        m = r.meter
        u = m.unit
        p = u.property
        latest_command = latest_relay_commands.get(m.id)
        relay_state = reconcile_live_relay_command_state(
            m,
            latest_command,
            r.status_word,
            r.ts,
            is_fresh=status["measurement_is_fresh"],
        )
        confirmed_state = relay_state["confirmed_state"]

        payload.append(
            {
                "meter_id": m.id,
                "is_online": status["is_online"],
                "connection_state": status["connection_state"],
                "is_connected": status["is_connected"],
                "measurement_is_fresh": status["measurement_is_fresh"],
                "last_contact_at": _ts_iso(status["last_contact_at"]),
                "last_measurement_at": _ts_iso(status["last_measurement_at"]),
                "power_status": (m.power_status or "OFF").upper(),
                "relay_confirmed": bool(confirmed_state),
                "relay_confirmed_state": confirmed_state or "",
                "relay_confirmed_at": _ts_iso(r.ts) if confirmed_state else None,
                "relay_command_status": relay_state["status"],
                "relay_command_desired_state": relay_state["desired_state"],
                "relay_command_error": relay_state["error"],
                "relay_operation_label": relay_state["operation_label"],
                "relay_indicator_label": relay_state["indicator_label"],
                "relay_indicator_class": relay_state["indicator_class"],
                # values that map to your table columns
                "property_name": p.property_name or "",
                "property_short": (p.property_name or "")[:8],
                "unit_number": m.display_location_name,
                "tenant_name": tenant_info.get(u.id, {}).get("name", "Vacant"),
                "tenant_id": tenant_info.get(u.id, {}).get("tenant_id"),
                "meter_number": m.meter_number or "",
                "meter_role": m.meter_role,
                "meter_role_display": m.get_meter_role_display(),
                "updated_ts": _ts_iso(r.ts),
                # optional: pre-formatted display strings
                "source_ip": status["source_ip"] or r.source_ip or "",
                "port": status["source_port"] or r.source_port or "",
                "balance": _fmt(r.balance, 2),
                "total_energy": _fmt(r.total_energy, 3),
                "forward_energy": _fmt(
                    r.forward_active_energy_kwh
                    if r.forward_active_energy_kwh is not None else r.total_energy, 3
                ),
                "reverse_energy": _fmt(r.reverse_active_energy_kwh, 3),
                "voltage_a": _fmt(r.voltage_a, 1),
                "current_a": _fmt(r.current_a, 3),
                "total_power": _fmt(r.total_power, 3),
                "pf_total": _fmt(r.pf_total, 3),
            }
        )

    return JsonResponse({"rows": payload, "online_minutes": online_threshold_minutes()})
