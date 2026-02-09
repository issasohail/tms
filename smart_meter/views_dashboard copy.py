from __future__ import annotations
from django.http import JsonResponse, Http404
import re
from collections import defaultdict
# adjust app label if different
from invoices.models import Invoice, InvoiceItem, ItemCategory
from properties.models import Unit
from django.shortcuts import render
from django.db.models import Sum, Q
from datetime import date
from leases.models import Lease
# adjust app label if needed
from invoices.models import Invoice, InvoiceItem, ItemCategory
from smart_meter.models import Meter, MeterReading
from django.db.models import Sum
from calendar import monthrange
from datetime import date, datetime
from django.urls import reverse
from django.utils.http import urlencode
from decimal import Decimal
from weasyprint import HTML
from collections import OrderedDict, defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal, ROUND_HALF_UP
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.decorators import permission_required

from django.db.models import Min, Max
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from properties.models import Property, Unit
from smart_meter.models import Meter, MeterReading, LiveReading, MeterBalance
import json
from django.utils.safestring import mark_safe

KW_RE = re.compile(r"total\s*usage\s*[:=]\s*([0-9]+(?:\.[0-9]+)?)", re.I)
try:
    from leases.models import Lease
except Exception:
    Lease = None

BILLING_CURRENCY = "Rs."
ONLINE_MINUTES = 10


def _aware_range(start_d: date, end_d: date):
    tz = timezone.get_current_timezone()
    start_naive = datetime.combine(start_d, time.min)
    end_naive = datetime.combine(end_d, time.max)
    start_dt = timezone.make_aware(start_naive, tz)
    end_dt = timezone.make_aware(end_naive, tz) + timedelta(microseconds=1)
    return start_dt, end_dt, tz


def _period_key_and_label(ts_local: datetime, granularity: str):
    if granularity == "hourly":
        key = ts_local.replace(minute=0, second=0, microsecond=0)
        label = key.strftime("%b %d %H:00")
    elif granularity == "monthly":
        key = date(ts_local.year, ts_local.month, 1)
        label = key.strftime("%b %Y")
    else:  # daily
        key = ts_local.date()
        label = key.strftime("%b %d, %Y")
    return key, label


# smart_meter/views_dashboard.py


def _chain_rows_from_groups(groups, prev_end: Decimal | None = None):
    """
    Build rows so that each row's start_kwh equals the previous row's end_kwh.
    If prev_end is None (no earlier reading), we fall back to the group's min.
    """
    rows = []
    last_end = prev_end  # carry-forward baseline

    for _, g in groups.items():
        gmin = Decimal(g["min"])
        gmax = Decimal(g["max"])

        start_kwh = gmin if last_end is None else last_end
        end_kwh = gmax

        usage = end_kwh - start_kwh
        if usage < 0:
            # handle meter reset/replacement gracefully
            usage = Decimal("0")

        rows.append({
            "period_label": g["label"],
            "start_kwh": start_kwh,
            "end_kwh": end_kwh,
            "usage": usage,
        })
        last_end = end_kwh

    return rows


def _fmt0(n: Decimal | float | int) -> str:
    """Round half up to 0 decimals and format with comma."""
    d = Decimal(str(n)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    return f"{int(d):,}"


def _per_meter_series(meters_qs, start_d: date, end_d: date, granularity: str):
    """
    Returns:
      labels_asc, datasets, combined_rows_asc, totals_dict
    """
    start_dt, end_dt, tz = _aware_range(start_d, end_d)

    meter_to_rows = {}
    key_union = set()

    for m in meters_qs:
        qs = (MeterReading.objects
              .filter(meter=m, ts__gte=start_dt, ts__lt=end_dt)
              .order_by("ts")
              .values("ts", "total_energy"))

        # NEW: get the last reading *before* the start window to seed Begin kWh
        prev_snap = (MeterReading.objects
                     .filter(meter=m, ts__lt=start_dt)
                     .order_by("-ts")
                     .values("total_energy")
                     .first())
        prev_end = Decimal(
            str(prev_snap["total_energy"])) if prev_snap else None

        groups = OrderedDict()
        for r in qs:
            ts_local = timezone.localtime(r["ts"], tz)
            key, label = _period_key_and_label(ts_local, granularity)
            val = Decimal(str(r["total_energy"] or "0"))
            if key not in groups:
                groups[key] = {"min": val, "max": val, "label": label}
            else:
                groups[key]["max"] = val

        base_rows = _chain_rows_from_groups(groups, prev_end=prev_end)

        rows_with_key = []
        for k, g in groups.items():
            row = next(
                (x for x in base_rows if x["period_label"] == g["label"]), None)
            if row:
                r2 = dict(row)
                r2["period_key"] = k
                rows_with_key.append(r2)

        meter_to_rows[m] = rows_with_key
        for r in rows_with_key:
            key_union.add(r["period_key"])

    if not key_union:
        return [], [], [], {
            "total_kwh": Decimal("0"),
            "usage_charges": Decimal("0"),
            "service_charges": Decimal("0"),
            "grand_total": Decimal("0"),
        }

    # Oldest → Latest
    keys_sorted = sorted(key_union)

    def _label_from_key(k):
        if isinstance(k, datetime):
            return k.strftime("%b %d %H:00")
        elif isinstance(k, date):
            if k.day == 1:
                return k.strftime("%b %Y")
            return k.strftime("%b %d, %Y")
        return str(k)

    labels_sorted = [_label_from_key(k) for k in keys_sorted]

    datasets = []
    combined_rows = []
    total_kwh = Decimal("0")
    usage_charges = Decimal("0")
    service_total = Decimal("0")

    for m in meters_qs:
        unit_number = getattr(m.unit, "unit_number", "")
        legend_label = f"{unit_number} / {m.meter_number}"
        key_to_row = {r["period_key"]: r for r in meter_to_rows.get(m, [])}
        series = []
        for k in keys_sorted:
            row = key_to_row.get(k)
            series.append(float(row["usage"]) if row else None)

        datasets.append({
            "label": legend_label,
            "data": series,
            "meterNumber": m.meter_number,
            "unitNumber": unit_number,

        })

        rate = Decimal(str(m.unit_rate or "0"))
        svc = Decimal(str(m.service_charges or "0"))

        tenant_name = "vacant"
        if Lease:
            try:
                overlap = (Lease.objects
                           .filter(unit=m.unit,
                                   start_date__lte=end_d,
                                   end_date__gte=start_d)
                           .order_by("-start_date")
                           .first())
                if overlap and getattr(overlap, "tenant", None):
                    tn = (overlap.tenant.get_full_name()
                          if hasattr(overlap.tenant, "get_full_name")
                          else str(overlap.tenant))
                    tenant_name = tn
                else:
                    today = timezone.localdate()
                    active = (Lease.objects
                              .filter(unit=m.unit,
                                      start_date__lte=today,
                                      end_date__gte=today)
                              .order_by("-start_date")
                              .first())
                    if active and getattr(active, "tenant", None):
                        tn = (active.tenant.get_full_name()
                              if hasattr(active.tenant, "get_full_name")
                              else str(active.tenant))
                        tenant_name = tn
            except Exception:
                tenant_name = "vacant"

        for k in keys_sorted:
            row = key_to_row.get(k)
            if not row:
                continue
            usage_amt = (row["usage"] * rate).quantize(Decimal("0.01"))
            service_amt = svc if granularity == "monthly" else Decimal("0.00")
            total_amt = (usage_amt + service_amt).quantize(Decimal("0.01"))

            combined_rows.append({
                "meter_id": m.id,
                "meter_number": m.meter_number,
                "unit_number": unit_number,
                "property_name": getattr(m.unit.property, "property_name", ""),
                "tenant_name": tenant_name,
                "period_label": _label_from_key(k),
                "period_key": k,
                "start_kwh": row["start_kwh"],
                "end_kwh": row["end_kwh"],
                "usage": row["usage"],
                "unit_rate": rate,
                "usage_amount": usage_amt,
                "service_charges": service_amt,
                "total_amount": total_amt,
            })
            total_kwh += row["usage"]
            usage_charges += usage_amt
            service_total += service_amt

    # Sort by period asc, then Unit, then Meter (for ALL meters default)
    combined_rows.sort(key=lambda x: (
        x["period_key"], x["unit_number"], x["meter_number"]))

    totals = {
        "total_kwh": total_kwh,
        "usage_charges": usage_charges,
        "service_charges": service_total,
        "grand_total": (usage_charges + service_total).quantize(Decimal("0.01")),
    }
    return labels_sorted, datasets, combined_rows, totals


def _export_filename(request, base: str, ext: str) -> str:
    meter_id = (request.GET.get("meter") or "").strip()
    start = (request.GET.get("start") or "").strip()
    end = (request.GET.get("end") or "").strip()

    if meter_id:
        try:
            mn = Meter.objects.values_list(
                "meter_number", flat=True).get(id=meter_id)
        except Meter.DoesNotExist:
            mn = "unknown"
        return f"meter-data-{mn}-{start or 'start'}_to_{end or 'end'}.{ext}"
    return f"All-meters-{start or 'start'}_to_{end or 'end'}.{ext}"


def _export_rows(request):
    today = timezone.localdate()
    start_date = date(today.year, today.month, 1)
    end_date = today
    if request.GET.get("start") and request.GET.get("end"):
        try:
            start_date = date.fromisoformat(request.GET["start"])
            end_date = date.fromisoformat(request.GET["end"])
        except Exception:
            pass

    report_type = (request.GET.get("report_type") or "daily").strip().lower()
    if report_type not in ("hourly", "daily", "monthly"):
        report_type = "daily"

    prop_id = (request.GET.get("property") or "").strip()
    unit_id = (request.GET.get("unit") or "").strip()
    meter_id = (request.GET.get("meter") or "").strip()

    units_qs = Unit.objects.all()
    if prop_id:
        units_qs = units_qs.filter(property_id=prop_id)

    meters_qs = Meter.objects.select_related("unit", "unit__property")
    if unit_id:
        meters_qs = meters_qs.filter(unit_id=unit_id)
    elif prop_id:
        meters_qs = meters_qs.filter(unit__property_id=prop_id)

    selected_meters = meters_qs.order_by("meter_number")
    if meter_id:
        selected_meters = selected_meters.filter(id=meter_id)

    labels, datasets, rows, totals = _per_meter_series(
        selected_meters, start_date, end_date, report_type)
    return report_type, rows, totals, start_date, end_date, prop_id, meter_id


def _user_whatsapp(request):
    # Adjust these attribute names to your auth/profile model
    u = getattr(request, 'user', None)
    if not u:
        return ''
    for attr in ("whatsapp", "whatsapp_number", "phone_number", "mobile", "profile.whatsapp", "profile.phone_number"):
        try:
            parts = attr.split('.')
            obj = u
            for p in parts:
                obj = getattr(obj, p)
            if obj:
                return str(obj)
        except Exception:
            continue
    return ''


@login_required
def energy_dashboard(request):
    wa = ""
    if request.user.is_authenticated:
        wa = getattr(request.user, "whatsapp_number", "") or ""

    today = timezone.localdate()
    start_date = today
    end_date = today

    if request.GET.get("start") and request.GET.get("end"):
        try:
            start_date = date.fromisoformat(request.GET["start"])
            end_date = date.fromisoformat(request.GET["end"])
        except Exception:
            pass

    report_type = (request.GET.get("report_type") or "daily").strip().lower()
    if report_type not in ("hourly", "daily", "monthly"):
        report_type = "daily"

    prop_id = (request.GET.get("property") or "").strip()
    unit_id = (request.GET.get("unit") or "").strip()
    meter_id = (request.GET.get("meter") or "").strip()

    all_properties = Property.objects.all().order_by("property_name")

    units_qs = Unit.objects.all()
    if prop_id:
        units_qs = units_qs.filter(property_id=prop_id)
    filtered_units = units_qs.order_by("unit_number")

    meters_qs = Meter.objects.select_related("unit", "unit__property")
    if unit_id:
        meters_qs = meters_qs.filter(unit_id=unit_id)
    elif prop_id:
        meters_qs = meters_qs.filter(unit__property_id=prop_id)
    filtered_meters = meters_qs.order_by("meter_number")

    selected_meters = filtered_meters
    if meter_id:
        selected_meters = selected_meters.filter(id=meter_id)

    per_meter_mode = bool(meter_id)

    selected_meter = None
    live_panel = {}
    if per_meter_mode:
        selected_meter = get_object_or_404(Meter, id=meter_id)
        lr = LiveReading.objects.filter(meter=selected_meter).first()
        if lr:
            last_ts = lr.ts
            voltage_a = lr.voltage_a
            current_a = lr.current_a
            total_energy = lr.total_energy
        else:
            snap = MeterReading.objects.filter(
                meter=selected_meter).order_by("-ts").first()
            last_ts = snap.ts if snap else None
            voltage_a = getattr(snap, "voltage_a", None) if snap else None
            current_a = getattr(snap, "current_a", None) if snap else None
            total_energy = getattr(snap, "total_energy",
                                   None) if snap else None

        cutoff_dt = timezone.now() - timedelta(minutes=ONLINE_MINUTES)
        is_online = bool(lr and lr.ts >= cutoff_dt)
        bal_obj = MeterBalance.objects.filter(unit=selected_meter.unit).first()

        live_panel = {
            "unit_name": getattr(selected_meter.unit, "unit_number", ""),
            "meter_number": selected_meter.meter_number,
            "is_online": is_online,
            "last_ts": last_ts,
            "voltage_a": voltage_a,
            "current_a": current_a,
            "total_energy": total_energy,
            "unit_rate": selected_meter.unit_rate,
            "balance": getattr(bal_obj, "balance", None),
        }

    labels, datasets, table_rows, totals = _per_meter_series(
        selected_meters, start_date, end_date, report_type
    )

    online_count = offline_count = 0
    online_set = set()
    if not per_meter_mode:
        cutoff = timezone.now() - timedelta(minutes=ONLINE_MINUTES)
        # get latest live-reading timestamps for all selected meters
        live_map = {lr.meter_id: lr.ts for lr in LiveReading.objects.filter(
            meter__in=selected_meters)}
        for mid in selected_meters.values_list("id", flat=True):
            ts = live_map.get(mid)
            if ts and ts >= cutoff:
                online_count += 1
                online_set.add(mid)
            else:
                offline_count += 1

    # Monthly heading for all meters
    month_heading = ""
    if report_type == "monthly":
        if start_date.year == end_date.year and start_date.month == end_date.month:
            month_heading = start_date.strftime("%B %Y")
        else:
            month_heading = f"{start_date.strftime('%b %Y')} – {end_date.strftime('%b %Y')}"

    # Property name for heading when filtering by property
    property_name = None
    if prop_id:
        try:
            property_name = Property.objects.values_list(
                "property_name", flat=True).get(id=prop_id)
        except Property.DoesNotExist:
            property_name = None

    # Build display-safe rows with rounding (0 decimals) and comma formatting
    wa = _user_whatsapp(request)
    rows_disp = []
    for r in table_rows:
        rows_disp.append({
            **r,
            "whatsapp": wa,
            "is_online": (r["meter_id"] in online_set) if not per_meter_mode else None,
            "start_kwh": _fmt0(r["start_kwh"]),
            "end_kwh": _fmt0(r["end_kwh"]),
            "usage": _fmt0(r["usage"]),
            "unit_rate": _fmt0(r["unit_rate"]),
            "usage_amount": _fmt0(r["usage_amount"]),
            "service_charges": _fmt0(r["service_charges"]),
            "total_amount": _fmt0(r["total_amount"]),
        })

    totals_disp = {
        "total_kwh": _fmt0(totals["total_kwh"]),
        "usage_charges": _fmt0(totals["usage_charges"]),
        "service_charges": _fmt0(totals["service_charges"]),
        "grand_total": _fmt0(totals["grand_total"]),
    }

    context = {
        "all_properties": all_properties,
        "filtered_units": filtered_units,
        "filtered_meters": filtered_meters,
        "meters": filtered_meters,
        "current_property": prop_id,
        "current_unit": unit_id,
        "current_meter": meter_id,
        "report_type": report_type,
        "start_date": start_date,
        "end_date": end_date,
        "online_minutes": ONLINE_MINUTES,

        "labels": labels,
        "datasets": datasets,
        "rows": table_rows,
        "rows_disp": rows_disp,
        "totals": totals,
        "totals_disp": totals_disp,

        "selected_meter": selected_meter,
        "live_panel": live_panel,
        "online_count": online_count,
        "offline_count": offline_count,
        "currency": BILLING_CURRENCY,
        "month_heading": month_heading,
        "property_name": property_name,
    }
    return render(request, "smart_meter/dashboard.html", context)


# ---------- Exports ----------
@permission_required("accounts.can_export_energy", raise_exception=True)
def energy_export_csv(request):
    report_type, rows, totals, start_date, end_date, prop_id, meter_id = _export_rows(
        request)
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{_export_filename(request, "energy", "csv")}"'

    import csv
    w = csv.writer(resp)
    header = [
        "S/N", "Meter #", "Unit", "Property", "Tenant", "WhatsApp", "Period",
        "Begin (kWh)", "End (kWh)", "Usage (kWh)", "Rate (Rs/kWh)", "Usage Charges (Rs)"
    ]
    if report_type == "monthly":
        header += ["Service Charges (Rs)", "Total (Rs)"]
    w.writerow(header)

    wa = _user_whatsapp(request)
    for i, r in enumerate(rows, start=1):
        base = [
            i,
            r["meter_number"],
            r["unit_number"],
            r["property_name"],
            r.get("tenant_name", ""),
            wa,
            r["period_label"],
            _fmt0(r["start_kwh"]),
            _fmt0(r["end_kwh"]),
            _fmt0(r["usage"]),
            _fmt0(r["unit_rate"]),
            _fmt0(r["usage_amount"]),
        ]
        if report_type == "monthly":
            base += [_fmt0(r["service_charges"]), _fmt0(r["total_amount"])]
        w.writerow(base)
    return resp


@permission_required("accounts.can_export_energy", raise_exception=True)
def energy_export_xlsx(request):
    report_type, rows, totals, start_date, end_date, prop_id, meter_id = _export_rows(
        request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Energy"

    header = [
        "S/N", "Meter #", "Unit", "Property", "Tenant", "WhatsApp", "Period",
        "Begin (kWh)", "End (kWh)", "Usage (kWh)", "Rate (Rs/kWh)", "Usage Charges (Rs)"
    ]
    if report_type == "monthly":
        header += ["Service Charges (Rs)", "Total (Rs)"]
    ws.append(header)

    wa = _user_whatsapp(request)
    for i, r in enumerate(rows, start=1):
        base = [
            i,
            r["meter_number"],
            r["unit_number"],
            r["property_name"],
            r.get("tenant_name", ""),
            wa,
            r["period_label"],
            _fmt0(r["start_kwh"]),
            _fmt0(r["end_kwh"]),
            _fmt0(r["usage"]),
            _fmt0(r["unit_rate"]),
            _fmt0(r["usage_amount"]),
        ]
        if report_type == "monthly":
            base += [_fmt0(r["service_charges"]), _fmt0(r["total_amount"])]
        ws.append(base)

    for col in ws.columns:
        try:
            max_len = max(len(str(c.value))
                          if c.value is not None else 0 for c in col)
        except ValueError:
            max_len = 10
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 32)

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{_export_filename(request, "energy", "xlsx")}"'
    return resp


@permission_required("accounts.can_export_energy", raise_exception=True)
def energy_export_pdf(request):
    # Build the same dataset used by the dashboard, but locally for PDF
    report_type, rows, totals, start_date, end_date, prop_id, meter_id = _export_rows(
        request)

    # Heading bits
    month_heading = ""
    if report_type == "monthly":
        if start_date.year == end_date.year and start_date.month == end_date.month:
            month_heading = start_date.strftime("%B %Y").upper()
        else:
            month_heading = f"{start_date.strftime('%b %Y').upper()} – {end_date.strftime('%b %Y').upper()}"

    property_name = None
    if prop_id:
        property_name = (
            Property.objects.filter(id=prop_id)
            .values_list("property_name", flat=True)
            .first()
        )

    # Display-ready rows (0 decimals + comma formatting) + WhatsApp
    wa = _user_whatsapp(request)
    rows_disp = [
        {
            **r,
            "whatsapp": wa,
            "start_kwh": _fmt0(r["start_kwh"]),
            "end_kwh": _fmt0(r["end_kwh"]),
            "usage": _fmt0(r["usage"]),
            "unit_rate": _fmt0(r["unit_rate"]),
            "usage_amount": _fmt0(r["usage_amount"]),
            "service_charges": _fmt0(r["service_charges"]),
            "total_amount": _fmt0(r["total_amount"]),
        }
        for r in rows
    ]

    totals_disp = {
        "total_kwh": _fmt0(totals["total_kwh"]),
        "usage_charges": _fmt0(totals["usage_charges"]),
        "service_charges": _fmt0(totals["service_charges"]),
        "grand_total": _fmt0(totals["grand_total"]),
    }

    context = {
        "rows_disp": rows_disp,
        "totals_disp": totals_disp,
        "currency": BILLING_CURRENCY,
        "report_type": report_type,
        "start_date": start_date,
        "end_date": end_date,
        "property_name": property_name,
        "month_heading": month_heading,
        "current_meter": meter_id,  # template checks this in a heading condition
    }

    html = render_to_string(
        "smart_meter/dashboard_pdf.html", context, request=request)
    pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf()

    # Reuse the standard filename helper
    fname = _export_filename(request, "energy", "pdf")
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


def energy_chart_page(request):
    # Reuse the same data as dashboard
    # You already compute everything in energy_dashboard; replicate the core by calling helpers
    today = timezone.localdate()
    start_date = date(today.year, today.month, 1)
    end_date = today
    try:
        if request.GET.get("start"):
            start_date = date.fromisoformat(request.GET["start"])
        if request.GET.get("end"):
            end_date = date.fromisoformat(request.GET["end"])
    except Exception:
        pass

    report_type = (request.GET.get("report_type") or "daily").strip().lower()
    if report_type not in ("hourly", "daily", "monthly"):
        report_type = "daily"

    prop_id = (request.GET.get("property") or "").strip()
    unit_id = (request.GET.get("unit") or "").strip()
    meter_id = (request.GET.get("meter") or "").strip()

    all_properties = Property.objects.all().order_by("property_name")
    units_qs = Unit.objects.all()
    if prop_id:
        units_qs = units_qs.filter(property_id=prop_id)
    filtered_units = units_qs.order_by("unit_number")

    meters_qs = Meter.objects.select_related("unit", "unit__property")
    if unit_id:
        meters_qs = meters_qs.filter(unit_id=unit_id)
    elif prop_id:
        meters_qs = meters_qs.filter(unit__property_id=prop_id)
    filtered_meters = meters_qs.order_by("meter_number")

    selected_meters = filtered_meters
    if meter_id:
        selected_meters = selected_meters.filter(id=meter_id)

    labels, datasets, table_rows, totals = _per_meter_series(
        selected_meters, start_date, end_date, report_type
    )

    # Optional monthly heading + property name (same as dashboard)
    month_heading = ""
    if report_type == "monthly":
        if start_date.year == end_date.year and start_date.month == end_date.month:
            month_heading = start_date.strftime("%B %Y")
        else:
            month_heading = f"{start_date.strftime('%b %Y')} – {end_date.strftime('%b %Y')}"

    property_name = None
    if prop_id:
        try:
            property_name = Property.objects.values_list(
                "property_name", flat=True).get(id=prop_id)
        except Property.DoesNotExist:
            property_name = None

    # Display-friendly rows/totals
    wa = _user_whatsapp(request)
    rows_disp = [{
        **r,
        "whatsapp": wa,
        "start_kwh": _fmt0(r["start_kwh"]),
        "end_kwh": _fmt0(r["end_kwh"]),
        "usage": _fmt0(r["usage"]),
        "unit_rate": _fmt0(r["unit_rate"]),
        "usage_amount": _fmt0(r["usage_amount"]),
        "service_charges": _fmt0(r["service_charges"]),
        "total_amount": _fmt0(r["total_amount"]),
    } for r in table_rows]

    totals_disp = {
        "total_kwh": _fmt0(totals["total_kwh"]),
        "usage_charges": _fmt0(totals["usage_charges"]),
        "service_charges": _fmt0(totals["service_charges"]),
        "grand_total": _fmt0(totals["grand_total"]),
    }

    # Back link with same filters
    back_qs = request.GET.urlencode()
    back_url = f"{reverse('smart_meter:energy_dashboard')}?{back_qs}" if back_qs else reverse(
        'smart_meter:energy_dashboard')

    ctx = {
        "all_properties": all_properties,
        "filtered_units": filtered_units,
        "filtered_meters": filtered_meters,
        "meters": filtered_meters,
        "current_property": prop_id,
        "current_unit": unit_id,
        "current_meter": meter_id,
        "report_type": report_type,
        "start_date": start_date,
        "end_date": end_date,

        "labels": labels,
        "datasets": datasets,
        "rows": table_rows,
        "rows_disp": rows_disp,
        "totals": totals,
        "totals_disp": totals_disp,
        "currency": BILLING_CURRENCY,
        "month_heading": month_heading,
        "property_name": property_name,

        "back_url": back_url,
    }
    return render(request, "smart_meter/chart_only.html", ctx)

# views_dashboard.py


BILLING_CURRENCY = "Rs."  # already present in this module in your repo


def _month_bounds(yyyy_mm: str | None):
    """Return (start_date, end_date) for selected month; defaults to current month."""
    today = timezone.localdate()
    if not yyyy_mm:
        y, m = today.year, today.month
    else:
        y, m = map(int, yyyy_mm.split("-"))
    start = date(y, m, 1)
    end = date(y, m, monthrange(y, m)[1])
    return start, end


def _prev_month_bounds(any_day: date):
    m = any_day.month - 1 or 12
    y = any_day.year - 1 if any_day.month == 1 else any_day.year
    start = date(y, m, 1)
    end = date(y, m, monthrange(y, m)[1])
    return start, end


# smart_meter/views_dashboard.py


BILLING_SUMMARY_DEFAULT_COLUMNS = [
    "Unit",
    "Tenant",
    "Month",
    "Rent",
    "Society Maintenance",
    "Water",
    "Internet",
    "Electric (prev)",
    "Other",
    "Total",
]


def _first_day(d: date) -> date:
    return date(d.year, d.month, 1)


def _last_day(d: date) -> date:
    return date(d.year, d.month, monthrange(d.year, d.month)[1])


def _prev_month(d: date) -> date:
    y = d.year
    m = d.month - 1
    if m == 0:
        m = 12
        y -= 1
    return date(y, m, 1)


def _cat_lookup():
    """
    Build a case-insensitive lookup of ItemCategory names -> id.
    """
    by_name = {}
    for c in ItemCategory.objects.filter(is_active=True):
        by_name[c.name.strip().lower()] = c.id
    return by_name


def _money(x):
    return Decimal(x or 0).quantize(Decimal("0.01"))

# views_dashboard.py


# PDF (WeasyPrint already used elsewhere in this file)
try:
    from weasyprint import HTML
except Exception:
    HTML = None  # graceful fallback

from smart_meter.models import Bill  # electric bills (prev month)  # noqa

BILLING_CURRENCY = "Rs"  # adjust your symbol if needed


def _ym_from_qs(request):
    """Read ?month=YYYY-MM (defaults to current month)."""
    today = timezone.localdate()
    y = today.year
    m = today.month
    mstr = (request.GET.get("month") or "").strip()
    if mstr:
        try:
            y, m = [int(x) for x in mstr.split("-")]
        except Exception:
            pass
    return y, m


def _bounds_for_month(y, m):
    start = timezone.datetime(y, m, 1).date()
    end = timezone.datetime(y, m, monthrange(y, m)[1]).date()
    return start, end


def _prev_year_month(y, m):
    return (y - 1, 12) if m == 1 else (y, m - 1)


def _zero():
    return Decimal("0.00")


def _fmt0(n: Decimal | float | int) -> str:
    """Round half-up to 0 decimals and format with comma (for HTML/PDF only)."""
    d = Decimal(str(n or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return f"{int(d):,}"


def _sum_items(invoice_qs, categories: list[str]) -> Decimal:
    """Sum invoice items in given categories for invoices queryset."""
    if not categories:
        return _zero()
    cats = [c.strip().lower() for c in categories]
    qs = (InvoiceItem.objects
          .filter(invoice__in=invoice_qs, category__name__in=categories)
          .aggregate(t=Sum("amount")))
    return qs["t"] or _zero()


def _sum_items_excluding(invoice_qs, excluded: list[str]) -> Decimal:
    exc = excluded or []
    qs = (InvoiceItem.objects
          .filter(invoice__in=invoice_qs)
          .exclude(category__name__in=exc)
          .aggregate(t=Sum("amount")))
    return qs["t"] or _zero()


def _lease_balance_upto(lease: Lease, upto_date):
    """Invoices - Payments up to a date."""
    inv = (lease.invoices
           .filter(issue_date__lte=upto_date)
           .aggregate(t=Sum("amount"))["t"] or _zero())

    # Try a general payments relation first; if a Payments app exists, we’ll use it
    pay = _zero()
    try:
        pay = (lease.payments.filter(date__lte=upto_date)
               .aggregate(t=Sum("amount"))["t"] or _zero())
    except Exception:
        try:
            from payments.models import Payment as LeasePayment
            pay = (LeasePayment.objects.filter(lease=lease, date__lte=upto_date)
                   .aggregate(t=Sum("amount"))["t"] or _zero())
        except Exception:
            pass
    return inv - pay


# --- Electric kWh parser (from item description) ---
KW_RE = re.compile(r"total\s*usage\s*[:=]\s*([0-9]+(?:\.[0-9]+)?)", re.I)


def _parse_kwh_from_desc(desc):
    if not desc:
        return Decimal("0")
    m = KW_RE.search(desc)
    return Decimal(m.group(1)) if m else Decimal("0")


# --- Category normalization ---
# Canonical keys are lowercased slugs
CANON_SYNONYMS = {
    "rent": {"rent"},
    "society maintenance": {"society maintenance", "society maint.", "maintenance", "maintenance charges"},
    "internet": {"internet", "broadband", "wifi"},
    "electric": {"electric", "electricity", "electric bill", "electricity bill", "power"},
}
CANON_LABELS = {
    "rent": "Rent",
    "society maintenance": "Society Maint.",
    "internet": "Internet",
    "electric": "Electric",
    # pseudo / computed:
    "electric_kwh": "Electric kWh",
    "other": "Other",
    "total": "Total",
    "prev_balance": "Prev Balance",
    "total_balance": "Total Balance",
}


def _cat_to_name_lower(cat_obj) -> str:
    if cat_obj is None:
        return ""
    if isinstance(cat_obj, str):
        base = cat_obj.strip().lower()
    else:
        # ForeignKey like ItemCategory
        for attr in ("name", "title", "label", "code"):
            val = getattr(cat_obj, attr, None)
            if isinstance(val, str) and val.strip():
                base = val.strip().lower()
                break
        else:
            base = str(cat_obj).strip().lower()

    # map to canonical if in synonyms
    for canon, syns in CANON_SYNONYMS.items():
        if base in syns:
            return canon
    return base  # use the raw normalized label for custom categories


def billing_summary(request):
    """
    Page: Billing Summary — grouped by Property.
    For the selected month, summarize by unit:
      Rent, Society Maintenance, Internet, Electric (Rs), Electric (kWh), Other, Total
    If a category has no data, show 0.
    """
    # ✅ parse month safely without touching `request`
    month_str = (request.GET.get("month") or "").strip()

    if month_str:
        try:
            y_str, m_str = month_str.split("-", 1)
            y, m = int(y_str), int(m_str)
            if not (1 <= m <= 12):
                raise ValueError("bad month")
        except Exception:
            today = timezone.localdate()
            y, m = today.year, today.month
    else:
        today = timezone.localdate()
        y, m = today.year, today.month

    start, end = _bounds_for_month(y, m)

    # previous month bounds still useful for balances; keep if you still show those cols
    py, pm = _prev_year_month(y, m)
    pstart, pend = _bounds_for_month(py, pm)

    leases = (
        Lease.objects
        .select_related("tenant", "unit", "unit__property")
        .filter(status="active")
        .order_by("unit__property__property_name", "unit__unit_number")
    )

    groups = defaultdict(list)
    grand = defaultdict(Decimal)

    # Preload all invoices for the month once (optional but nice)
    inv_all = (
        Invoice.objects
        .filter(issue_date__gte=start, issue_date__lte=end)
        .select_related("lease")
    )

    # Map lease_id -> invoices
    inv_by_lease = defaultdict(list)
    for inv in inv_all:
        inv_by_lease[inv.lease_id].append(inv)

    # If you have an InvoiceItem model:
    # We’ll fetch all items for those invoices (avoids N+1)
    invoice_ids = [i.id for i in inv_all]
    items_by_invoice = defaultdict(list)

    if invoice_ids:
        for it in (InvoiceItem.objects
                   .filter(invoice_id__in=invoice_ids)
                   .select_related("category")            # <-- important
                   .only("invoice_id", "amount", "description", "category__name")):
            items_by_invoice[it.invoice_id].append(it)

    ZERO = _zero()

    # Canonical names your DB uses (case-sensitive per your note)
    CAT_RENT = "Rent"
    CAT_SOC = "Society Maintenance"
    CAT_NET = "Internet"
    CAT_ELEC = {"Electric", "Electricity"}  # include both if used

    for lease in leases:
        prop = lease.unit.property
        unit = lease.unit
        inv_month = inv_by_lease.get(lease.id, [])

        # Sums for this unit in the selected month
        rent = ZERO
        society = ZERO
        internet = ZERO
        electric_rs = ZERO
        electric_kwh = ZERO
        other = ZERO

        # Tally from line items
        for inv in inv_month:
            items = items_by_invoice.get(inv.id, [])
            for it in items:
                cat = _cat_to_name_lower(it.category)   # from earlier fix
                amt = Decimal(it.amount or 0)
                if cat in {"rent"}:
                    rent += amt
                elif cat in {"society maintenance", "society maint.", "maintenance", "maintenance charges"}:
                    society += amt
                elif cat in {"internet", "broadband", "wifi"}:
                    internet += amt
                elif cat in {"electric", "electricity", "electric bill", "electricity bill", "power"}:
                    electric_rs += amt
                    electric_kwh += _parse_kwh_from_desc(
                        getattr(it, "description", ""))
                else:
                    other += amt

        total = rent + society + internet + electric_rs + other

        # Balances (optional: same as before)
        prev_balance = _lease_balance_upto(
            lease, pend)   # end of previous month
        total_balance = _lease_balance_upto(
            lease, end)   # end of selected month

        row = {
            "lease": lease,
            "property": prop,
            "unit": unit,
            "tenant": lease.tenant,

            "rent": rent,
            "society": society,
            "internet": internet,
            "electric_rs": electric_rs,
            "electric_kwh": electric_kwh,
            "other": other,

            "total": total,
            "prev_balance": prev_balance,
            "total_balance": total_balance,
        }
        groups[prop.id].append(row)

        # accumulate grand totals
        for k in ("rent", "society", "internet", "electric_rs", "electric_kwh",
                  "other", "total", "prev_balance", "total_balance"):
            grand[k] = grand.get(k, ZERO) + row[k]

    # Build data structure for template: ordered by property name
    props = Property.objects.filter(
        id__in=groups.keys()).order_by("property_name")

    grouped = []
    sn_counter = 1
    for prop in props:
        rows = groups[prop.id]
        subtotal = defaultdict(Decimal)
        for r in rows:
            for k in ("rent", "society", "internet", "electric_rs", "electric_kwh",
                      "other", "total", "prev_balance", "total_balance"):
                subtotal[k] = subtotal.get(k, ZERO) + r[k]
        for r in rows:
            r["sn"] = sn_counter
            sn_counter += 1

        grouped.append({"property": prop, "rows": rows, "subtotal": subtotal})

    context = {
        "year": y,
        "month": m,
        "start": start,
        "end": end,
        "groups": grouped,
        "grand": grand,
        "currency": BILLING_CURRENCY,
    }
    return render(request, "smart_meter/billing_summary.html", context)


@permission_required("accounts.can_export_energy", raise_exception=True)
def billing_summary_export_excel(request):
    """Excel export (grouped, with subtotals + grand total)."""
    y, m = _ym_from_qs(request)
    # Reuse page data
    resp = billing_summary(request)
    ctx = resp.context_data if hasattr(resp, "context_data") else resp.context

    wb = Workbook()
    ws = wb.active
    ws.title = f"Billing {y}-{m:02d}"

    # header
    headers = [
        "S.N.", "Unit", "Tenant", "Rent", "Society Maint.", "Water", "Internet",
        f"Electric (prev {ctx['prev_label']})", "Electric kWh (prev)", "Other",
        "Total", "Prev Balance", "Total Balance"
    ]
    ws.append(headers)

    # rows
    for g in ctx["groups"]:
        ws.append([f"Property: {g['property'].property_name}"])
        for r in g["rows"]:
            ws.append([
                r["sn"],
                r["unit"].unit_number,
                getattr(r["tenant"], "name", str(r["tenant"])),
                float(r["rent"]),
                float(r["society"]),
                float(r["water"]),
                float(r["internet"]),
                float(r["electric_prev"]),
                float(r["electric_prev_kwh"]),
                float(r["other"]),
                float(r["total"]),
                float(r["prev_balance"]),
                float(r["total_balance"]),
            ])
        # subtotal row
        st = g["subtotal"]
        ws.append([
            "", "Subtotal", "",
            float(st["rent"]), float(st["society"]), float(
                st["water"]), float(st["internet"]),
            float(st["electric_prev"]), "", float(st["other"]),
            float(st["total"]), float(
                st["prev_balance"]), float(st["total_balance"])
        ])

    # grand total
    gt = ctx["grand"]
    ws.append([
        "", "GRAND TOTAL", "",
        float(gt["rent"]), float(gt["society"]), float(
            gt["water"]), float(gt["internet"]),
        float(gt["electric_prev"]), "", float(gt["other"]),
        float(gt["total"]), float(
            gt["prev_balance"]), float(gt["total_balance"])
    ])

    # Let user tweak column widths in Excel—no strict widths here.
    # But make numbers look like numbers (we already wrote floats).

    from io import BytesIO
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    fn = f"billing_summary_{y}-{m:02d}.xlsx"
    resp = HttpResponse(
        out.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{fn}"'
    return resp


@permission_required("accounts.can_export_energy", raise_exception=True)
def billing_summary_export_pdf(request):
    """PDF export (landscape)."""
    if HTML is None:
        return HttpResponse("WeasyPrint not installed on server.", status=500)

    # Reuse the page context but render a print-optimized template
    y, m = _ym_from_qs(request)
    page = billing_summary(request)
    ctx = page.context_data if hasattr(page, "context_data") else page.context

    html = render_to_string(
        "smart_meter/billing_summary_pdf.html", ctx, request=request)
    pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf()

    fn = f"billing_summary_{y}-{m:02d}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{fn}"'
    return resp

    """
    Page + Export:
      - ?month=YYYY-MM (defaults to current month)
      - ?export=excel|pdf
      - ?widths=comma,separated,ints  (apply to table & exports)
    """
    today = timezone.localdate()
    month_str = (request.GET.get("month")
                 or f"{today.year:04d}-{today.month:02d}")[:7]
    year, mon = [int(p) for p in month_str.split("-")]
    selected_month = date(year, mon, 1)

    # For tenancy/invoice range
    start = _first_day(selected_month)
    end = _last_day(selected_month)

    # Electric is previous month
    electric_month = _prev_month(selected_month)

    # Optional column widths
    widths_param = (request.GET.get("widths") or "").strip()
    try:
        widths = [int(x)
                  for x in widths_param.split(",")] if widths_param else []
    except Exception:
        widths = []

    # Build category map (case-insensitive)
    cats = _cat_lookup()
    cat_rent = cats.get("rent")
    cat_society = cats.get("society maintenance") or cats.get(
        "society_maintenance")
    cat_water = cats.get("water") or cats.get("water charges")
    cat_internet = cats.get("internet")
    cat_electric = cats.get("electric") or cats.get(
        "electric bill") or cats.get("electricity")

    # Find leases that overlap selected month
    overlapping_leases = (
        Lease.objects
        .filter(start_date__lte=end, end_date__gte=start)
        .select_related("tenant", "unit", "unit__property")
        .order_by("unit__property__property_name", "unit__unit_number", "tenant__id")
    )

    # Fetch invoices **issued** in selected month for those leases
    invs = (
        Invoice.objects
        .filter(lease__in=overlapping_leases, issue_date__gte=start, issue_date__lte=end)
        .prefetch_related("items", "lease__tenant", "lease__unit", "lease__unit__property")
        .order_by("lease__unit__property__property_name", "lease__unit__unit_number")
    )

    rows = []
    for inv in invs:
        lease = inv.lease
        unit = lease.unit
        tenant_name = getattr(lease.tenant, "name", None) or str(lease.tenant)

        # Sum by category (case-insensitive names)
        rent = Decimal("0.00")
        society = Decimal("0.00")
        water = Decimal("0.00")
        internet = Decimal("0.00")
        electric = Decimal("0.00")
        grand_total = Decimal("0.00")

        for it in inv.items.all():
            amt = _money(it.amount)
            grand_total += amt
            cid = it.category_id if it.category_id else None
            # Route into buckets
            if cid == cat_rent:
                rent += amt
            elif cid == cat_society:
                society += amt
            elif cid == cat_water:
                water += amt
            elif cid == cat_internet:
                internet += amt
            elif cid == cat_electric:
                electric += amt
            else:
                # count toward "Other" later via grand_total - known buckets
                pass

        other = (grand_total - (rent + society + water +
                 internet + electric)).quantize(Decimal("0.01"))

        rows.append({
            "unit": f"{unit.property.property_name}-{unit.unit_number}",
            "tenant": tenant_name,
            "month_label": selected_month.strftime("%b %Y"),
            "rent": rent,
            "society": society,
            "water": water,
            "internet": internet,
            "electric": electric,  # represents previous month’s usage, included on this invoice
            "other": other,
            "total": grand_total,
            "electric_month_label": electric_month.strftime("%b %Y"),
        })

    # ----- EXPORTS -----
    export = (request.GET.get("export") or "").lower().strip()
    if export == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Billing Summary"

        headers = BILLING_SUMMARY_DEFAULT_COLUMNS
        ws.append(headers)

        # Apply column widths, if provided
        for idx, w in enumerate(widths, start=1):
            try:
                ws.column_dimensions[get_column_letter(
                    idx)].width = max(5, int(w))
            except Exception:
                pass

        for r in rows:
            ws.append([
                r["unit"],
                r["tenant"],
                r["month_label"],
                float(r["rent"]),
                float(r["society"]),
                float(r["water"]),
                float(r["internet"]),
                float(r["electric"]),
                float(r["other"]),
                float(r["total"]),
            ])

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"Billing-Summary-{month_str}.xlsx"
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp

    if export == "pdf":
        html = render_to_string("smart_meter/billing_summary_pdf.html", {
            "rows": rows,
            "month_str": month_str,
            "headers": BILLING_SUMMARY_DEFAULT_COLUMNS,
            "widths": widths,
            # pass electric previous-month label for header tooltip
            "electric_note": f"Electric shows prior month (e.g., {electric_month.strftime('%b %Y')})",
        })
        pdf = HTML(
            string=html, base_url=request.build_absolute_uri("/")).write_pdf()
        resp = HttpResponse(pdf, content_type="application/pdf")
        filename = f"Billing-Summary-{month_str}.pdf"
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    # ----- NORMAL PAGE RENDER -----
    return render(request, "smart_meter/billing_summary.html", {
        "rows": rows,
        "month_str": month_str,
        "headers": BILLING_SUMMARY_DEFAULT_COLUMNS,
        "widths": widths,
        "electric_month_label": electric_month.strftime("%b %Y"),
    })
