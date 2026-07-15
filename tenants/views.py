# tenants/views.py
import csv
import json
import logging
import re
import uuid
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from urllib.parse import quote

import openpyxl
from django.apps import apps
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core import signing
from django.core.exceptions import ValidationError
from django.core.mail import EmailMessage
from django.core.signing import BadSignature, SignatureExpired
from django.db.models import (
    Case,
    DecimalField,
    Exists,
    F,
    OuterRef,
    Prefetch,
    Q,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Coalesce, Replace
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import (
    reverse,
    reverse_lazy,
)
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    UpdateView,
)
from django_tables2 import SingleTableView
from django_tables2.export.views import ExportMixin
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from weasyprint import HTML

from core.models import GlobalSettings
from core.utils.identity import format_cnic, format_phone, normalize_cnic, normalize_phone, validate_cnic
from invoices.models import Invoice, RecurringCharge, SecurityDepositTransaction
from leases.models import (
    Lease,
    LeaseFamilyMember,
    LeaseRelationshipType,
    LeaseVehicle,
    LeaseVehicleType,
)
from leases.models_renewal import LeaseRenewal
from leases.services.lease_expiry import attach_lease_expiry_countdown
from leases.services.vehicle_submissions import (
    create_pending_vehicle_submissions_from_post,
)
from leases.whatsapp import build_whatsapp_url
from payments.models import Payment
from properties.models import Property, Unit
from tenants.models import Tenant
from whatsapp.models import TrustedDeviceRegistry, WhatsAppExternalLinkToken
from whatsapp.services.external_links import record_external_link_access

from .forms import (
    TenantForm,
    TenantPreRegistrationLinkForm,
    TenantPublicRegistrationForm,
    TenantRegistrationSubmissionReviewForm,
)
from .models import (
    PendingRegistrationPerson,
    Tenant,
    TenantRegistrationSubmission,
)
from .tables import (
    LedgerTable,  # We'll create this later
    TenantTable,
)

TENANT_REGISTRATION_MAX_AGE = 60 * 60 * 24 * 7
TENANT_REGISTRATION_SALT = "tenants.registration-link"


def tenant_registration_token(tenant):
    return signing.dumps({"tenant_id": tenant.pk}, salt=TENANT_REGISTRATION_SALT)


def _tenant_from_registration_token(token):
    data = signing.loads(
        token,
        salt=TENANT_REGISTRATION_SALT,
        max_age=TENANT_REGISTRATION_MAX_AGE,
    )
    return get_object_or_404(Tenant, pk=data["tenant_id"])


def _split_registration_name(name):
    parts = (name or "").strip().split()
    if not parts:
        return "Tenant", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


# ===================== TENANT DETAIL INLINE UPDATE ADDITIONS =====================
# Copy/paste this block into tenants/views.py near your other tenant_detail helper views.
# Required import to add near the top if not already present:
# from django.core.exceptions import ValidationError


@login_required
@require_POST
def tenant_inline_update(request, pk):
    """AJAX single-field update from tenant_detail.html.

    Permission: user must have tenants.change_tenant.
    This endpoint intentionally accepts only whitelisted fields.
    """
    if not request.user.has_perm("tenants.change_tenant"):
        return JsonResponse({"ok": False, "error": "Permission denied."}, status=403)

    tenant = get_object_or_404(Tenant, pk=pk)
    field_name = (request.POST.get("field") or "").strip()
    raw_value = request.POST.get("value", "")

    allowed_fields = {
        "prefix",
        "first_name",
        "relation",
        "last_name",
        "email",
        "phone",
        "phone2",
        "phone3",
        "cnic",
        "occupation",
        "employer_name",
        "employer_phone",
        "reference_name_1",
        "reference_phone_1",
        "reference_relation_1",
        "reference_name_2",
        "reference_phone_2",
        "reference_relation_2",
        "nationality",
        "city",
        "province",
        "country",
        "address",
        "temporary_address",
        "permanent_address",
        "working_address",
        "gender",
        "date_of_birth",
        "emergency_contact_name",
        "emergency_contact_phone",
        "emergency_contact_relation",
        "number_of_family_member",
        "family_member_adults",
        "family_member_children",
        "nadra_family_no",
        "is_active",
        "notes",
        "police_verification_status",
        "police_verification_date",
        "police_verification_remarks",
        "police_verification_follow_up_date",
    }

    if field_name not in allowed_fields:
        return JsonResponse(
            {"ok": False, "error": "This field cannot be updated inline."}, status=400
        )

    try:
        model_field = Tenant._meta.get_field(field_name)
        if model_field.get_internal_type() == "DateField":
            value = parse_date(raw_value) if raw_value else None
            if raw_value and value is None:
                return JsonResponse(
                    {"ok": False, "error": "Invalid date format."}, status=400
                )
        elif model_field.get_internal_type() == "BooleanField":
            value = str(raw_value).lower() in {"1", "true", "yes", "on"}
        elif model_field.get_internal_type() in {
            "PositiveIntegerField",
            "IntegerField",
        }:
            value = int(raw_value or 0)
        else:
            value = raw_value.strip()

        setattr(tenant, field_name, value)
        tenant.full_clean()
        tenant.save()

        normalized_value = getattr(tenant, field_name)
        display_value = normalized_value
        if field_name == "cnic":
            display_value = format_cnic(normalized_value)
        elif field_name in {
            "phone",
            "phone2",
            "phone3",
            "employer_phone",
            "reference_phone_1",
            "reference_phone_2",
            "emergency_contact_phone",
        }:
            display_value = format_phone(normalized_value)
        elif model_field.get_internal_type() == "DateField" and display_value:
            display_value = display_value.strftime("%b %d, %Y")
        elif model_field.choices:
            display_value = getattr(tenant, f"get_{field_name}_display")()
        elif model_field.get_internal_type() == "BooleanField":
            display_value = "Active" if display_value else "Inactive"
        else:
            display_value = str(display_value or "-")

        return JsonResponse(
            {
                "ok": True,
                "field": field_name,
                "value": normalized_value,
                "display": display_value,
                "display_value": display_value,
            }
        )
    except ValidationError as exc:
        return JsonResponse(
            {
                "ok": False,
                "error": exc.message_dict
                if hasattr(exc, "message_dict")
                else exc.messages,
            },
            status=400,
        )
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)


@login_required
@require_POST
def tenant_document_replace(request, pk):
    """AJAX document/photo replacement from tenant_detail.html.

    The Tenant model already controls upload naming through upload_to and compresses
    tenant files in save(), so this view does not duplicate filename logic.
    """
    if not request.user.has_perm("tenants.change_tenant"):
        return JsonResponse({"ok": False, "error": "Permission denied."}, status=403)

    tenant = get_object_or_404(Tenant, pk=pk)
    field_name = (request.POST.get("field") or "").strip()
    upload = request.FILES.get("file")

    allowed_file_fields = {
        "photo": "Tenant Photo",
        "cnic_front": "CNIC Front",
        "cnic_back": "CNIC Back",
        "police_verification_document": "Police Verification Document",
    }

    if field_name not in allowed_file_fields:
        return JsonResponse(
            {"ok": False, "error": "This document field cannot be replaced."},
            status=400,
        )
    if not upload:
        return JsonResponse({"ok": False, "error": "No file was uploaded."}, status=400)

    setattr(tenant, field_name, upload)
    try:
        tenant.full_clean(exclude=["photo_crop", "cnic_front_crop", "cnic_back_crop"])
        tenant.save()
    except ValidationError as exc:
        return JsonResponse(
            {
                "ok": False,
                "error": exc.message_dict
                if hasattr(exc, "message_dict")
                else exc.messages,
            },
            status=400,
        )

    file_field = getattr(tenant, field_name)
    return JsonResponse(
        {
            "ok": True,
            "field": field_name,
            "label": allowed_file_fields[field_name],
            "url": file_field.url if file_field else "",
        }
    )


# In TenantDetailView.get_context_data(), add this before return context:
# context["can_inline_update_tenant"] = self.request.user.has_perm("tenants.change_tenant")


def _create_new_registration_shell(form=None):
    cleaned_data = getattr(form, "cleaned_data", {}) if form else {}
    first_name, last_name = _split_registration_name(
        cleaned_data.get("name") or "New Registration"
    )
    notes = (cleaned_data.get("notes") or "").strip()
    tenant = Tenant.objects.create(
        first_name=first_name,
        last_name=last_name or "Registration",
        phone=(cleaned_data.get("phone") or "").strip(),
        email=(cleaned_data.get("email") or "").strip(),
        cnic="",
        is_active=False,
        notes=notes or "Created from quick tenant registration.",
    )
    if cleaned_data.get("interested_in"):
        tenant.interested_in.set(cleaned_data["interested_in"])
    return tenant


def _tenant_list_public_registration_payload(request):
    """
    Public no-login link for a brand-new tenant registration.

    This is not tied to an existing tenant.
    When opened, it creates a temporary tenant shell and redirects
    to the existing secure public registration form.
    """
    link = request.build_absolute_uri(reverse("tenants:tenant_public_registration_new"))

    message = (
        f"Please complete your tenant registration using this public link:\n\n{link}"
    )

    return {
        "link": link,
        "whatsapp_url": f"https://wa.me/?text={quote(message)}",
    }


def tenant_public_registration_new(request):
    """
    No-login entry point for a new tenant with no prior tenant record.

    It creates a temporary tenant shell only when the public link is opened,
    then sends the user to the existing token-based public registration form.
    """
    tenant = _create_new_registration_shell()
    token = tenant_registration_token(tenant)
    return redirect("tenants:tenant_public_registration", token=token)


def _registration_link_payload(request, tenant):
    link = request.build_absolute_uri(
        reverse(
            "tenants:tenant_public_registration",
            args=[tenant_registration_token(tenant)],
        )
    )
    settings_obj = GlobalSettings.get_solo()
    message = (
        "Please complete your tenant registration using this secure link:\n\n"
        f"{link}\n\n"
        f"This link is valid for {TENANT_REGISTRATION_MAX_AGE // (60 * 60 * 24)} days."
    )
    user_phone = getattr(request.user, "whatsapp_number", "") or getattr(
        settings_obj, "whatsapp_number", ""
    )
    return {
        "tenant_id": tenant.pk,
        "link": link,
        "masked_link": "Registration link ready",
        "whatsapp_url": build_whatsapp_url(
            user_phone,
            message,
            country_code=getattr(settings_obj, "country_code", "+92"),
        ),
    }


def _family_members_from_post(post):
    rows = {}
    for key, value in post.items():
        if not key.startswith("family-"):
            continue
        parts = key.split("-", 2)
        if len(parts) != 3:
            continue
        rows.setdefault(parts[1], {})[parts[2]] = (value or "").strip()
    return [row for row in rows.values() if row.get("name") or row.get("cnic")]


def _relationship_type_from_value(value):
    if not value:
        return None
    LeaseRelationshipType = apps.get_model("leases", "LeaseRelationshipType")
    try:
        return LeaseRelationshipType.objects.filter(
            pk=int(value), is_active=True
        ).first()
    except (TypeError, ValueError):
        return LeaseRelationshipType.objects.filter(code=value, is_active=True).first()


def _family_relationship_defaults(value):
    relationship_type = _relationship_type_from_value(value)
    if relationship_type:
        return {
            "relationship_type": relationship_type,
            "relationship": relationship_type.code[:30],
        }
    return {"relationship": (value or "other")[:30]}


def _tenant_detail_target_lease(tenant, lease_id):
    queryset = Lease.objects.filter(tenant=tenant).order_by("-start_date", "-id")
    if lease_id:
        return get_object_or_404(queryset, pk=lease_id)
    return queryset.filter(status="active").first() or queryset.first()


def _age_from_dob(dob):
    if not dob:
        return None
    if hasattr(dob, "date"):
        dob = dob.date()
    today = timezone.localdate()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))


def _family_counts(links):
    adults = 0
    children = 0
    for link in links:
        age = _age_from_dob(getattr(link.family_member, "date_of_birth", None))
        if age is not None and age < 18:
            children += 1
        else:
            adults += 1
    return {
        "family_adults": adults,
        "family_children": children,
        "family_total": adults + children,
    }


def _apply_family_members_from_submission(tenant, family_members):
    LeaseFamilyMember = apps.get_model("leases", "LeaseFamilyMember")
    lease = (
        tenant.current_lease
        or Lease.objects.filter(tenant=tenant).order_by("-start_date", "-id").first()
    )
    if not lease:
        return 0

    saved = 0
    for row in family_members or []:
        full_name = (row.get("name") or "").strip()
        cnic = (row.get("cnic") or "").strip()
        if not full_name and not cnic:
            continue
        first_name, last_name = _split_registration_name(full_name or "Family Member")
        cnic_digits = normalize_cnic(cnic)
        family_tenant = (
            Tenant.objects.filter(cnic_digits=cnic_digits).first()
            if cnic_digits
            else None
        )
        if not family_tenant:
            family_tenant = Tenant.objects.create(
                first_name=first_name,
                last_name=last_name,
                cnic=cnic,
                phone=row.get("phone") or "",
                occupation=row.get("occupation") or "",
                nationality=row.get("nationality") or "Pakistani",
                email=row.get("email") or "",
                notes=row.get("notes") or "",
            )
        else:
            changed = []
            for field, value in {
                "phone": row.get("phone"),
                "occupation": row.get("occupation"),
                "nationality": row.get("nationality"),
                "email": row.get("email"),
                "notes": row.get("notes"),
            }.items():
                if value and getattr(family_tenant, field) != value:
                    setattr(family_tenant, field, value)
                    changed.append(field)
            if changed:
                family_tenant.save(update_fields=changed)

        dob = parse_date(row.get("dob") or "")
        if dob and not family_tenant.date_of_birth:
            family_tenant.date_of_birth = dob
            family_tenant.save(update_fields=["date_of_birth"])

        relationship_type = _relationship_type_from_value(
            row.get("relationship_type") or row.get("relationship")
        )
        defaults = {
            "is_adult": True if not dob else (_age_from_dob(dob) or 0) >= 18,
            "lives_with_tenant": True,
        }
        if relationship_type:
            defaults["relationship_type"] = relationship_type
            defaults["relationship"] = relationship_type.code[:30]
        else:
            defaults["relationship"] = row.get("relationship") or "other"

        LeaseFamilyMember.objects.update_or_create(
            lease=lease,
            primary_tenant=tenant,
            family_member=family_tenant,
            defaults=defaults,
        )
        saved += 1
    return saved


def tenant_public_registration_update(request, token):
    try:
        tenant = _tenant_from_registration_token(token)
    except SignatureExpired:
        return render(request, "tenants/public_registration_expired.html", status=410)
    except BadSignature:
        raise Http404("Invalid registration link")

    whatsapp_link = (
        WhatsAppExternalLinkToken.objects.filter(
            link_type=WhatsAppExternalLinkToken.LINK_TENANT_REGISTRATION,
            tenant=tenant,
            is_active=True,
        )
        .order_by("-created_at")
        .first()
    )
    if whatsapp_link and whatsapp_link.is_valid:
        record_external_link_access(
            request, whatsapp_link, TrustedDeviceRegistry.USER_TYPE_TENANT
        )

    initial = {
        "prefix": tenant.prefix,
        "first_name": tenant.first_name,
        "relation": tenant.relation,
        "last_name": tenant.last_name,
        "email": tenant.email,
        "phone": tenant.phone,
        "phone2": tenant.phone2,
        "phone3": tenant.phone3,
        "cnic": tenant.cnic,
        "occupation": tenant.occupation,
        "employer_name": tenant.employer_name,
        "employer_phone": tenant.employer_phone,
        "employer_address": tenant.employer_address,
        "reference_name_1": tenant.reference_name_1,
        "reference_phone_1": tenant.reference_phone_1,
        "reference_relation_1": tenant.reference_relation_1,
        "reference_name_2": tenant.reference_name_2,
        "reference_phone_2": tenant.reference_phone_2,
        "reference_relation_2": tenant.reference_relation_2,
        "nationality": tenant.nationality,
        "city": tenant.city,
        "province": tenant.province,
        "country": tenant.country,
        "gender": tenant.gender,
        "date_of_birth": tenant.date_of_birth if tenant.date_of_birth else None,
        "address": tenant.address,
        "temporary_address": tenant.temporary_address,
        "permanent_address": tenant.permanent_address,
        "working_address": tenant.working_address,
        "emergency_contact_name": tenant.emergency_contact_name,
        "emergency_contact_phone": tenant.emergency_contact_phone,
        "emergency_contact_relation": tenant.emergency_contact_relation,
        "number_of_family_member": tenant.number_of_family_member,
        "interested_in": tenant.interested_in.all(),
        "notes": tenant.notes,
    }
    if request.method == "POST":
        form = TenantPublicRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            submitted_data = form.cleaned_data.copy()
            submitted_data["interested_in"] = [
                item.pk for item in submitted_data.get("interested_in", [])
            ]
            for date_field in ["date_of_birth"]:
                if submitted_data.get(date_field):
                    submitted_data[date_field] = submitted_data[date_field].isoformat()
            for file_field in ["photo", "cnic_front", "cnic_back"]:
                submitted_data.pop(file_field, None)
            submitted_data["family_members"] = _family_members_from_post(request.POST)
            submission = TenantRegistrationSubmission.objects.create(
                tenant=tenant,
                submitted_data=submitted_data,
                photo=request.FILES.get("photo"),
                cnic_front=request.FILES.get("cnic_front"),
                cnic_back=request.FILES.get("cnic_back"),
            )
            role_prefixes = {
                "family": PendingRegistrationPerson.ROLE_FAMILY,
                "proposer": PendingRegistrationPerson.ROLE_PROPOSER,
                "seconder": PendingRegistrationPerson.ROLE_SECONDER,
                "witness1": PendingRegistrationPerson.ROLE_WITNESS_1,
                "witness2": PendingRegistrationPerson.ROLE_WITNESS_2,
            }
            for prefix, role in role_prefixes.items():
                indexes = range(0, 20) if prefix == "family" else range(0, 1)
                for index in indexes:
                    base = f"{prefix}-{index}-" if prefix == "family" else f"{prefix}-"
                    first_name = (request.POST.get(base + "first_name") or "").strip()
                    last_name = (request.POST.get(base + "last_name") or "").strip()
                    if prefix == "family" and not (first_name or last_name):
                        full_name = (request.POST.get(base + "name") or "").strip()
                        first_name, last_name = (
                            _split_registration_name(full_name)
                            if full_name
                            else ("", "")
                        )
                    cnic = (request.POST.get(base + "cnic") or "").strip()
                    if not any(
                        (first_name, last_name, cnic, request.POST.get(base + "phone"))
                    ):
                        continue
                    person = PendingRegistrationPerson.objects.create(
                        submission=submission,
                        role=role,
                        first_name=first_name,
                        last_name=last_name,
                        father_husband_name=(
                            request.POST.get(base + "father_husband_name") or ""
                        ).strip(),
                        cnic=cnic,
                        phone=(request.POST.get(base + "phone") or "").strip(),
                        date_of_birth=parse_date(
                            request.POST.get(base + "date_of_birth") or request.POST.get(base + "dob") or ""
                        ),
                        address=(request.POST.get(base + "address") or "").strip(),
                        relationship=(
                            request.POST.get(base + "relationship") or ""
                        ).strip(),
                        relationship_type_id=(
                            request.POST.get(base + "relationship_type") or None
                        ),
                        photo=request.FILES.get(base + "photo"),
                        cnic_front=request.FILES.get(base + "cnic_front"),
                        cnic_back=request.FILES.get(base + "cnic_back"),
                    )
                    from tenants.services.registration_workflow import proposed_changes

                    person.proposed_updates = proposed_changes(person)
                    person.save(update_fields=["proposed_updates", "updated_at"])
            vehicle_rows, vehicle_errors = create_pending_vehicle_submissions_from_post(
                request,
                tenant=tenant,
                pending_tenant_submission=submission,
                source="tenant_registration",
            )
            if vehicle_errors:
                submission.delete()
                for error in vehicle_errors:
                    form.add_error(None, error)
            else:
                vehicle_message = (
                    "Vehicle information submitted and waiting for staff approval."
                    if vehicle_rows
                    else ""
                )
                return render(
                    request,
                    "tenants/public_registration_submitted.html",
                    {
                        "tenant": tenant,
                        "submission": submission,
                        "submitted_name": f"{submitted_data.get('first_name', '')} {submitted_data.get('last_name', '')}".strip(),
                        "submitted_phone": submitted_data.get("phone") or "",
                        "vehicle_message": vehicle_message,
                        "submitted_photo": submission.photo,
                    },
                )
    else:
        form = TenantPublicRegistrationForm(initial=initial)
    existing_family = []
    try:
        lease = tenant.current_lease
        if lease:
            existing_family = lease.family_members.select_related("family_member").all()
    except Exception:
        existing_family = []
    relationship_types = (
        apps.get_model("leases", "LeaseRelationshipType")
        .objects.filter(is_active=True)
        .order_by("sort_order", "name")
    )
    return render(
        request,
        "tenants/public_registration_form.html",
        {
            "tenant": tenant,
            "form": form,
            "existing_family": existing_family,
            "relationship_types": relationship_types,
            "vehicle_types": LeaseVehicleType.objects.filter(is_active=True).order_by(
                "sort_order", "name"
            ),
        },
    )


@require_POST
@login_required
def tenant_pre_registration_link_create(request):
    form = TenantPreRegistrationLinkForm(request.POST or None)
    if not form.is_valid():
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": False, "errors": form.errors}, status=400)
        messages.error(
            request, "Could not create registration. Please check the required fields."
        )
        return redirect("tenants:tenant_list")

    tenant = _create_new_registration_shell(form)
    payload = _registration_link_payload(request, tenant)
    payload["success"] = True
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse(payload)
    messages.success(request, "New tenant registration link generated.")
    return redirect(f"{reverse('tenants:tenant_list')}?new_registration={tenant.pk}")


class TenantRegistrationSubmissionListView(LoginRequiredMixin, ListView):
    model = TenantRegistrationSubmission
    template_name = "tenants/registration_submission_list.html"
    context_object_name = "submissions"

    def get_queryset(self):
        return super().get_queryset().select_related("tenant")


class TenantRegistrationSubmissionDetailView(LoginRequiredMixin, DetailView):
    model = TenantRegistrationSubmission
    template_name = "tenants/registration_submission_detail.html"
    context_object_name = "submission"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = TenantRegistrationSubmissionReviewForm(instance=self.object)
        context["registration_link"] = self.request.build_absolute_uri(
            reverse(
                "tenants:tenant_public_registration",
                args=[tenant_registration_token(self.object.tenant)],
            )
        )
        context["pending_vehicle_submissions"] = (
            self.object.pending_vehicle_submissions.select_related("vehicle_type")
        )
        context["pending_people"] = self.object.pending_people.select_related(
            "matched_tenant", "processed_tenant"
        )
        return context


@login_required
@require_POST
def tenant_registration_submission_review(request, pk):
    submission = get_object_or_404(TenantRegistrationSubmission, pk=pk)
    form = TenantRegistrationSubmissionReviewForm(request.POST, instance=submission)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.reviewed_by = request.user
        obj.reviewed_at = timezone.now()
        obj.save()
        if obj.status == "approved":
            tenant = obj.tenant
            tenant.is_active = True
            allowed = [
                "prefix",
                "first_name",
                "relation",
                "last_name",
                "email",
                "phone",
                "phone2",
                "phone3",
                "cnic",
                "occupation",
                "employer_name",
                "employer_phone",
                "employer_address",
                "reference_name_1",
                "reference_phone_1",
                "reference_relation_1",
                "reference_name_2",
                "reference_phone_2",
                "reference_relation_2",
                "nationality",
                "city",
                "province",
                "country",
                "gender",
                "date_of_birth",
                "address",
                "temporary_address",
                "permanent_address",
                "working_address",
                "emergency_contact_name",
                "emergency_contact_phone",
                "emergency_contact_relation",
                "number_of_family_member",
                "family_member_adults",
                "family_member_children",
                "nadra_family_no",
                "notes",
            ]
            for field in allowed:
                value = obj.submitted_data.get(field, getattr(tenant, field))
                if field == "date_of_birth" and value:
                    value = parse_date(value)
                setattr(tenant, field, value)
            if obj.photo:
                tenant.photo = obj.photo
            if obj.cnic_front:
                tenant.cnic_front = obj.cnic_front
            if obj.cnic_back:
                tenant.cnic_back = obj.cnic_back
            tenant.save()
            if "interested_in" in obj.submitted_data:
                tenant.interested_in.set(obj.submitted_data.get("interested_in") or [])
            family_count = _apply_family_members_from_submission(
                tenant,
                obj.submitted_data.get("family_members", []),
            )
            if family_count:
                messages.success(
                    request, f"{family_count} family member relationship(s) saved."
                )
            messages.success(
                request, "Tenant registration update approved and applied."
            )
        else:
            messages.success(request, "Tenant registration submission updated.")
    else:
        messages.error(request, "Could not update registration submission.")
    return redirect("tenants:registration_submission_detail", pk=submission.pk)


def get_units_by_property(request):
    property_id = request.GET.get("property_id")
    units = Unit.objects.filter(property_id=property_id).order_by("unit_number")
    data = {
        "units": [{"id": unit.id, "unit_number": unit.unit_number} for unit in units]
    }
    return JsonResponse(data)


@login_required
@require_POST
def tenant_vehicle_add(request, pk):
    tenant = get_object_or_404(Tenant, pk=pk)
    lease = _tenant_detail_target_lease(tenant, request.POST.get("lease"))
    if not lease:
        messages.error(request, "Create a lease before adding a vehicle.")
        return redirect("tenants:tenant_detail", pk=tenant.pk)

    vehicle_type = get_object_or_404(
        LeaseVehicleType,
        pk=request.POST.get("vehicle_type"),
        is_active=True,
    )
    registration_number = (request.POST.get("registration_number") or "").strip()
    if not registration_number:
        messages.error(request, "Vehicle registration number is required.")
        return redirect(
            f"{reverse('tenants:tenant_detail', args=[tenant.pk])}#tenantVehicles"
        )

    year = None
    year_raw = (request.POST.get("year") or "").strip()
    if year_raw:
        try:
            year = int(year_raw)
        except (TypeError, ValueError):
            messages.error(request, "Vehicle year must be a number.")
            return redirect(
                f"{reverse('tenants:tenant_detail', args=[tenant.pk])}#tenantVehicles"
            )

    try:
        LeaseVehicle.objects.create(
            lease=lease,
            tenant=tenant,
            vehicle_type=vehicle_type,
            registration_number=registration_number,
            make=(request.POST.get("make") or "").strip(),
            model=(request.POST.get("model") or "").strip(),
            color=(request.POST.get("color") or "").strip(),
            year=year,
            owner_name=(request.POST.get("owner_name") or "").strip(),
            owner_cnic=(request.POST.get("owner_cnic") or "").strip(),
            parking_slot=(request.POST.get("parking_slot") or "").strip(),
            vehicle_photo=request.FILES.get("vehicle_photo"),
            registration_book_photo=request.FILES.get("registration_book_photo"),
            notes=(request.POST.get("notes") or "").strip(),
        )
    except Exception as exc:
        messages.error(request, f"Vehicle not added: {exc}")
    else:
        messages.success(request, "Vehicle added.")
    return redirect(
        f"{reverse('tenants:tenant_detail', args=[tenant.pk])}#tenantVehicles"
    )


@login_required
@require_POST
def tenant_family_add(request, pk):
    tenant = get_object_or_404(Tenant, pk=pk)
    lease = _tenant_detail_target_lease(tenant, request.POST.get("lease"))
    if not lease:
        messages.error(request, "Create a lease before adding family members.")
        return redirect("tenants:tenant_detail", pk=tenant.pk)

    family_member_id = request.POST.get("family_member")
    relation = (request.POST.get("relation") or "").strip()
    if not family_member_id:
        messages.error(request, "Please select a family member tenant.")
        return redirect(
            f"{reverse('tenants:tenant_detail', args=[tenant.pk])}#tenantFamily"
        )
    family_member = get_object_or_404(Tenant, pk=family_member_id)
    if family_member.pk == tenant.pk:
        messages.error(request, "A tenant cannot be added as their own family member.")
        return redirect(
            f"{reverse('tenants:tenant_detail', args=[tenant.pk])}#tenantFamily"
        )

    link, created = LeaseFamilyMember.objects.get_or_create(
        lease=lease,
        primary_tenant=tenant,
        family_member=family_member,
        defaults=_family_relationship_defaults(relation),
    )
    if not created and relation:
        defaults = _family_relationship_defaults(relation)
        changed = []
        if (
            defaults.get("relationship_type")
            and link.relationship_type_id != defaults["relationship_type"].pk
        ):
            link.relationship_type = defaults["relationship_type"]
            changed.append("relationship_type")
        if (
            defaults.get("relationship")
            and link.relationship != defaults["relationship"]
        ):
            link.relationship = defaults["relationship"]
            changed.append("relationship")
        if changed:
            link.save(update_fields=changed)
    messages.success(request, f"{family_member.get_full_name()} added to family.")
    return redirect(
        f"{reverse('tenants:tenant_detail', args=[tenant.pk])}#tenantFamily"
    )


@login_required
@require_POST
def tenant_family_create_and_add(request, pk):
    tenant = get_object_or_404(Tenant, pk=pk)
    lease = _tenant_detail_target_lease(tenant, request.POST.get("lease"))
    if not lease:
        messages.error(request, "Create a lease before adding family members.")
        return redirect("tenants:tenant_detail", pk=tenant.pk)

    full_name = (request.POST.get("full_name") or "").strip()
    first_name = (request.POST.get("first_name") or "").strip()
    last_name = (request.POST.get("last_name") or "").strip()
    relation = (request.POST.get("relation") or "").strip()
    cnic = (request.POST.get("cnic") or "").strip()
    phone = (request.POST.get("phone") or "").strip()
    date_of_birth = (request.POST.get("date_of_birth") or "").strip()
    if full_name and not first_name:
        first_name, last_name = _split_registration_name(full_name)
    if not first_name:
        messages.error(request, "Family member name is required.")
        return redirect(
            f"{reverse('tenants:tenant_detail', args=[tenant.pk])}#tenantFamily"
        )
    if not last_name:
        last_name = "Family"
    if not relation:
        messages.error(request, "Relationship is required.")
        return redirect(
            f"{reverse('tenants:tenant_detail', args=[tenant.pk])}#tenantFamily"
        )

    cnic_digits = normalize_cnic(cnic)
    family_member = (
        Tenant.objects.filter(cnic_digits=cnic_digits).first() if cnic_digits else None
    )
    if not family_member:
        family_member = Tenant(
            first_name=first_name,
            last_name=last_name,
            cnic=cnic,
            phone=phone or None,
            gender=(request.POST.get("gender") or "M"),
        )
        if date_of_birth:
            family_member.date_of_birth = date_of_birth
        for field_name in ("photo", "cnic_front", "cnic_back"):
            uploaded = request.FILES.get(field_name)
            if uploaded:
                setattr(family_member, field_name, uploaded)
        family_member.save()

    LeaseFamilyMember.objects.get_or_create(
        lease=lease,
        primary_tenant=tenant,
        family_member=family_member,
        defaults=_family_relationship_defaults(relation),
    )
    messages.success(request, f"{family_member.get_full_name()} added to family.")
    return redirect(
        f"{reverse('tenants:tenant_detail', args=[tenant.pk])}#tenantFamily"
    )


class TenantDetailView(LoginRequiredMixin, DetailView):
    model = Tenant
    template_name = "tenants/tenant_detail.html"
    context_object_name = "tenant"

    def get_queryset(self):
        # PERF: tenant detail loops over leases and then touches unit/property/payments/invoices.
        lease_qs = (
            Lease.objects.select_related(
                "unit",
                "unit__property",
            )
            .prefetch_related(
                Prefetch(
                    "payments",
                    queryset=Payment.objects.select_related(
                        "lease",
                        "lease__tenant",
                        "lease__unit",
                        "lease__unit__property",
                    ),
                ),
                Prefetch(
                    "invoices",
                    queryset=Invoice.objects.select_related(
                        "lease",
                        "lease__tenant",
                        "lease__unit",
                        "lease__unit__property",
                    ),
                ),
            )
            .order_by("-start_date", "-id")
        )
        return (
            super()
            .get_queryset()
            .prefetch_related(Prefetch("leases", queryset=lease_qs))
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = self.object

        # PERF: this detail page repeats tenant lease/invoice/payment lookups; reuse selected querysets.
        all_leases = list(tenant.leases.all())
        lease_ids = [item.id for item in all_leases]

        # Get active lease if available
        lease = next((item for item in all_leases if item.status == "active"), None)
        tenant.lease = lease

        # Safely initialize all values
        invoices = Invoice.objects.none()
        payments = Payment.objects.none()

        money_field = DecimalField(max_digits=12, decimal_places=2)
        zero = Decimal("0.00")
        invoice_totals = {}
        payment_totals = {}
        security_totals = {}
        if lease_ids:
            invoice_totals = {
                row["lease_id"]: row["total"] or zero
                for row in (
                    Invoice.objects.filter(lease_id__in=lease_ids)
                    .values("lease_id")
                    .annotate(total=Coalesce(Sum("amount"), zero))
                )
            }
            payment_totals = {
                row["lease_id"]: row["total"] or zero
                for row in (
                    Payment.objects.filter(lease_id__in=lease_ids)
                    .values("lease_id")
                    .annotate(
                        total=Coalesce(
                            Sum(
                                Case(
                                    When(
                                        detail__isnull=False,
                                        then=F("detail__lease_amount"),
                                    ),
                                    default=F("amount"),
                                    output_field=money_field,
                                )
                            ),
                            zero,
                        )
                    )
                )
            }
            for row in (
                SecurityDepositTransaction.objects.filter(lease_id__in=lease_ids)
                .values("lease_id", "type")
                .annotate(total=Coalesce(Sum("amount"), zero))
            ):
                security_totals.setdefault(row["lease_id"], {})[row["type"]] = (
                    row["total"] or zero
                )

        tenant.current_balance = Decimal("0.00")
        for item in all_leases:
            lease_balance = invoice_totals.get(item.id, zero) - payment_totals.get(
                item.id, zero
            )
            sd = security_totals.get(item.id, {})
            security_balance = (
                (getattr(item, "security_deposit", None) or zero)
                - sd.get("PAYMENT", zero)
                - sd.get("ADJUST", zero)
            )
            if security_balance < zero:
                security_balance = zero
            item.list_balance = lease_balance
            item.list_security_due = security_balance
            item.tenant_detail_lease_balance = lease_balance
            item.tenant_detail_security_balance = security_balance
            item.tenant_detail_total_balance = lease_balance + security_balance
            tenant.current_balance += item.tenant_detail_total_balance

        # Defaults for the “active lease” tables
        invoices = Invoice.objects.none()
        payments = Payment.objects.none()
        total_invoices_active = 0
        total_payments_active = 0

        # If there is an active lease, load its items for the table ONLY,
        # but DO NOT override tenant.current_balance anymore.
        if lease:
            invoices = sorted(
                list(lease.invoices.all()), key=lambda inv: inv.issue_date, reverse=True
            )
            payments = sorted(
                list(lease.payments.all()),
                key=lambda payment: payment.payment_date,
                reverse=True,
            )

            total_invoices_active = sum((inv.amount or 0 for inv in invoices), 0)
            total_payments_active = sum(
                (payment.amount or 0 for payment in payments), 0
            )

        # After you compute tenant.current_balance in TenantDetailView
        context["all_leases"] = all_leases
        context["leases"] = all_leases
        context["leases_total_balance"] = (
            tenant.current_balance
        )  # all lease total balances, active and inactive
        family_target_leases = [
            item for item in all_leases if item.status == "active"
        ] or all_leases

        context.update(
            {
                "invoices": invoices,
                "payments": payments,
                # active-lease totals for that section
                "total_invoices": total_invoices_active,
                # active-lease totals for that section
                "total_payments": total_payments_active,
                "all_leases": all_leases,  # for the partial loop
                "leases": all_leases,
                "family_target_leases": family_target_leases,
                "family_target_lease_count": len(family_target_leases),
                # <-- expose tenant-wide balance explicitly
                "current_balance": tenant.current_balance,
            }
        )
        context["registration_link"] = self.request.build_absolute_uri(
            reverse(
                "tenants:tenant_public_registration",
                args=[tenant_registration_token(tenant)],
            )
        )
        registration_message = (
            f"Hello {tenant.get_full_name()},\n\n"
            "Please complete or update your tenant registration using the secure link below:\n\n"
            f"{context['registration_link']}\n\n"
            f"This link will expire in {TENANT_REGISTRATION_MAX_AGE // (60 * 60 * 24)} days.\n\n"
            "Thank you."
        )
        settings_obj = GlobalSettings.get_solo()
        context["registration_whatsapp_url"] = build_whatsapp_url(
            tenant.phone or tenant.phone2 or tenant.phone3 or "",
            registration_message,
            country_code=getattr(settings_obj, "country_code", "+92"),
        )
        family_links = []
        family_lease_links = []
        try:
            LeaseFamilyMember = apps.get_model("leases", "LeaseFamilyMember")

            # Family members under this tenant's own lease
            family_links = list(
                LeaseFamilyMember.objects.filter(primary_tenant=tenant)
                .select_related("lease", "family_member")
                .order_by(
                    "sort_order",
                    "family_member__first_name",
                    "family_member__last_name",
                )
            )

            # Leases where this tenant is listed as someone else's family member
            family_lease_links = list(
                LeaseFamilyMember.objects.filter(family_member=tenant)
                .select_related(
                    "lease",
                    "lease__tenant",
                    "lease__unit",
                    "lease__unit__property",
                    "primary_tenant",
                    "relationship_type",
                )
                .order_by("-lease__start_date", "-lease__id")
            )
        except Exception:
            family_links = []
            family_lease_links = []

        context["family_members"] = family_links
        context["family_lease_links"] = family_lease_links
        context["tenant_vehicles"] = (
            LeaseVehicle.objects.filter(Q(tenant=tenant) | Q(lease_id__in=lease_ids))
            .select_related(
                "lease", "lease__unit", "lease__unit__property", "vehicle_type"
            )
            .distinct()
            .order_by("vehicle_type__sort_order", "registration_number")
        )
        context["vehicle_types"] = LeaseVehicleType.objects.filter(
            is_active=True
        ).order_by("sort_order", "name")
        context["relationship_types"] = (
            apps.get_model("leases", "LeaseRelationshipType")
            .objects.filter(is_active=True)
            .order_by("sort_order", "name")
        )
        context["tenant_family_options"] = Tenant.objects.exclude(
            pk=tenant.pk
        ).order_by("first_name", "last_name")
        context.update(_family_counts(family_links))
        from tenants.services.role_history import tenant_role_history

        context["role_history_rows"] = tenant_role_history(tenant)

        def get_object(self, queryset=None):
            tenant = super().get_object(queryset)
            print(
                f"Retrieved tenant: ID={tenant.pk}, Name={tenant.first_name} {tenant.last_name}"
            )
            return tenant

        print(f"Tenant: {tenant}")
        print(f"Active leases: {tenant.leases.filter(status='active').exists()}")
        print(f"Found lease: {tenant.lease}")
        return context

    def get_object(self, queryset=None):
        # Get the tenant object
        tenant = super().get_object(queryset)
        # Debug print
        print(f"Tenant PK: {tenant.pk}, Name: {tenant.first_name} {tenant.last_name}")
        return tenant


class TenantRoleHistoryView(LoginRequiredMixin, DetailView):
    model = Tenant
    template_name = "tenants/tenant_role_history.html"
    context_object_name = "tenant"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from tenants.services.role_history import tenant_role_history

        role = self.request.GET.get("role") or None
        allowed = {None, "family_member", "proposer", "seconder", "witness"}
        if role not in allowed:
            role = None
        rows = tenant_role_history(self.object, role)
        sort = self.request.GET.get("sort", "status_date")
        direction = self.request.GET.get("direction", "desc")
        keymap = {
            "role_type": lambda r: r.role_type,
            "lease": lambda r: r.lease.pk,
            "tenant": lambda r: r.related_tenant.get_full_name(),
            "property_unit": lambda r: (str(r.property), str(r.unit)),
            "lease_period": lambda r: (r.lease_start, r.lease_end),
            "balance": lambda r: r.balance,
            "status_date": lambda r: r.status_date,
        }
        rows.sort(
            key=keymap.get(sort, keymap["status_date"]), reverse=direction != "asc"
        )
        context.update(
            {
                "role_history_rows": rows,
                "role_filter": role or "",
                "sort": sort,
                "direction": direction,
            }
        )
        return context


class TenantCreateView(LoginRequiredMixin, CreateView):
    model = Tenant
    form_class = TenantForm
    template_name = "tenants/tenant_form.html"

    def get_success_url(self):
        return reverse("tenants:tenant_detail", kwargs={"pk": self.object.pk})

    def get_initial(self):
        """Set initial values from URL parameters"""
        initial = super().get_initial()
        unit_id = self.request.GET.get("unit")
        if unit_id:
            initial["unit"] = unit_id
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Create form has no saved tenant yet, so no lease/family records can be linked here.
        context["all_leases"] = []
        context["leases"] = []
        context["family_target_leases"] = []
        context["family_target_lease_count"] = 0
        context["existing_family"] = []
        context["family_members"] = []
        context["tenant_family_options"] = []

        # Needed by the shared family template/dropdowns, even if family add is hidden on create.
        context["relationship_types"] = LeaseRelationshipType.objects.filter(
            is_active=True
        ).order_by("sort_order", "name")

        return context

    def form_valid(self, form):
        messages.success(self.request, "Tenant was created successfully!")
        return super().form_valid(form)


class TenantUpdateView(LoginRequiredMixin, UpdateView):
    model = Tenant
    form_class = TenantForm
    template_name = "tenants/tenant_form.html"
    success_url = reverse_lazy("tenants:tenant_list")

    def form_valid(self, form):
        messages.success(self.request, "Tenant was updated successfully!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = self.object

        all_leases = (
            Lease.objects.filter(tenant=tenant)
            .select_related("unit", "unit__property")
            .order_by("-start_date", "-id")
        )

        family_target_leases = [
            lease for lease in all_leases if lease.status == "active"
        ] or list(all_leases)

        family_members = (
            LeaseFamilyMember.objects.filter(primary_tenant=tenant)
            .select_related(
                "lease",
                "lease__unit",
                "lease__unit__property",
                "family_member",
                "relationship_type",
            )
            .order_by(
                "sort_order",
                "family_member__first_name",
                "family_member__last_name",
            )
        )

        existing_family_ids = family_members.values_list("family_member_id", flat=True)

        context["all_leases"] = all_leases
        context["leases"] = all_leases
        context["family_target_leases"] = family_target_leases
        context["family_target_lease_count"] = len(family_target_leases)
        context["existing_family"] = family_members
        context["family_members"] = family_members

        context["tenant_family_options"] = (
            Tenant.objects.exclude(pk=tenant.pk)
            .exclude(pk__in=existing_family_ids)
            .order_by("first_name", "last_name")
        )

        context["relationship_types"] = LeaseRelationshipType.objects.filter(
            is_active=True
        ).order_by("sort_order", "name")

        return context


class TenantDeleteView(LoginRequiredMixin, DeleteView):
    model = Tenant
    template_name = "tenants/tenant_confirm_delete.html"
    success_url = reverse_lazy("tenants:tenant_list")

    def delete(self, request, *args, **kwargs):
        """Add success message when tenant is deleted"""
        messages.success(request, "Tenant was deleted successfully!")
        return super().delete(request, *args, **kwargs)


def create_export(self, export_format):
    queryset = self.get_queryset()
    filename = f"tenants_{timezone.now().strftime('%Y%m%d_%H%M%S')}"

    if export_format == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'

        writer = csv.writer(response)
        writer.writerow(["ID", "Name", "Phone", "Property", "Unit", "Rent", "Balance"])

        for tenant in queryset:
            lease = tenant.current_lease
            writer.writerow(
                [
                    tenant.id,
                    f"{tenant.first_name} {tenant.last_name}",
                    format_phone(tenant.phone),
                    lease.unit.property.property_name if lease else "",
                    lease.unit.unit_number if lease else "",
                    lease.get_total_payment if lease else "",
                    lease.get_balance if lease else "",
                ]
            )
        return response

    elif export_format == "xlsx":
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tenants"

        headers = ["ID", "Name", "Phone", "Property", "Unit", "Rent", "Balance"]
        ws.append(headers)

        for tenant in queryset:
            lease = tenant.current_lease
            ws.append(
                [
                    tenant.id,
                    f"{tenant.first_name} {tenant.last_name}",
                    format_phone(tenant.phone),
                    lease.unit.property.property_name if lease else "",
                    lease.unit.unit_number if lease else "",
                    lease.get_total_payment if lease else "",
                    lease.get_balance if lease else "",
                ]
            )

        buffer = BytesIO()
        wb.save(buffer)
        response.write(buffer.getvalue())
        return response

    elif export_format == "pdf":
        html_string = render_to_string(
            "tenants/tenant_list_pdf.html",
            {"tenants": queryset, "date": timezone.now().strftime("%Y-%m-%d")},
        )

        html = HTML(string=html_string, base_url=self.request.build_absolute_uri())
        pdf = html.write_pdf()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
        return response

    return None


def create_tenant(request):
    if request.method == "POST":
        form = TenantForm(request.POST)
        if form.is_valid():
            tenant = form.save()
            # Get unit from form data if available
            unit = (
                form.cleaned_data.get("unit") if "unit" in form.cleaned_data else None
            )

            if unit:  # Only create lease if unit is provided
                Lease.objects.create(
                    tenant=tenant,
                    unit=unit,
                    start_date=timezone.now().date(),
                    # ... other lease fields
                )
            return redirect("success_url")
    else:
        form = TenantForm()
    return render(request, "tenant_form.html", {"form": form})


def generate_agreement(request, lease_id):
    lease = get_object_or_404(Lease, pk=lease_id)
    html_string = render_to_string("leases/agreement_template.html", {"lease": lease})

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="Lease_{lease.id}.pdf"'

    HTML(string=html_string).write_pdf(response)
    return response


# views.py


class BalanceDetailView(DetailView):
    template_name = "balance_detail.html"

    def get_object(self):
        if "tenant_id" in self.kwargs:
            return Tenant.objects.get(pk=self.kwargs["tenant_id"])
        return Lease.objects.get(pk=self.kwargs["lease_id"])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obj = self.get_object()

        if isinstance(obj, Tenant):
            context["invoices"] = Invoice.objects.filter(lease__tenant=obj)
            context["payments"] = Payment.objects.filter(lease__tenant=obj)
        else:
            context["invoices"] = obj.invoices.all()
            context["payments"] = obj.payments.all()

        return context


def print_tenant_view(request):
    tenant_ids = request.GET.get("ids", "").split(",")
    tenants = Tenant.objects.filter(id__in=tenant_ids)
    return render(request, "admin/tenant_print.html", {"tenants": tenants})


def ledger_pdf(request, tenant_id):
    """Generate PDF ledger for tenant"""
    tenant = get_object_or_404(Tenant, pk=tenant_id)

    try:
        # Get transactions using your TenantLedgerView
        transactions = TenantLedgerView().get_queryset()

        # Calculate financials
        total_paid = sum(t["amount"] for t in transactions if t["amount"] > 0)
        total_owed = sum(-t["amount"] for t in transactions if t["amount"] < 0)
        balance = total_paid - total_owed

        context = {
            "tenant": tenant,
            "transactions": transactions,
            "total_paid": total_paid,
            "total_owed": total_owed,
            "balance": balance,
            "date": timezone.now().date(),
        }

        # Render HTML template
        html_string = render_to_string("leases/ledger_pdf.html", context)

        # Add CSS styling (optional - can also be in the template)
        css_string = """
        body { font-family: Arial; font-size: 12px; }
        h1 { color: #333; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }
        .header { display: flex; justify-content: space-between; }
        .totals { margin-top: 20px; font-weight: bold; }
        """

        # Generate PDF
        html = HTML(string=html_string)
        pdf_content = html.write_pdf(stylesheets=[CSS(string=css_string)])

        # Create response
        response = HttpResponse(pdf_content, content_type="application/pdf")
        response["Content-Disposition"] = (
            f"attachment; filename=tenant_{tenant.id}_ledger_{timezone.now().date()}.pdf"
        )
        return response

    except Exception as e:
        # Log the error (you can use logging module)
        print(f"PDF generation error: {str(e)}")
        return HttpResponse(f"Failed to generate PDF: {str(e)}", status=500)


# leases/views.py or tenants/views.py


logger = logging.getLogger(__name__)


@login_required
def send_ledger(request, lease_id):
    if request.method != "POST":
        return JsonResponse(
            {"status": "error", "message": "Invalid request"}, status=400
        )

    lease = get_object_or_404(Lease, pk=lease_id)
    tenant = lease.tenant

    if not tenant.email:
        return JsonResponse(
            {"status": "error", "message": "Tenant has no email address"}, status=400
        )

    try:
        # Generate PDF using the lease-based view
        view = LeaseLedgerView()
        view.request = request
        view.kwargs = {"lease_id": lease_id}
        transactions = view.get_queryset()

        context = {
            "lease": lease,
            "tenant": tenant,
            "transactions": transactions,
            "current_balance": lease.get_balance(),
            "current_start_date": request.GET.get("start_date"),
            "current_end_date": request.GET.get("end_date"),
        }

        html_string = render_to_string("tenants/ledger_pdf_export.html", context)
        pdf = HTML(string=html_string).write_pdf()

        # Create email
        subject = f"Rent Ledger for {tenant.get_full_name()}"
        body = render_to_string(
            "tenants/ledger_email_body.txt",
            {
                "tenant": tenant,
                "lease": lease,
                "current_balance": transactions[-1]["balance"] if transactions else 0,
                "start_date": request.GET.get("start_date"),
                "end_date": request.GET.get("end_date"),
            },
        )

        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[tenant.email],
            reply_to=[settings.DEFAULT_FROM_EMAIL],
        )
        email.attach(
            f"ledger_{tenant.id}_{timezone.now().date()}.pdf", pdf, "application/pdf"
        )
        email.send()

        return JsonResponse(
            {
                "status": "success",
                "message": "Ledger sent successfully",
                "email": tenant.email,
            }
        )

    except Exception as e:
        logger.error(f"Failed to send ledger: {str(e)}", exc_info=True)
        return JsonResponse(
            {"status": "error", "message": f"Failed to send ledger: {str(e)}"},
            status=500,
        )


# tenants/views.py


# tenants/views.py


class LeaseLedgerView(LoginRequiredMixin, SingleTableView):
    table_class = LedgerTable
    template_name = "tenants/lease_ledger.html"
    context_object_name = "table"

    def dispatch(self, request, *args, **kwargs):
        Lease = apps.get_model("leases", "Lease")
        self.lease = get_object_or_404(Lease, pk=kwargs["lease_id"])
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        transactions = []
        balance = Decimal("0.00")

        # Process invoices
        for index, invoice in enumerate(
            self.lease.invoices.all().order_by("issue_date")
        ):
            transactions.append(
                {
                    "index": index,
                    "date": invoice.issue_date,
                    "type": "Invoice",
                    "description": invoice.description,
                    "amount": -invoice.amount,
                    "balance": None,
                }
            )

        # Process payments
        for index, payment in enumerate(
            self.lease.payments.all().order_by("payment_date"), start=len(transactions)
        ):
            transactions.append(
                {
                    "index": index,
                    "date": payment.payment_date,
                    "type": "Payment",
                    "description": payment.reference_number or "Payment",
                    "amount": payment.amount,
                    "balance": None,
                }
            )

        return sorted(transactions, key=lambda x: x["date"])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "lease": self.lease,
                "tenant": self.lease.tenant,
                "current_balance": self.lease.get_balance(),
                "export_formats": ["csv", "xlsx", "pdf"],
            }
        )
        return context


def tenant_search(request):
    search_term = request.GET.get("q", "").strip()
    identity_term = normalize_phone(search_term)
    cnic_term = normalize_cnic(search_term)

    tenants = Tenant.objects.all()
    if search_term:
        search_query = (
            Q(first_name__icontains=search_term)
            | Q(last_name__icontains=search_term)
            | Q(email__icontains=search_term)
        )
        if identity_term:
            search_query |= Q(phone__icontains=identity_term) | Q(phone2__icontains=identity_term) | Q(phone3__icontains=identity_term)
        if cnic_term:
            search_query |= Q(cnic_digits__icontains=cnic_term)
        tenants = tenants.filter(search_query)
    tenants = tenants.distinct().order_by("first_name", "last_name")[:20]

    results = []
    for tenant in tenants:
        lease = tenant.current_lease
        result = {
            "id": tenant.id,
            "text": tenant.get_full_name(),
            "detail_url": reverse("tenants:tenant_detail", args=[tenant.pk]),
            "property": "",
            "unit": "",
            "balance": "",
        }
        if lease and lease.unit:
            balance_getter = getattr(lease, "get_balance_due", None) or getattr(
                lease, "get_balance", None
            )
            balance = balance_getter() if callable(balance_getter) else balance_getter
            result.update(
                {
                    "property": lease.unit.property.property_name
                    if lease.unit.property
                    else "",
                    "unit": lease.unit.unit_number,
                    "balance": balance if balance is not None else "",
                }
            )
        results.append(result)

    return JsonResponse({"results": results})


class TenantLedgerView(LoginRequiredMixin, SingleTableView):
    table_class = LedgerTable
    template_name = "tenants/tenant_ledger.html"
    context_object_name = "table"

    def get_queryset(self):
        tenant_id = self.request.GET.get("tenant", self.kwargs.get("pk"))
        if not tenant_id:
            return []

        tenant = get_object_or_404(Tenant, pk=tenant_id)
        transactions = []
        balance = Decimal("0.00")

        # Get date filters
        date_range = self.request.GET.get("date_range", "")
        start_date = self.request.GET.get("start_date")
        end_date = self.request.GET.get("end_date")

        # Convert string dates to date objects
        if start_date:
            try:
                start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                start_date = None
        if end_date:
            try:
                end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                end_date = None

        # Get transactions
        invoices = Invoice.objects.filter(lease__tenant=tenant)
        payments = Payment.objects.filter(lease__tenant=tenant)

        # Apply date filters
        if start_date:
            invoices = invoices.filter(issue_date__gte=start_date)
            payments = payments.filter(payment_date__gte=start_date)
        if end_date:
            invoices = invoices.filter(issue_date__lte=end_date)
            payments = payments.filter(payment_date__lte=end_date)

        # Process invoices
        for index, invoice in enumerate(invoices.order_by("issue_date")):
            transactions.append(
                {
                    "index": index,
                    "transaction_date": invoice.issue_date,
                    "type": "Invoice",
                    "description": invoice.description,
                    "amount": -invoice.amount,
                    "balance": None,  # Will be calculated in the table
                }
            )

        # Process payments
        for index, payment in enumerate(
            payments.order_by("payment_date"), start=len(transactions)
        ):
            transactions.append(
                {
                    "index": index,
                    "transaction_date": payment.payment_date,
                    "type": "Payment",
                    "description": payment.reference_number or f"Payment {payment.id}",
                    "amount": payment.amount,
                    "balance": None,  # Will be calculated in the table
                }
            )

        # Sort all transactions by date
        transactions.sort(key=lambda x: x["transaction_date"])
        return transactions

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant_id = self.request.GET.get("tenant", self.kwargs.get("pk"))
        tenant = get_object_or_404(Tenant, pk=tenant_id) if tenant_id else None
        lease = tenant.leases.filter(status="active").first() if tenant else None

        # Calculate current balance
        current_balance = Decimal("0.00")
        if tenant and lease:
            total_invoices = Invoice.objects.filter(lease=lease).aggregate(
                total=Sum("amount")
            )["total"] or Decimal("0.00")
            total_payments = Payment.objects.filter(lease=lease).aggregate(
                total=Sum("amount")
            )["total"] or Decimal("0.00")
            current_balance = total_payments - total_invoices

        context.update(
            {
                "tenant": tenant,
                "lease": lease,
                "current_balance": current_balance,
                "current_date_range": self.request.GET.get("date_range", ""),
                "current_start_date": self.request.GET.get("start_date", ""),
                "current_end_date": self.request.GET.get("end_date", ""),
                "export_formats": ["csv", "xlsx", "pdf"],
            }
        )
        return context


def get_units_by_property(request):
    property_id = request.GET.get("property_id")
    units = Unit.objects.filter(property_id=property_id).order_by("unit_number")
    data = {"units": [{"id": u.id, "unit_number": u.unit_number} for u in units]}
    return JsonResponse(data)


# tenants/views.py (update_tenant_field)

@login_required
@require_POST
def update_tenant_field(request, tenant_id):
    try:
        if not request.user.has_perm("tenants.change_tenant"):
            return JsonResponse(
                {"success": False, "error": "Permission denied"}, status=403
            )
        tenant = Tenant.objects.get(pk=tenant_id)
        field = request.POST.get("field")
        value = (request.POST.get("value") or "").strip()

        if field not in [
            "first_name",
            "last_name",
            "email",
            "phone",
            "cnic",
            "address",
        ]:
            return JsonResponse({"success": False, "error": "Invalid field"})

        if field == "cnic":
            digits = normalize_cnic(value)
            try:
                validate_cnic(digits)
            except ValidationError as exc:
                return JsonResponse(
                    {"success": False, "error": exc.messages[0]}
                )

            # Duplicate check
            qs = Tenant.objects.exclude(pk=tenant.pk).filter(cnic_digits=digits)
            if digits and qs.exists():
                return JsonResponse(
                    {
                        "success": False,
                        "error": "A tenant with this CNIC already exists.",
                    }
                )

            value = digits
        elif field == "phone":
            value = normalize_phone(value)
        setattr(tenant, field, value)
        update_fields = [field, "cnic_digits"] if field == "cnic" else [field]
        tenant.save(update_fields=update_fields)
        return JsonResponse({
            "success": True,
            "value": getattr(tenant, field) or "",
            "display_value": format_cnic(getattr(tenant, field)) if field == "cnic" else format_phone(getattr(tenant, field)) if field == "phone" else getattr(tenant, field),
        })
    except Tenant.DoesNotExist:
        return JsonResponse({"success": False, "error": "Tenant not found"})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


# at top of tenants/views.py


class TenantListView(LoginRequiredMixin, ExportMixin, SingleTableView):
    model = Tenant
    template_name = "tenants/tenant_list.html"
    table_class = TenantTable
    context_object_name = "tenants"
    paginate_by = 50
    export_formats = ["csv", "xlsx", "pdf"]  # Add this line

    def get_queryset(self):
        today = timezone.now().date()
        show_inactive = self.request.GET.get("show_inactive") == "on"

        lease_prefetch_queryset = (
            Lease.objects.all()
            if show_inactive
            else Lease.objects.filter(
                status="active",
                start_date__lte=today,
                end_date__gte=today,
            )
        )

        active_leases = Prefetch(
            "leases",
            queryset=lease_prefetch_queryset.select_related("unit__property").order_by(
                "-start_date", "-id"
            ),
            to_attr="active_leases",
        )

        queryset = (
            super().get_queryset().prefetch_related(active_leases, "interested_in")
        )

        tenant_id = self.request.GET.get("tenant")
        phone = self.request.GET.get("phone")
        property_id = self.request.GET.get("property")
        unit_id = self.request.GET.get("unit")
        interest_ids = [
            value for value in self.request.GET.getlist("interested_in") if value
        ]

        tenant_status = self.request.GET.get("tenant_status")
        family_member = self.request.GET.get("family_member") == "1"
        potential_tenant = self.request.GET.get("potential_tenant") == "1"
        on_notice = self.request.GET.get("on_notice") == "1"
        sort = self.request.GET.get("sort")

        has_any_filter = any(
            [
                tenant_id,
                phone,
                property_id,
                unit_id,
                show_inactive,
                family_member,
                potential_tenant,
                on_notice,
                interest_ids,
                tenant_status,
            ]
        )
        if not has_any_filter:
            tenant_status = "active"

        if tenant_id:
            return queryset.filter(id=tenant_id).order_by("first_name", "last_name")

        if phone:
            normalized_phone = normalize_phone(phone)
            normalized_cnic = normalize_cnic(phone)
            phone_query = Q(first_name__icontains=phone) | Q(last_name__icontains=phone)
            if normalized_phone:
                phone_query |= Q(phone__icontains=normalized_phone) | Q(phone2__icontains=normalized_phone) | Q(phone3__icontains=normalized_phone)
            if normalized_cnic:
                phone_query |= Q(cnic_digits__icontains=normalized_cnic)
            queryset = queryset.filter(phone_query)

        if interest_ids:
            queryset = queryset.filter(interested_in__id__in=interest_ids).distinct()

        if potential_tenant:
            queryset = queryset.filter(
                interested_in__isnull=False, is_active=False
            ).distinct()

        lease_filter = Lease.objects.filter(
            tenant_id=OuterRef("pk"),
            status="active",
            start_date__lte=today,
            end_date__gte=today,
        )

        family_filter = LeaseFamilyMember.objects.filter(
            family_member_id=OuterRef("pk"),
            lease__status="active",
            lease__start_date__lte=today,
            lease__end_date__gte=today,
        )

        if property_id:
            lease_filter = lease_filter.filter(unit__property_id=property_id)
            family_filter = family_filter.filter(lease__unit__property_id=property_id)

        if unit_id:
            lease_filter = lease_filter.filter(unit_id=unit_id)
            family_filter = family_filter.filter(lease__unit_id=unit_id)

        if on_notice:
            notice_until = today + timezone.timedelta(days=60)
            lease_filter = lease_filter.filter(
                end_date__gte=today, end_date__lte=notice_until
            )
            family_filter = family_filter.filter(
                lease__end_date__gte=today, lease__end_date__lte=notice_until
            )

        queryset = queryset.annotate(
            has_matching_lease=Exists(lease_filter),
            is_active_family_member=Exists(family_filter),
        )

        if family_member:
            queryset = queryset.filter(is_active_family_member=True)

        if tenant_status == "active":
            queryset = queryset.filter(is_active=True).filter(
                Q(has_matching_lease=True) | Q(is_active_family_member=True)
            )
        elif tenant_status == "inactive":
            queryset = queryset.exclude(
                is_active=True,
                has_matching_lease=True,
            ).exclude(
                is_active=True,
                is_active_family_member=True,
            )
        elif property_id or unit_id or on_notice:
            queryset = queryset.filter(
                Q(has_matching_lease=True) | Q(is_active_family_member=True)
            )
        elif not show_inactive and not interest_ids and not potential_tenant:
            queryset = queryset.filter(is_active=True).filter(
                Q(has_matching_lease=True) | Q(is_active_family_member=True)
            )

        if sort == "lease":
            queryset = queryset.order_by(
                "leases__unit__property__property_name",
                "leases__unit__unit_number",
                "first_name",
                "last_name",
            )
        elif sort == "-lease":
            queryset = queryset.order_by(
                "-leases__unit__property__property_name",
                "-leases__unit__unit_number",
                "first_name",
                "last_name",
            )
        else:
            queryset = queryset.order_by("first_name", "last_name")

        return queryset.distinct()

    def _attach_lease_totals(self, tenants):
        leases = []
        for tenant in tenants:
            lease = tenant.current_lease
            if lease:
                leases.append(lease)

        lease_ids = [lease.id for lease in leases]
        if not lease_ids:
            return

        zero = Decimal("0.00")
        decimal_field = DecimalField(max_digits=12, decimal_places=2)
        zero_db = Value(zero, output_field=decimal_field)

        invoice_totals = {
            row["lease_id"]: row["total"] or zero
            for row in (
                Invoice.objects.filter(lease_id__in=lease_ids)
                .values("lease_id")
                .annotate(total=Coalesce(Sum("amount"), zero_db))
            )
        }

        payment_totals = {
            row["lease_id"]: row["total"] or zero
            for row in (
                Payment.objects.filter(lease_id__in=lease_ids)
                .values("lease_id")
                .annotate(
                    total=Coalesce(
                        Sum(
                            Case(
                                When(
                                    detail__isnull=False,
                                    then=F("detail__lease_amount"),
                                ),
                                default=F("amount"),
                                output_field=decimal_field,
                            )
                        ),
                        zero_db,
                    )
                )
            )
        }

        security_rows = (
            SecurityDepositTransaction.objects.filter(
                lease_id__in=lease_ids, type__in=["PAYMENT", "ADJUST"]
            )
            .values("lease_id", "type")
            .annotate(total=Coalesce(Sum("amount"), zero_db))
        )
        security_totals = {}
        for row in security_rows:
            security_totals.setdefault(row["lease_id"], {})[row["type"]] = (
                row["total"] or zero
            )

        property_ids = [
            lease.unit.property_id
            for lease in leases
            if getattr(lease, "unit_id", None)
        ]
        recurring_rows = (
            RecurringCharge.objects.filter(active=True, kind="FIXED")
            .filter(
                Q(scope="GLOBAL")
                | Q(scope="PROPERTY", property_id__in=property_ids)
                | Q(scope="LEASE", lease_id__in=lease_ids)
            )
            .values("scope", "lease_id", "property_id")
            .annotate(total=Coalesce(Sum("amount"), zero_db))
        )

        recurring_global = zero
        recurring_by_property = {}
        recurring_by_lease = {}
        for row in recurring_rows:
            if row["scope"] == "GLOBAL":
                recurring_global += row["total"] or zero
            elif row["scope"] == "PROPERTY":
                recurring_by_property[row["property_id"]] = recurring_by_property.get(
                    row["property_id"], zero
                ) + (row["total"] or zero)
            elif row["scope"] == "LEASE":
                recurring_by_lease[row["lease_id"]] = recurring_by_lease.get(
                    row["lease_id"], zero
                ) + (row["total"] or zero)

        today = timezone.localdate()
        for lease in leases:
            attach_lease_expiry_countdown(lease, today=today)
            balance = invoice_totals.get(lease.id, zero) - payment_totals.get(
                lease.id, zero
            )
            paid_in = security_totals.get(lease.id, {}).get("PAYMENT", zero)
            adjust = security_totals.get(lease.id, {}).get("ADJUST", zero)
            security_due = max(
                (lease.security_deposit or zero) - paid_in - adjust, zero
            )

            recurring_total = (
                recurring_global
                + recurring_by_property.get(
                    getattr(lease.unit, "property_id", None), zero
                )
                + recurring_by_lease.get(lease.id, zero)
            )
            lease_monthly = Decimal(
                str(getattr(lease, "get_total_payment", zero) or zero)
            )

            lease.cached_monthly_payment = (
                recurring_total if recurring_total > 0 else lease_monthly
            )
            lease._cached_get_balance = balance
            lease._cached_security_due = security_due

    def get_context_data(self, **kwargs):
        # This template renders its own HTML table, so skip SingleTableView's
        # unused table context and avoid a second pagination/count query.
        context = ListView.get_context_data(self, **kwargs)
        property_id = self.request.GET.get("property")
        page_tenants = list(context.get("object_list") or context.get("tenants") or [])
        self._attach_lease_totals(page_tenants)

        today = timezone.now().date()
        LeaseFamilyMember = apps.get_model("leases", "LeaseFamilyMember")

        family_relations = (
            LeaseFamilyMember.objects.filter(
                family_member_id__in=[tenant.pk for tenant in page_tenants],
                lease__status="active",
                lease__start_date__lte=today,
                lease__end_date__gte=today,
            )
            .select_related(
                "lease__unit__property",
                "primary_tenant",
                "relationship_type",
            )
            .order_by("family_member_id", "-lease__start_date", "-id")
        )
        family_map = {}
        for relation in family_relations:
            family_map.setdefault(relation.family_member_id, relation)
        for tenant in page_tenants:
            tenant.current_family_relationship = family_map.get(tenant.pk)

        tenant_ids = [tenant.pk for tenant in page_tenants]
        role_targets = {
            tenant_id: {
                "family_member": [],
                "proposer": [],
                "seconder": [],
                "witness": [],
            }
            for tenant_id in tenant_ids
        }
        role_target_urls = {
            tenant_id: {key: set() for key in targets}
            for tenant_id, targets in role_targets.items()
        }

        def add_role_target(tenant_id, role_key, lease, renewal=None):
            if not tenant_id or tenant_id not in role_targets:
                return
            if renewal is None:
                detail_url = reverse("leases:lease_detail", args=[lease.pk])
                period_start, period_end = lease.start_date, lease.end_date
                destination_type = "Lease"
            else:
                detail_url = reverse(
                    "leases:lease_history_detail", args=[lease.pk, renewal.pk]
                )
                period_start, period_end = renewal.start_date, renewal.end_date
                destination_type = f"Renewal #{renewal.renewal_number}"
            if detail_url in role_target_urls[tenant_id][role_key]:
                return
            role_target_urls[tenant_id][role_key].add(detail_url)
            role_targets[tenant_id][role_key].append(
                {
                    "url": detail_url,
                    "label": (
                        f"{lease.unit.property.property_name}-"
                        f"{lease.unit.unit_number} · {destination_type} #{lease.pk}"
                        if renewal is None
                        else (
                            f"{lease.unit.property.property_name}-"
                            f"{lease.unit.unit_number} · {destination_type}"
                        )
                    ),
                    "period": f"{period_start:%Y-%m-%d} to {period_end:%Y-%m-%d}",
                }
            )

        for link in (
            LeaseFamilyMember.objects.filter(family_member_id__in=tenant_ids)
            .select_related("lease__unit__property")
            .order_by("lease__start_date", "lease_id")
        ):
            add_role_target(link.family_member_id, "family_member", link.lease)

        role_leases = (
            Lease.objects.filter(
                Q(proposer_id__in=tenant_ids)
                | Q(seconder_id__in=tenant_ids)
                | Q(witness1_tenant_id__in=tenant_ids)
                | Q(witness2_tenant_id__in=tenant_ids)
            )
            .select_related("unit__property")
            .order_by("start_date", "id")
        )
        for role_lease in role_leases:
            add_role_target(role_lease.proposer_id, "proposer", role_lease)
            add_role_target(role_lease.seconder_id, "seconder", role_lease)
            add_role_target(role_lease.witness1_tenant_id, "witness", role_lease)
            add_role_target(role_lease.witness2_tenant_id, "witness", role_lease)

        role_renewals = (
            LeaseRenewal.objects.filter(
                Q(witness1_tenant_id__in=tenant_ids)
                | Q(witness2_tenant_id__in=tenant_ids)
            )
            .select_related("lease__unit__property")
            .order_by("start_date", "id")
        )
        for renewal in role_renewals:
            add_role_target(
                renewal.witness1_tenant_id, "witness", renewal.lease, renewal
            )
            add_role_target(
                renewal.witness2_tenant_id, "witness", renewal.lease, renewal
            )

        role_labels = (
            ("family_member", "Family Member"),
            ("proposer", "Proposer"),
            ("seconder", "Seconder"),
            ("witness", "Witness"),
        )
        for tenant in page_tenants:
            tenant.role_counts = {
                key: len(role_targets[tenant.pk][key]) for key, _label in role_labels
            }
            tenant.role_badges = [
                {
                    "key": key,
                    "label": label,
                    "targets": role_targets[tenant.pk][key],
                    "count": tenant.role_counts[key],
                }
                for key, label in role_labels
                if role_targets[tenant.pk][key]
            ]
            tenant.role_total = sum(tenant.role_counts.values())

        # Add all tenants for the tenant dropdown
        context["all_tenants"] = Tenant.objects.only(
            "id", "first_name", "last_name"
        ).order_by("first_name", "last_name")

        # Add all properties for the property dropdown
        context["properties"] = Property.objects.only("id", "property_name").order_by(
            "property_name"
        )

        # Add all units (for when no property is selected)
        context["all_units"] = Unit.objects.only("id", "unit_number").order_by(
            "unit_number"
        )

        # Add filtered units (for when a property is selected)
        context["filtered_units"] = (
            Unit.objects.filter(property_id=property_id)
            .only("id", "unit_number")
            .order_by("unit_number")
            if property_id
            else []
        )

        context["current_tenant"] = self.request.GET.get("tenant")
        context["current_phone"] = self.request.GET.get("phone")
        context["current_property"] = property_id
        context["current_unit"] = self.request.GET.get("unit")
        context["show_inactive"] = bool(self.request.GET.get("show_inactive"))
        context["interest_types"] = (
            apps.get_model("tenants", "TenantInterestType")
            .objects.filter(is_active=True)
            .order_by("sort_order", "name")
        )
        context["current_interested_in"] = self.request.GET.getlist("interested_in")
        context["pre_registration_form"] = TenantPreRegistrationLinkForm()
        context["registration_link_days"] = TENANT_REGISTRATION_MAX_AGE // (
            60 * 60 * 24
        )
        context["public_registration_payload"] = (
            _tenant_list_public_registration_payload(self.request)
        )
        context["pending_registration_count"] = (
            TenantRegistrationSubmission.objects.filter(status="pending").count()
        )
        active_lease_exists = Lease.objects.filter(
            tenant_id=OuterRef("pk"), status="active"
        )

        today = timezone.now().date()
        notice_until = today + timezone.timedelta(days=60)

        notice_lease_exists = Lease.objects.filter(
            tenant_id=OuterRef("pk"),
            status="active",
            end_date__gte=today,
            end_date__lte=notice_until,
        )

        all_tenant_qs = Tenant.objects.all()

        context["total_tenant_count"] = all_tenant_qs.count()

        context["active_tenant_count"] = (
            all_tenant_qs.annotate(has_active_lease=Exists(active_lease_exists))
            .filter(has_active_lease=True)
            .count()
        )

        context["inactive_tenant_count"] = (
            all_tenant_qs.annotate(has_active_lease=Exists(active_lease_exists))
            .filter(has_active_lease=False)
            .count()
        )

        context["on_notice_count"] = (
            all_tenant_qs.annotate(on_notice=Exists(notice_lease_exists))
            .filter(on_notice=True)
            .count()
        )

        active_family_exists = LeaseFamilyMember.objects.filter(
            family_member_id=OuterRef("pk"),
            lease__status="active",
            lease__start_date__lte=today,
            lease__end_date__gte=today,
        )
        context["family_member_tenant_count"] = (
            all_tenant_qs.annotate(is_active_family_member=Exists(active_family_exists))
            .filter(is_active_family_member=True)
            .count()
        )

        context["potential_tenant_count"] = (
            all_tenant_qs.filter(interested_in__isnull=False).distinct().count()
        )
        new_registration_id = self.request.GET.get("new_registration")
        if new_registration_id:
            try:
                context["new_registration_payload"] = _registration_link_payload(
                    self.request,
                    Tenant.objects.get(pk=new_registration_id),
                )
            except Tenant.DoesNotExist:
                context["new_registration_payload"] = None

        return context

    def create_export(self, export_format):
        queryset = self.get_queryset()
        filename = f"tenants_{timezone.now().strftime('%Y%m%d_%H%M%S')}"

        try:
            if export_format == "csv":
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    f'attachment; filename="{filename}.csv"'
                )

                writer = csv.writer(response)
                writer.writerow(
                    [
                        "ID",
                        "First Name",
                        "Last Name",
                        "Phone",
                        "Property",
                        "Unit",
                        "Rent",
                        "Balance",
                    ]
                )

                for tenant in queryset:
                    lease = tenant.current_lease
                    writer.writerow(
                        [
                            tenant.id,
                            tenant.first_name,
                            tenant.last_name,
                            format_phone(tenant.phone),
                            lease.unit.property.property_name if lease else "",
                            lease.unit.unit_number if lease else "",
                            lease.get_total_payment if lease else "",
                            lease.get_balance if lease else "",
                        ]
                    )
                return response

            elif export_format == "xlsx":
                try:
                    tenants = list(queryset)
                    self._attach_lease_totals(tenants)

                    output = BytesIO()
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Tenants"

                    # ===== REPORT TITLE =====
                    last_column = 19
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
                    title_cell = ws.cell(row=1, column=1, value="TENANT LIST REPORT")
                    title_cell.font = Font(bold=True, color="FFFFFF", size=16)
                    title_cell.fill = PatternFill("solid", fgColor="1F4E78")
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[1].height = 28

                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
                    info_cell = ws.cell(
                        row=2,
                        column=1,
                        value=(
                            f"Generated: {timezone.localtime().strftime('%d-%b-%Y %I:%M %p')}"
                            f"    |    Total records: {len(tenants)}"
                        ),
                    )
                    info_cell.font = Font(italic=True, color="44546A", size=10)
                    info_cell.fill = PatternFill("solid", fgColor="D9EAF7")
                    info_cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[2].height = 20

                    # ===== HEADER FORMATTING =====
                    headers = [
                        "SN",
                        "First\nName",
                        "Last\nName",
                        "Phone",
                        "Email",
                        "Property\nName",
                        "Unit\nNo.",
                        "Lease\nEnd Date",
                        "Monthly\nRent",
                        "Current\nBalance",
                        "Family\nMembers",
                        "Gender",
                        "Emergency\nContact",
                        "Emergency\nPhone",
                        "Photo",
                        "CNIC\nFront",
                        "CNIC\nBack",
                        "Address\n(Full)",
                        "Notes\n(Detailed)",
                    ]

                    header_row = 4
                    thin_gray = Side(style="thin", color="A6A6A6")
                    cell_border = Border(
                        left=thin_gray,
                        right=thin_gray,
                        top=thin_gray,
                        bottom=thin_gray,
                    )

                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=header_row, column=col_num, value=header)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )
                        cell.font = Font(bold=True, color="FFFFFF", size=10)
                        cell.fill = PatternFill("solid", fgColor="4472C4")
                        cell.border = cell_border

                    # ===== COLUMN WIDTHS =====
                    col_widths = {
                        "A": 3,
                        "B": 12,
                        "C": 12,
                        "D": 12,
                        "E": 16,
                        "F": 15,
                        "G": 10,
                        "H": 12,
                        "I": 12,
                        "J": 15,
                        "K": 7,
                        "L": 8,
                        "M": 13,
                        "N": 12,
                        "O": 15,
                        "P": 22,
                        "Q": 22,
                        "R": 20,
                        "S": 20,
                    }
                    for col, width in col_widths.items():
                        ws.column_dimensions[col].width = width

                    # ===== ROW HEIGHTS =====
                    ws.row_dimensions[header_row].height = 32
                    data_row_height = 72  # Data rows (for images)

                    # ===== CELL STYLES =====
                    center_style = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    wrap_style = Alignment(
                        vertical="center", wrap_text=True, horizontal="left"
                    )

                    # ===== DATA POPULATION =====
                    for idx, tenant in enumerate(tenants, start=header_row + 1):
                        lease = tenant.current_lease
                        ws.row_dimensions[idx].height = data_row_height

                        # Basic data
                        data = [
                            idx - header_row,  # SN
                            tenant.first_name,
                            tenant.last_name,
                            format_phone(tenant.phone),
                            tenant.email,
                            lease.unit.property.property_name if lease else "",
                            lease.unit.unit_number if lease else "",
                            lease.end_date
                            if lease
                            else None,
                            lease.get_total_payment if lease else None,
                            lease.get_balance if lease else None,
                            tenant.number_of_family_member,
                            tenant.get_gender_display(),
                            tenant.emergency_contact_name,
                            format_phone(tenant.emergency_contact_phone),
                            "",  # Photo placeholder
                            "",  # CNIC Front placeholder
                            "",  # CNIC Back placeholder
                            tenant.address or "-",
                            tenant.notes or "-",
                        ]

                        # Apply styles to each cell
                        for col_num, value in enumerate(data, 1):
                            cell = ws.cell(row=idx, column=col_num, value=value)
                            cell.alignment = (
                                wrap_style if col_num in [18, 19] else center_style
                            )
                            cell.border = cell_border
                            if idx % 2 == 0:
                                cell.fill = PatternFill("solid", fgColor="F2F6FC")
                            if col_num in [9, 10]:
                                cell.number_format = '"Rs." #,##0.00;[Red]-"Rs." #,##0.00'
                            elif col_num == 8 and value:
                                cell.number_format = "dd-mmm-yyyy"

                        # ===== IMAGE HANDLING =====
                        image_data = [
                            ("O", tenant.photo, 90, 90),  # Photo
                            # CNIC Front (adjusted ratio)
                            ("P", tenant.cnic_front, 150, 90),
                            # CNIC Back (adjusted ratio)
                            ("Q", tenant.cnic_back, 150, 90),
                        ]

                        for col, image, width, height in image_data:
                            if image:
                                try:
                                    image_path = image.path
                                    if not image_path:
                                        continue
                                    img = Image(image_path)
                                    img.width = width
                                    img.height = height
                                    ws.add_image(img, f"{col}{idx}")
                                except (OSError, ValueError, TypeError):
                                    ws[f"{col}{idx}"] = "Image"
                                    ws[f"{col}{idx}"].alignment = center_style

                    # ===== FINAL TOUCHES =====
                    last_row = max(header_row, header_row + len(tenants))
                    ws.auto_filter.ref = f"A{header_row}:S{last_row}"
                    ws.freeze_panes = f"A{header_row + 1}"
                    ws.sheet_view.showGridLines = False
                    ws.page_setup.orientation = "landscape"
                    ws.page_setup.paperSize = ws.PAPERSIZE_A3
                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 0
                    ws.sheet_properties.pageSetUpPr.fitToPage = True
                    ws.print_title_rows = f"1:{header_row}"
                    ws.print_area = f"A1:S{last_row}"
                    ws.oddFooter.center.text = "Page &P of &N"
                    ws.oddFooter.right.text = "Tenant Management System"
                    wb.save(output)
                    output.seek(0)

                    response = HttpResponse(
                        output.getvalue(),
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    response["Content-Disposition"] = (
                        f'attachment; filename="Tenants_Export_{timezone.localdate().strftime("%Y%m%d")}.xlsx"'
                    )
                    return response

                except Exception as e:
                    logger.error(f"Excel export error: {str(e)}", exc_info=True)
                    messages.error(
                        self.request, "Failed to generate Excel file. Please try again."
                    )
                    return redirect("tenants:tenant_list")

            elif export_format == "pdf":
                # Get pagination data to match HTML view
                queryset = self.get_queryset()
                request = self.request

                # Build filter description
                filter_description = []

                # Property filter
                if property_id := request.GET.get("property"):
                    try:
                        property = Property.objects.get(id=property_id)
                        filter_description.append(f"Property: {property.property_name}")
                    except Property.DoesNotExist:
                        pass

                # Unit filter
                if unit_id := request.GET.get("unit"):
                    try:
                        unit = Unit.objects.get(id=unit_id)
                        filter_description.append(f"Unit: {unit.unit_number}")
                    except Unit.DoesNotExist:
                        pass

                # Tenant filter
                if tenant_id := request.GET.get("tenant"):
                    try:
                        tenant = Tenant.objects.get(id=tenant_id)
                        filter_description.append(
                            f"Tenant: {tenant.first_name} {tenant.last_name}"
                        )
                    except Tenant.DoesNotExist:
                        pass

                # Phone filter
                if phone := request.GET.get("phone"):
                    filter_description.append(f"Phone: {phone}")

                # Inactive filter
                show_inactive = request.GET.get("show_inactive") == "on"
                status_text = "Including Inactive" if show_inactive else "Active Only"
                filter_description.append(f"Status: {status_text}")

                # Combine all filters
                filter_text = (
                    " | ".join(filter_description)
                    if filter_description
                    else "All Tenants"
                )

                html_string = render_to_string(
                    "tenants/tenant_list_pdf.html",
                    {
                        "tenants": queryset,  # Full filtered list
                        "date": timezone.now().strftime("%Y-%m-%d"),
                        "filter_text": filter_text,  # Make sure this is passed
                    },
                )

                html = HTML(
                    string=html_string, base_url=self.request.build_absolute_uri()
                )
                pdf = html.write_pdf()

                response = HttpResponse(pdf, content_type="application/pdf")
                response["Content-Disposition"] = (
                    f'attachment; filename="{filename}.pdf"'
                )
                return response
        except Exception as e:
            logger.error(f"Export failed: {str(e)}")
            messages.error(self.request, f"Export failed: {str(e)}")
            return redirect("tenants:tenant_list")

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get("export")
        if export_format in self.export_formats:
            return self.create_export(export_format)
        return super().get(request, *args, **kwargs)


# tenants/views.py


@login_required
@require_POST
def tenant_status_toggle(request):
    try:
        data = json.loads(request.body.decode("utf-8") or "{}")
        tenant = get_object_or_404(Tenant, pk=data.get("id"))
        tenant.is_active = str(data.get("is_active")) in ["1", "true", "True"]
        tenant.save(update_fields=["is_active", "updated_at"])
        return JsonResponse({"success": True, "is_active": tenant.is_active})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


def tenant_ajax_update(request):
    if (
        request.method == "POST"
        and request.headers.get("x-requested-with") == "XMLHttpRequest"
    ):
        pk = request.POST.get("pk")
        name = request.POST.get("name")
        value = request.POST.get("value")

        if not all([pk, name, value]):
            return JsonResponse(
                {"status": "error", "message": "Missing required parameters"},
                status=400,
            )

        try:
            tenant = Tenant.objects.get(pk=pk)
            setattr(tenant, name, value)
            tenant.save()
            normalized_value = getattr(tenant, name)
            payload = {"status": "success"}
            if name == "cnic":
                payload["value"] = normalized_value
                payload["display_value"] = format_cnic(normalized_value)
            elif name in {
                "phone", "phone2", "phone3", "employer_phone",
                "reference_phone_1", "reference_phone_2",
                "emergency_contact_phone",
            }:
                payload["value"] = normalized_value
                payload["display_value"] = format_phone(normalized_value)
            return JsonResponse(payload)
        except Tenant.DoesNotExist:
            return JsonResponse(
                {"status": "error", "message": "Tenant not found"}, status=404
            )
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

    return JsonResponse({"status": "error", "message": "Invalid request"}, status=400)


@login_required
@require_POST
def tenant_lead_inline_update(request, pk):
    tenant = get_object_or_404(Tenant.objects.prefetch_related("interested_in"), pk=pk)
    try:
        data = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON."}, status=400)

    field = data.get("field")
    if field not in {"phone", "interested_in"}:
        return JsonResponse(
            {"success": False, "error": "Field is not editable."}, status=400
        )

    if field == "phone":
        tenant.phone = normalize_phone(data.get("value"))
        tenant.save(update_fields=["phone"])
    else:
        raw_ids = data.get("value") or []
        try:
            interest_ids = [int(value) for value in raw_ids if str(value).strip()]
        except (TypeError, ValueError):
            return JsonResponse(
                {"success": False, "error": "Invalid interest type."}, status=400
            )
        TenantInterestType = apps.get_model("tenants", "TenantInterestType")
        interests = TenantInterestType.objects.filter(
            pk__in=interest_ids, is_active=True
        )
        tenant.interested_in.set(interests)

    tenant = Tenant.objects.prefetch_related("interested_in").get(pk=tenant.pk)
    return JsonResponse(
        {
            "success": True,
            "tenant": {
                "id": tenant.pk,
                "phone": tenant.phone or "",
                "phone_display": format_phone(tenant.phone),
                "interested_in": [
                    {"id": item.pk, "name": item.name}
                    for item in tenant.interested_in.all()
                ],
            },
        }
    )
